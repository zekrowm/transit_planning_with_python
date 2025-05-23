"""
Script Name:
    bus_block_stop_sequence_printable.py

Purpose:
    Generates printable, stop-by-stop Excel schedules for each
    vehicle block defined in GTFS data. Merges trip, stop time,
    stop, and route information. Supports optional filtering by
    service ID and route short name. Primarily intended for field
    operations and audits.

Inputs:
    1. GTFS data files (specifically requires trips.txt,
       stop_times.txt, stops.txt, routes.txt, calendar.txt)
       located in the folder specified by GTFS_FOLDER_PATH.

Outputs:
    1. Individual Excel (.xlsx) files, one per vehicle block found
       after applying filters. Each file contains a formatted schedule
       with columns for field data entry (Actual Time, Boardings, etc.).
       Files are saved to the folder specified by BASE_OUTPUT_PATH.

Dependencies:
    pandas, openpyxl
"""

import math
import os

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_FOLDER_PATH = r"C:\Path\To\Your\Input\Folder"  # <<< EDIT HERE
BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output\Folder"  # <<< EDIT HERE

REQUIRED_GTFS_FILES = [
    "trips.txt",
    "stop_times.txt",
    "stops.txt",
    "routes.txt",
    "calendar.txt",
]

# If you only want certain service IDs or route short names, specify them here:
FILTER_SERVICE_IDS = []  # e.g. ['1', '2'] or [] to include all
FILTER_ROUTE_SHORT_NAMES = []  # e.g. ['101', '202'] or [] to include all

# Placeholder values for printing:
MISSING_TIME = "________"
MISSING_VALUE = "_____"

# Maximum column width for neat Excel formatting:
MAX_COLUMN_WIDTH = 35

# =============================================================================
# FUNCTIONS
# =============================================================================

# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(files=None, dtype=str):
    """
    Loads GTFS files into pandas DataFrames from a path defined externally
    (GTFS_FOLDER_PATH).

    Parameters:
        files (list[str], optional): GTFS filenames to load. Default is all
            standard GTFS files:
            [
                "agency.txt", "stops.txt", "routes.txt", "trips.txt",
                "stop_times.txt", "calendar.txt", "calendar_dates.txt",
                "fare_attributes.txt", "fare_rules.txt", "feed_info.txt",
                "frequencies.txt", "shapes.txt", "transfers.txt"
            ]
        dtype (str or dict, optional): Pandas dtype to use. Default is str.

    Returns:
        dict[str, pd.DataFrame]: Dictionary keyed by file name without extension.

    Raises:
        FileNotFoundError: If GTFS_FOLDER_PATH doesn't exist or if any required
            file is missing.
        ValueError: If a file is empty or there's a parsing error.
        Exception: For any unexpected error during loading.
    """

    # Check if GTFS_FOLDER_PATH exists
    if not os.path.exists(GTFS_FOLDER_PATH):
        raise FileNotFoundError(f"The directory '{GTFS_FOLDER_PATH}' does not exist.")

    # Default to all standard GTFS files if none were specified
    if files is None:
        files = [
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        ]

    # Check for missing files
    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(GTFS_FOLDER_PATH, file_name))
    ]
    if missing:
        raise FileNotFoundError(
            f"Missing GTFS files in '{GTFS_FOLDER_PATH}': {', '.join(missing)}"
        )

    # Load files into DataFrames
    data = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(GTFS_FOLDER_PATH, file_name)

        try:
            df = pd.read_csv(file_path, dtype=dtype)
            data[key] = df
            print(f"Loaded {file_name} ({len(df)} records).")

        except pd.errors.EmptyDataError:
            raise ValueError(f"File '{file_name}' is empty.")
        except pd.errors.ParserError as err:
            raise ValueError(f"Parser error in '{file_name}': {err}")
        except Exception as err:
            raise Exception(f"Error loading '{file_name}': {err}")

    return data


# -----------------------------------------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------------------------------------


def time_to_seconds(time_str):
    """
    Converts a 'HH:MM:SS' or 'HH:MM' string into total seconds.
    Handles hours >= 24 by rolling over (e.g., 25:10:00 -> 1:10:00).
    """
    if pd.isnull(time_str):
        return math.nan

    parts = time_str.strip().split(":")
    if len(parts) < 2:
        return math.nan

    try:
        hours = int(parts[0]) % 24  # Roll over hours >= 24
        minutes = int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
    except ValueError:
        return math.nan

    return hours * 3600 + minutes * 60 + seconds


def format_hhmm(total_seconds):
    """
    Given a time in total seconds, returns a 'HH:MM' string (24-hour).
    If invalid, returns an empty string.
    """
    if pd.isnull(total_seconds) or total_seconds < 0:
        return ""
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"


# -----------------------------------------------------------------------------
# OTHER FUNCTIONS
# -----------------------------------------------------------------------------


def export_to_excel(data_frame, output_file):
    """
    Exports a DataFrame to an Excel file and applies basic formatting:
    - Left alignment
    - Sets column widths
    - Text wrapping for headers
    """
    if data_frame.empty:
        print(f"No data to export to {output_file}")
        return

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Write DataFrame to Excel and access the worksheet object for formatting
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        data_frame.to_excel(writer, index=False, sheet_name="Schedule")
        worksheet = writer.sheets["Schedule"]

        # Adjust columns
        for col_i, col_name in enumerate(data_frame.columns, 1):
            col_letter = get_column_letter(col_i)

            # Header alignment and text wrap
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Data alignment
            for row_i in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_i}"]
                cell.alignment = Alignment(horizontal="left")

            # Set column width based on max content length, capped at MAX_COLUMN_WIDTH
            max_len = max(len(str(col_name)), 10)  # Minimum width
            for row_i in range(2, worksheet.max_row + 1):
                val = worksheet[f"{col_letter}{row_i}"].value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            worksheet.column_dimensions[col_letter].width = min(
                max_len + 2, MAX_COLUMN_WIDTH
            )

    print(f"Exported: {output_file}")


def filter_data(trips_df, stop_times_df, routes_df):
    """
    Merges and filters trips/routes, then filters stop_times accordingly.
    Returns updated trips_df, stop_times_df.
    """

    # Merge route_short_name into trips
    routes_subset = routes_df[["route_id", "route_short_name"]]
    trips_df = trips_df.merge(routes_subset, on="route_id", how="left")

    # Apply Route Filtering
    if FILTER_ROUTE_SHORT_NAMES:
        blocks_for_selected_routes = (
            trips_df[trips_df["route_short_name"].isin(FILTER_ROUTE_SHORT_NAMES)][
                "block_id"
            ]
            .dropna()
            .unique()
        )
        if len(blocks_for_selected_routes) == 0:
            print("No blocks found with the specified route short names.")
            return pd.DataFrame(), pd.DataFrame()

        trips_df = trips_df[trips_df["block_id"].isin(blocks_for_selected_routes)]

    # Apply Service ID Filtering
    if FILTER_SERVICE_IDS:
        trips_df = trips_df[trips_df["service_id"].isin(FILTER_SERVICE_IDS)]

    # Filter stop_times to only include relevant trips
    stop_times_df = stop_times_df[stop_times_df["trip_id"].isin(trips_df["trip_id"])]

    return trips_df, stop_times_df


def prepare_stop_times(trips_df, stop_times_df, stops_df):
    """
    Adds and formats columns in stop_times_df with scheduling and stop info.
    Returns the updated stop_times_df.
    """
    # If 'timepoint' does not exist, create a new column with 0.
    if "timepoint" not in stop_times_df.columns:
        stop_times_df["timepoint"] = 0
    else:
        # Convert to numeric, fill NaN with 0
        stop_times_df["timepoint"] = (
            pd.to_numeric(stop_times_df["timepoint"], errors="coerce")
            .fillna(0)
            .astype(int)
        )

    # Merge essential trip columns into stop_times
    needed_trip_cols = ["trip_id", "block_id", "route_short_name", "direction_id"]
    stop_times_df = stop_times_df.merge(
        trips_df[needed_trip_cols], on="trip_id", how="left"
    )

    # Convert arrival/departure times to seconds and format
    stop_times_df["arrival_seconds"] = stop_times_df["arrival_time"].apply(
        time_to_seconds
    )
    stop_times_df["departure_seconds"] = stop_times_df["departure_time"].apply(
        time_to_seconds
    )
    stop_times_df["scheduled_time_hhmm"] = stop_times_df["departure_seconds"].apply(
        format_hhmm
    )

    # Merge in stop names
    stop_name_map = stops_df.set_index("stop_id")["stop_name"].to_dict()
    stop_times_df["stop_name"] = (
        stop_times_df["stop_id"].map(stop_name_map).fillna("Unknown Stop")
    )

    # Sort by block, trip, and stop_sequence
    stop_times_df = stop_times_df.dropna(subset=["block_id"])
    stop_times_df["stop_sequence"] = pd.to_numeric(
        stop_times_df["stop_sequence"], errors="coerce"
    )
    stop_times_df = stop_times_df.dropna(subset=["stop_sequence"])
    stop_times_df.sort_values(["block_id", "trip_id", "stop_sequence"], inplace=True)

    return stop_times_df


def export_blocks(stop_times_df):
    """
    Groups rows by block_id and exports each block to a separate Excel file.
    """
    all_blocks = stop_times_df["block_id"].unique()
    print(f"Found {len(all_blocks)} blocks to export.\n")

    for block_id in all_blocks:
        block_subset = stop_times_df[stop_times_df["block_id"] == block_id].copy()
        if block_subset.empty:
            continue

        # For each trip_id within this block, find earliest departure
        first_departures = (
            block_subset.groupby("trip_id")["departure_seconds"]
            .min()
            .reset_index(name="trip_start_seconds")
        )
        first_departures["trip_start_hhmm"] = first_departures[
            "trip_start_seconds"
        ].apply(format_hhmm)
        block_subset = block_subset.merge(first_departures, on="trip_id", how="left")

        block_subset["Trip Start Time"] = block_subset["trip_start_hhmm"]

        # Select and rename columns for clarity
        out_cols = [
            "block_id",
            "route_short_name",
            "direction_id",
            "trip_id",
            "Trip Start Time",
            "stop_sequence",
            "timepoint",
            "stop_id",
            "stop_name",
            "scheduled_time_hhmm",
        ]
        final_df = block_subset[out_cols].copy()
        final_df.rename(
            columns={
                "block_id": "Block ID",
                "route_short_name": "Route",
                "direction_id": "Direction",
                "trip_id": "Trip ID",
                "stop_sequence": "Stop Sequence",
                "timepoint": "Timepoint",
                "stop_id": "Stop ID",
                "stop_name": "Stop Name",
                "scheduled_time_hhmm": "Scheduled Time",
            },
            inplace=True,
        )

        # Insert placeholders
        final_df["Actual Time"] = MISSING_TIME
        final_df["Boardings"] = MISSING_VALUE
        final_df["Alightings"] = MISSING_VALUE
        final_df["Comments"] = MISSING_VALUE

        # Reorder columns to place 'Timepoint' after 'Stop Sequence'
        final_df = final_df[
            [
                "Block ID",
                "Route",
                "Direction",
                "Trip ID",
                "Trip Start Time",
                "Stop Sequence",
                "Timepoint",
                "Stop ID",
                "Stop Name",
                "Scheduled Time",
                "Actual Time",
                "Boardings",
                "Alightings",
                "Comments",
            ]
        ]

        final_df.sort_values(
            by=["Trip Start Time", "Trip ID", "Stop Sequence"], inplace=True
        )

        filename = f"block_{block_id}_schedule_printable.xlsx"
        output_path = os.path.join(BASE_OUTPUT_PATH, filename)
        export_to_excel(final_df, output_path)


# =============================================================================
# MAIN
# =============================================================================


def main():
    """
    Main function to orchestrate loading, filtering, processing,
    and exporting of printable bus block schedules.
    """
    print("========================================================")
    print(" GTFS Block Schedule Printable Generator")
    print("========================================================")
    print(f"Input GTFS Folder: {BASE_INPUT_PATH}")
    print(f"Output Folder:     {BASE_OUTPUT_PATH}")
    if FILTER_ROUTE_SHORT_NAMES:
        print(f"Filtering for Routes: {FILTER_ROUTE_SHORT_NAMES}")
    if FILTER_SERVICE_IDS:
        print(f"Filtering for Service IDs: {FILTER_SERVICE_IDS}")

    try:
        # --- Load required GTFS data using the INLINED function ---
        # The dtype=str ensures IDs are read correctly without scientific notation etc.
        gtfs_data = load_gtfs_data(files=REQUIRED_GTFS_FILES, dtype=str)

        # --- Get DataFrames from the loaded dictionary ---
        trips_df = gtfs_data.get("trips")
        stop_times_df = gtfs_data.get("stop_times")
        stops_df = gtfs_data.get("stops")
        routes_df = gtfs_data.get("routes")

        # --- Validate that essential dataframes were loaded ---
        if (
            trips_df is None
            or stop_times_df is None
            or stops_df is None
            or routes_df is None
        ):
            # The loader should have raised an error before this, but double-check
            raise ValueError(
                "Failed to load one or more essential GTFS files (trips, stop_times, stops, routes)."
            )

        # --- Filter Data based on Configuration ---
        trips_df, stop_times_df = filter_data(trips_df, stop_times_df, routes_df)

        # Exit if filtering removed all data
        if trips_df.empty or stop_times_df.empty:
            print("\nNo data remains after filtering. No files will be generated.")
            return  # Exit gracefully

        # --- Prepare stop_times data for export ---
        prepared_stop_times = prepare_stop_times(trips_df, stop_times_df, stops_df)

        # Exit if preparation resulted in no data
        if prepared_stop_times.empty:
            print("\nNo data remains after preparation. No files will be generated.")
            return  # Exit gracefully

        # --- Export each block to a separate Excel file ---
        export_blocks(prepared_stop_times)

        print("\n========================================================")
        print(" Script finished successfully.")
        print("========================================================")

    # --- Error Handling ---
    except FileNotFoundError as fnf_err:
        print("\n-------------------- ERROR -------------------------")
        print(f" File Not Found Error: {fnf_err}")
        print(" Please ensure the BASE_INPUT_PATH is correct and")
        print(f" the following files exist inside it: {', '.join(REQUIRED_GTFS_FILES)}")
        print("----------------------------------------------------")
    except ValueError as val_err:
        print("\n-------------------- ERROR -------------------------")
        print(f" Data Error: {val_err}")
        print(" Please check the format and content of your GTFS files.")
        print("----------------------------------------------------")
    except TypeError as type_err:
        print("\n-------------------- ERROR -------------------------")
        print(f" Configuration Error: {type_err}")
        print(" Ensure REQUIRED_GTFS_FILES list is correctly defined.")
        print("----------------------------------------------------")
    except MemoryError as mem_err:
        print("\n-------------------- ERROR -------------------------")
        print(f" Memory Error: {mem_err}")
        print(" The script ran out of memory, likely due to a very large GTFS file.")
        print(" Consider running on a machine with more RAM or optimizing the script.")
        print("----------------------------------------------------")
    except Exception as err:
        print("\n-------------------- UNEXPECTED ERROR -------------------------")
        print(f" An unexpected error occurred: {err}")
        print(" Please review the error message and the script.")
        # Optional: Uncomment below for full error details during debugging
        # import traceback
        # traceback.print_exc()
        print("---------------------------------------------------------------")

    finally:
        # This block runs whether there was an error or not
        print("\nExiting script.")


if __name__ == "__main__":
    main()
