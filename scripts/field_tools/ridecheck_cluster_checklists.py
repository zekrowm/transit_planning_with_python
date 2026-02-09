"""Generate Excel field-audit checklists from a GTFS feed.

This module produces printable checklists for on-street audits by slicing GTFS
stop-times data by schedule type (e.g., weekday) and user-defined stop clusters
(e.g., station complexes). Each (cluster × schedule × time-window) combination
is exported as a formatted Excel workbook with placeholders for observed times,
bus numbers, and comments.

Workflow:
    1. Validate presence of required GTFS text files.
    2. Load trips, stop_times, routes, stops, and calendar.
    3. Resolve cluster definitions to canonical ``stop_id`` values
       (optionally via ``stop_code``).
    4. Optionally flag non-cluster stops within a configured distance of any
       cluster stop.
    5. For each schedule type:
         a. Select active service_ids based on calendar day flags.
         b. Build a joined table (stop_times → trips → routes → stops).
         c. Extract rows for each cluster and normalize times.
         d. Insert auditing placeholders and export Excel outputs.

Outputs:
    * One full-day workbook per (cluster × schedule).
    * Optional time-window-specific workbooks.
    * Optional CSV report of nearby non-cluster stops.

Assumptions:
    * GTFS feed is internally consistent.
    * Calendar columns are 0/1 flags.
    * Clusters are defined in this configuration block using either ``stop_id``
      or ``stop_code``.
"""

from __future__ import annotations

import logging
import math
import os
from collections.abc import Mapping, Sequence
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

# Output directory for generated Excel checklists and nearby-stop QA reports.
# Use a local or network path that field staff can access.
BASE_OUTPUT_PATH = r"R:\transit\field_checks\YYYY_MM_checklists"

# Input directory containing a complete GTFS feed (trips.txt, stop_times.txt,
# routes.txt, stops.txt, calendar.txt).
BASE_INPUT_PATH = r"R:\transit\gtfs\connector_YYYY_MM_DD"

# How CLUSTERS are specified:
# - "stop_id"  -> CLUSTERS lists are stop_id values from stops.txt
# - "stop_code" -> CLUSTERS lists are stop_code values; we map them to stop_id.
STOP_IDENTIFIER_FIELD = "stop_code"  # or "stop_id"

# Columns to force as strings (others infer automatically)
DTYPE_DICT: Dict[str, type] = {
    "stop_id": str,
    "trip_id": str,
    "route_id": str,
    "service_id": str,
    "stop_code": str,
}

# Required GTFS files
GTFS_FILES: List[str] = [
    "trips.txt",
    "stop_times.txt",
    "routes.txt",
    "stops.txt",
    "calendar.txt",
]

# Clusters keyed on STOP_IDENTIFIER_FIELD.
# If STOP_IDENTIFIER_FIELD == "stop_id", these must be stop_id values.
# If STOP_IDENTIFIER_FIELD == "stop_code", these must be stop_code values.
CLUSTERS: Dict[str, List[str]] = {
    "Sample Transit Center": [
        "1001",
        "1002",
        "1003",
    ],
    "Sample Park & Ride": [
        "2001",
        "2002",
    ],
    # "Another Hub": ["3001", "3002", "3003"],
}

# Schedule types and which calendar columns must be 1/True
SCHEDULE_TYPES: Dict[str, List[str]] = {
    "Weekday": ["monday", "tuesday", "wednesday", "thursday", "friday"],
    # "Saturday": ["saturday"],
    # "Sunday": ["sunday"],
}

# Time windows per schedule type
TIME_WINDOWS: Dict[str, Dict[str, Tuple[str, str]]] = {
    "Weekday": {
        "morning": ("06:00", "09:59"),
        "afternoon": ("14:00", "17:59"),
    },
    # "Saturday": { ... },
    # "Sunday": { ... },
}

# Route_short_name values to bold in the Excel output
SPECIAL_ROUTES: List[str] = [
    # "101",
    # "202",
]

# Nearby-stop QA buffer
NEARBY_STOP_BUFFER_FT: int = 500

# Internal constants
_FEET_TO_METERS = 0.3048
_EARTH_RADIUS_M = 6_371_000.0

# =============================================================================
# FILE / IO HELPERS
# =============================================================================


def validate_input_directory(base_input_path: str, gtfs_files: Iterable[str]) -> None:
    """Ensure base_input_path exists and contains all GTFS_FILES."""
    if not os.path.isdir(base_input_path):
        raise FileNotFoundError(f"Input directory does not exist: {base_input_path}")

    missing: List[str] = []
    for file_name in gtfs_files:
        path = os.path.join(base_input_path, file_name)
        if not os.path.isfile(path):
            missing.append(file_name)

    if missing:
        raise FileNotFoundError(
            f"Missing required GTFS file(s) in {base_input_path}: {', '.join(missing)}"
        )


def create_output_directory(base_output_path: str) -> None:
    """Create base_output_path if needed (idempotent)."""
    os.makedirs(base_output_path, exist_ok=True)


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_folder_path: Absolute or relative path to the folder
            containing the GTFS feed.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Folder missing or one of *files* not present.
        ValueError: Empty file or CSV parser failure.
        RuntimeError: Generic OS error while reading a file.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data: dict[str, pd.DataFrame] = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info("Loaded %s (%d records).", file_name, len(df))

        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

    return data





# =============================================================================
# IDENTIFIER / CLUSTER RESOLUTION
# =============================================================================


def _normalize_identifier_list(values: Iterable[str | int]) -> List[str]:
    """Return a list of str identifiers from a heterogeneous list."""
    return [str(v) for v in values]


def build_cluster_stop_ids(
    stops: pd.DataFrame,
    clusters: Dict[str, List[str]],
    identifier_field: str,
) -> Dict[str, List[str]]:
    """Resolve cluster definitions to canonical stop_id lists.

    Clusters may be keyed on stop_id or stop_code, but this function always returns
    a mapping of cluster_name -> list of stop_id strings that exist in `stops`.

    Args:
        stops: GTFS stops table with at least stop_id and optionally stop_code.
        clusters: Mapping of cluster_name -> identifiers (stop_id or stop_code).
        identifier_field: "stop_id" or "stop_code".

    Returns:
        Dictionary mapping cluster_name -> list of canonical stop_id values (as str).

    Raises:
        ValueError: If identifier_field is invalid or critical columns are missing.
    """
    if identifier_field not in {"stop_id", "stop_code"}:
        raise ValueError("identifier_field must be 'stop_id' or 'stop_code'.")

    if "stop_id" not in stops.columns:
        raise ValueError("stops dataframe must contain 'stop_id' column.")

    stops_local = stops.copy()
    stops_local["stop_id"] = stops_local["stop_id"].astype(str)

    if identifier_field == "stop_id":
        # Simple case: cluster identifiers are stop_id.
        cluster_to_ids: Dict[str, List[str]] = {}
        existing_stop_ids: Set[str] = set(stops_local["stop_id"].astype(str))

        for cname, ids in clusters.items():
            id_list = _normalize_identifier_list(ids)
            matched = [sid for sid in id_list if sid in existing_stop_ids]
            missing = [sid for sid in id_list if sid not in existing_stop_ids]
            if missing:
                logging.warning(
                    "Warning: cluster '%s' references stop_id(s) not found in stops.txt: %s",
                    cname,
                    missing,
                )
            if not matched:
                logging.warning(
                    "Warning: cluster '%s' has no valid stop_ids after "
                    "resolution; it will be skipped.",
                    cname,
                )
            cluster_to_ids[cname] = matched

        return cluster_to_ids

    # identifier_field == "stop_code"
    if "stop_code" not in stops_local.columns:
        raise ValueError("identifier_field is 'stop_code' but stops.txt has no 'stop_code' column.")

    stops_local["stop_code"] = stops_local["stop_code"].astype(str)

    # Map stop_code -> list of stop_id (in case of duplicates)
    code_to_ids: Dict[str, List[str]] = {}
    for _, row in stops_local[["stop_code", "stop_id"]].dropna().iterrows():
        code = row["stop_code"]
        sid = row["stop_id"]
        code_to_ids.setdefault(code, []).append(sid)

    cluster_to_ids = {}
    for cname, codes in clusters.items():
        codes_str = _normalize_identifier_list(codes)
        resolved_ids: List[str] = []
        for code in codes_str:
            if code not in code_to_ids:
                logging.warning(
                    "Warning: cluster '%s' references stop_code '%s' not found in stops.txt.",
                    cname,
                    code,
                )
                continue
            mapped_ids = sorted(set(code_to_ids[code]))
            if len(mapped_ids) > 1:
                logging.warning(
                    "Warning: stop_code '%s' for cluster '%s' maps to "
                    "multiple stop_ids %s; all will be included.",
                    code,
                    cname,
                    mapped_ids,
                )
            resolved_ids.extend(mapped_ids)

        resolved_ids = sorted(set(resolved_ids))
        if not resolved_ids:
            logging.warning(
                "Warning: cluster '%s' has no valid stop_ids after resolving "
                "stop_code; it will be skipped.",
                cname,
            )
        cluster_to_ids[cname] = resolved_ids

    return cluster_to_ids


# =============================================================================
# TIME HELPERS
# =============================================================================


def normalize_gtfs_time_to_hhmm(time_str: str) -> str:
    """Normalize GTFS HH:MM[:SS] (possibly ≥ 24h) to 'HH:MM' 0–23h.

    Args:
        time_str: GTFS time string (e.g. '27:15:00', '05:07', '06:30:30').

    Returns:
        Time represented as 'HH:MM' with hours modulo 24.

    Raises:
        ValueError: If time_str is not parseable.
    """
    if pd.isna(time_str):
        raise ValueError("Cannot normalize NaN time string.")

    parts = str(time_str).split(":")
    if len(parts) < 2:
        raise ValueError(f"Invalid time string: {time_str!r}")

    hours = int(parts[0])
    minutes = int(parts[1])

    hours = hours % 24
    return f"{hours:02d}:{minutes:02d}"


# =============================================================================
# NEARBY STOP QA
# =============================================================================


def _haversine_m(
    lat1: float,
    lon1: float,
    lat2: pd.Series,
    lon2: pd.Series,
) -> pd.Series:
    """Vectorized haversine distance from one point to many points (meters)."""
    lat1r = math.radians(lat1)
    lon1r = math.radians(lon1)

    lat2r = lat2.astype(float).map(math.radians)
    lon2r = lon2.astype(float).map(math.radians)

    dlat = lat2r - lat1r
    dlon = lon2r - lon1r

    a = (dlat / 2.0).map(math.sin) ** 2 + math.cos(lat1r) * (lat2r.map(math.cos)) * (
        (dlon / 2.0).map(math.sin) ** 2
    )

    c = 2 * a.map(math.sqrt).map(lambda x: math.asin(min(1.0, x)))
    return _EARTH_RADIUS_M * c


def find_nearby_stops_for_clusters(
    stops: pd.DataFrame,
    cluster_stop_ids: Dict[str, List[str]],
    buffer_ft: int,
    verbose: bool = True,
) -> pd.DataFrame:
    """Find stops within buffer_ft of any cluster stop but not in that cluster.

    Args:
        stops: GTFS stops table with at least stop_id, stop_name, stop_lat, stop_lon.
        cluster_stop_ids: Mapping of cluster_name -> list of stop_id (canonical).
        buffer_ft: Distance threshold in feet.
        verbose: Whether to print summary warnings.

    Returns:
        DataFrame with columns:
            ['cluster', 'anchor_stop_id', 'anchor_stop_name',
             'nearby_stop_id', 'nearby_stop_name', 'distance_ft']
    """
    required = {"stop_id", "stop_name", "stop_lat", "stop_lon"}
    missing = required.difference(stops.columns)
    if missing:
        raise ValueError(f"stops is missing columns: {sorted(missing)}")

    stops_local = stops.copy()
    stops_local["stop_id"] = stops_local["stop_id"].astype(str)

    buffer_m = buffer_ft * _FEET_TO_METERS
    records: List[Dict[str, object]] = []

    for cname, id_list in cluster_stop_ids.items():
        if not id_list:
            if verbose:
                logging.info("Skipping nearby-stop QA for cluster '%s' (no stop_ids).", cname)
            continue

        anchor_ids: Set[str] = set(id_list)
        anchors = stops_local[stops_local["stop_id"].isin(anchor_ids)]
        others = stops_local[~stops_local["stop_id"].isin(anchor_ids)]

        if anchors.empty:
            if verbose:
                logging.warning(
                    "Warning: no anchor stops found for cluster '%s' in stops.txt.", cname
                )
            continue

        for _, arow in anchors.iterrows():
            dists_m = _haversine_m(
                float(arow["stop_lat"]),
                float(arow["stop_lon"]),
                others["stop_lat"],
                others["stop_lon"],
            )
            mask = dists_m <= buffer_m
            if not mask.any():
                continue

            near = others.loc[mask, ["stop_id", "stop_name"]].copy()
            near["distance_ft"] = (dists_m.loc[mask] / _FEET_TO_METERS).round(1)

            for _, nrow in near.iterrows():
                records.append(
                    {
                        "cluster": cname,
                        "anchor_stop_id": arow["stop_id"],
                        "anchor_stop_name": arow["stop_name"],
                        "nearby_stop_id": nrow["stop_id"],
                        "nearby_stop_name": nrow["stop_name"],
                        "distance_ft": float(nrow["distance_ft"]),
                    }
                )

        if verbose:
            cnt = sum(1 for r in records if r["cluster"] == cname)
            if cnt > 0:
                logging.warning(
                    "Warning: cluster '%s' has %d nearby stop(s) within "
                    "%d ft that are not in the cluster.",
                    cname,
                    cnt,
                    buffer_ft,
                )

    result = pd.DataFrame.from_records(
        records,
        columns=[
            "cluster",
            "anchor_stop_id",
            "anchor_stop_name",
            "nearby_stop_id",
            "nearby_stop_name",
            "distance_ft",
        ],
    )

    if verbose and not result.empty and os.path.isdir(BASE_OUTPUT_PATH):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_csv = os.path.join(BASE_OUTPUT_PATH, f"nearby_stop_warnings_{ts}.csv")
        try:
            result.sort_values(["cluster", "anchor_stop_id", "distance_ft"]).to_csv(
                out_csv, index=False
            )
            logging.info("Nearby stop report written to %s", out_csv)
        except PermissionError:
            logging.warning("Warning: could not write nearby stop report (permission denied).")

    return result


# =============================================================================
# EXCEL EXPORT
# =============================================================================


def export_to_excel(df: pd.DataFrame, output_file: str) -> None:
    """Write DataFrame to Excel with basic formatting and route highlighting."""
    bold_font = Font(bold=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]

        # Header left-aligned
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="left")

        # Bold rows for special routes
        if "route_short_name" in df.columns and SPECIAL_ROUTES:
            route_idx = int(df.columns.get_loc("route_short_name")) + 1
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=route_idx).value
                if str(val) in SPECIAL_ROUTES:
                    for c in ws[row_idx]:
                        c.font = bold_font

        # Auto-fit column widths
        for idx, col_name in enumerate(df.columns, start=1):
            max_len_data = 0
            if not df[col_name].empty:
                max_len_data = df[col_name].astype(str).map(len).max()
            max_len = max(max_len_data, len(str(col_name))) + 2
            ws.column_dimensions[get_column_letter(idx)].width = max_len


def prepend_sample_row(df: pd.DataFrame, cluster_name: str, schedule_name: str) -> pd.DataFrame:
    """Prepend a sample row 5 min before the first trip, if possible."""
    if df.empty:
        logging.warning(
            "Warning: empty dataframe for %s (%s); cannot prepend sample row.",
            cluster_name,
            schedule_name,
        )
        return df

    if "arrival_time" not in df.columns or "departure_time" not in df.columns:
        logging.warning(
            "Warning: dataframe for %s (%s) missing "
            "'arrival_time' or 'departure_time'; cannot prepend sample row.",
            cluster_name,
            schedule_name,
        )
        return df

    sample = {col: "" for col in df.columns}
    try:
        first_arr_str = df["arrival_time"].iloc[0]
        first_dep_str = df["departure_time"].iloc[0]

        first_arr = pd.to_datetime(first_arr_str, format="%H:%M")
        first_dep = pd.to_datetime(first_dep_str, format="%H:%M")

        sample_arr = (first_arr - pd.Timedelta(minutes=5)).strftime("%H:%M")
        sample_dep = (first_dep - pd.Timedelta(minutes=5)).strftime("%H:%M")

        updates = {
            "route_short_name": "SAMPLE",
            "trip_headsign": "Sample Trip",
            "arrival_time": sample_arr,
            "act_arrival": sample_arr,
            "departure_time": sample_dep,
            "act_departure": sample_dep,
            "comments": "Please use 24-hour HH:MM format",
        }
        for k, v in updates.items():
            if k in sample:
                sample[k] = v

        return pd.concat([pd.DataFrame([sample]), df], ignore_index=True)
    except Exception as exc:  # noqa: BLE001
        logging.error(
            "Error creating sample row for %s (%s): %s. Proceeding without sample row.",
            cluster_name,
            schedule_name,
            exc,
        )
        return df


def process_cluster_slice(
    cluster_data: pd.DataFrame,
    cluster_name: str,
    schedule_name: str,
    base_output_path: str,
    time_windows: Optional[Dict[str, Dict[str, Tuple[str, str]]]] = None,
) -> None:
    """Transform one cluster×schedule slice and export Excel (full + windows)."""
    if cluster_data.empty:
        logging.info("No data for %s (%s); skipping.", cluster_name, schedule_name)
        return

    df = cluster_data.copy()

    # Normalize times
    for col in ["arrival_time", "departure_time"]:
        if col in df.columns:
            df[col] = df[col].astype(str).map(normalize_gtfs_time_to_hhmm)

    # Sort by arrival_time if possible
    try:
        df["arrival_sort"] = pd.to_datetime(df["arrival_time"], format="%H:%M")
        df = df.sort_values("arrival_sort").drop(columns=["arrival_sort"])
    except Exception as exc:  # noqa: BLE001
        logging.warning(
            "Warning: could not sort by arrival_time for %s (%s): %s",
            cluster_name,
            schedule_name,
            exc,
        )

    # Placeholders
    df.insert(int(df.columns.get_loc("arrival_time")) + 1, "act_arrival", "________")
    df.insert(int(df.columns.get_loc("departure_time")) + 1, "act_departure", "________")

    if "sequence_long" in df.columns:
        df.loc[df["sequence_long"] == "start", "act_arrival"] = "__XXXX__"
        df.loc[df["sequence_long"] == "last", "act_departure"] = "__XXXX__"

    df["bus_number"] = "________"
    df["comments"] = "________________"

    # Column ordering
    desired_first = [
        "route_short_name",
        "trip_headsign",
        "stop_sequence",
        "sequence_long",
        "stop_id",
        "stop_name",
        "arrival_time",
        "act_arrival",
        "departure_time",
        "act_departure",
        "block_id",
        "bus_number",
        "comments",
    ]
    existing_first = [c for c in desired_first if c in df.columns]
    missing_first = [c for c in desired_first if c not in df.columns]
    if missing_first:
        logging.warning(
            "Warning: expected columns missing for %s (%s): %s",
            cluster_name,
            schedule_name,
            missing_first,
        )

    other_cols = [c for c in df.columns if c not in existing_first]
    df = df[existing_first + other_cols]

    # Drop internal / noisy GTFS columns
    to_drop = [
        "shape_dist_traveled",
        "shape_id",
        "route_id",
        "service_id",
        "trip_id",
        "timepoint",
        "direction_id",
        "stop_headsign",
        "pickup_type",
        "drop_off_type",
        "wheelchair_accessible",
        "bikes_allowed",
        "trip_short_name",
        "stop_code",
    ]
    df = df.drop(columns=[c for c in to_drop if c in df.columns], errors="ignore")

    if not os.path.isdir(base_output_path):
        logging.error(
            "Error: output directory %s does not exist; cannot write Excel.", base_output_path
        )
        return

    # Full-day export
    full_path = os.path.join(base_output_path, f"{cluster_name}_{schedule_name}_data.xlsx")
    df_full = prepend_sample_row(df.copy(), cluster_name, schedule_name)
    export_to_excel(df_full, full_path)
    logging.info("Exported %s (%s) full-day to %s", cluster_name, schedule_name, full_path)

    # Time-window exports
    if time_windows and schedule_name in time_windows:
        for win_name, (start_s, end_s) in time_windows[schedule_name].items():
            try:
                st = pd.to_datetime(start_s, format="%H:%M").time()
                et = pd.to_datetime(end_s, format="%H:%M").time()
                atimes = pd.to_datetime(df["arrival_time"], format="%H:%M", errors="coerce").dt.time

                mask = pd.notnull(atimes) & (atimes >= st) & (atimes <= et)
                subset = df.loc[mask].copy()

                if subset.empty:
                    logging.info("  No %s data for %s (%s).", win_name, cluster_name, schedule_name)
                    continue

                path_win = os.path.join(
                    base_output_path,
                    f"{cluster_name}_{schedule_name}_{win_name}_data.xlsx",
                )
                subset_with_sample = prepend_sample_row(
                    subset, cluster_name, f"{schedule_name} {win_name}"
                )
                export_to_excel(subset_with_sample, path_win)
                logging.info(
                    "  Exported %s for %s (%s) to %s",
                    win_name,
                    cluster_name,
                    schedule_name,
                    path_win,
                )
            except Exception as exc:  # noqa: BLE001
                logging.error(
                    "  Error processing window %s for %s (%s): %s",
                    win_name,
                    cluster_name,
                    schedule_name,
                    exc,
                )


# =============================================================================
# MAIN ORCHESTRATION
# =============================================================================


def generate_gtfs_checklists() -> None:
    """End-to-end orchestration for GTFS field checklist generation."""
    # 1. Basic validation and load
    validate_input_directory(BASE_INPUT_PATH, GTFS_FILES)
    create_output_directory(BASE_OUTPUT_PATH)

    gtfs_data = load_gtfs_data(BASE_INPUT_PATH, files=GTFS_FILES, dtype=DTYPE_DICT)
    trips = gtfs_data["trips"]
    stop_times = gtfs_data["stop_times"]
    routes = gtfs_data["routes"]
    stops = gtfs_data["stops"]
    calendar = gtfs_data["calendar"]

    # Normalize ids as strings where relevant
    for df in (trips, stop_times, routes, stops, calendar):
        if "route_id" in df.columns:
            df["route_id"] = df["route_id"].astype(str)
        if "trip_id" in df.columns:
            df["trip_id"] = df["trip_id"].astype(str)
        if "stop_id" in df.columns:
            df["stop_id"] = df["stop_id"].astype(str)
        if "service_id" in df.columns:
            df["service_id"] = df["service_id"].astype(str)

    # 2. Resolve cluster definitions to canonical stop_ids
    cluster_stop_ids = build_cluster_stop_ids(
        stops=stops,
        clusters=CLUSTERS,
        identifier_field=STOP_IDENTIFIER_FIELD,
    )

    # Drop clusters that ended up with no valid stop_ids
    cluster_stop_ids = {cname: ids for cname, ids in cluster_stop_ids.items() if ids}

    if not cluster_stop_ids:
        logging.info("No valid clusters after resolving identifiers; nothing to process.")
        return

    # 3. Nearby-stop QA (optional but helpful)
    _nearby_df = find_nearby_stops_for_clusters(
        stops=stops,
        cluster_stop_ids=cluster_stop_ids,
        buffer_ft=NEARBY_STOP_BUFFER_FT,
        verbose=True,
    )
    if _nearby_df.empty:
        logging.info(
            "No nearby non-cluster stops found within %d ft of any cluster.",
            NEARBY_STOP_BUFFER_FT,
        )

    # 4. Process each schedule
    for schedule_name, days in SCHEDULE_TYPES.items():
        logging.info("\nProcessing schedule: %s", schedule_name)

        # Filter calendar to services that run on all requested days
        for day_col in days:
            if day_col not in calendar.columns:
                raise ValueError(
                    f"calendar.txt missing required column '{day_col}' "
                    f"for schedule '{schedule_name}'."
                )

        service_mask = calendar[days].astype(bool).all(axis=1)
        relevant_services = calendar.loc[service_mask, "service_id"].astype(str)

        if relevant_services.empty:
            logging.info(
                "No service_ids active for schedule '%s'; skipping schedule.", schedule_name
            )
            continue

        trips_filtered = trips[trips["service_id"].isin(relevant_services)]
        if trips_filtered.empty:
            logging.info(
                "No trips for schedule '%s' after service_id filter; skipping schedule.",
                schedule_name,
            )
            continue

        # Build master joined table for this schedule
        merged = (
            stop_times.merge(
                trips_filtered,
                on="trip_id",
                how="inner",
            )
            .merge(
                routes[["route_id", "route_short_name"]],
                on="route_id",
                how="left",
            )
            .merge(
                stops[["stop_id", "stop_name", "stop_lat", "stop_lon"]],
                on="stop_id",
                how="left",
            )
        )

        # Mark start/last in sequence_long
        merged["stop_sequence"] = merged["stop_sequence"].astype(int)
        merged["sequence_long"] = "middle"
        merged.loc[merged["stop_sequence"] == 1, "sequence_long"] = "start"
        max_seq = merged.groupby("trip_id")["stop_sequence"].transform("max")
        merged.loc[merged["stop_sequence"] == max_seq, "sequence_long"] = "last"

        # 5. Cluster loops
        for cname, id_list in cluster_stop_ids.items():
            logging.info("  Processing cluster '%s' for %s", cname, schedule_name)

            cluster_slice = merged[merged["stop_id"].isin(id_list)]
            if cluster_slice.empty:
                logging.info(
                    "  No data found for cluster '%s' on schedule '%s'; skipping.",
                    cname,
                    schedule_name,
                )
                continue

            process_cluster_slice(
                cluster_data=cluster_slice,
                cluster_name=cname,
                schedule_name=schedule_name,
                base_output_path=BASE_OUTPUT_PATH,
                time_windows=TIME_WINDOWS,
            )

    logging.info("\nAll clusters and schedules processed.")


def main() -> None:
    """Script entry point."""
    logging.basicConfig(level=logging.INFO)
    generate_gtfs_checklists()


if __name__ == "__main__":
    main()
