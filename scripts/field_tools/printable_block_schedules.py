"""Generate printable Excel schedules for each vehicle block in a GTFS feed.

The script reads the five core GTFS tables—``trips``, ``stop_times``, ``stops``,
``routes`` and ``calendar``—and optionally filters them by *service ID* and/or
*route short name*.  For every vehicle ``block_id`` that survives filtering it
produces a nicely-formatted ``.xlsx`` file ready for field auditing.

Typical usage
-------------
Run from the command line, an ArcGIS Pro Python toolbox, or a notebook.

Key Features
------------
- Loads GTFS text files into ``pandas`` DataFrames with robust error handling.
- Converts ``HH:MM(:SS)`` time strings to seconds (and back) safely.
- Applies ergonomic Excel formatting via ``openpyxl`` (column widths, wrapping).
- Inserts placeholders for handwritten field notes (actual time, boardings, etc.).
"""

from __future__ import annotations

import logging
import math
import os
from collections.abc import Mapping, Sequence
from typing import Any, Optional, Union

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_FOLDER_PATH = r"C:\Path\To\Your\Input\Folder"  # <<< EDIT HERE
BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output\Folder"  # <<< EDIT HERE

REQUIRED_GTFS_FILES = [
    "trips.txt",
    "stop_times.txt",
    "stops.txt",
    "routes.txt",
    "calendar.txt",
]

# If you only want certain service IDs or route short names, specify them here:
FILTER_SERVICE_IDS: list[str] = []  # e.g. ["WKD", "SAT"]
FILTER_ROUTE_SHORT_NAMES: list[str] = []  # e.g. ["101", "202"]

# Placeholder values for printing:
MISSING_TIME = "________"
MISSING_VALUE = "_____"

# Maximum column width for neat Excel formatting:
MAX_COLUMN_WIDTH = 35

# =============================================================================
# FUNCTIONS
# =============================================================================


def time_to_seconds(time_str: str) -> Union[float, int]:
    """Convert a ``HH:MM`` or ``HH:MM:SS`` string to total seconds.

    Args:
        time_str: Time string *or* ``NaN``; may exceed 24 h (e.g.,
            ``'25:10:00'`` → 1:10 a.m. next day).

    Returns:
        Non-negative number of seconds, or :pydata:`math.nan` on failure.
    """
    if pd.isnull(time_str):
        return math.nan

    parts = time_str.strip().split(":")
    if len(parts) < 2:
        return math.nan

    try:
        hours = int(parts[0]) % 24  # Roll over hours >= 24
        minutes = int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
    except ValueError:
        return math.nan

    return hours * 3600 + minutes * 60 + seconds


def format_hhmm(total_seconds: Union[int, float]) -> str:
    """Render seconds since midnight as a ``HH:MM`` string.

    Args:
        total_seconds: Seconds since 00:00.  Negative or ``NaN`` returns
            an empty string.

    Returns:
        Two-digit hour and minute representation (24-hour clock).
    """
    if pd.isnull(total_seconds) or total_seconds < 0:
        return ""
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"


# -----------------------------------------------------------------------------
# OTHER FUNCTIONS
# -----------------------------------------------------------------------------


def export_to_excel(data_frame: pd.DataFrame, output_file: str) -> None:
    """Write *data_frame* to an Excel file with basic styling.

    The sheet is named **Schedule** and receives:

    * Left-aligned cells.
    * Word-wrapped headers.
    * Column widths sized to longest cell (capped by ``MAX_COLUMN_WIDTH``).

    Args:
        data_frame: Tidy table to export; must be non-empty.
        output_file: Full path of the ``.xlsx`` file to create.

    Notes:
        ``os.makedirs`` is called with *exist_ok=True* so nested output
        folders are created automatically.
    """
    if data_frame.empty:
        logging.info("No data to export to %s", output_file)
        return

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Write DataFrame to Excel and access the worksheet object for formatting
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        data_frame.to_excel(writer, index=False, sheet_name="Schedule")
        worksheet = writer.sheets["Schedule"]

        # Adjust columns
        for col_i, col_name in enumerate(data_frame.columns, 1):
            col_letter = get_column_letter(col_i)

            # Header alignment and text wrap
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Data alignment
            for row_i in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_i}"]
                cell.alignment = Alignment(horizontal="left")

            # Set column width based on max content length, capped at MAX_COLUMN_WIDTH
            max_len = max(len(str(col_name)), 10)  # Minimum width
            for row_i in range(2, worksheet.max_row + 1):
                val = worksheet[f"{col_letter}{row_i}"].value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            worksheet.column_dimensions[col_letter].width = min(max_len + 2, MAX_COLUMN_WIDTH)

    logging.info("Exported: %s", output_file)


def filter_data(
    trips_df: pd.DataFrame, stop_times_df: pd.DataFrame, routes_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Apply route and service filters, propagating them to stop times.

    Args:
        trips_df: Parsed **trips.txt** table.
        stop_times_df: Parsed **stop_times.txt** table.
        routes_df: Parsed **routes.txt** table (for ``route_short_name``).

    Returns:
        ``(filtered_trips, filtered_stop_times)``.

    Raises:
        KeyError: If required columns are missing.

    Warning:
        If the global constants ``FILTER_ROUTE_SHORT_NAMES`` or
        ``FILTER_SERVICE_IDS`` remove every trip, the function returns
        two **empty** DataFrames.
    """
    # Merge route_short_name into trips
    routes_subset = routes_df[["route_id", "route_short_name"]]
    trips_df = trips_df.merge(routes_subset, on="route_id", how="left")

    # Apply Route Filtering
    if FILTER_ROUTE_SHORT_NAMES:
        blocks_for_selected_routes = (
            trips_df[trips_df["route_short_name"].isin(FILTER_ROUTE_SHORT_NAMES)]["block_id"]
            .dropna()
            .unique()
        )
        if len(blocks_for_selected_routes) == 0:
            logging.info("No blocks found with the specified route short names.")
            return pd.DataFrame(), pd.DataFrame()

        trips_df = trips_df[trips_df["block_id"].isin(blocks_for_selected_routes)]

    # Apply Service ID Filtering
    if FILTER_SERVICE_IDS:
        trips_df = trips_df[trips_df["service_id"].isin(FILTER_SERVICE_IDS)]

    # Filter stop_times to only include relevant trips
    stop_times_df = stop_times_df[stop_times_df["trip_id"].isin(trips_df["trip_id"])]

    return trips_df, stop_times_df


def prepare_stop_times(
    trips_df: pd.DataFrame, stop_times_df: pd.DataFrame, stops_df: pd.DataFrame
) -> pd.DataFrame:
    """Enrich and tidy ``stop_times`` for Excel export.

    Steps
    -----
    1. Ensure a numeric ``timepoint`` column (create if absent).
    2. Attach ``block_id``, ``route_short_name`` and ``direction_id``.
    3. Convert arrival/departure times → seconds → ``HH:MM`` format.
    4. Map ``stop_id`` → human-readable stop names.
    5. Sort by ``block_id``, ``trip_id``, ``stop_sequence``.

    Args:
        trips_df: Output of :pyfunc:`filter_data`.
        stop_times_df: Ditto.
        stops_df: Parsed **stops.txt** table.

    Returns:
        Cleaned ``stop_times`` DataFrame ready for grouping by block.
    """
    # If 'timepoint' does not exist, create a new column with 0.
    if "timepoint" not in stop_times_df.columns:
        stop_times_df["timepoint"] = 0
    else:
        # Convert to numeric, fill NaN with 0
        stop_times_df["timepoint"] = (
            pd.to_numeric(stop_times_df["timepoint"], errors="coerce").fillna(0).astype(int)
        )

    # Merge essential trip columns into stop_times
    needed_trip_cols = ["trip_id", "block_id", "route_short_name", "direction_id"]
    stop_times_df = stop_times_df.merge(trips_df[needed_trip_cols], on="trip_id", how="left")

    # Convert arrival/departure times to seconds and format
    stop_times_df["arrival_seconds"] = stop_times_df["arrival_time"].apply(time_to_seconds)
    stop_times_df["departure_seconds"] = stop_times_df["departure_time"].apply(time_to_seconds)
    stop_times_df["scheduled_time_hhmm"] = stop_times_df["departure_seconds"].apply(format_hhmm)

    # Merge in stop names
    stop_name_map = stops_df.set_index("stop_id")["stop_name"].to_dict()
    stop_times_df["stop_name"] = stop_times_df["stop_id"].map(stop_name_map).fillna("Unknown Stop")

    # Sort by block, trip, and stop_sequence
    stop_times_df = stop_times_df.dropna(subset=["block_id"])
    stop_times_df["stop_sequence"] = pd.to_numeric(stop_times_df["stop_sequence"], errors="coerce")
    stop_times_df = stop_times_df.dropna(subset=["stop_sequence"])
    stop_times_df.sort_values(["block_id", "trip_id", "stop_sequence"], inplace=True)

    return stop_times_df


def export_blocks(stop_times_df: pd.DataFrame) -> None:
    """Generate one Excel schedule per vehicle block.

    Args:
        stop_times_df: Prepared stop times (see
            :pyfunc:`prepare_stop_times`).  Must include the columns
            produced earlier (``block_id``, ``scheduled_time_hhmm``,
            etc.).

    Side Effects:
        Writes ``block_<id>_schedule_printable.xlsx`` to
        ``BASE_OUTPUT_PATH``; creates the folder tree if needed.
    """
    all_blocks = stop_times_df["block_id"].unique()
    logging.info("Found %d blocks to export.\n", len(all_blocks))

    for block_id in all_blocks:
        block_subset = stop_times_df[stop_times_df["block_id"] == block_id].copy()
        if block_subset.empty:
            continue

        # For each trip_id within this block, find earliest departure
        first_departures = (
            block_subset.groupby("trip_id")["departure_seconds"]
            .min()
            .reset_index(name="trip_start_seconds")
        )
        first_departures["trip_start_hhmm"] = first_departures["trip_start_seconds"].apply(
            format_hhmm
        )
        block_subset = block_subset.merge(first_departures, on="trip_id", how="left")

        block_subset["Trip Start Time"] = block_subset["trip_start_hhmm"]

        # Select and rename columns for clarity
        out_cols = [
            "block_id",
            "route_short_name",
            "direction_id",
            "trip_id",
            "Trip Start Time",
            "stop_sequence",
            "timepoint",
            "stop_id",
            "stop_name",
            "scheduled_time_hhmm",
        ]
        final_df = block_subset[out_cols].copy()
        final_df.rename(
            columns={
                "block_id": "Block ID",
                "route_short_name": "Route",
                "direction_id": "Direction",
                "trip_id": "Trip ID",
                "stop_sequence": "Stop Sequence",
                "timepoint": "Timepoint",
                "stop_id": "Stop ID",
                "stop_name": "Stop Name",
                "scheduled_time_hhmm": "Scheduled Time",
            },
            inplace=True,
        )

        # Insert placeholders
        final_df["Actual Time"] = MISSING_TIME
        final_df["Boardings"] = MISSING_VALUE
        final_df["Alightings"] = MISSING_VALUE
        final_df["Comments"] = MISSING_VALUE

        # Reorder columns to place 'Timepoint' after 'Stop Sequence'
        final_df = final_df[
            [
                "Block ID",
                "Route",
                "Direction",
                "Trip ID",
                "Trip Start Time",
                "Stop Sequence",
                "Timepoint",
                "Stop ID",
                "Stop Name",
                "Scheduled Time",
                "Actual Time",
                "Boardings",
                "Alightings",
                "Comments",
            ]
        ]

        final_df.sort_values(by=["Trip Start Time", "Trip ID", "Stop Sequence"], inplace=True)

        filename = f"block_{block_id}_schedule_printable.xlsx"
        output_path = os.path.join(BASE_OUTPUT_PATH, filename)
        export_to_excel(final_df, output_path)


# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_folder_path: Absolute or relative path to the folder
            containing the GTFS feed.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Folder missing or one of *files* not present.
        ValueError: Empty file or CSV parser failure.
        RuntimeError: Generic OS error while reading a file.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data: dict[str, pd.DataFrame] = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info("Loaded %s (%d records).", file_name, len(df))

        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

    return data


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """Command-line entry point.

    Orchestrates:

    * Logging configuration.
    * Data ingestion via :pyfunc:`load_gtfs_data`.
    * Optional filtering (:pyfunc:`filter_data`).
    * Data preparation (:pyfunc:`prepare_stop_times`).
    * Per-block Excel export (:pyfunc:`export_blocks`).

    The function traps anticipated exceptions and logs them with useful
    context before exiting with a non-zero status.
    """
    # --------------------------------------------------------------
    # Configure logging *inside* main (repo style)
    # --------------------------------------------------------------
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    logging.info("========================================================")
    logging.info("GTFS Block Schedule Printable Generator")
    logging.info("Input GTFS Folder: %s", GTFS_FOLDER_PATH)
    logging.info("Output Folder:     %s", BASE_OUTPUT_PATH)
    if FILTER_ROUTE_SHORT_NAMES:
        logging.info("Filtering for Routes: %s", FILTER_ROUTE_SHORT_NAMES)
    if FILTER_SERVICE_IDS:
        logging.info("Filtering for Service IDs: %s", FILTER_SERVICE_IDS)

    try:
        gtfs_data = load_gtfs_data(
            gtfs_folder_path=GTFS_FOLDER_PATH,
            files=REQUIRED_GTFS_FILES,
            dtype=str,
        )

        trips_df = gtfs_data["trips"]
        stop_times_df = gtfs_data["stop_times"]
        stops_df = gtfs_data["stops"]
        routes_df = gtfs_data["routes"]

        trips_df, stop_times_df = filter_data(trips_df, stop_times_df, routes_df)
        if trips_df.empty or stop_times_df.empty:
            logging.warning("No data remains after filtering – no files generated.")
            return

        prepared = prepare_stop_times(trips_df, stop_times_df, stops_df)
        if prepared.empty:
            logging.warning("No data remains after preparation – no files generated.")
            return

        export_blocks(prepared)
        logging.info("Script finished successfully.")

    except (OSError, ValueError, RuntimeError) as err:
        logging.error("%s", err)
    except Exception as err:  # catch-all for unforeseen issues
        logging.exception("Unexpected error: %s", err)
    finally:
        logging.info("Exiting script.")


if __name__ == "__main__":
    main()
