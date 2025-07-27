"""GTFS KPI & Change Analyzer.

Key features
------------
* **Explicit `service_id` groups** – you map “Weekday”, “Saturday”, “Sunday”
  (or anything else) to the exact `service_id` values used in your feed.
* **calendar_dates aware** – the script merges `calendar.txt` and
  `calendar_dates.txt` so it only analyses trips that truly run on a chosen
  date.
* **Revenue‑hour tracking** – scheduled platform hours per user‑defined group,
  with up/down change flags between sign‑ups.
* **Coverage polygons** – 0.25‑mile buffered stop coverage per route
  (GeoPandas and Shapely are required imports).
* No command‑line arguments – everything lives in the CONFIGURATION block.
"""

from __future__ import annotations

import logging
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Mapping, Sequence, Tuple

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from shapely.geometry import MultiPolygon, Polygon, LineString
from shapely.ops import unary_union

# =============================================================================
# CONFIGURATION
# =============================================================================

SIGNUPS: list[tuple[str, Path]] = [
    # (label, path_to_gtfs_folder)
#    ("Oct25", Path(r"Path\To\Your\GTFS_Folder")),
#    ("Jan25", Path(r"Path\To\Your\GTFS_Folder")),
    ("May25", Path(r"Path\To\Your\GTFS_Folder)),
]

OUTPUT_DIR: Path = Path(r"Path\To\Your\GTFS_Folder")
EXPORT_COVERAGE: bool = True  # set False to skip shapefile export

# ------------------------------------------------------------------
# Explicit mapping: Schedule name → tuple of GTFS service_id values
# ------------------------------------------------------------------
SERVICE_ID_GROUPS: dict[str, tuple[str, ...]] = {
    "Weekday": ("3",),
    "Saturday": ("1",),
    "Sunday": ("4",),
    "Holiday": ("2",),  # ← trailing comma makes it a tuple
}

# The revenue‑hour table can use a *different* grouping if desired.
REVENUE_HOUR_GROUPS: dict[str, tuple[str, ...]] = SERVICE_ID_GROUPS.copy()

# Routes that should never appear in any output
ROUTE_FILTER_OUT: set[str] = {"9999A", "9999B", "9999C"}

# Coverage parameters
COVERAGE_BUFFER_MILES: float = 0.25
GEOM_CHANGE_THRESHOLD: float = 0.05  # 5 % of area

# Time‑of‑day blocks for headway reporting
TIME_BLOCKS_CFG: Mapping[str, Tuple[str, str]] = {
    "am": ("04:00", "09:00"),
    "midday": ("09:00", "15:00"),
    "pm": ("15:00", "21:00"),
    "night": ("21:00", "28:00"),
}

GTFS_FILES: Tuple[str, ...] = (
    "routes.txt",
    "trips.txt",
    "stop_times.txt",
    "calendar.txt",
    "calendar_dates.txt",
    "stops.txt",
    "shapes.txt",  # <-- Add this line
)

# -----------------------------------------------------------------------------
LOGGER = logging.getLogger("GTFS‑Analyzer")
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(message)s")

# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass(frozen=True)
class Signup:
    """A GTFS sign‑up (aka service change) bundle."""

    label: str
    path: Path


@dataclass
class RouteSignature:
    """Snapshot of geometry + key stats for a route in one sign‑up."""

    geometry: Polygon | MultiPolygon | None
    signature_str: str
    rev_hours: dict[str, float]


@dataclass
class SignupSummary:
    """Condensed information for later sign‑up comparison."""

    routes: Dict[str, RouteSignature]  # key = route_short_name
    stops: set[str]


# =============================================================================
# GENERIC HELPERS
# =============================================================================
def check_input_files(base_dir: Path) -> None:
    """Raise if any required GTFS file is missing."""
    missing = [f for f in GTFS_FILES if not (base_dir / f).exists()]
    if missing:
        raise FileNotFoundError(f"Missing GTFS file(s) in {base_dir}: {', '.join(missing)}")


def load_gtfs(base_dir: Path) -> Dict[str, pd.DataFrame]:
    """Load all GTFS text files into a dict of DataFrames."""
    data: Dict[str, pd.DataFrame] = {}
    for fname in GTFS_FILES:
        fp = base_dir / fname
        data[fp.stem] = pd.read_csv(fp, low_memory=False)
        LOGGER.info("Loaded %-15s | %6d rows", fname, len(data[fp.stem]))
    return data


# ------------------------------------------------------------------
# Build service_id → set[YYYYMMDD] of all active dates
# ------------------------------------------------------------------
def build_service_calendar(
    calendar_df: pd.DataFrame,
    calendar_dates_df: pd.DataFrame,
) -> dict[str, set[str]]:
    """Merge calendar and calendar_dates into active-date sets per service_id."""
    # ── enforce string keys so look-ups succeed everywhere else ─────────
    calendar_df = calendar_df.copy()
    calendar_df["service_id"] = calendar_df["service_id"].astype(str)

    calendar_dates_df = calendar_dates_df.copy()
    calendar_dates_df["service_id"] = calendar_dates_df["service_id"].astype(str)
    # -------------------------------------------------------------------
    svc_dates: dict[str, set[str]] = {}

    for row in calendar_df.itertuples(index=False):
        # --- Start of Added Code ---
        # Create a date range from start_date to end_date
        start = pd.to_datetime(str(row.start_date), format="%Y%m%d")
        end = pd.to_datetime(str(row.end_date), format="%Y%m%d")
        rng = pd.date_range(start, end)

        # Create a boolean mask for active days of the week
        days_active = [
            row.monday, row.tuesday, row.wednesday, row.thursday,
            row.friday, row.saturday, row.sunday
        ]
        # (rng.dayofweek maps Monday=0, Sunday=6)
        mask = [days_active[d] == 1 for d in rng.dayofweek]
        # --- End of Added Code ---

        # Apply the mask to the date range to get active dates
        svc_dates[row.service_id] = set(rng[mask].strftime("%Y%m%d"))

    # Add or remove dates based on calendar_dates.txt
    for cd in calendar_dates_df.itertuples(index=False):
        svc_set = svc_dates.setdefault(cd.service_id, set())
        if cd.exception_type == 1:
            svc_set.add(str(cd.date))
        elif cd.exception_type == 2:
            svc_set.discard(str(cd.date))

    return svc_dates


def parse_time_blocks(cfg: Mapping[str, Tuple[str, str]]):
    """Convert TIME_BLOCKS_CFG into timedelta ranges."""
    return {
        label: (
            pd.Timedelta(hours=int(start.split(":")[0]), minutes=int(start.split(":")[1])),
            pd.Timedelta(hours=int(end.split(":")[0]), minutes=int(end.split(":")[1])),
        )
        for label, (start, end) in cfg.items()
    }


def assign_time_block(td: pd.Timedelta, blocks):
    """Return the block name whose range contains *td*."""
    for name, (start, end) in blocks.items():
        if start <= td < end:
            return name
    return "other"


def format_timedelta(td: pd.Timedelta | None) -> str | None:
    """Convert timedelta to HH:MM string or None."""
    if td is None or pd.isna(td):
        return None
    secs = int(td.total_seconds())
    hours, mins = divmod(secs, 3600)
    mins //= 60
    return f"{hours:02d}:{mins:02d}"


# =============================================================================
# REVENUE‑HOUR LOGIC
# =============================================================================
def compute_revenue_hours(
    data: Dict[str, pd.DataFrame],
    groups: Mapping[str, tuple[str, ...]],
) -> pd.DataFrame:
    """Return route-level revenue-hours (platform hrs) per user group."""
    st = data["stop_times"].copy()
    st[["arrival_time", "departure_time"]] = st[["arrival_time", "departure_time"]].apply(
        pd.to_timedelta,
        errors="coerce",
    )

    # .copy() avoids SettingWithCopyWarning when adding dtype
    trips = data["trips"][["trip_id", "route_id", "service_id"]].copy()
    trips["service_id"] = trips["service_id"].astype(str)

    valid_sids = {sid for sids in groups.values() for sid in sids}
    trips = trips[trips["service_id"].isin(valid_sids)]

    st = (
        st.groupby("trip_id", as_index=False)
        .agg(run_min=("departure_time", lambda s: (s.max() - s.min()).total_seconds() / 60))
        .merge(trips, on="trip_id", how="inner")
    )

    frames: list[pd.DataFrame] = []
    for grp_label, sids in groups.items():
        sids = (sids,) if isinstance(sids, str) else sids
        subset = st[st["service_id"].isin(sids)]
        if subset.empty:
            continue
        agg = (
            subset.groupby("route_id", as_index=False)["run_min"]
            .sum()
            # The .round(1) method was added to the line below
            .assign(group=grp_label, rev_hours=lambda df: (df["run_min"] / 60).round(1))
        )
        frames.append(agg[["route_id", "group", "rev_hours"]])

    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    

# ── normalise SERVICE_ID_GROUPS / REVENUE_HOUR_GROUPS ─────────────────────────
def _normalize_group_map(
    mapping: dict[str, tuple[str, ...] | str | Sequence[str]],
) -> dict[str, tuple[str, ...]]:
    """Ensure every dict value is a tuple[str, ...]."""
    out: dict[str, tuple[str, ...]] = {}
    for key, val in mapping.items():
        if isinstance(val, (list, tuple, set)):
            out[key] = tuple(map(str, val))
        else:  # scalar
            out[key] = (str(val),)
    return out


# =============================================================================
# KPI GENERATION
# =============================================================================
def _modal_headway(departs: pd.Series) -> float | None:
    """Return the mode of headway (minutes) or None."""
    diffs = departs.sort_values().diff().dropna().dt.total_seconds().div(60)
    return diffs.mode()[0] if not diffs.empty else None


def _calc_trip_times(group: pd.DataFrame) -> pd.Series:
    """Return first_trip, last_trip, trips for *group*."""
    t = group["departure_time"].sort_values()
    return pd.Series(
        {
            "first_trip": format_timedelta(t.min()),
            "last_trip": format_timedelta(t.max()),
            "trips": len(t),
        }
    )


def _save_excel(df: pd.DataFrame, fp: Path, sheet: str) -> None:
    """Write *df* to an XLSX file with modest formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(df.columns.to_list())
    for row in df.itertuples(index=False):
        ws.append(row)
    for idx, col in enumerate(ws.columns, start=1):
        width = max(len(str(c.value)) for c in col if c.value) + 2
        ws.column_dimensions[get_column_letter(idx)].width = width
        for cell in col:
            cell.alignment = Alignment(horizontal="center")
    wb.save(fp)
    LOGGER.info("Saved %s", fp)


def build_kpi_tables(
    data: Dict[str, pd.DataFrame],
    outdir: Path,
    label: str,
    groups: Mapping[str, tuple[str, ...]],
    coverage: Dict[str, Polygon | MultiPolygon],
) -> None:
    """Generate KPI Excel files for every schedule group in *groups*."""
    blocks = parse_time_blocks(TIME_BLOCKS_CFG)
    routes = data["routes"]
    trips = data["trips"].copy()
    trips["service_id"] = trips["service_id"].astype(str)
    st = data["stop_times"].copy()
    st["departure_time"] = pd.to_timedelta(st["departure_time"], errors="coerce")
    st = st.dropna(subset=["departure_time"])
    
    FEET_PER_MILE = 5280
    area_df = pd.DataFrame(
        [(rsn, geom.area / FEET_PER_MILE**2) for rsn, geom in coverage.items()],
        columns=["route_short_name", "coverage_sq_mi"],
    )
    
    shape_stats_df = pd.DataFrame()
    if "shapes" in data:
        shape_stats_df = _calculate_shape_stats(data)
    else:
        LOGGER.warning("shapes.txt not found. Cannot calculate route length or stop counts.")
    
    rev_hrs = compute_revenue_hours(data, groups)
    rev_pivot = (
        rev_hrs.pivot(index="route_id", columns="group", values="rev_hours")
        .rename(columns=lambda c: f"{c}_rev_hours")
        .reset_index()
        if not rev_hrs.empty
        else pd.DataFrame(columns=["route_id"])
    )

    svc_calendar = build_service_calendar(data["calendar"], data["calendar_dates"])

    for sched_name, svc_ids_raw in groups.items():
        svc_ids = tuple(str(s) for s in svc_ids_raw)
        trips_sub = trips[trips["service_id"].isin(svc_ids)]
        if trips_sub.empty:
            LOGGER.info("No trips for %s %s – skipped", label, sched_name)
            continue
        
        # --- Start of Corrected Section ---
        # This block defines `trips_filt` and must come BEFORE it is used.
        candidate_dates = (
            pd.Series(svc_ids)
            .map(lambda sid: sorted(svc_calendar.get(sid, [])))
            .explode()
            .dropna()
            .unique()
        )
        if candidate_dates.size == 0:
            LOGGER.warning("No active dates for %s – skipped", sched_name)
            continue
        target_date = candidate_dates[0]

        valid_sids = {sid for sid in svc_ids if target_date in svc_calendar.get(sid, set())}
        trips_filt = trips_sub[trips_sub["service_id"].isin(valid_sids)]
        if trips_filt.empty:
            LOGGER.info("No trips valid on %s for %s – skipped", target_date, sched_name)
            continue
        
        # Now that `trips_filt` is defined, we can calculate modal shapes.
        modal_shapes_df = pd.DataFrame()
        if not shape_stats_df.empty:
            modal_shapes = trips_filt.groupby(["route_id", "direction_id"])["shape_id"].apply(
                lambda s: s.mode()[0] if not s.mode().empty else None
            ).dropna()
            modal_shapes_df = modal_shapes.reset_index(name="shape_id")
            modal_shapes_df = modal_shapes_df.merge(shape_stats_df, on="shape_id", how="left")
        # --- End of Corrected Section ---
            
        st_first = (
            st.merge(trips_filt[["trip_id", "route_id", "direction_id"]], on="trip_id")
            .query("stop_sequence == 1")
        )
        st_first["time_block"] = st_first["departure_time"].apply(
            lambda td: assign_time_block(td, blocks)
        )
        st_first = st_first[st_first["time_block"] != "other"]
        if st_first.empty:
            continue

        kpi = (
            st_first.groupby(["route_id", "direction_id"], group_keys=False).apply(_calc_trip_times).reset_index()
            .merge(routes[["route_id", "route_short_name", "route_long_name"]], on="route_id")
        )

        head = (
            st_first.groupby(["route_id", "direction_id", "time_block"])["departure_time"]
            .apply(_modal_headway)
            .unstack()
        )
        kpi = (
            kpi.join(head, on=["route_id", "direction_id"])
            .merge(rev_pivot, on="route_id", how="left")
            .sort_values(["route_short_name", "direction_id"])
        )
        
        if not modal_shapes_df.empty:
            kpi = kpi.merge(modal_shapes_df, on=["route_id", "direction_id"], how="left")
            if "length_mi" in kpi.columns:
                kpi["length_mi"] = kpi["length_mi"].round(2)

        kpi = kpi.merge(area_df, on="route_short_name", how="left")
        if "coverage_sq_mi" in kpi.columns:
            kpi["coverage_sq_mi"] = kpi["coverage_sq_mi"].round(2)

        _save_excel(kpi, outdir / f"{label}_route_kpi_{sched_name.upper()}.xlsx", "KPIs")


# =============================================================================
# COVERAGE & SIGNATURES
# =============================================================================
def _build_route_coverage(data: Dict[str, pd.DataFrame]) -> Dict[str, Polygon | MultiPolygon]:
    """Buffer every stop 0.25 mi and dissolve per route."""
    stops = data["stops"]
    trips = data["trips"]
    routes = data["routes"]
    st = data["stop_times"][["trip_id", "stop_id"]]

    rt_map = trips.merge(routes[["route_id", "route_short_name"]], on="route_id")
    st = (
        st.merge(rt_map[["trip_id", "route_short_name"]])
        .merge(stops[["stop_id", "stop_lat", "stop_lon"]])
    )
    st = st[~st["route_short_name"].isin(ROUTE_FILTER_OUT)].dropna()

    gdf = gpd.GeoDataFrame(
        st, geometry=gpd.points_from_xy(st.stop_lon, st.stop_lat), crs="EPSG:4326"
    ).to_crs("EPSG:2248")

    buf_ft = COVERAGE_BUFFER_MILES * 5280
    return {rt: unary_union([pt.buffer(buf_ft) for pt in sdf.geometry]) for rt, sdf in gdf.groupby("route_short_name")}


def _geom_change(a: Polygon | MultiPolygon | None, b: Polygon | MultiPolygon | None) -> str:
    """Describe how geometry changed between sign‑ups."""
    if (a is None or a.is_empty) and (b and not b.is_empty):
        return "Geography expanded"
    if (b is None or b.is_empty) and (a and not a.is_empty):
        return "Geography contracted"
    if not a or not b or (a.is_empty and b.is_empty):
        return "No geographic change"

    delta = b.area - a.area
    if abs(delta) / a.area > GEOM_CHANGE_THRESHOLD:
        return "Geography expanded" if delta > 0 else "Geography contracted"
    if (a ^ b).area > 1e-9:  # symmetric difference non‑zero
        return "Geography modified"
    return "No geographic change"


def _rev_change(old: dict[str, float], new: dict[str, float], tol: float = 1e-6) -> str:
    """Return human‑readable delta of revenue hours."""
    msgs: list[str] = []
    for grp in REVENUE_HOUR_GROUPS:
        delta = new.get(grp, 0.0) - old.get(grp, 0.0)
        if abs(delta) > tol:
            msgs.append(f"Group {grp} revenue‑hrs {'up' if delta > 0 else 'down'} {abs(delta):.1f}")
    return "; ".join(msgs)


def summarise_signup(
    data: Dict[str, pd.DataFrame], 
    coverage: Dict[str, Polygon | MultiPolygon] # <-- Updated signature
) -> SignupSummary:
    """Create a condensed summary used later for sign-up comparison."""
    routes_df = data["routes"][["route_id", "route_short_name"]]
    routes = routes_df["route_short_name"].unique()
    stops_set: set[str] = set(data["stops"]["stop_id"].astype(str))

    # The 'coverage = _build_route_coverage(data)' line is now removed from here.

    # headway snippet for schedule signature
    st = data["stop_times"].copy()
    st["departure_time"] = pd.to_timedelta(st["departure_time"], errors="coerce")
    st = st[st["stop_sequence"] == 1].dropna(subset=["departure_time"])
    st = st.merge(data["trips"][["trip_id", "route_id", "direction_id"]])
    st = st.merge(routes_df)

    head = (
        st.groupby(["route_short_name", "direction_id"])["departure_time"]
        .apply(_modal_headway)
        .reset_index()
    )

    rev_df = compute_revenue_hours(data, REVENUE_HOUR_GROUPS)
    rev_by_route = rev_df.groupby("route_id").apply(lambda g: g.set_index("group")["rev_hours"].to_dict())
    route_id_lookup = routes_df.set_index("route_short_name")["route_id"].to_dict()

    sigs: Dict[str, RouteSignature] = {}
    for rt in routes:
        h = head[head.route_short_name == rt].to_string(index=False)
        sigs[rt] = RouteSignature(
            geometry=coverage.get(rt),
            signature_str=h,
            rev_hours=rev_by_route.get(route_id_lookup.get(rt), {}),
        )
    return SignupSummary(sigs, stops_set)


def _calculate_shape_stats(data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Pre-computes length and stop count for each unique shape_id."""
    LOGGER.info("Calculating route lengths and stop counts per shape...")
    shapes = data["shapes"]
    trips = data["trips"]
    stop_times = data["stop_times"]

    # 1. Calculate shape lengths in miles
    shapes = shapes.sort_values(["shape_id", "shape_pt_sequence"])
    lines = shapes.groupby("shape_id")[["shape_pt_lon", "shape_pt_lat"]].apply(
        lambda x: LineString(x.values)
    )
    shapes_gdf = gpd.GeoDataFrame(geometry=lines, crs="EPSG:4326").reset_index()
    shapes_gdf["length_mi"] = (
        shapes_gdf.to_crs("EPSG:2248").geometry.length / 5280
    )
    length_df = shapes_gdf[["shape_id", "length_mi"]].copy()

    # 2. Calculate modal number of stops per shape
    trip_stops = stop_times[["trip_id", "stop_id"]].merge(
        trips[["trip_id", "shape_id"]], on="trip_id"
    )
    trip_stop_counts = trip_stops.groupby("trip_id")["stop_id"].nunique().reset_index(name="num_stops")
    shape_stop_counts = trip_stop_counts.merge(trips[["trip_id", "shape_id"]], on="trip_id")
    modal_stops_df = shape_stop_counts.groupby("shape_id")["num_stops"].apply(
        lambda s: s.mode()[0] if not s.mode().empty else 0
    ).reset_index()

    # 3. Combine and return
    final_stats = length_df.merge(modal_stops_df, on="shape_id", how="left")
    final_stats["num_stops"] = final_stats["num_stops"].fillna(0).astype(int)
    
    LOGGER.info("Finished calculating shape stats.")
    return final_stats


def compare_signups(prev: SignupSummary, curr: SignupSummary) -> pd.DataFrame:
    """Return a DataFrame describing route‑level changes between two sign‑ups."""
    all_routes = sorted(set(prev.routes) | set(curr.routes))
    rows: list[tuple[str, str]] = []

    for rt in all_routes:
        if rt not in prev.routes:
            rows.append((rt, "Route created"))
            continue
        if rt not in curr.routes:
            rows.append((rt, "Route eliminated"))
            continue

        changes: list[str] = []
        g_change = _geom_change(prev.routes[rt].geometry, curr.routes[rt].geometry)
        if g_change != "No geographic change":
            changes.append(g_change)

        if prev.routes[rt].signature_str != curr.routes[rt].signature_str:
            changes.append("Schedule/headway change")

        rev_msg = _rev_change(prev.routes[rt].rev_hours, curr.routes[rt].rev_hours)
        if rev_msg:
            changes.append(rev_msg)

        rows.append((rt, ", ".join(changes) or "No change"))

    rows.append(
        (
            "[SYSTEM]",
            f"Stops added: {len(curr.stops - prev.stops)} | "
            f"Stops deleted: {len(prev.stops - curr.stops)}",
        )
    )
    return pd.DataFrame(rows, columns=["route_short_name", "change"])


# Normalise group maps once at import time
SERVICE_ID_GROUPS = _normalize_group_map(SERVICE_ID_GROUPS)
REVENUE_HOUR_GROUPS = _normalize_group_map(REVENUE_HOUR_GROUPS)

# =============================================================================
# MAIN ORCHESTRATION
# =============================================================================

def main() -> None:
    """Entry point."""
    signup_objs = [Signup(label, path) for label, path in SIGNUPS]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    summaries: dict[str, SignupSummary] = {}

    for s in signup_objs:
        LOGGER.info("=== Processing %s ===", s.label)
        check_input_files(s.path)
        data = load_gtfs(s.path)

        # --- OPTIMIZATION ---
        # Calculate coverage once and reuse it
        LOGGER.info("Calculating route coverage polygons...")
        coverage = _build_route_coverage(data)
        LOGGER.info("Coverage calculation complete.")
        
        # Pass 'coverage' as the final argument
        build_kpi_tables(data, OUTPUT_DIR, s.label, SERVICE_ID_GROUPS, coverage)

        if EXPORT_COVERAGE:
            gdf = gpd.GeoDataFrame(
                {"route_short_name": list(coverage), "geometry": list(coverage.values())},
                crs="EPSG:2248",
            )
            gdf["area_sq_mi"] = gdf["geometry"].area / (5280**2)
            gdf.to_file(OUTPUT_DIR / f"{s.label}_coverage.shp")
            LOGGER.info("Coverage shapefile exported for %s", s.label)

        # Pass 'coverage' as the final argument
        summaries[s.label] = summarise_signup(data, coverage)

    # ── compare successive sign-ups ─────────────────────────────────────
    if len(signup_objs) >= 2:
        comp_frames: list[pd.DataFrame] = []
        for prev, curr in zip(signup_objs[:-1], signup_objs[1:]):
            df = compare_signups(summaries[prev.label], summaries[curr.label])
            df["signup"] = f"{prev.label} ➜ {curr.label}"
            comp_frames.append(df)

        comp = (
            pd.concat(comp_frames, ignore_index=True)
            .pivot(index="route_short_name", columns="signup", values="change")
            .reset_index()
        )
        _save_excel(comp, OUTPUT_DIR / "service_change_comparison.xlsx", "Comparison")

    LOGGER.info("Processing complete. Outputs are in %s", OUTPUT_DIR)


if __name__ == "__main__":
    try:
        main()
    except Exception:  # pylint: disable=broad-except
        LOGGER.exception("Unhandled exception")
        sys.exit(1)
