"""Assign scheduled bus trips to bays within a cluster to minimize capacity conflicts.

Two solver flavours are supported:
- Greedy – fast heuristic; usually “good enough”.
- Integer Programming (PuLP) – slower but can eliminate more conflicts.

Typical usage:
- ArcPro notebook
- Jupyter notebook
"""

import os
from typing import Dict, List, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    import pulp

    PULP_AVAILABLE = True
except ImportError:
    PULP_AVAILABLE = False

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

INPUT_DIR = r"Path\To\Your\Input_Folder"
OUTPUT_DIR = r"Path\To\Your\Output_Folder"

USE_PULP = False
USE_GREEDY = True

CLUSTER_DEFINITIONS = {
    "Metro": {
        "single_bay_stops": ["2373"],
        "double_bay_stops": ["2832"],
        "triple_bay_stops": [],
        "overflow_bays": [],
    },
}

ALLOW_ROUTE_DIRECTION_SPLIT = False
ROUTE_TO_SPECIFIC_STOPS: Dict[Tuple[str, str], List[str]] = {}
ROUTE_PAIRS_TOGETHER: Set[Tuple[str, str]] = set()
ROUTE_TO_LAYOVER_BAYS: Dict[str, List[str]] = {}
ALLOW_LAYOVER_AT_REAL_STOPS = False

PASSENGER_SERVICE_STATUSES = {"ARRIVE", "DEPART", "ARRIVE/DEPART", "LOADING"}
ARRIVE_STATUSES = {"ARRIVE", "ARRIVE/DEPART"}
DEPART_STATUSES = {"DEPART", "ARRIVE/DEPART"}
LAYOVER_STATUSES = {"LAYOVER", "DWELL", "LONG BREAK", "LOADING"}

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def build_stop_capacities(cluster_info: Dict[str, List[str]]) -> Dict[str, int]:
    """Map stop IDs to their integer bay capacities.

    Args:
        cluster_info: Dictionary with stop ID lists for keys 'single_bay_stops',
            'double_bay_stops', 'triple_bay_stops', and 'overflow_bays'.

    Returns:
        A dictionary mapping string stop IDs to their integer capacity values.
    """
    stop_caps = {}
    for stop_id in cluster_info.get("single_bay_stops", []):
        stop_caps[str(stop_id)] = 1
    for stop_id in cluster_info.get("double_bay_stops", []):
        stop_caps[str(stop_id)] = 2
    for stop_id in cluster_info.get("triple_bay_stops", []):
        stop_caps[str(stop_id)] = 3
    for ovf in cluster_info.get("overflow_bays", []):
        stop_caps[str(ovf)] = 1
    return stop_caps


def recompute_conflict_types(df_in: pd.DataFrame, cluster_info: Dict[str, List[str]]) -> List[str]:
    """Classify each row with 'STOP', 'CLUSTER', 'BOTH', or 'NONE' based on capacity limits.

    Args:
        df_in: Input DataFrame with 'Timestamp', 'Block', 'AssignedStop', and 'Status'.
        cluster_info: Dictionary of stop bay definitions for the cluster.

    Returns:
        A list of conflict classifications corresponding to each row in the input.
    """
    df = df_in.copy()
    stop_caps = build_stop_capacities(cluster_info)

    # Overall cluster capacity
    single_ct = len(cluster_info.get("single_bay_stops", []))
    double_ct = len(cluster_info.get("double_bay_stops", []))
    triple_ct = len(cluster_info.get("triple_bay_stops", []))
    overflow_ct = len(cluster_info.get("overflow_bays", []))
    cluster_capacity = single_ct + (2 * double_ct) + (3 * triple_ct) + overflow_ct

    # Timestamps that exceed cluster capacity
    presence_df = df[~df["Status"].isna()]
    presence_count = presence_df.groupby("Timestamp")["Block"].count()
    cluster_conf_set = {ts for ts, ct in presence_count.items() if ct > cluster_capacity}

    # Stop-level conflicts for passenger-service statuses
    pass_df = df[df["Status"].str.upper().isin(PASSENGER_SERVICE_STATUSES)]
    group_stop = pass_df.groupby(["AssignedStop", "Timestamp"])
    stop_conf_set = set()
    for (sid, ts), grp in group_stop:
        cap = stop_caps.get(str(sid), 1)
        if len(grp) > cap:
            stop_conf_set.add((sid, ts))

    # Build final conflict col
    out_list = []
    for _, row in df.iterrows():
        ts = row["Timestamp"]
        sid = row.get("AssignedStop", None)
        if pd.isna(sid) or not sid:
            out_list.append("NONE")
            continue

        c_conf = ts in cluster_conf_set
        s_conf = (sid, ts) in stop_conf_set
        if c_conf and s_conf:
            out_list.append("BOTH")
        elif c_conf:
            out_list.append("CLUSTER")
        elif s_conf:
            out_list.append("STOP")
        else:
            out_list.append("NONE")

    return out_list


def count_conflicts_by_routedir(
    df: pd.DataFrame, conflict_col: str = "ConflictType_Recalc"
) -> pd.DataFrame:
    """Count conflict-labeled minutes for each (Route, Direction) pair.

    Args:
        df: DataFrame containing a conflict label column.
        conflict_col: Name of the column indicating conflict classification.

    Returns:
        A summary DataFrame with total conflict minutes by route and direction.
    """
    tmp = df.copy()
    tmp["HasConflict"] = tmp[conflict_col].isin(["STOP", "CLUSTER", "BOTH"])
    grp = tmp.groupby(["Route", "Direction"])["HasConflict"].sum().reset_index()
    grp.rename(columns={"HasConflict": "ConflictMinutes"}, inplace=True)
    return grp


# --------------------------------------------------------------------------------------------------
# GREEDY SOLVER
# --------------------------------------------------------------------------------------------------


def solve_bus_assignment_greedy(
    df: pd.DataFrame, cluster_info: Dict[str, List[str]]
) -> pd.DataFrame:
    """Assign buses to stops using a greedy first-fit strategy.

    Args:
        df: Input DataFrame of stop-time presence events.
        cluster_info: Dictionary of stop bay capacity definitions.

    Returns:
        A copy of the input DataFrame with the 'AssignedStop' column added.
    """
    stop_caps = build_stop_capacities(cluster_info)
    all_stops = list(stop_caps.keys())

    used_stops_by_bus = {}
    occupancy = {}

    out_df = df.copy()
    out_df.sort_values(["Timestamp", "Block", "Trip ID"], inplace=True)
    out_df.reset_index(drop=True, inplace=True)

    assigned_list = []

    for _, row in out_df.iterrows():
        t = row["Timestamp"]
        blk = row["Block"]
        trip = row["Trip ID"]
        route = str(row["Route"])
        direct = str(row["Direction"])
        stat = str(row["Status"]).upper()

        bus_key = (blk, trip)
        if bus_key not in used_stops_by_bus:
            used_stops_by_bus[bus_key] = set()

        primary_stop_id = str(row["Stop ID"]) if pd.notna(row["Stop ID"]) else ""

        if stat == "DEPART":
            # must depart from primary
            candidates = [primary_stop_id]
        else:
            candidates = list(all_stops)
            if (route, direct) in ROUTE_TO_SPECIFIC_STOPS:
                candidates = ROUTE_TO_SPECIFIC_STOPS[(route, direct)]

            used_already = used_stops_by_bus[bus_key]
            if len(used_already) >= 3:
                candidates = [s for s in candidates if s in used_already]
            else:
                candidates = sorted(candidates, key=lambda s: (s not in used_already))

        chosen = None
        for s in candidates:
            cap = stop_caps.get(s, 1)
            occ = occupancy.get((s, t), 0)
            if occ < cap:
                chosen = s
                break

        if chosen is None:
            chosen = primary_stop_id

        occupancy[(chosen, t)] = occupancy.get((chosen, t), 0) + 1
        used_stops_by_bus[bus_key].add(chosen)
        assigned_list.append(chosen)

    out_df["AssignedStop"] = assigned_list
    return out_df


# --------------------------------------------------------------------------------------------------
# PULP-BASED SOLVER
# --------------------------------------------------------------------------------------------------


def solve_bus_assignment_pulp(df: pd.DataFrame, cluster_info: Dict[str, List[str]]) -> pd.DataFrame:
    """Assign buses to stops using an integer programming formulation via PuLP.

    Args:
        df: Input DataFrame with timestamped bus presence records.
        cluster_info: Dictionary defining stop bay capacities.

    Returns:
        A DataFrame with an 'AssignedStop' column based on the optimized assignment.

    Raises:
        ImportError: If PuLP is not available in the environment.
    """
    if not PULP_AVAILABLE:
        raise ImportError("PuLP is not available.")
    from pulp import (
        LpBinary,
        LpInteger,
        LpMinimize,
        LpProblem,
        LpVariable,
        lpSum,
        value,
    )

    stop_caps = build_stop_capacities(cluster_info)
    all_stops = list(stop_caps.keys())

    df["BusKey"] = list(zip(df["Block"], df["Trip ID"]))
    bus_keys = df["BusKey"].unique().tolist()
    times = sorted(df["Timestamp"].unique().tolist())

    # presence
    bus_info = {}
    for b in bus_keys:
        bdf = df[df["BusKey"] == b]
        bdict = {}
        for _, row in bdf.iterrows():
            t = row["Timestamp"]
            st = str(row["Status"]).upper()
            bdict[t] = {
                "status": st,
                "primary_stop_id": (str(row["Stop ID"]) if pd.notna(row["Stop ID"]) else None),
                "requires_capacity": st in PASSENGER_SERVICE_STATUSES,
            }
        bus_info[b] = bdict

    model = LpProblem("BusBayAssignment", LpMinimize)
    X = LpVariable.dicts("X", (bus_keys, times, all_stops), cat=LpBinary)

    # capacity constraints
    for s in all_stops:
        cap = stop_caps[s]
        for t in times:
            model += (lpSum(X[b][t][s] for b in bus_keys) <= cap, f"Cap_{s}_{t}")

    # presence constraints
    for b in bus_keys:
        for t in times:
            present = t in bus_info[b]
            if present:
                model += (lpSum(X[b][t][s] for s in all_stops) == 1, f"Pres_{b}_{t}")
            else:
                for s in all_stops:
                    model += (X[b][t][s] == 0, f"NoPres_{b}_{t}_{s}")

    # each bus uses at most 3 stops in total
    Y = LpVariable.dicts("Y", (bus_keys, all_stops), cat=LpBinary)
    for b in bus_keys:
        for s in all_stops:
            for t in times:
                model += (X[b][t][s] <= Y[b][s], f"Link_{b}_{s}_{t}")
        model += (lpSum(Y[b][s] for s in all_stops) <= 3, f"Max3_{b}")

    # must depart from primary stop
    for b in bus_keys:
        for t, inf in bus_info[b].items():
            if inf["status"] == "DEPART":
                must = inf["primary_stop_id"]
                if must and (must in all_stops):
                    model += (X[b][t][must] == 1, f"Depart_{b}_{t}")

    # OverCap variables for objective
    OverCap = LpVariable.dicts("OverCap", (all_stops, times), lowBound=0, cat=LpInteger)
    for s in all_stops:
        cap = stop_caps[s]
        for t in times:
            model += (
                OverCap[s][t] >= lpSum(X[b][t][s] for b in bus_keys) - cap,
                f"OC_{s}_{t}",
            )

    # Minimize total over-capacity across all stops/times
    model += lpSum(OverCap[s][t] for s in all_stops for t in times)

    print("\nSolving bus assignment via PuLP...")
    model.solve()
    print(f" -> Solver status: {model.status}, {pulp.LpStatus[model.status]}")
    print(f" -> Objective value: {value(model.objective)}")

    assigned_list = []
    for idx, row in df.iterrows():
        b = row["BusKey"]
        t = row["Timestamp"]
        if t not in bus_info[b]:
            assigned_list.append(None)
            continue
        chosen_s = None
        for s in all_stops:
            val = X[b][t][s].varValue
            if val and val > 0.5:
                chosen_s = s
                break
        assigned_list.append(chosen_s)

    out_df = df.copy()
    out_df["AssignedStop"] = assigned_list
    return out_df


# ==================================================================================================
# MAIN
# ==================================================================================================


def main() -> None:
    """Execute the full assignment and output pipeline for each defined cluster.

    This function:
        - Loads cluster-level presence data from Excel.
        - Applies either a greedy or IP-based assignment algorithm.
        - Recomputes conflict classifications.
        - Outputs annotated Excel files with assignments and summaries.

    Raises:
        FileNotFoundError: If required input files are missing.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Decide solver
    if USE_PULP and not PULP_AVAILABLE:
        print("Warning: USE_PULP=True but PuLP is not installed; using Greedy.")
    if USE_PULP and PULP_AVAILABLE:
        solver_method = "pulp"
    elif USE_GREEDY:
        solver_method = "greedy"
    else:
        solver_method = "greedy"
        print("No valid solver flags => defaulting to GREEDY.")

    for cluster_name, cinfo in CLUSTER_DEFINITIONS.items():
        safe_name = cluster_name.replace(" ", "_")
        step2_xlsx = os.path.join(INPUT_DIR, f"{safe_name}_Conflicts.xlsx")
        if not os.path.exists(step2_xlsx):
            print(f"Step2 file not found for cluster {cluster_name}, skipping.")
            continue

        df_before = pd.read_excel(step2_xlsx, sheet_name="AllStops")
        if df_before.empty:
            print(f"No data for {cluster_name}, skipping.")
            continue

        print(f"=== {cluster_name} => running {solver_method.upper()} solver ===")
        # 1) Solve assignment
        if solver_method == "pulp":
            df_after = solve_bus_assignment_pulp(df_before, cinfo)
        else:
            df_after = solve_bus_assignment_greedy(df_before, cinfo)

        # 2) Recompute conflicts (before vs after)
        df_before_annot = df_before.copy()
        df_before_annot["AssignedStop"] = df_before_annot["Stop ID"].fillna("")
        df_before_annot["ConflictType_Recalc"] = recompute_conflict_types(df_before_annot, cinfo)

        df_after_annot = df_after.copy()
        df_after_annot["ConflictType_Recalc"] = recompute_conflict_types(df_after_annot, cinfo)

        # 3) Build row-level “Before/After” file
        df_before_cols = df_before_annot[
            [
                "Block",
                "Trip ID",
                "Timestamp",
                "Route",
                "Direction",
                "Stop ID",
                "Stop Name",
            ]
        ].copy()
        df_before_cols.rename(
            columns={
                "Stop ID": "AssignedBeforeStopID",
                "Stop Name": "AssignedBeforeStopName",
            },
            inplace=True,
        )

        df_after_cols = df_after_annot[
            [
                "Block",
                "Trip ID",
                "Timestamp",
                "Route",
                "Direction",
                "Status",
                "AssignedStop",
                "ConflictType_Recalc",
            ]
        ].copy()
        df_after_cols.rename(
            columns={
                "AssignedStop": "AssignedAfterStopID",
                "ConflictType_Recalc": "ConflictType_RecalcAfter",
            },
            inplace=True,
        )

        df_after_merge = pd.merge(
            df_before_cols,
            df_after_cols,
            on=["Block", "Trip ID", "Timestamp", "Route", "Direction"],
            how="left",
        )
        df_after_merge["AssignedAfterStopName"] = ""

        detail_out = os.path.join(OUTPUT_DIR, f"{safe_name}_BeforeAfter.xlsx")
        with pd.ExcelWriter(detail_out, engine="openpyxl") as writer:
            df_before_annot.to_excel(writer, sheet_name="BeforeAssignment", index=False)
            df_after_merge.to_excel(writer, sheet_name="AfterAssignment", index=False)

        # NEW/UPDATED: BOLD ROWS WITH CONFLICTS (AfterAssignment sheet).
        # If you want to bold in the "BeforeAssignment" sheet as well,
        # just repeat the same logic but check "ConflictType_Recalc" there.
        wb = load_workbook(detail_out)
        ws = wb["AfterAssignment"]

        # Find the column index where "ConflictType_RecalcAfter" is located
        conflict_col_index = None
        header_row = 1  # by default, row 1 is the header in to_excel
        for cell in ws[header_row]:
            if cell.value == "ConflictType_RecalcAfter":
                conflict_col_index = cell.column  # 1-based index
                break

        if conflict_col_index is not None:
            # Iterate through rows below the header (start at row=2)
            for row_idx in range(header_row + 1, ws.max_row + 1):
                conflict_val = ws.cell(row=row_idx, column=conflict_col_index).value
                # Bold the entire row if conflict is not NONE (i.e., STOP, CLUSTER, BOTH)
                if conflict_val and conflict_val != "NONE":
                    for col_idx in range(1, ws.max_column + 1):
                        cell_to_bold = ws.cell(row=row_idx, column=col_idx)
                        cell_to_bold.font = Font(bold=True)

        wb.save(detail_out)
        print(
            f" -> Wrote row-level detail file (with bold conflicts): {os.path.basename(detail_out)}"
        )

        # 4) Build a route+direction conflict summary
        conf_before = count_conflicts_by_routedir(df_before_annot, "ConflictType_Recalc")
        conf_before.rename(columns={"ConflictMinutes": "ConflictBefore"}, inplace=True)

        conf_after = count_conflicts_by_routedir(df_after_annot, "ConflictType_Recalc")
        conf_after.rename(columns={"ConflictMinutes": "ConflictAfter"}, inplace=True)

        df_conf = pd.merge(conf_before, conf_after, on=["Route", "Direction"], how="outer").fillna(
            0
        )
        df_conf["ConflictBefore"] = df_conf["ConflictBefore"].astype(int)
        df_conf["ConflictAfter"] = df_conf["ConflictAfter"].astype(int)

        # 5) Multi-bucket approach for ARRIVE/DEPART in the final assignment
        def classify_buckets(status_str: str) -> Set[str]:
            result = set()
            s_up = status_str.upper()
            if s_up in ARRIVE_STATUSES:
                result.add("ARRIVE")
            if s_up in DEPART_STATUSES:
                result.add("DEPART")
            if not result:
                result.add("LAYOVER")
            return result

        rows_exploded = []
        for i, row in df_after_annot.iterrows():
            cats = classify_buckets(str(row["Status"]))
            for c in cats:
                new_row = row.copy()
                new_row["StatusBucket"] = c
                rows_exploded.append(new_row)

        df_exploded = pd.DataFrame(rows_exploded)

        grouped = (
            df_exploded.groupby(["Route", "Direction", "StatusBucket"])["AssignedStop"]
            .agg(lambda g: sorted(x for x in set(g) if x not in (None, "")))
            .reset_index()
        )

        pivoted = grouped.pivot_table(
            index=["Route", "Direction"],
            columns="StatusBucket",
            values="AssignedStop",
            aggfunc=lambda x: x,
        ).reset_index()

        pivoted.columns.name = None
        col_map = {
            "ARRIVE": "ArriveStops",
            "DEPART": "DepartStops",
            "LAYOVER": "LayoverStops",
        }
        pivoted.rename(columns=col_map, inplace=True)

        for c in ["ArriveStops", "DepartStops", "LayoverStops"]:
            if c in pivoted.columns:
                pivoted[c] = pivoted[c].apply(
                    lambda lst: ",".join(lst) if isinstance(lst, list) else ""
                )

        df_summary = pd.merge(df_conf, pivoted, on=["Route", "Direction"], how="outer").fillna("")

        summary_out = os.path.join(OUTPUT_DIR, f"{safe_name}_Summary.xlsx")
        with pd.ExcelWriter(summary_out, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Route_Summary", index=False)

        print(f" -> Wrote route-level summary file: {os.path.basename(summary_out)}")
        print("====================================================")


if __name__ == "__main__":
    main()
