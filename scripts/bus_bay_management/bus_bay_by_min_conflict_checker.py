"""
Script Name:
    bus_bay_by_min_conflict_checker.py

Purpose:
    Analyzes block-level bus data to detect and report scheduling
    conflicts at transit clusters based on defined bay capacities
    (single, double, triple, overflow) and bus statuses. Identifies
    both cluster-level and stop-level over-capacity situations.
    Designed as Step 2 in a transit analysis pipeline following
    block_status_by_minute_generator.py.

Inputs:
    1. Block-level transit data spreadsheets (XLSX format) from the
       directory specified by BLOCK_OUTPUT_FOLDER.
    2. Configuration constants defined in the script: 
       CLUSTER_DEFINITIONS (stop IDs, bay types per cluster),
    3. PRESENCE_STATUSES (statuses indicating bus presence), 
       PASSENGER_SERVICE_STATUSES (statuses indicating bay occupancy).

Outputs:
    1. Excel conflict analysis reports (one file per defined cluster)
       saved to the directory specified by CLUSTER_CONFLICT_OUTPUT_FOLDER.
    2. Each report includes an 'AllStops' summary sheet and 
       individual sheets for each stop/bay within the cluster, 
       highlighting rows with detected conflicts (CLUSTER, STOP, BOTH).

Dependencies:
    1. Libraries: pandas, openpyxl
"""

import os

import pandas as pd
from openpyxl.styles import Font

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

# Folder containing your block-level XLSX files from Step 1
BLOCK_OUTPUT_FOLDER = r"\\Path\To\Your\Input_Folder"

# Where to save the cluster conflict outputs
CLUSTER_CONFLICT_OUTPUT_FOLDER = r"\\Path\To\Your\Output_Folder"

# Dictionary of clusters, including single_bay, double_bay, triple_bay, and overflow.
# Each key is the cluster name, the value is a dict with lists:
#   'single_bay_stops' -> capacity 1 each
#   'double_bay_stops' -> capacity 2 each
#   'triple_bay_stops' -> capacity 3 each
#   'overflow_bays'    -> capacity 1 each
CLUSTER_DEFINITIONS = {
    "Park & Ride": {
        "single_bay_stops": ["3882", "3881"],
        "double_bay_stops": [],
        "triple_bay_stops": [],
        "overflow_bays": [],
    },
    "Metro": {
        "single_bay_stops": ["2373"],
        "double_bay_stops": ["2832"],
        "triple_bay_stops": [],
        "overflow_bays": ["layover_bay_A", "layover_bay_B"],
    },
}

# Define which statuses indicate bus presence in the cluster
PRESENCE_STATUSES = {
    "ARRIVE",
    "DEPART",
    "ARRIVE/DEPART",
    "DWELL",
    "LOADING",
    "LAYOVER",
    "LONG BREAK",
}

# Define which statuses indicate the bus is physically occupying a stop bay
PASSENGER_SERVICE_STATUSES = {"ARRIVE", "DEPART", "ARRIVE/DEPART", "LOADING"}

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================

# --------------------------------------------------------------------------------------------------
# CAPACITY AND STOP-LIST BUILDING BASED ON BAY COUNTS
# --------------------------------------------------------------------------------------------------

def get_all_official_stops(cinfo):
    """
    Return a combined list of all official stops in this cluster
    (regardless of single, double, or triple bay).
    """
    return (
        cinfo.get("single_bay_stops", [])
        + cinfo.get("double_bay_stops", [])
        + cinfo.get("triple_bay_stops", [])
    )


def build_cluster_capacities():
    """
    Each cluster’s capacity = sum of:
      - (# single-bay stops) * 1
      - (# double-bay stops) * 2
      - (# triple-bay stops) * 3
      + (# overflow bays) * 1
    """
    capacities = {}
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        n_single = len(cinfo.get("single_bay_stops", []))
        n_double = len(cinfo.get("double_bay_stops", []))
        n_triple = len(cinfo.get("triple_bay_stops", []))
        n_overflow = len(cinfo.get("overflow_bays", []))

        cluster_cap = (1 * n_single) + (2 * n_double) + (3 * n_triple) + (1 * n_overflow)
        capacities[cname] = cluster_cap
    return capacities


def build_stop_capacities():
    """
    Returns a dict {stop_id: capacity}, where:
      - single-bay = capacity 1
      - double-bay = capacity 2
      - triple-bay = capacity 3
      - overflow-bay = capacity 1
    """
    stop_caps = {}
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        for stop_id in cinfo.get("single_bay_stops", []):
            stop_caps[str(stop_id)] = 1
        for stop_id in cinfo.get("double_bay_stops", []):
            stop_caps[str(stop_id)] = 2
        for stop_id in cinfo.get("triple_bay_stops", []):
            stop_caps[str(stop_id)] = 3
        for ovf in cinfo.get("overflow_bays", []):
            stop_caps[str(ovf)] = 1
    return stop_caps


# --------------------------------------------------------------------------------------------------
# CORE CONFLICT-DETECTION LOGIC
# --------------------------------------------------------------------------------------------------

def normalize_stop_id(stop_id):
    """Convert stop IDs like '2956.0' -> '2956'. Handles NaN gracefully."""
    if pd.isna(stop_id):
        return None
    sid_str = str(stop_id).strip()
    if sid_str.endswith(".0"):
        sid_str = sid_str[:-2]
    return sid_str


def assign_cluster_name(df_in):
    """
    Add a 'ClusterName' column to indicate which cluster the row’s stop is in.
    If a stop appears in multiple clusters (unlikely), it will match the first found.
    """
    df_out = df_in.copy()
    df_out["ClusterName"] = None

    # Build a map of cluster -> set_of_stop_ids
    cluster_map = {}
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        official_stops = get_all_official_stops(cinfo)
        overflow = cinfo.get("overflow_bays", [])
        all_stops = official_stops + overflow
        cluster_map[cname] = set(map(str, all_stops))

    for cname, stop_set in cluster_map.items():
        mask = df_out["Stop ID"].isin(stop_set)
        df_out.loc[mask, "ClusterName"] = cname

    return df_out


def find_cluster_conflicts(df_in):
    """
    Return a set of (cluster_name, timestamp) where the number of buses present
    in that cluster at that time > cluster's capacity.
    """
    cluster_caps = build_cluster_capacities()
    conflict_set = set()

    present_df = df_in[df_in["Status"].isin(PRESENCE_STATUSES)].copy()
    present_df = present_df.dropna(subset=["ClusterName"])  # ignore rows w/o cluster

    group = present_df.groupby(["ClusterName", "Timestamp"])
    for (cname, ts), grp in group:
        cap = cluster_caps.get(cname, 1)
        # If more vehicles than capacity
        if len(grp) > cap:
            conflict_set.add((cname, ts))

    return conflict_set


def find_stop_conflicts(df_in):
    """
    Return a set of (stop_id, timestamp) where the number of buses
    in passenger-service statuses at that stop > that stop’s capacity.
    """
    stop_caps = build_stop_capacities()
    conflict_set = set()

    pass_df = df_in[df_in["Status"].isin(PASSENGER_SERVICE_STATUSES)].copy()
    pass_df = pass_df[pass_df["Stop ID"].notna()]

    group = pass_df.groupby(["Stop ID", "Timestamp"])
    for (sid, ts), grp in group:
        cap = stop_caps.get(sid, 1)
        if len(grp) > cap:
            conflict_set.add((sid, ts))

    return conflict_set


def annotate_conflicts(df_in, cluster_conflicts, stop_conflicts):
    """
    Append a 'ConflictType' column with values:
      - "NONE" (no conflict)
      - "CLUSTER" (cluster conflict only)
      - "STOP" (stop conflict only)
      - "BOTH" (both cluster & stop conflict)
    """
    df_out = df_in.copy()
    conflict_types = []

    for idx, row in df_out.iterrows():
        cname = row["ClusterName"]
        ts = row["Timestamp"]
        sid = row["Stop ID"]

        has_cluster_conf = pd.notna(cname) and (cname, ts) in cluster_conflicts
        has_stop_conf = sid is not None and (sid, ts) in stop_conflicts

        if has_cluster_conf and has_stop_conf:
            conflict_types.append("BOTH")
        elif has_cluster_conf:
            conflict_types.append("CLUSTER")
        elif has_stop_conf:
            conflict_types.append("STOP")
        else:
            conflict_types.append("NONE")

    df_out["ConflictType"] = conflict_types
    return df_out


# --------------------------------------------------------------------------------------------------
# I/O AND EXCEL WRITING LOGIC
# --------------------------------------------------------------------------------------------------

def gather_block_spreadsheets(block_folder):
    """
    Read all 'block_*.xlsx' spreadsheets from Step 1 in `block_folder`,
    and concatenate them into a single DataFrame.
    """
    all_files = [
        f
        for f in os.listdir(block_folder)
        if f.lower().endswith(".xlsx") and f.startswith("block_")
    ]
    if not all_files:
        raise FileNotFoundError(f"No block_*.xlsx files found in {block_folder}.")

    big_df_list = []
    for fname in all_files:
        path = os.path.join(block_folder, fname)
        temp_df = pd.read_excel(path)
        temp_df["FileName"] = fname
        big_df_list.append(temp_df)

    df_combined = pd.concat(big_df_list, ignore_index=True)
    print(f"Loaded {len(df_combined)} total rows from Step 1 block XLSX files.")
    return df_combined


def run_step2_conflict_detection():
    """
    1) Read block-level spreadsheets (Step 1 output).
    2) Normalize data, assign clusters, find conflicts with multi-bay logic.
    3) For each cluster, create an output Excel file:
       - "AllStops" sheet with all cluster events
       - One sheet per official stop or overflow bay, listing only that stop's events
       - Rows with conflicts are bolded.
    """
    print("=== Step 2: Conflict detection and per-cluster output (multi-bay) ===")
    os.makedirs(CLUSTER_CONFLICT_OUTPUT_FOLDER, exist_ok=True)

    # 1) Gather Step 1 data
    df = gather_block_spreadsheets(BLOCK_OUTPUT_FOLDER)

    # Basic checks/cleanup
    required_cols = [
        "Timestamp",
        "Trip ID",
        "Block",
        "Route",
        "Direction",
        "Stop ID",
        "Stop Name",
        "Stop Sequence",
        "Arrival Time",
        "Departure Time",
        "Status",
    ]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns in block-level data: {missing_cols}")

    # Normalize Stop ID text
    df["Stop ID"] = df["Stop ID"].apply(normalize_stop_id)
    # Ensure string Timestamps
    df["Timestamp"] = df["Timestamp"].astype(str).str.strip()

    # 2) Assign cluster, detect conflicts, annotate
    df = assign_cluster_name(df)
    cluster_conflicts = find_cluster_conflicts(df)
    stop_conflicts = find_stop_conflicts(df)
    df_annotated = annotate_conflicts(df, cluster_conflicts, stop_conflicts)

    # 3) Write results per cluster, each cluster => single XLSX with multiple sheets
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        # Combine all official stops + overflow for that cluster
        official_stops = get_all_official_stops(cinfo)
        overflow_stops = cinfo.get("overflow_bays", [])
        all_cluster_stops = official_stops + overflow_stops

        sub = df_annotated[df_annotated["ClusterName"] == cname].copy()
        if sub.empty:
            print(f"No rows found for cluster '{cname}'. Skipping.")
            continue

        safe_name = cname.replace(" ", "_")
        out_path = os.path.join(CLUSTER_CONFLICT_OUTPUT_FOLDER, f"{safe_name}_Conflicts.xlsx")
        print(f"Building conflict output for cluster '{cname}' => {out_path}")

        # Sort by timestamp, then by block or stop ID
        sub.sort_values(["Timestamp", "Block", "Stop ID"], inplace=True)

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # 3a) "AllStops" sheet
            sub.to_excel(writer, sheet_name="AllStops", index=False)

            # Bold the conflict rows in "AllStops"
            conflict_col_index = sub.columns.get_loc("ConflictType") + 1  # +1 for 1-based indexing
            worksheet_all = writer.sheets["AllStops"]
            for row_idx in range(2, len(sub) + 2):  # data starts on row 2
                conflict_val = worksheet_all.cell(row=row_idx, column=conflict_col_index).value
                if conflict_val != "NONE":
                    # Bold entire row
                    for col_idx in range(1, len(sub.columns) + 1):
                        cell = worksheet_all.cell(row=row_idx, column=col_idx)
                        cell.font = Font(bold=True)

            # 3b) One sheet per stop
            for stop_id in all_cluster_stops:
                sid_str = str(stop_id)
                stop_df = sub[sub["Stop ID"] == sid_str].copy()
                if stop_df.empty:
                    # no usage => skip or write empty sheet
                    continue

                # Make a sheet name that is safe in Excel (<=31 chars)
                sheet_name = f"Stop_{sid_str}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

                stop_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Bold the conflict rows in each stop’s sheet
                conflict_col_index_stop = stop_df.columns.get_loc("ConflictType") + 1
                worksheet_stop = writer.sheets[sheet_name]
                for row_idx in range(2, len(stop_df) + 2):
                    conflict_val = worksheet_stop.cell(
                        row=row_idx, column=conflict_col_index_stop
                    ).value
                    if conflict_val != "NONE":
                        for col_idx in range(1, len(stop_df.columns) + 1):
                            cell = worksheet_stop.cell(row=row_idx, column=col_idx)
                            cell.font = Font(bold=True)

        print(f" -> Completed writing {out_path}")

    # Final conflict summary stats
    print(f"\nDistinct cluster-conflict points: {len(cluster_conflicts)}")
    print(f"Distinct stop-conflict points: {len(stop_conflicts)}")
    print("Step 2 complete.")


# ==================================================================================================
# MAIN
# ==================================================================================================

def main():
    """
    Main entry point for the conflict detection script.

    1) Reads and combines block-level spreadsheets (the Step 1 outputs)
       from the configured BLOCK_OUTPUT_FOLDER.
    2) Normalizes stop IDs, assigns each record to a transit cluster,
       and identifies potential conflicts based on cluster- and stop-level bay capacities.
    3) For each cluster, writes an Excel file containing:
       - An "AllStops" sheet with every record for that cluster, highlighting conflict rows.
       - Individual sheets for each official stop or overflow bay, also highlighting conflicts.
    4) Prints a final summary of how many cluster-level and stop-level conflict points were detected.

    This function relies on the global configuration variables (e.g., CLUSTER_DEFINITIONS,
    BLOCK_OUTPUT_FOLDER, CLUSTER_CONFLICT_OUTPUT_FOLDER, PRESENCE_STATUSES, PASSENGER_SERVICE_STATUSES),
    which define the transit clusters, their bay capacities, and file locations.

    No arguments are accepted and nothing is returned; all results are saved directly to disk.
    """
    run_step2_conflict_detection()


if __name__ == "__main__":
    main()
