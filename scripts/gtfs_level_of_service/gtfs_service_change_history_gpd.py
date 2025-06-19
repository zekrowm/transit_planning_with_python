"""Analyzes chronological GTFS datasets to detect public transit service changes.

This script processes multiple GTFS signups, computes route coverage areas,
generates headway and schedule summaries, and compares signups to identify
new, eliminated, expanded, or modified routes.

Typical usage:
    - Update MULTIPLE_GTFS_CONFIGS with GTFS dataset paths and labels.
    - Run from ArcPro or a notebook to generate outputs.

Inputs:
    - GTFS files: routes, trips, stop_times, calendar, calendar_dates, stops.
    - Signup-specific GTFS folders and metadata.
    - Time blocks and day-type configurations.

Outputs:
    - Per-signup Excel files with route schedules and headways.
    - Per-signup shapefiles of route coverage areas.
    - One final Excel file comparing changes across signups.
"""

import os
from datetime import timedelta

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from shapely.ops import unary_union
from pandas import DataFrame
from shapely.geometry import Polygon

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

# 1) List of GTFS dataset inputs, each item is a dict:
#    - "name": a label (e.g., "September_2024_signup")
#    - "path": a folder path containing the GTFS files for that signup
MULTIPLE_GTFS_CONFIGS = [
    {"name": "September_2024_signup", "path": r"\\path\to\SEPT_2024\\"},
    {"name": "October_2025_signup", "path": r"\\path\to\OCT_2025\\"},
    # Add as many as you wish, in chronological order
]

# 2) Optional routes to filter out from all outputs/comparisons
ROUTE_FILTER_OUT = ["9999A", "9999B", "9999C"]  # If empty => no filtering

# 3) Output directory for Excel and shapefile outputs
OUTPUT_PATH = r"\\your_file_path\here\\"

# 4) GTFS files to load
gtfs_files = [
    "routes.txt",
    "trips.txt",
    "stop_times.txt",
    "calendar.txt",
    "calendar_dates.txt",
    "stops.txt",  # Needed for lat/lon for geometry
]

# 5) Base name for the output Excel file (for each schedule type)
OUTPUT_EXCEL = "route_schedule_headway_with_modes.xlsx"

# 6) Name of the final multi-signup comparison spreadsheet
COMPARISON_EXCEL = "service_change_comparison.xlsx"

# 7) Define time blocks with start/end times in 'HH:MM'
time_blocks_config = {
    "am": ("04:00", "09:00"),
    "midday": ("09:00", "15:00"),
    "pm": ("15:00", "21:00"),
    "night": ("21:00", "28:00"),  # 28:00 = 04:00 next day
}

# 8) Define multiple schedule types and their corresponding days
schedule_types = {
    "Weekday": ["monday", "tuesday", "wednesday", "thursday", "friday"],
    "Saturday": ["saturday"],
    "Sunday": ["sunday"],
}

# 9) Geo configuration
GEO_CRS = "EPSG:2248"  # A projected CRS for the DC region. Adjust as needed.
STOP_BUFFER_DISTANCE_MILES = 0.25  # 0.25-mile buffer for coverage

# 10) For "geography expanded" / "geography contracted" classification
GEOM_CHANGE_THRESHOLD = 0.05  # 5% area difference

# -----------------------------------------------------------------------------
# GLOBAL STORAGE
# -----------------------------------------------------------------------------

ALL_SIGNUP_FINAL_DATA: dict[str, list[DataFrame]] = {}  # { label: [DataFrame, DataFrame, ...] }
ALL_SIGNUP_COVERAGES: dict[str, dict[str, Polygon]] = {}  # { label: { route_short_name: polygon } }

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def check_input_files(base_path, files):
    """Verify that the input directory and all required GTFS files exist."""
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"The input directory {base_path} does not exist.")
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        if not os.path.exists(file_path):
            raise FileNotFoundError(
                f"The required GTFS file {file_name} does not exist in {base_path}."
            )


def load_gtfs_data(base_path, files):
    """Load GTFS data from specified files into a dictionary of pandas DataFrames."""
    data = {}
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        data_name = file_name.replace(".txt", "")
        try:
            data[data_name] = pd.read_csv(file_path)
            print(f"Loaded {file_name} with {len(data[data_name])} records.")
        except Exception as error:  # NOTE: Catching a broad exception can be OK here.
            raise Exception(f"Error loading {file_name}: {error}") from error
    return data


def parse_time_blocks(time_blocks_str):
    """Convert the human-readable time block strings into start/end timedeltas for each named time-block."""
    parsed_blocks = {}
    for block_name, (start_str, end_str) in time_blocks_str.items():
        start_h, start_m = map(int, start_str.split(":"))
        end_h, end_m = map(int, end_str.split(":"))
        parsed_blocks[block_name] = (
            timedelta(hours=start_h, minutes=start_m),
            timedelta(hours=end_h, minutes=end_m),
        )
    return parsed_blocks


def assign_time_block(time_delta, blocks):
    """Given a time_delta (time of day) and pre-parsed blocks, determine which block (am, midday, pm, night, or 'other')."""
    for block_name, (start, end) in blocks.items():
        if start <= time_delta < end:
            return block_name
    return "other"


def format_timedelta(time_delta):
    """Format a timedelta object as HH:MM string, or None if invalid."""
    if pd.isna(time_delta):
        return None
    total_seconds = int(time_delta.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{hours:02}:{minutes:02}"


def find_large_break(trip_times):
    """Check if there's a 3+ hour break in the trip_times array between 10:00 and 14:00."""
    late_morning = pd.Timedelta(hours=10)
    early_afternoon = pd.Timedelta(hours=14)
    midday_trips = trip_times[
        (trip_times >= late_morning) & (trip_times <= early_afternoon)
    ]
    midday_trips = midday_trips.reset_index(drop=True)
    if len(midday_trips) < 2:
        return False
    for idx in range(1, len(midday_trips)):
        if (midday_trips[idx] - midday_trips[idx - 1]) > pd.Timedelta(hours=3):
            return True
    return False


def calculate_trip_times(group):
    """Calculate first and last trip times, and if there's a long midday gap, separate AM and PM times. Return a series with relevant info."""
    trip_times = group["departure_time"].sort_values()
    first_trip = trip_times.min()
    last_trip = trip_times.max()
    trips_count = len(trip_times)

    if first_trip >= pd.Timedelta(hours=15):
        return pd.Series(
            {
                "first_trip_time": format_timedelta(first_trip),
                "last_trip_time": format_timedelta(last_trip),
                "am_last_trip_time": None,
                "pm_first_trip_time": format_timedelta(first_trip),
                "trips": trips_count,
            }
        )

    if last_trip <= pd.Timedelta(hours=10):
        return pd.Series(
            {
                "first_trip_time": format_timedelta(first_trip),
                "last_trip_time": format_timedelta(last_trip),
                "am_last_trip_time": format_timedelta(last_trip),
                "pm_first_trip_time": None,
                "trips": trips_count,
            }
        )

    if find_large_break(trip_times):
        am_last_trip = trip_times[trip_times < pd.Timedelta(hours=10)].max()
        pm_first_trip = trip_times[trip_times > pd.Timedelta(hours=14)].min()
        return pd.Series(
            {
                "first_trip_time": format_timedelta(first_trip),
                "last_trip_time": format_timedelta(last_trip),
                "am_last_trip_time": format_timedelta(am_last_trip),
                "pm_first_trip_time": format_timedelta(pm_first_trip),
                "trips": trips_count,
            }
        )

    return pd.Series(
        {
            "first_trip_time": format_timedelta(first_trip),
            "last_trip_time": format_timedelta(last_trip),
            "am_last_trip_time": None,
            "pm_first_trip_time": None,
            "trips": trips_count,
        }
    )


def calculate_headways(departure_times):
    """Given all departure_times for a route/direction/time_block, calculate and return the most common (mode) headway in minutes."""
    sorted_times = departure_times.sort_values()
    headway_values = (
        sorted_times.diff().dropna().apply(lambda x: x.total_seconds() / 60)
    )
    if headway_values.empty:
        return None
    return headway_values.mode()[0]


def process_headways(merged_data):
    """Group data by route/direction/time_block and apply calculate_headways, then store them in a dict."""
    headways = (
        merged_data.groupby(
            ["route_short_name", "route_long_name", "direction_id", "time_block"]
        )["departure_time"]
        .apply(calculate_headways)
        .reset_index()
    )

    headway_dict = {
        "weekday_am_headway": {},
        "weekday_midday_headway": {},
        "weekday_pm_headway": {},
        "weekday_night_headway": {},
    }

    for _, row in headways.iterrows():
        key = (row["route_short_name"], row["route_long_name"], row["direction_id"])
        block = row["time_block"]
        val = row["departure_time"]
        if block == "am":
            headway_dict["weekday_am_headway"][key] = val
        elif block == "midday":
            headway_dict["weekday_midday_headway"][key] = val
        elif block == "pm":
            headway_dict["weekday_pm_headway"][key] = val
        elif block == "night":
            headway_dict["weekday_night_headway"][key] = val
    return headway_dict


def merge_headways(trip_times_df, headway_dict):
    """Merge the computed headways from process_headways() back into the trip_times_df."""

    def get_val_from_dict(row, block_type):
        return headway_dict[block_type].get(
            (row["route_short_name"], row["route_long_name"], row["direction_id"]), None
        )

    trip_times_df["weekday_am_headway"] = trip_times_df.apply(
        lambda r: get_val_from_dict(r, "weekday_am_headway"), axis=1
    )
    trip_times_df["weekday_midday_headway"] = trip_times_df.apply(
        lambda r: get_val_from_dict(r, "weekday_midday_headway"), axis=1
    )
    trip_times_df["weekday_pm_headway"] = trip_times_df.apply(
        lambda r: get_val_from_dict(r, "weekday_pm_headway"), axis=1
    )
    trip_times_df["weekday_night_headway"] = trip_times_df.apply(
        lambda r: get_val_from_dict(r, "weekday_night_headway"), axis=1
    )
    return trip_times_df


def save_to_excel(final_data, output_dir, output_file):
    """Save the final_data DataFrame to an Excel file, auto-sizing columns and centering text in each cell."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Route_Schedule_Headway"

    headers = final_data.columns.tolist()
    worksheet.append(headers)

    for row in final_data.itertuples(index=False, name=None):
        worksheet.append(row)

    for col_cells in worksheet.columns:
        max_length = (
            max(len(str(cell.value)) if cell.value else 0 for cell in col_cells) + 2
        )
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = max_length
        for cell in col_cells:
            cell.alignment = Alignment(horizontal="center")

    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, output_file)
    workbook.save(file_path)
    print(f"Final data successfully saved to {file_path}")


def build_coverage_polygons(gtfs_data, route_filter_out, label):
    """Build combined coverage polygon based on buffer for each route.

    For each route, gather all stops used by that route,
    buffer them 0.25 miles, dissolve into a single coverage polygon.

    Returns a dictionary: coverage_polygons[route_short_name] = polygon.

    Also, exports a single shapefile per signup (named with the label)
    containing all route coverages.

    Routes in route_filter_out are skipped.
    """
    if "stops" not in gtfs_data:
        print("No stops.txt found, cannot build coverage polygons.")
        return {}

    stops_df = gtfs_data["stops"].copy()
    if not {"stop_id", "stop_lat", "stop_lon"}.issubset(stops_df.columns):
        print("stops.txt is missing stop_id, stop_lat, or stop_lon columns.")
        return {}

    routes_df = gtfs_data["routes"]
    trips_df = gtfs_data["trips"]
    stop_times_df = gtfs_data["stop_times"]

    trips_with_routes = pd.merge(
        trips_df[["trip_id", "route_id"]],
        routes_df[["route_id", "route_short_name"]],
        on="route_id",
        how="left",
    )
    stimes_merged = pd.merge(
        stop_times_df[["trip_id", "stop_id"]],
        trips_with_routes[["trip_id", "route_short_name"]],
        on="trip_id",
        how="left",
    )
    if route_filter_out:
        stimes_merged = stimes_merged[
            ~stimes_merged["route_short_name"].isin(route_filter_out)
        ]
    full_merged = pd.merge(
        stimes_merged,
        stops_df[["stop_id", "stop_lat", "stop_lon"]],
        on="stop_id",
        how="left",
    ).dropna(subset=["route_short_name", "stop_lat", "stop_lon"])

    # Convert to GeoDataFrame (assume GTFS stops are in EPSG:4326)
    gdf = gpd.GeoDataFrame(
        full_merged,
        geometry=gpd.points_from_xy(full_merged["stop_lon"], full_merged["stop_lat"]),
        crs="EPSG:4326",
    )

    # Reproject to user-defined CRS
    gdf = gdf.to_crs(GEO_CRS)

    coverage_polygons = {}
    grouped = gdf.groupby("route_short_name")
    for route_name, group_df in grouped:
        # Buffer each point: convert 0.25 miles to meters (~1609.34m per mile)
        buffer_dist = STOP_BUFFER_DISTANCE_MILES * 1609.34
        buffered = [pt.buffer(buffer_dist) for pt in group_df.geometry]
        union_poly = unary_union(buffered)
        coverage_polygons[route_name] = union_poly

    # Create a GeoDataFrame from the coverage_polygons dict
    routes_list = []
    polys_list = []
    for route_name, poly in coverage_polygons.items():
        routes_list.append(route_name)
        polys_list.append(poly)
    coverage_gdf = gpd.GeoDataFrame(
        {"route_short_name": routes_list, "geometry": polys_list}, crs=GEO_CRS
    )
    out_folder = os.path.join(OUTPUT_PATH, "coverages")
    os.makedirs(out_folder, exist_ok=True)
    out_path = os.path.join(out_folder, f"{label}_coverages.shp")
    coverage_gdf.to_file(out_path)
    print(f"Saved coverage shapefile for {label} to {out_path}")

    return coverage_polygons


def process_schedule_type(schedule_type, days, data, label):
    """Process a single schedule type, save final results, and export Excel file."""
    calendar_df = data["calendar"]
    trips_df = data["trips"]
    routes_df = data["routes"]
    stop_times_df = data["stop_times"]

    print(f"Processing schedule: {schedule_type} for {label}")

    # 1. Create mask for services running on specified days
    mask = pd.Series([True] * len(calendar_df))
    for day in days:
        mask &= calendar_df[day] == 1
    relevant_service_ids = calendar_df[mask]["service_id"]
    if relevant_service_ids.empty:
        print(f"No services found for {schedule_type}. Skipping.\n")
        return

    # 2. Filter trips
    trips_filtered = trips_df[trips_df["service_id"].isin(relevant_service_ids)]
    if trips_filtered.empty:
        print(f"No trips found for {schedule_type}. Skipping.\n")
        return

    # 3. Merge routes & trip info
    trip_info = trips_filtered[
        ["trip_id", "route_id", "service_id", "direction_id", "block_id"]
    ].merge(
        routes_df[["route_id", "route_short_name", "route_long_name"]],
        on="route_id",
        how="left",
    )
    if ROUTE_FILTER_OUT:
        trip_info = trip_info[~trip_info["route_short_name"].isin(ROUTE_FILTER_OUT)]
    if trip_info.empty:
        print(
            f"All routes filtered out or no data for {schedule_type} ({label}). Skipping.\n"
        )
        return
    print(
        f"Merged trip info has {len(trip_info)} records for {schedule_type} ({label}).\n"
    )

    # 4. Merge with stop_times
    merged_data = pd.merge(
        stop_times_df[["trip_id", "departure_time", "stop_sequence"]],
        trip_info,
        on="trip_id",
        how="inner",
    )
    print(
        f"Merged data has {len(merged_data)} records for {schedule_type} ({label}).\n"
    )

    # 5. Filter to starting stops
    merged_data = merged_data[merged_data["stop_sequence"] == 1]
    print(
        f"Filtered starting trips count: {len(merged_data)} for {schedule_type} ({label})\n"
    )
    if merged_data.empty:
        return

    # 6. Convert departure_time to timedelta
    merged_data["departure_time"] = pd.to_timedelta(
        merged_data["departure_time"], errors="coerce"
    )
    merged_data.dropna(subset=["departure_time"], inplace=True)
    if merged_data.empty:
        return

    # 7. Assign time blocks
    time_blocks = parse_time_blocks(time_blocks_config)
    merged_data["time_block"] = merged_data["departure_time"].apply(
        lambda x: assign_time_block(x, time_blocks)
    )
    merged_data = merged_data[merged_data["time_block"] != "other"]
    print(
        f"Trips after filtering 'other': {len(merged_data)} for {schedule_type} ({label})\n"
    )
    if merged_data.empty:
        return

    # 8. Determine interlined routes via block_id
    block_to_routes = (
        trip_info.groupby("block_id")["route_short_name"]
        .apply(lambda s: set(s.dropna()))
        .to_dict()
    )
    interlined_routes_map = {}
    for block_id, route_set in block_to_routes.items():
        for route_name in route_set:
            # NOTE: rename 'blk' -> 'block_id' or use _blk if truly unused
            interlined_routes_map.setdefault(route_name, set()).update(
                route_set - {route_name}
            )

    # 9. Group by route/direction and calculate trip times
    trip_times = (
        merged_data.groupby(["route_short_name", "route_long_name", "direction_id"])
        .apply(calculate_trip_times)
        .reset_index()
    )

    # 10. Calculate headways
    headway_dict = process_headways(merged_data)

    # 11. Merge headways
    final_data = merge_headways(trip_times, headway_dict)

    # 12. Add interlined_routes column
    def get_interlined(route_name):
        """Build a comma-separated string of routes that share the same block_id."""
        return ", ".join(sorted(interlined_routes_map.get(route_name, [])))

    final_data["interlined_routes"] = final_data["route_short_name"].apply(
        get_interlined
    )

    # 13. Save to Excel
    safe_label = label.replace(" ", "_")
    out_excel = f"{safe_label}_{schedule_type}_{OUTPUT_EXCEL}"
    save_to_excel(final_data, OUTPUT_PATH, out_excel)

    # 14. Store final_data for later comparison
    ALL_SIGNUP_FINAL_DATA.setdefault(label, []).append(final_data)


def process_gtfs_dataset(gtfs_path, label):
    """Load, process, and export GTFS data for a given signup label."""
    if not gtfs_path.strip():
        print(f"Empty path provided for {label}. Skipping.")
        return

    print(f"Checking input files for {label} GTFS...")
    check_input_files(gtfs_path, gtfs_files)
    print("All input files are present.\n")

    print(f"Loading GTFS data for {label}...")
    data = load_gtfs_data(gtfs_path, gtfs_files)
    print("GTFS data loaded.\n")

    # Build coverage polygons and export a single shapefile for this signup.
    coverage_polygons = build_coverage_polygons(data, ROUTE_FILTER_OUT, label)
    ALL_SIGNUP_COVERAGES[label] = coverage_polygons

    # Process each schedule type.
    for sch_type, days in schedule_types.items():
        process_schedule_type(sch_type, days, data, label)


def build_route_signatures_for_signup(final_data_list, coverage_dict):
    """Combine schedule-based data and coverage geometry for each route into a dict."""
    combined_df = (
        pd.concat(final_data_list, ignore_index=True)
        if final_data_list
        else pd.DataFrame()
    )
    route_signatures = {}

    if not combined_df.empty:
        grouped = combined_df.groupby("route_short_name")
        for route_name, group_df in grouped:
            all_interlined = set()
            for interlined_str in group_df["interlined_routes"].dropna():
                splitted = [s.strip() for s in interlined_str.split(",") if s.strip()]
                all_interlined.update(splitted)
            interlined_str = ", ".join(sorted(all_interlined))

            skip_cols = {"route_short_name", "interlined_routes"}
            tmp = group_df.sort_values(by=["direction_id", "route_long_name"]).fillna(
                ""
            )
            keep_cols = [col for col in tmp.columns if col not in skip_cols]
            row_strings = []
            for row in tmp[keep_cols].itertuples(index=False, name=None):
                row_str = "|".join(str(x) for x in row)
                row_strings.append(row_str)
            others_str = "\n".join(row_strings)

            geom = coverage_dict.get(route_name, None)
            route_signatures[route_name] = {
                "interlining": interlined_str,
                "others": others_str,
                "geometry": geom,
            }
    else:
        # If there's no final_data at all, still store geometry if available
        for route_name, poly in coverage_dict.items():
            route_signatures[route_name] = {
                "interlining": "",
                "others": "",
                "geometry": poly,
            }

    return route_signatures


def classify_geometry_change(poly_old, poly_new, threshold=GEOM_CHANGE_THRESHOLD):
    """Compare coverage polygons to see if the route's coverage is expanded/contracted/modified."""
    if (poly_old is None or poly_old.is_empty) and (poly_new and not poly_new.is_empty):
        return "Geography expanded"
    if (poly_new is None or poly_new.is_empty) and (poly_old and not poly_old.is_empty):
        return "Geography contracted"
    if (poly_old is None or poly_old.is_empty) and (
        poly_new is None or poly_new.is_empty
    ):
        return "No geographic change"

    area_old = poly_old.area
    area_new = poly_new.area
    if area_old == 0 and area_new == 0:
        return "No geographic change"
    diff = area_new - area_old
    if area_old > 0:
        frac = diff / area_old
    else:
        return "Geography expanded"

    if frac > threshold:
        return "Geography expanded"
    if frac < -threshold:
        return "Geography contracted"

    sym_diff_area = (poly_old ^ poly_new).area
    if sym_diff_area > 1e-9:
        return "Geography modified"

    return "No geographic change"


def compare_signups_detailed(labels_in_order, all_signups_data):
    """Build a DataFrame comparing each signup to the previous one, labeling each route."""
    # Gather all routes from all signups
    all_routes = set()
    for lab in labels_in_order:
        all_routes.update(all_signups_data[lab].keys())
    all_routes = sorted(all_routes)

    # Prepare the output DataFrame: one column per signup, one row per route
    comparison_df = pd.DataFrame(index=all_routes, columns=labels_in_order, dtype=str)

    # For the first signup, there's no previous version => label everything "No change"
    if labels_in_order:
        first_label = labels_in_order[0]
        for rt in all_routes:
            comparison_df.loc[rt, first_label] = "No change"

    # Compare each subsequent signup to the previous
    for idx in range(1, len(labels_in_order)):
        prev_lab = labels_in_order[idx - 1]
        curr_lab = labels_in_order[idx]

        prev_data = all_signups_data[prev_lab]
        curr_data = all_signups_data[curr_lab]

        for rt in all_routes:
            changes = []

            in_prev = rt in prev_data
            in_curr = rt in curr_data

            # 1) Route eliminated
            if in_prev and not in_curr:
                changes.append("Route eliminated")

            # 2) Route created
            elif not in_prev and in_curr:
                changes.append("Route created")

            # 3) If route is in both signups, check geometry/interlining/schedule
            elif in_prev and in_curr:
                poly_old = prev_data[rt]["geometry"]
                poly_new = curr_data[rt]["geometry"]
                inter_old = prev_data[rt]["interlining"]
                inter_new = curr_data[rt]["interlining"]
                others_old = prev_data[rt]["others"]
                others_new = curr_data[rt]["others"]

                # Geometry comparison
                geom_change = classify_geometry_change(poly_old, poly_new)
                if geom_change != "No geographic change":
                    changes.append(geom_change)

                # Interlining change
                if inter_old != inter_new:
                    changes.append("Interlining change")

                # Schedule or "others" change
                if others_old != others_new:
                    changes.append("Other change")

            # If we found no changes, label "No change"
            if not changes:
                comparison_df.loc[rt, curr_lab] = "No change"
            else:
                comparison_df.loc[rt, curr_lab] = ", ".join(changes)

    return comparison_df


def save_comparison_to_excel(comparison_df, output_path, filename):
    """Save the final route-by-route comparison DataFrame to an Excel file."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Comparison"

    headers = ["route_short_name"] + comparison_df.columns.tolist()
    worksheet.append(headers)

    for route_name in comparison_df.index:
        row_values = [route_name] + comparison_df.loc[route_name].tolist()
        worksheet.append(row_values)

    for col_cells in worksheet.columns:
        max_length = (
            max(len(str(cell.value)) if cell.value else 0 for cell in col_cells) + 2
        )
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = max_length
        for cell in col_cells:
            cell.alignment = Alignment(horizontal="center")

    os.makedirs(output_path, exist_ok=True)
    file_path = os.path.join(output_path, filename)
    workbook.save(file_path)
    print(f"Comparison data saved to {file_path}")


# ==================================================================================================
# MAIN
# ==================================================================================================


def main():
    """Main entry point orchestrates script."""
    # Process each GTFS config in order
    for cfg in MULTIPLE_GTFS_CONFIGS:
        label = cfg["name"]
        path = cfg["path"]
        print(f"\n========== Processing {label} ========== \n")
        try:
            process_gtfs_dataset(path, label)
        except (
            Exception
        ) as exc:  # NOTE: This is broad, but often necessary for top-level
            print(f"Error processing {label}: {exc}")

    # Build route-level signatures (including coverage geometry) for each signup
    all_signups_data = {}
    signup_labels = list(ALL_SIGNUP_FINAL_DATA.keys()) + list(
        ALL_SIGNUP_COVERAGES.keys()
    )
    for label in set(signup_labels):
        final_data_list = ALL_SIGNUP_FINAL_DATA.get(label, [])
        coverage_dict = ALL_SIGNUP_COVERAGES.get(label, {})
        route_signs = build_route_signatures_for_signup(final_data_list, coverage_dict)
        all_signups_data[label] = route_signs

    # Compare signups in chronological order
    labels_in_order = [cfg["name"] for cfg in MULTIPLE_GTFS_CONFIGS]
    comparison_df = compare_signups_detailed(labels_in_order, all_signups_data)
    save_comparison_to_excel(comparison_df, OUTPUT_PATH, COMPARISON_EXCEL)

    print("\nAll GTFS sets processed successfully!")
    print("Detailed service change comparison file has been generated.")
    print("Coverage shapefiles exported to coverages/ subfolder.")


if __name__ == "__main__":
    main()
