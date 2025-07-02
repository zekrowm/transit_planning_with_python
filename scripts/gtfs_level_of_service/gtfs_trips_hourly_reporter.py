"""Generate Excel reports of trip counts per time interval from a GTFS feed.

The module loads GTFS text files, filters and aggregates trips by route,
direction, service, and user-defined time bins (e.g., hourly), then
exports the results to one Excel workbook per
*route × direction × service* combination.

Typical usage is in ArcGIS Pro or Jupyter.

The constants in the *CONFIGURATION* block control input paths, routes,
directions, time-bin width, and service filtering.
"""

import datetime as dt
import logging
import os
from typing import Any, Mapping, Optional  # noqa: ANN401  (re-exported for typing)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

# Input directory containing GTFS files
BASE_INPUT_PATH = r"\\Your\Folder\Path\Here\\"

# Output directory for the Excel file
BASE_OUTPUT_PATH = r"\\Your\Folder\Path\Here\\"

# GTFS files to load
GTFS_FILES = [
    "trips.txt",
    "stop_times.txt",
    "routes.txt",
    "calendar.txt",
    # 'stops.txt',  # Uncomment if needed
]

# Routes and directions to process
ROUTE_DIRECTIONS = [
    {
        "route_short_name": "101",
        "direction_id": 0,
    },  # Process only direction 0 for route 101
    {
        "route_short_name": "202",
        "direction_id": None,
    },  # Process all directions for route 202
]

# New Configuration: Time Interval in Minutes
TIME_INTERVAL_MINUTES = 60  # Users can change this to 30, 15, etc.

# Choose one service to analyse (None → run each service separately)
SERVICE_ID = "3"  # Replace with your desired service_id value from calendar.txt

# =============================================================================
# FUNCTIONS
# =============================================================================


def fix_time_format(
    time_str: str | pd.NA | None,
) -> str | pd.NA | None:
    """Normalize GTFS HH:MM:SS strings.

    * Adds a leading zero when the hour has one digit.
    * Converts post-24 h times (e.g. ``25:15:00``) to 24-hour rollover
      by subtracting 24 from the hour component.

    Args:
        time_str: Raw time string from ``stop_times.txt`` or
            :pydata:`pandas.NA`.

    Returns:
        The cleaned time string, or ``pd.NA`` if *time_str* is null-like.
    """
    if pd.isna(time_str):
        return time_str
    parts = time_str.split(":")

    # Add leading zero if the hour is a single digit
    if len(parts[0]) == 1:
        parts[0] = "0" + parts[0]

    # Correct times where the hour exceeds 23 (indicating next day service)
    if int(parts[0]) >= 24:
        parts[0] = str(int(parts[0]) - 24).zfill(2)

    return ":".join(parts)


def get_time_bin(
    t: dt.time,
    interval: int,
) -> str:
    """Map a :class:`datetime.time` to a formatted time-bin label.

    Args:
        t: The time of day to classify.
        interval: Width of each time bin in minutes
            (must divide 1 440 evenly).

    Returns:
        A label such as ``"08:00-08:59"``.
    """
    total_minutes = t.hour * 60 + t.minute
    bin_start = (total_minutes // interval) * interval
    bin_end = bin_start + interval - 1
    if bin_end >= 1440:
        bin_end -= 1440  # Wrap around if necessary
    start_hour, start_min = divmod(bin_start, 60)
    end_hour, end_min = divmod(bin_end, 60)
    return f"{str(start_hour).zfill(2)}:{str(start_min).zfill(2)}-{str(end_hour).zfill(2)}:{str(end_min).zfill(2)}"


def process_and_export(
    data: dict[str, pd.DataFrame],
    route_dirs: list[dict],
    output_path: str,
    interval_minutes: int,
    service_id: str | None = None,
) -> None:
    """Create “trips-per-time-bin” Excel workbooks from a GTFS feed.

    Args:
        data: Output of :func:`load_gtfs_data`.
        route_dirs: Iterable of route/direction filters, e.g.
            ``[{"route_short_name": "101", "direction_id": 0}, …]``.
        output_path: Folder in which the Excel files will be written. The
            directory is created if it does not exist.
        interval_minutes: Width of each time bin (e.g. ``60`` for hourly).
        service_id: If supplied, restrict processing to that
            ``service_id``; otherwise iterate over every ``service_id``
            present in the feed.

    Returns:
        None. Side-effect: Excel files are written to *output_path*.

    Raises:
        ValueError: If *interval_minutes* does not evenly divide 1 440.
    """
    trips = data["trips"]
    stop_times = data["stop_times"]
    routes = data["routes"]

    # ── Apply service filter ────────────────────────────────────────────────
    if service_id is not None:
        trips_filtered = trips[trips["service_id"] == str(service_id)]
        logging.info("Analysing only service_id=%s (%d trips).", service_id, len(trips_filtered))
    else:
        trips_filtered = trips.copy()
        logging.info("No service_id specified – will iterate over each one.")

    # ── Merge the required tables ───────────────────────────────────────────
    merged = stop_times.merge(trips_filtered, on="trip_id", how="inner").merge(
        routes[["route_id", "route_short_name"]], on="route_id", how="left"
    )

    # Cast key fields to numeric so comparisons work (dtype=str in loader)
    merged["direction_id"] = pd.to_numeric(merged["direction_id"], errors="coerce")
    merged["stop_sequence"] = pd.to_numeric(merged["stop_sequence"], errors="coerce")

    # ── Fix & parse time columns ────────────────────────────────────────────
    for col in ("arrival_time", "departure_time"):
        merged[col] = (
            merged[col]
            .apply(fix_time_format)
            .str.strip()
            .pipe(pd.to_datetime, format="%H:%M:%S", errors="coerce")
            .dt.time
        )

    # ── Pre-compute full list of time-bin labels ────────────────────────────
    time_bins = [
        f"{h:02d}:{m:02d}-{(h + (m + interval_minutes - 1) // 60) % 24:02d}:"
        f"{(m + interval_minutes - 1) % 60:02d}"
        for h in range(24)
        for m in range(0, 60, interval_minutes)
    ]

    # ── Loop over requested routes/directions ───────────────────────────────
    for rd in route_dirs:
        r_short = rd["route_short_name"]
        d_id = rd["direction_id"]

        sel = merged[merged["route_short_name"] == r_short]
        if d_id is not None:
            sel = sel[sel["direction_id"] == d_id]

        # first stop of each trip
        starts = sel[sel["stop_sequence"] == 1].dropna(subset=["departure_time"])

        # Determine which service_ids to iterate over
        sid_iter = [service_id] if service_id is not None else starts["service_id"].unique()

        for sid in sid_iter:
            cur = starts if service_id is not None else starts[starts["service_id"] == sid]

            if cur.empty:
                logging.info(
                    "No trips for route=%s dir=%s service=%s – skipping.",
                    r_short,
                    d_id,
                    sid,
                )
                continue

            cur = cur.assign(
                time_bin=cur["departure_time"].apply(lambda t: get_time_bin(t, interval_minutes))
            )

            trips_per_bin = (
                cur.groupby("time_bin")
                .size()
                .reindex(time_bins, fill_value=0)
                .reset_index()
                .rename(columns={0: "trip_count"})
            )

            _export_to_excel(
                trips_per_bin,
                output_path,
                interval_minutes,
                r_short,
                d_id,
                service_id=sid if service_id is None else service_id,
            )


def _export_to_excel(
    df: pd.DataFrame,
    output_path: str,
    interval_minutes: int,
    route_short: str,
    direction_id: int | None = None,
    service_id: str | None = None,
) -> None:
    """Write a single *trips-per-bin* table to an Excel workbook.

    Args:
        df: Two-column DataFrame with ``time_bin`` and ``trip_count``.
        output_path: Destination directory (created if necessary).
        interval_minutes: Width of each time bin, replicated in the file name.
        route_short: ``route_short_name`` used in the file/work-sheet title.
        direction_id: ``direction_id`` (``0`` or ``1``) or ``None`` for all.
        service_id: ``service_id`` or ``None`` if the feed was pre-filtered.

    Returns:
        None. The workbook is saved to *output_path*.

    Notes:
        This helper is intentionally private; callers should invoke
        :func:`process_and_export` instead of calling this function directly.
    """
    wb = Workbook()
    ws = wb.active

    # Create a title reflecting route/direction/service
    if service_id:
        if direction_id is not None:
            ws.title = f"Service_{service_id}_Route_{route_short}_Dir_{direction_id}"
            file_name = (
                f"Trips_Per_{interval_minutes}Min_"
                f"Service_{service_id}_Route_{route_short}_Dir_{direction_id}.xlsx"
            )
        else:
            ws.title = f"Service_{service_id}_Route_{route_short}_All_Dirs"
            file_name = (
                f"Trips_Per_{interval_minutes}Min_"
                f"Service_{service_id}_Route_{route_short}_All_Directions.xlsx"
            )
    else:
        if direction_id is not None:
            ws.title = f"Route_{route_short}_Dir_{direction_id}"
            file_name = (
                f"Trips_Per_{interval_minutes}Min_Route_{route_short}_Dir_{direction_id}.xlsx"
            )
        else:
            ws.title = f"Route_{route_short}_All_Directions"
            file_name = f"Trips_Per_{interval_minutes}Min_Route_{route_short}_All_Directions.xlsx"

    # Write headers
    ws.append(df.columns.tolist())

    # Write data rows
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    # Adjust column widths and alignments
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col) + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length
        for cell in col:
            cell.alignment = Alignment(horizontal="center")

    # Ensure output directory exists
    os.makedirs(output_path, exist_ok=True)

    # Save the workbook
    output_file = os.path.join(output_path, file_name)
    wb.save(output_file)
    logging.info(
        "Trips per %d minutes for %s successfully exported!",
        interval_minutes,
        file_name,
    )


# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[list[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into a dictionary of DataFrames.

    Args:
        gtfs_folder_path (str): Absolute or relative path to the directory
            containing GTFS text files.
        files (list[str] | None): Explicit list of GTFS filenames to load.
            If ``None``, the full standard GTFS set is read.
        dtype (str | Mapping[str, Any]): Value forwarded to
            :pyfunc:`pandas.read_csv` to control column dtypes;
            defaults to ``str``.

    Returns:
        dict[str, pandas.DataFrame]: Mapping of file stem → DataFrame.
        For example, ``data["trips"]`` contains *trips.txt*.

    Raises:
        OSError: The folder does not exist or a required file is missing.
        ValueError: A file is empty or malformed.
        RuntimeError: An OS-level error occurs while reading.
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = [
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        ]

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info(f"Loaded {file_name} ({len(df)} records).")

        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc
    return data


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """Run the end-to-end pipeline from CLI or “Run ▶” button.

    1. Configure logging.
    2. Validate the time-interval width.
    3. Load GTFS tables via :func:`load_gtfs_data`.
    4. Generate and export reports via :func:`process_and_export`.

    Any uncaught exception is logged with full traceback before
    the script exits with a non-zero status.
    """
    # ─── Logging setup (moved here) ───────────────────────────────────────────
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  [%(levelname)s]  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        force=True,  # overrides any library defaults
    )
    # ─────────────────────────────────────────────────────────────────────────

    try:
        # Validate interval
        if TIME_INTERVAL_MINUTES <= 0 or 1440 % TIME_INTERVAL_MINUTES:
            raise ValueError("TIME_INTERVAL_MINUTES must divide evenly into 1,440.")

        # Use the *new* loader (it will raise if files/dir are missing)
        data = load_gtfs_data(
            BASE_INPUT_PATH,
            files=GTFS_FILES,
            dtype=str,  # keep everything as strings – avoids TZ parsing issues
        )

        # Run the core logic
        process_and_export(
            data,
            ROUTE_DIRECTIONS,
            BASE_OUTPUT_PATH,
            TIME_INTERVAL_MINUTES,
            service_id=SERVICE_ID,  # NEW ARG
        )

    except Exception as exc:  # catch-all so the stack-trace hits your log
        logging.exception("Fatal error: %s", exc)


if __name__ == "__main__":
    main()
