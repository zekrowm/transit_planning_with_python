"""Compute route-level transit KPIs from GTFS feed.

Designed for ArcGIS Pro or notebook use. Identifies interlined routes and
outputs Excel reports by schedule type (Weekday, Saturday, Sunday).

Outputs:
    - One Excel file per schedule type with calculated KPIs for each route.
"""

from __future__ import annotations

import logging
from datetime import timedelta
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_INPUT_PATH = Path(r"\\your_folder_path\here\\")
OUTPUT_PATH = Path(r"\\your_folder_path\here\\")

GTFS_FILES: List[str] = [
    "routes.txt",
    "trips.txt",
    "stop_times.txt",
    "calendar.txt",
    "calendar_dates.txt",
]

OUTPUT_EXCEL = "route_schedule_headway_with_modes.xlsx"

TIME_BLOCKS_CONFIG: Mapping[str, Tuple[str, str]] = {
    "AM Early": ("04:00", "05:59"),
    "AM Peak": ("06:00", "08:59"),
    "Midday": ("09:00", "14:59"),
    "PM Peak": ("15:00", "17:59"),
    "PM Late": ("18:00", "20:59"),
    "PM Nite": ("21:00", "23:59"),
    "Other": ("24:00", "27:59"),
}

SCHEDULE_TYPES: Mapping[str, List[str]] = {
    "Weekday": ["monday", "tuesday", "wednesday", "thursday", "friday"],
    "Saturday": ["saturday"],
    "Sunday": ["sunday"],
}

FAKE_ROUTES: List[str] = ["9999A", "9999B", "9999C"]

# One of {"meters", "feet", "miles"}
DISTANCE_UNIT = "meters"

# =============================================================================
# FUNCTIONS
# =============================================================================
def distance_conversion_factor(unit: str) -> float:
    """Return a multiplier to convert *unit* → miles."""
    match unit.lower():
        case "meters":
            return 1 / 1609.344
        case "feet":
            return 1 / 5280.0
        case "miles":
            return 1.0
        case _:
            raise ValueError(
                f"Unknown distance unit '{unit}'. Valid: meters, feet, miles."
            )


def check_input_files(base_path: Path, files: List[str]) -> None:
    """Ensure *base_path* exists and contains each file in *files*."""
    if not base_path.exists():
        raise FileNotFoundError(f"Input directory {base_path} does not exist.")
    missing = [f for f in files if not (base_path / f).exists()]
    if missing:
        raise FileNotFoundError(f"Missing GTFS file(s): {', '.join(missing)}")


def load_gtfs_data(base_path: Path, files: List[str]) -> Dict[str, pd.DataFrame]:
    """Load each GTFS text file into a DataFrame dict keyed by stem."""
    data: Dict[str, pd.DataFrame] = {}
    for fname in files:
        path = base_path / fname
        stem = path.stem
        df = pd.read_csv(path, low_memory=False)
        data[stem] = df
        LOGGER.info("Loaded %-20s | %6d rows", fname, len(df))
    return data


def parse_time_blocks(
    tb_config: Mapping[str, Tuple[str, str]]
) -> Dict[str, Tuple[pd.Timedelta, pd.Timedelta]]:
    """Convert HH:MM strings in *tb_config* to timedelta bounds."""
    parsed: Dict[str, Tuple[pd.Timedelta, pd.Timedelta]] = {}
    for name, (start, end) in tb_config.items():
        h1, m1 = map(int, start.split(":"))
        h2, m2 = map(int, end.split(":"))
        parsed[name] = (timedelta(hours=h1, minutes=m1), timedelta(hours=h2, minutes=m2))
    return parsed


def format_timedelta(td: Optional[pd.Timedelta]) -> Optional[str]:
    """Return 'HH:MM' for *td*, or None if *td* is NaT/None."""
    if td is None or pd.isna(td):
        return None
    total_secs = int(td.total_seconds())
    hours, remainder = divmod(total_secs, 3600)
    minutes = remainder // 60
    return f"{hours:02d}:{minutes:02d}"


def modal_headway(departures: pd.Series) -> Optional[float]:
    """Compute the modal headway in minutes from a series of datetimes."""
    sorted_dep = departures.sort_values()
    diffs = sorted_dep.diff().dropna().apply(lambda x: x.total_seconds() / 60.0)
    return None if diffs.empty else diffs.mode()[0]


def save_df_to_excel(
    df: pd.DataFrame, out_path: Path, sheet_name: str = "Route_Schedule_Headway"
) -> None:
    """Write *df* to an Excel workbook at *out_path* with basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(df.columns.to_list())
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = max_len + 3
        for cell in col_cells:
            cell.alignment = Alignment(horizontal="center")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    LOGGER.info("Saved %s", out_path)


def assign_time_block(
    td: pd.Timedelta, blocks: Mapping[str, Tuple[pd.Timedelta, pd.Timedelta]]
) -> str:
    """Return the time-block name for *td*, or 'other' if none match."""
    for name, (start, end) in blocks.items():
        if start <= td <= end:
            return name
    return "other"


def build_interlining_map(
    trips: pd.DataFrame, routes: pd.DataFrame
) -> Dict[str, set]:
    """Return route_short_name → set of interlined route_short_names."""
    merged = trips.merge(
        routes[["route_id", "route_short_name"]],
        on="route_id",
        how="left",
    )
    block_map = (
        merged.groupby("block_id")["route_short_name"]
        .apply(lambda s: set(s.dropna()))
        .to_dict()
    )
    interlines: Dict[str, set] = {}
    for rt_set in block_map.values():
        for rt in rt_set:
            interlines.setdefault(rt, set()).update(rt_set - {rt})
    return interlines


def validate_trip_bins(row: pd.Series, bin_cols: List[str]) -> None:
    """Warn if sum of *bin_cols* ≠ total_trips (handles duplicate columns)."""
    binned_sum = 0
    for col in bin_cols:
        if col in row.index:
            val = row[col]
            if isinstance(val, pd.Series):
                val = val.iloc[0]
            if pd.notnull(val):
                binned_sum += int(val)
    total = row.get("total_trips")
    if pd.notnull(total) and binned_sum != total:
        LOGGER.warning(
            "Trip bin mismatch | route=%s dir=%s | total=%s bins=%s",
            row.get("route_short_name"),
            row.get("direction_id"),
            int(total),
            binned_sum,
        )


def calculate_trip_metrics(stop_times_sub: pd.DataFrame) -> pd.Series:
    """Return run time (min) and distance (miles) for one trip."""
    arrival = stop_times_sub["arrival_time"].max()
    depart = stop_times_sub["departure_time"].min()

    run_min: Optional[float]
    if pd.notnull(arrival) and pd.notnull(depart):
        run_td = arrival - depart
        run_min = run_td.total_seconds() / 60.0
    else:
        run_min = None

    dist = (
        stop_times_sub["shape_dist_traveled"].max()
        if "shape_dist_traveled" in stop_times_sub.columns
        else None
    )

    return pd.Series({"trip_run_minutes": run_min, "trip_distance_miles": dist})

# -----------------------------------------------------------------------------
# PREPROCESSING & WORKBOOK BUILDERS
# -----------------------------------------------------------------------------

def preprocess(
    data: Dict[str, pd.DataFrame], dist_factor: float
) -> Tuple[
    pd.DataFrame,
    pd.DataFrame,
    Dict[str, set],
    Dict[str, Tuple[pd.Timedelta, pd.Timedelta]],
]:
    """Filter, compute trip metrics, build interlines & parse time blocks."""
    routes = data["routes"].copy()
    trips = data["trips"]
    stop_times = data["stop_times"]

    routes = routes[~routes["route_short_name"].isin(FAKE_ROUTES)]
    LOGGER.info("%d routes remain after removing FAKE_ROUTES", len(routes))

    stop_times["arrival_time"] = pd.to_timedelta(
        stop_times["arrival_time"], errors="coerce"
    )
    stop_times["departure_time"] = pd.to_timedelta(
        stop_times["departure_time"], errors="coerce"
    )

    if "shape_dist_traveled" in stop_times.columns:
        stop_times["shape_dist_traveled"] *= dist_factor
        LOGGER.info("shape_dist_traveled converted to miles (unit=%s).", DISTANCE_UNIT)

    # Trip-level metrics (with group_keys=False to avoid deprecation warning)
    trip_metrics = (
        stop_times.groupby("trip_id", group_keys=False, as_index=False)
        .apply(calculate_trip_metrics)
        .merge(trips[["trip_id", "route_id", "direction_id"]], on="trip_id", how="left")
    )

    agg = {
        "trip_run_minutes": "mean",
        "trip_distance_miles": ["mean", "median"],
    }
    rdm = (
        trip_metrics.groupby(["route_id", "direction_id"], as_index=False)
        .agg(agg)
    )
    rdm.columns = [
        "route_id",
        "direction_id",
        "avg_run_minutes",
        "avg_distance_miles",
        "median_distance_miles",
    ]
    rdm["avg_speed_mph"] = (
        rdm["avg_distance_miles"] / (rdm["avg_run_minutes"] / 60.0)
    )
    # ⚡ Round avg_run_minutes to 1 decimal
    rdm["avg_run_minutes"] = rdm["avg_run_minutes"].round(1)

    rdm = rdm.merge(
        routes[["route_id", "route_short_name", "route_long_name"]],
        on="route_id",
        how="left",
    )

    interlines = build_interlining_map(trips, routes)
    time_blocks = parse_time_blocks(TIME_BLOCKS_CONFIG)

    # Save back modified stop_times for use in schedule builder
    data["stop_times"] = stop_times

    return routes, rdm, interlines, time_blocks


def build_schedule_book(
    sched_name: str,
    weekdays: List[str],
    data: Dict[str, pd.DataFrame],
    routes: pd.DataFrame,
    rdm: pd.DataFrame,
    interlines: Dict[str, set],
    time_blocks: Dict[str, Tuple[pd.Timedelta, pd.Timedelta]],
) -> None:
    """Build and save the Excel workbook for one schedule type."""
    trips = data["trips"]
    stop_times = data["stop_times"]
    calendar = data["calendar"]

    mask = calendar[weekdays].eq(1).all(axis=1)
    svc_ids = set(calendar.loc[mask, "service_id"])

    trips_filt = trips[
        trips["service_id"].isin(svc_ids) & trips["route_id"].isin(routes["route_id"])
    ]
    if trips_filt.empty:
        LOGGER.warning("No trips for %s - skipping", sched_name)
        return

    st_first = stop_times[
        stop_times["trip_id"].isin(trips_filt["trip_id"])
        & (stop_times["stop_sequence"] == 1)
    ].dropna(subset=["departure_time"])

    info = trips_filt[["trip_id", "route_id", "direction_id"]].merge(
        routes[["route_id", "route_short_name", "route_long_name"]],
        on="route_id",
        how="left",
    )
    st_first = st_first.merge(info, on="trip_id", how="left")
    if st_first.empty:
        LOGGER.warning("No first-stop data for %s - skipping", sched_name)
        return

    st_first["time_block"] = st_first["departure_time"].apply(
        lambda td: assign_time_block(td, time_blocks)
    )

    # Headways
    head = (
        st_first.groupby(
            ["route_id", "direction_id", "time_block"], group_keys=False
        )["departure_time"]
        .apply(modal_headway)
        .unstack("time_block")
        .rename(columns=lambda c: f"{c.lower().replace(' ', '_')}_headway")
        .reset_index()
    )

    # Trip counts
    counts = (
        st_first.groupby(
            ["route_id", "direction_id", "time_block"], group_keys=False
        )["trip_id"]
        .nunique()
        .unstack("time_block", fill_value=0)
        .rename(columns=lambda c: f"{c.lower().replace(' ', '_')}_trips")
        .reset_index()
    )

    # First, last, total
    def first_last(grp: pd.DataFrame) -> pd.Series:
        first_d = grp["departure_time"].min()
        last_d = grp["departure_time"].max()
        return pd.Series(
            {
                "first_trip_time": format_timedelta(first_d),
                "last_trip_time": format_timedelta(last_d),
                "total_trips": grp["trip_id"].nunique(),
            }
        )

    spans = (
        st_first.groupby(
            ["route_id", "direction_id"], group_keys=False
        )
        .apply(first_last)
        .reset_index()
    )

    # Merge KPIs
    kpi = (
        spans.merge(head, on=["route_id", "direction_id"], how="left")
        .merge(counts, on=["route_id", "direction_id"], how="left")
        .merge(rdm, on=["route_id", "direction_id"], how="left")
    )

    kpi["interlined_routes"] = kpi["route_short_name"].map(
        lambda r: ", ".join(sorted(interlines.get(r, set())))
    )

    # Service span HH:MM
    kpi["service_span"] = kpi.apply(
        lambda r: format_timedelta(
            pd.to_timedelta(f"{r['last_trip_time']}:00")
            - pd.to_timedelta(f"{r['first_trip_time']}:00")
        )
        if pd.notnull(r["first_trip_time"]) and pd.notnull(r["last_trip_time"])
        else None,
        axis=1,
    )

    # Avg run time HH:MM
    kpi["avg_run_time"] = kpi["avg_run_minutes"].apply(
        lambda m: format_timedelta(timedelta(minutes=m)) if pd.notnull(m) else None
    )

    # Round other numeric columns to 1 decimal
    for col in ("avg_distance_miles", "median_distance_miles", "avg_speed_mph"):
        if col in kpi.columns:
            kpi[col] = kpi[col].round(1)

    # Validate bins
    bin_cols = [f"{tb.lower().replace(' ', '_')}_trips" for tb in TIME_BLOCKS_CONFIG]
    kpi.apply(lambda row: validate_trip_bins(row, bin_cols), axis=1)

    # Reorder columns
    desired = [
        "route_short_name",
        "route_long_name",
        "direction_id",
        "interlined_routes",
        "service_span",
        "first_trip_time",
        "last_trip_time",
        "total_trips",
        *bin_cols,
        *(f"{tb.lower().replace(' ', '_')}_headway" for tb in TIME_BLOCKS_CONFIG),
        "avg_run_minutes",
        "avg_run_time",
        "avg_distance_miles",
        "median_distance_miles",
        "avg_speed_mph",
    ]
    cols = [c for c in desired if c in kpi.columns] + [
        c for c in kpi.columns if c not in desired
    ]
    kpi = kpi[cols]

    save_df_to_excel(kpi, OUTPUT_PATH / f"{sched_name}_{OUTPUT_EXCEL}")


# =============================================================================
# MAIN
# =============================================================================
def main() -> None:
    """Entrypoint."""
    try:
        factor = distance_conversion_factor(DISTANCE_UNIT)
        check_input_files(GTFS_INPUT_PATH, GTFS_FILES)
        data = load_gtfs_data(GTFS_INPUT_PATH, GTFS_FILES)

        routes, rdm, interlines, time_blocks = preprocess(data, factor)

        for sched, days in SCHEDULE_TYPES.items():
            build_schedule_book(
                sched, days, data, routes, rdm, interlines, time_blocks
            )

        LOGGER.info("All schedule types processed successfully.")
    except Exception as exc:  # pylint: disable=broad-except
        LOGGER.exception("Unhandled error: %s", exc)


if __name__ == "__main__":
    main()
