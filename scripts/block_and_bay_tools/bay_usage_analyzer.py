"""Detect transit-cluster and stop-level scheduling conflicts.

The script analyzes block-level transit data to flag times when more
buses are scheduled than a bay or cluster can physically handle.  It
reads the *Step 1* block-level XLSX outputs, calculates capacity
constraints from the user-defined ``CLUSTER_DEFINITIONS``, and writes a
detailed Excel workbook for each cluster showing every event and
highlighting any conflicts.

Typical use is from a Jupyter Notebook or ArcGIS Pro “Python” pane, but
the module can also be invoked directly from the command line.
"""

from __future__ import annotations

import logging
import os
from typing import Dict, List, Optional, Set, Tuple

from typing import cast

import pandas as pd
from openpyxl.styles import Font
from pandas import DataFrame

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

# Folder containing your block-level XLSX files from Step 1
BLOCK_OUTPUT_FOLDER: str = r"\\Path\To\Your\Input_Folder"

# Where to save the cluster-conflict outputs
CLUSTER_CONFLICT_OUTPUT_FOLDER: str = r"\\Path\To\Your\Output_Folder"

# Dictionary of clusters, including single_bay, double_bay, triple_bay, and overflow.
# Each key is the cluster name, the value is a dict with lists:
#   'single_bay_stops' -> capacity 1 each
#   'double_bay_stops' -> capacity 2 each
#   'triple_bay_stops' -> capacity 3 each
#   'overflow_bays'    -> capacity 1 each
CLUSTER_DEFINITIONS: Dict[str, Dict[str, List[str]]] = {
    "Park & Ride": {
        "single_bay_stops": ["3882", "3881"],
        "double_bay_stops": [],
        "triple_bay_stops": [],
        "overflow_bays": [],
    },
    "Metro": {
        "single_bay_stops": ["2373"],
        "double_bay_stops": ["2832"],
        "triple_bay_stops": [],
        "overflow_bays": ["layover_bay_A", "layover_bay_B"],
    },
}

# Define which statuses indicate bus presence in the cluster
PRESENCE_STATUSES: Set[str] = {
    "ARRIVE",
    "DEPART",
    "ARRIVE/DEPART",
    "DWELL",
    "LOADING",
    "LAYOVER",
    "LONG BREAK",
}

# Define which statuses indicate the bus is physically occupying a stop bay
PASSENGER_SERVICE_STATUSES: Set[str] = {
    "ARRIVE",
    "DEPART",
    "ARRIVE/DEPART",
    "LOADING",
}

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def get_all_official_stops(cinfo: Dict[str, List[str]]) -> List[str]:
    """Return a combined list of all *official* stops in a cluster.

    Regardless of whether the stop is single-, double-, or triple-bay,
    overflow bays are *excluded* here because they do not count toward
    the “official” set used in certain analyses.

    Args:
    ----
    cinfo :
        The cluster-definition dictionary for a single cluster.

    Returns:
    -------
    list[str]
        Concatenated list of all official stop IDs.
    """
    return (
        cinfo.get("single_bay_stops", [])
        + cinfo.get("double_bay_stops", [])
        + cinfo.get("triple_bay_stops", [])
    )


def build_cluster_capacities() -> Dict[str, int]:
    """Compute total bay-capacity for every cluster.

    Capacity is calculated as::

        capacity = (#single * 1) + (#double * 2) + (#triple * 3) + (#overflow * 1)

    Returns:
    -------
    dict[str, int]
        Mapping of ``cluster_name -> capacity`` (number of buses that
        can be simultaneously present).
    """
    capacities: Dict[str, int] = {}
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        n_single = len(cinfo.get("single_bay_stops", []))
        n_double = len(cinfo.get("double_bay_stops", []))
        n_triple = len(cinfo.get("triple_bay_stops", []))
        n_overflow = len(cinfo.get("overflow_bays", []))

        cluster_cap = (1 * n_single) + (2 * n_double) + (3 * n_triple) + (1 * n_overflow)
        capacities[cname] = cluster_cap
    return capacities


def build_stop_capacities() -> Dict[str, int]:
    """Generate a per-stop capacity dictionary.

    Returns:
    -------
    dict[str, int]
        ``{stop_id: capacity}`` where single-bay = 1, double-bay = 2,
        triple-bay = 3, and overflow = 1.
    """
    stop_caps: Dict[str, int] = {}
    for cinfo in CLUSTER_DEFINITIONS.values():
        for stop_id in cinfo.get("single_bay_stops", []):
            stop_caps[str(stop_id)] = 1
        for stop_id in cinfo.get("double_bay_stops", []):
            stop_caps[str(stop_id)] = 2
        for stop_id in cinfo.get("triple_bay_stops", []):
            stop_caps[str(stop_id)] = 3
        for ovf in cinfo.get("overflow_bays", []):
            stop_caps[str(ovf)] = 1
    return stop_caps


# --------------------------------------------------------------------------------------------------
# CORE CONFLICT-DETECTION LOGIC
# --------------------------------------------------------------------------------------------------


def normalize_stop_id(stop_id: object) -> Optional[str]:
    """Return a normalized stop-ID string.

    Converts ``2956.0`` → ``"2956"`` and gracefully handles *NaN*.

    Args:
    ----
    stop_id :
        Value from the ``Stop ID`` column, potentially numeric or NaN.

    Returns:
    -------
    str | None
        Cleaned stop ID or *None* if ``stop_id`` is NaN.
    """
    if pd.isna(stop_id):
        return None
    sid_str = str(stop_id).strip()
    if sid_str.endswith(".0"):
        sid_str = sid_str[:-2]
    return sid_str


def assign_cluster_name(df_in: DataFrame) -> DataFrame:
    """Add a ``ClusterName`` column indicating which cluster each stop belongs to.

    If a stop appears in multiple clusters (unlikely), the first match in
    ``CLUSTER_DEFINITIONS`` wins.

    Args:
    ----------
    df_in :
        Input DataFrame containing at least a ``Stop ID`` column.

    Returns:
    -------
    pandas.DataFrame
        Copy of ``df_in`` with the additional ``ClusterName`` column.
    """
    df_out = df_in.copy()
    df_out["ClusterName"] = None

    # Build a map of cluster → set(stop_ids)
    cluster_map: Dict[str, Set[str]] = {}
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        official_stops = get_all_official_stops(cinfo)
        overflow = cinfo.get("overflow_bays", [])
        all_stops = official_stops + overflow
        cluster_map[cname] = set(map(str, all_stops))

    for cname, stop_set in cluster_map.items():
        mask = df_out["Stop ID"].isin(stop_set)
        df_out.loc[mask, "ClusterName"] = cname

    return df_out


def find_cluster_conflicts(df_in: DataFrame) -> Set[Tuple[str, str]]:
    """Identify cluster-level conflicts.

    A conflict occurs when the number of *present* buses exceeds the
    cluster’s capacity at a specific timestamp.

    Args:
    ----------
    df_in :
        DataFrame that already contains a ``ClusterName`` column.

    Returns:
    -------
    set[tuple[str, str]]
        {(cluster_name, timestamp), …} representing each conflict point.
    """
    cluster_caps = build_cluster_capacities()
    conflict_set: Set[Tuple[str, str]] = set()

    present_df = df_in[df_in["Status"].isin(PRESENCE_STATUSES)].copy()
    present_df = present_df.dropna(subset=["ClusterName"])  # ignore rows w/o cluster

    group = present_df.groupby(["ClusterName", "Timestamp"])
    for (cname, ts), grp in group:
        cap = cluster_caps.get(cname, 1)
        if len(grp) > cap:  # More vehicles than capacity
            conflict_set.add((cname, ts))

    return conflict_set


def find_stop_conflicts(df_in: DataFrame) -> Set[Tuple[str, str]]:
    """Identify stop-level conflicts.

    A conflict occurs when the number of buses in passenger-service
    statuses at a stop exceeds that stop’s capacity at a given timestamp.

    Args:
    ----------
    df_in :
        Raw DataFrame containing ``Stop ID`` and ``Status`` columns.

    Returns:
    -------
    set[tuple[str, str]]
        {(stop_id, timestamp), …} representing each conflict point.
    """
    stop_caps = build_stop_capacities()
    conflict_set: Set[Tuple[str, str]] = set()

    pass_df = df_in[df_in["Status"].isin(PASSENGER_SERVICE_STATUSES)].copy()
    pass_df = pass_df[pass_df["Stop ID"].notna()]

    group = pass_df.groupby(["Stop ID", "Timestamp"])
    for (sid, ts), grp in group:
        cap = stop_caps.get(sid, 1)
        if len(grp) > cap:
            conflict_set.add((sid, ts))

    return conflict_set


def annotate_conflicts(
    df_in: DataFrame,
    cluster_conflicts: Set[Tuple[str, str]],
    stop_conflicts: Set[Tuple[str, str]],
) -> DataFrame:
    """Append a ``ConflictType`` column categorizing each row.

    ``ConflictType`` will be one of ``{"NONE", "CLUSTER", "STOP", "BOTH"}``.

    Args:
    ----------
    df_in :
        Input DataFrame (must include ``ClusterName``, ``Stop ID``,
        ``Timestamp``).
    cluster_conflicts :
        Set of cluster-level conflict keys as produced by
        :func:`find_cluster_conflicts`.
    stop_conflicts :
        Set of stop-level conflict keys as produced by
        :func:`find_stop_conflicts`.

    Returns:
    -------
    pandas.DataFrame
        Copy of ``df_in`` with the additional ``ConflictType`` column.
    """
    df_out = df_in.copy()
    conflict_types: List[str] = []

    for _, row in df_out.iterrows():
        cname = row["ClusterName"]
        ts = row["Timestamp"]
        sid = row["Stop ID"]

        has_cluster_conf = pd.notna(cname) and (cname, ts) in cluster_conflicts
        has_stop_conf = sid is not None and (sid, ts) in stop_conflicts

        if has_cluster_conf and has_stop_conf:
            conflict_types.append("BOTH")
        elif has_cluster_conf:
            conflict_types.append("CLUSTER")
        elif has_stop_conf:
            conflict_types.append("STOP")
        else:
            conflict_types.append("NONE")

    df_out["ConflictType"] = conflict_types
    return df_out


# --------------------------------------------------------------------------------------------------
# I/O AND EXCEL WRITING LOGIC
# --------------------------------------------------------------------------------------------------


def gather_block_spreadsheets(block_folder: str) -> DataFrame:
    """Read and concatenate every ``block_*.xlsx`` spreadsheet in *block_folder*.

    Args:
    ----------
    block_folder :
        Directory containing the Step 1 block-level output workbooks.

    Returns:
    -------
    pandas.DataFrame
        Combined DataFrame of all block files.

    Raises:
    ------
    FileNotFoundError
        If no eligible ``block_*.xlsx`` files are found.
    """
    all_files = [
        f
        for f in os.listdir(block_folder)
        if f.lower().endswith(".xlsx") and f.startswith("block_")
    ]
    if not all_files:
        raise FileNotFoundError(f"No block_*.xlsx files found in {block_folder}.")

    big_df_list: List[DataFrame] = []
    for fname in all_files:
        path = os.path.join(block_folder, fname)
        temp_df = pd.read_excel(path)
        temp_df["FileName"] = fname
        big_df_list.append(temp_df)

    df_combined = pd.concat(big_df_list, ignore_index=True)
    logging.info("Loaded %d total rows from Step 1 block XLSX files.", len(df_combined))
    return df_combined


def run_step2_conflict_detection() -> None:
    """Execute the full Step 2 conflict-detection workflow.

    Steps
    -----
    1. Read block-level spreadsheets from :pydata:`BLOCK_OUTPUT_FOLDER`.
    2. Normalise stop IDs, assign clusters, find conflicts.
    3. For each cluster, build an Excel workbook containing:
       * “AllStops” sheet (all events, bolding conflict rows).
       * One sheet per stop/overflow bay (same bolding).
    4. Print summary statistics.

    Raises:
    ------
    ValueError
        If required columns are missing from the input spreadsheets.
    """
    logging.info("=== Step 2: Conflict detection and per-cluster output (multi-bay) ===")
    os.makedirs(CLUSTER_CONFLICT_OUTPUT_FOLDER, exist_ok=True)

    # 1) Gather Step 1 data
    df = gather_block_spreadsheets(BLOCK_OUTPUT_FOLDER)

    # Basic checks/cleanup
    required_cols = [
        "Timestamp",
        "Trip ID",
        "Block",
        "Route",
        "Direction",
        "Stop ID",
        "Stop Name",
        "Stop Sequence",
        "Arrival Time",
        "Departure Time",
        "Status",
    ]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns in block-level data: {missing_cols}")

    # Normalize Stop ID text
    df["Stop ID"] = df["Stop ID"].apply(normalize_stop_id)
    # Ensure string Timestamps
    df["Timestamp"] = df["Timestamp"].astype(str).str.strip()

    # 2) Assign cluster, detect conflicts, annotate
    df = assign_cluster_name(df)
    cluster_conflicts = find_cluster_conflicts(df)
    stop_conflicts = find_stop_conflicts(df)
    df_annotated = annotate_conflicts(df, cluster_conflicts, stop_conflicts)

    # 3) Write results per cluster, each cluster → single XLSX with multiple sheets
    for cname, cinfo in CLUSTER_DEFINITIONS.items():
        # Combine all official stops + overflow for that cluster
        official_stops = get_all_official_stops(cinfo)
        overflow_stops = cinfo.get("overflow_bays", [])
        all_cluster_stops = official_stops + overflow_stops

        sub = df_annotated[df_annotated["ClusterName"] == cname].copy()
        if sub.empty:
            logging.info("No rows found for cluster '%s'. Skipping.", cname)
            continue

        safe_name = cname.replace(" ", "_")
        out_path = os.path.join(
            CLUSTER_CONFLICT_OUTPUT_FOLDER,
            f"{safe_name}_Conflicts.xlsx",
        )
        logging.info("Building conflict output for cluster '%s' → %s", cname, out_path)

        # Sort by timestamp, then by block or stop ID
        sub.sort_values(["Timestamp", "Block", "Stop ID"], inplace=True)

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # 3a) “AllStops” sheet
            sub.to_excel(writer, sheet_name="AllStops", index=False)

            # Bold the conflict rows in “AllStops”
            conflict_col_index = int(cast(int, sub.columns.get_loc("ConflictType"))) + 1  # 1-based
            worksheet_all = writer.sheets["AllStops"]
            for row_idx in range(2, len(sub) + 2):  # data start row 2
                conflict_val = worksheet_all.cell(
                    row=row_idx,
                    column=conflict_col_index,
                ).value
                if conflict_val != "NONE":
                    for col_idx in range(1, len(sub.columns) + 1):
                        cell = worksheet_all.cell(row=row_idx, column=col_idx)
                        cell.font = Font(bold=True)

            # 3b) One sheet per stop
            for stop_id in all_cluster_stops:
                sid_str = str(stop_id)
                stop_df = sub[sub["Stop ID"] == sid_str].copy()
                if stop_df.empty:
                    continue

                # Make a sheet name that is safe in Excel (≤31 chars)
                sheet_name = f"Stop_{sid_str}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

                stop_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Bold the conflict rows in each stop’s sheet
                conflict_col_index_stop = int(cast(int, stop_df.columns.get_loc("ConflictType"))) + 1
                worksheet_stop = writer.sheets[sheet_name]
                for row_idx in range(2, len(stop_df) + 2):
                    conflict_val = worksheet_stop.cell(
                        row=row_idx,
                        column=conflict_col_index_stop,
                    ).value
                    if conflict_val != "NONE":
                        for col_idx in range(1, len(stop_df.columns) + 1):
                            cell = worksheet_stop.cell(row=row_idx, column=col_idx)
                            cell.font = Font(bold=True)

        logging.info(" → Completed writing %s", out_path)

    # Final conflict summary stats
    logging.info("\nDistinct cluster-conflict points: %d", len(cluster_conflicts))
    logging.info("Distinct stop-conflict points: %d", len(stop_conflicts))
    logging.info("Step 2 complete.")


# ==================================================================================================
# MAIN
# ==================================================================================================


def main() -> None:
    """Entry point when executing this module as a script."""
    logging.basicConfig(level=logging.INFO)
    run_step2_conflict_detection()


if __name__ == "__main__":
    main()
