"""Processes stop-level ridership data from an Excel file.

Reads an input Excel file (RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS).XLSX),
optionally filters by route or stop ID, aggregates boardings and alightings by
stop and time period, and saves the results to a new Excel file. Aggregated data
can optionally be rounded or categorized into bins.

When a GTFS feed is supplied (GTFS_PATH), aggregation sheets are enriched with
LATITUDE/LONGITUDE columns and an additional ``Stops Clean`` sheet is produced
with Stop Code, Stop Name, Latitude, Longitude, Boardings, Alightings, and Total.
A ``Summary`` sheet is also written that compares the post-filter selection
against the full input dataset (stop counts and total ridership, with percents).

The script is designed for analysts and data scientists who need a quick and
repeatable tool for ad-hoc stop ridership data requests, and it is suitable
for use in environments like ArcGIS Pro or Jupyter Notebooks.
"""

from __future__ import annotations

import logging
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_FILE_PATH: Path = Path(r"Path\To\Your\RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS).XLSX")
OUTPUT_FILE_SUFFIX: str = "_processed"
OUTPUT_FILE_EXTENSION: str = ".xlsx"
# If OUTPUT_DIR is None ⇒ use same directory as INPUT_FILE_PATH
OUTPUT_DIR: Path | None = Path(r"Path\To\Output\Folder")  # e.g. r"C:\Data\Outputs"

# ROUTES = keep-only list   |  ROUTES_EXCLUDE = toss-out list
ROUTES: List[str] = []  # keep these (empty → keep all)
ROUTES_EXCLUDE: List[str] = []  # drop these (empty → drop none)

# If True → aggregate across all selected routes at the same stop (adds a "ROUTES" column).
# If False → keep routes separate in aggregations (adds/retains a "ROUTE_NAME" column).
AGGREGATE_ROUTES_TOGETHER: bool = True

# Optional STOP_IDS filter list
STOP_IDS: List[int] = []  # keep these (empty → keep all)

# Optional TIME_PERIOD aggregation list
TIME_PERIODS: List[str] = [
    #    "AM EARLY",
    "AM PEAK",
    #    "MIDDAY",
    "PM PEAK",
    #    "PM LATE",
    #    "PM NITE",
]  # empty ⇒ skip time-period breakdown

# If True → round ridership columns in "Original" sheet and (if bins are off)
# round aggregated totals.  If False → leave raw numeric precision.
APPLY_ROUNDING: bool = True

# If True → convert aggregated totals to textual bins; numeric rounding is
# suppressed.  If False → leave numeric (subject to APPLY_ROUNDING).
AGGREGATE_BIN_RANGES: bool = False

# Optional GTFS enrichment.
# GTFS_PATH may point to a stops.txt file, a GTFS .zip archive, or an
# unzipped GTFS directory. If None, GTFS enrichment is skipped.
GTFS_PATH: Path | None = None  # e.g. Path(r"C:\Data\gtfs.zip")
GTFS_JOIN_KEY: str = "stop_id"  # "stop_id" or "stop_code" — which GTFS field
                                # the ridership STOP_ID column should match

REQUIRED_COLUMNS: Sequence[str] = (
    "TIME_PERIOD",
    "ROUTE_NAME",
    "STOP",
    "STOP_ID",
    "BOARD_ALL",
    "ALIGHT_ALL",
)
COLUMNS_TO_RETAIN: Sequence[str] = (
    "ROUTE_NAME",
    "STOP",
    "STOP_ID",
    "BOARD_ALL",
    "ALIGHT_ALL",
)

LOG_LEVEL: int = logging.INFO  # DEBUG / INFO / WARNING / ERROR

# =============================================================================
# FUNCTIONS
# =============================================================================


def _join_unique_routes(x: Any) -> str:
    """Join unique route names."""
    return ", ".join(sorted(pd.Series(x).dropna().unique()))


def bin_ridership_value(value: float) -> str:
    """Convert a numeric ridership value into a categorical range.

    Args:
        value: The boarding/alighting count for a stop.

    Returns:
        A string bucket—``"0-4.9"``, ``"5-24.9"``, or ``"25 or more"``.
    """
    if value < 5:
        return "0-4.9"
    if value < 25:
        return "5-24.9"
    return "25 or more"


def aggregate_by_stop(
    data_subset: pd.DataFrame,
    *,
    aggregate_routes_together: bool,
) -> pd.DataFrame:
    """Aggregate boardings/alightings by stop, with optional cross-route rollup.

    When ``aggregate_routes_together`` is True, rows for different routes serving the
    same stop are summed together and a ``ROUTES`` column lists the unique routes.
    When False, routes are *not* combined; the output is keyed by ``ROUTE_NAME``,
    ``STOP``, and ``STOP_ID``. In that case, ``ROUTE_NAME`` is a regular column.

    Args:
        data_subset: DataFrame already filtered to rows of interest.
        aggregate_routes_together: If True, aggregate across routes at a stop.

    Returns:
        Aggregated DataFrame with ``BOARD_ALL_TOTAL`` and ``ALIGHT_ALL_TOTAL`` plus:
          * If together:  one row per (STOP, STOP_ID) and a ``ROUTES`` column.
          * If separate:  one row per (ROUTE_NAME, STOP, STOP_ID).
    """
    cols_needed: List[str] = [
        "STOP",
        "STOP_ID",
        "BOARD_ALL",
        "ALIGHT_ALL",
        "ROUTE_NAME",
    ]
    subset: pd.DataFrame = data_subset[cols_needed].copy()
    subset["ROUTE_NAME"] = subset["ROUTE_NAME"].astype(str).str.strip()

    if aggregate_routes_together:
        grouping_cols: List[str] = ["STOP", "STOP_ID"]
        aggregated: pd.DataFrame = (
            subset.groupby(grouping_cols, as_index=False)
            .agg(
                {
                    "BOARD_ALL": "sum",
                    "ALIGHT_ALL": "sum",
                    "ROUTE_NAME": _join_unique_routes,
                }
            )
            .rename(
                columns={
                    "BOARD_ALL": "BOARD_ALL_TOTAL",
                    "ALIGHT_ALL": "ALIGHT_ALL_TOTAL",
                    "ROUTE_NAME": "ROUTES",
                }
            )
        )
    else:
        # Keep routes separate in the aggregates.
        grouping_cols = ["ROUTE_NAME", "STOP", "STOP_ID"]
        aggregated = (
            subset.groupby(grouping_cols, as_index=False)
            .agg({"BOARD_ALL": "sum", "ALIGHT_ALL": "sum"})
            .rename(columns={"BOARD_ALL": "BOARD_ALL_TOTAL", "ALIGHT_ALL": "ALIGHT_ALL_TOTAL"})
        )

    return aggregated


def log_missing_stop_ids(
    requested_ids: Sequence[int], present_ids: pd.Series | Sequence[int]
) -> None:
    """Warn when any requested *STOP_ID* is absent from *present_ids*.

    Args:
        requested_ids: The ``STOP_IDS`` list provided in CONFIGURATION.
        present_ids:  A 1-D iterable (e.g. ``DataFrame['STOP_ID']``) of IDs that
            survived the filter chain.

    Raises:
        TypeError: If *present_ids* cannot be coerced to a set of ``int``.
    """
    if not requested_ids:  # nothing requested ⇒ nothing to warn about
        return

    try:
        missing: set[int] = set(requested_ids) - {int(x) for x in present_ids}
    except (ValueError, TypeError) as exc:
        raise TypeError(
            "Unable to evaluate present_ids when checking for missing STOP_IDs."
        ) from exc

    if missing:
        logging.warning(
            "The following requested STOP_IDs were not found in the processed "
            "dataset and therefore will not appear in the output: %s",
            sorted(missing),
        )
    else:
        logging.info("All requested STOP_IDs are present in the processed data.")


def build_selection_summary(
    full_data: pd.DataFrame,
    filtered_data: pd.DataFrame,
) -> pd.DataFrame:
    """Compute and log selection-vs-system totals.

    Compares the post-filter ``filtered_data`` against the full input
    ``full_data`` and reports unique stop counts and total boardings /
    alightings, both as raw values and as percentages of the system total.

    Args:
        full_data: The unfiltered DataFrame loaded from the input workbook.
        filtered_data: The DataFrame after route/stop/exclude filters
            have been applied.

    Returns:
        A DataFrame with one row per metric (Stops, Boardings, Alightings,
        Total Ridership) and columns ``Selected``, ``Total``, ``Percent``,
        suitable for writing to its own worksheet.
    """
    def _safe_pct(numerator: float, denominator: float) -> float:
        return (numerator / denominator * 100.0) if denominator else 0.0

    selected_stops: int = int(filtered_data["STOP_ID"].nunique())
    total_stops: int = int(full_data["STOP_ID"].nunique())
    pct_stops: float = _safe_pct(selected_stops, total_stops)

    selected_board: float = float(filtered_data["BOARD_ALL"].sum())
    total_board: float = float(full_data["BOARD_ALL"].sum())
    pct_board: float = _safe_pct(selected_board, total_board)

    selected_alight: float = float(filtered_data["ALIGHT_ALL"].sum())
    total_alight: float = float(full_data["ALIGHT_ALL"].sum())
    pct_alight: float = _safe_pct(selected_alight, total_alight)

    selected_ridership: float = selected_board + selected_alight
    total_ridership: float = total_board + total_alight
    pct_ridership: float = _safe_pct(selected_ridership, total_ridership)

    logging.info("=" * 72)
    logging.info("SELECTION SUMMARY (filtered vs. full dataset)")
    logging.info("=" * 72)
    logging.info(
        "Stops selected:        %s of %s  (%.2f%%)",
        f"{selected_stops:,}", f"{total_stops:,}", pct_stops,
    )
    logging.info(
        "Boardings selected:    %s of %s  (%.2f%%)",
        f"{selected_board:,.1f}", f"{total_board:,.1f}", pct_board,
    )
    logging.info(
        "Alightings selected:   %s of %s  (%.2f%%)",
        f"{selected_alight:,.1f}", f"{total_alight:,.1f}", pct_alight,
    )
    logging.info(
        "Total ridership:       %s of %s  (%.2f%%)",
        f"{selected_ridership:,.1f}", f"{total_ridership:,.1f}", pct_ridership,
    )
    logging.info("=" * 72)

    return pd.DataFrame(
        [
            {"Metric": "Stops (unique STOP_ID)",
             "Selected": selected_stops,
             "Total": total_stops,
             "Percent": round(pct_stops, 2)},
            {"Metric": "Boardings (BOARD_ALL sum)",
             "Selected": round(selected_board, 1),
             "Total": round(total_board, 1),
             "Percent": round(pct_board, 2)},
            {"Metric": "Alightings (ALIGHT_ALL sum)",
             "Selected": round(selected_alight, 1),
             "Total": round(total_alight, 1),
             "Percent": round(pct_alight, 2)},
            {"Metric": "Total Ridership (Board + Alight)",
             "Selected": round(selected_ridership, 1),
             "Total": round(total_ridership, 1),
             "Percent": round(pct_ridership, 2)},
        ]
    )


def read_excel_file(input_file: Path) -> pd.DataFrame:
    """Load an Excel workbook into a DataFrame.

    Args:
        input_file: Absolute or relative path to the workbook.

    Returns:
        The first sheet of *input_file* as a ``pandas.DataFrame``.

    Raises:
        SystemExit: If the file does not exist or cannot be parsed.
    """
    try:
        return pd.read_excel(input_file)
    except FileNotFoundError:
        logging.error("The file '%s' does not exist.", input_file)
        sys.exit(1)
    except ValueError as exc:  # pandas re-raises most Excel errors as ValueError
        logging.error("Error reading the Excel file: %s", exc)
        sys.exit(1)


def load_gtfs_stops(gtfs_path: Path) -> pd.DataFrame:
    """Load and normalize the GTFS stops table.

    Accepts a path to a ``stops.txt`` file, a GTFS ``.zip`` archive
    containing one, or an unzipped GTFS directory.

    Args:
        gtfs_path: Path to stops.txt, a GTFS .zip, or a GTFS directory.

    Returns:
        DataFrame with columns ``stop_id``, ``stop_code``, ``stop_name``
        (all stripped strings) and ``stop_lat``, ``stop_lon`` (floats).
        ``stop_code`` is filled with empty strings when absent from the feed.

    Raises:
        SystemExit: If the file cannot be located, opened, or is missing
            required GTFS columns.
    """
    try:
        if gtfs_path.is_file() and gtfs_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(gtfs_path) as zf:
                with zf.open("stops.txt") as f:
                    stops: pd.DataFrame = pd.read_csv(f, dtype=str)
        elif gtfs_path.is_file():
            stops = pd.read_csv(gtfs_path, dtype=str)
        elif gtfs_path.is_dir():
            stops = pd.read_csv(gtfs_path / "stops.txt", dtype=str)
        else:
            logging.error("GTFS path '%s' does not exist.", gtfs_path)
            sys.exit(1)
    except (FileNotFoundError, KeyError, OSError, zipfile.BadZipFile) as exc:
        logging.error("Error reading GTFS stops: %s", exc)
        sys.exit(1)

    required: set[str] = {"stop_id", "stop_name", "stop_lat", "stop_lon"}
    missing: set[str] = required - set(stops.columns)
    if missing:
        logging.error("GTFS stops.txt missing required columns: %s", sorted(missing))
        sys.exit(1)

    if "stop_code" not in stops.columns:
        stops["stop_code"] = ""

    for col in ("stop_id", "stop_code", "stop_name"):
        stops[col] = stops[col].fillna("").astype(str).str.strip()
    stops["stop_lat"] = pd.to_numeric(stops["stop_lat"], errors="coerce")
    stops["stop_lon"] = pd.to_numeric(stops["stop_lon"], errors="coerce")

    return stops[["stop_id", "stop_code", "stop_name", "stop_lat", "stop_lon"]]


def enrich_with_gtfs(
    df: pd.DataFrame,
    gtfs_stops: pd.DataFrame,
    join_key: str,
    *,
    stop_id_col: str = "STOP_ID",
) -> pd.DataFrame:
    """Left-join GTFS coordinates onto a DataFrame keyed by stop.

    Args:
        df: DataFrame containing a stop identifier column.
        gtfs_stops: Output of :func:`load_gtfs_stops`.
        join_key: ``"stop_id"`` or ``"stop_code"`` — which GTFS column to
            match against ``df[stop_id_col]``.
        stop_id_col: Column in ``df`` that carries the ridership stop ID.

    Returns:
        A copy of ``df`` with ``LATITUDE`` and ``LONGITUDE`` columns
        appended.  Stops with no GTFS match get NaN coordinates.

    Raises:
        ValueError: If ``join_key`` is not ``"stop_id"`` or ``"stop_code"``.
    """
    if join_key not in {"stop_id", "stop_code"}:
        raise ValueError(f"join_key must be 'stop_id' or 'stop_code', got '{join_key}'")

    out: pd.DataFrame = df.copy()
    out["_join_key"] = out[stop_id_col].astype(str).str.strip()

    gtfs_subset: pd.DataFrame = (
        gtfs_stops[[join_key, "stop_lat", "stop_lon"]]
        .drop_duplicates(subset=[join_key])
        .rename(columns={join_key: "_join_key",
                         "stop_lat": "LATITUDE",
                         "stop_lon": "LONGITUDE"})
    )

    return out.merge(gtfs_subset, on="_join_key", how="left").drop(columns=["_join_key"])


def build_clean_stops_sheet(
    filtered_data: pd.DataFrame,
    gtfs_stops: pd.DataFrame,
    join_key: str,
) -> pd.DataFrame:
    """Build a tidy per-stop summary enriched with GTFS attributes.

    Aggregates BOARD_ALL and ALIGHT_ALL across all rows of ``filtered_data``
    by STOP_ID (collapsing routes and time periods), joins authoritative
    ``stop_code``, ``stop_name``, and coordinates from GTFS, and adds a
    ``Total`` column.

    Args:
        filtered_data: Post-filter ridership rows. Must contain STOP_ID,
            BOARD_ALL, ALIGHT_ALL.
        gtfs_stops: Output of :func:`load_gtfs_stops`.
        join_key: ``"stop_id"`` or ``"stop_code"``.

    Returns:
        DataFrame with columns ``Stop Code``, ``Stop Name``, ``Latitude``,
        ``Longitude``, ``Boardings``, ``Alightings``, ``Total`` — one row
        per unique STOP_ID.
    """
    if join_key not in {"stop_id", "stop_code"}:
        raise ValueError(f"join_key must be 'stop_id' or 'stop_code', got '{join_key}'")

    per_stop: pd.DataFrame = (
        filtered_data.groupby("STOP_ID", as_index=False)
        .agg(Boardings=("BOARD_ALL", "sum"), Alightings=("ALIGHT_ALL", "sum"))
    )
    per_stop["_join_key"] = per_stop["STOP_ID"].astype(str).str.strip()

    # Build _join_key as a derived column so we don't collide when
    # join_key itself is "stop_code" (which is also kept as an output column).
    gtfs_subset: pd.DataFrame = (
        gtfs_stops.assign(_join_key=gtfs_stops[join_key])
        [["_join_key", "stop_code", "stop_name", "stop_lat", "stop_lon"]]
        .drop_duplicates(subset=["_join_key"])
        .rename(columns={"stop_code": "Stop Code",
                         "stop_name": "Stop Name",
                         "stop_lat": "Latitude",
                         "stop_lon": "Longitude"})
    )

    merged: pd.DataFrame = per_stop.merge(
        gtfs_subset, on="_join_key", how="left"
    ).drop(columns=["_join_key"])

    unmatched_mask: pd.Series = merged["Latitude"].isna()
    if unmatched_mask.any():
        missing_ids: List[Any] = merged.loc[unmatched_mask, "STOP_ID"].tolist()
        logging.warning(
            "%d stop(s) had no GTFS match on '%s' and will lack coordinates: %s",
            unmatched_mask.sum(), join_key, missing_ids,
        )

    merged["Total"] = merged["Boardings"].fillna(0) + merged["Alightings"].fillna(0)

    if APPLY_ROUNDING:
        for col in ("Boardings", "Alightings", "Total"):
            merged[col] = merged[col].round(1)

    return merged[["Stop Code", "Stop Name", "Latitude", "Longitude",
                   "Boardings", "Alightings", "Total"]]


def verify_required_columns(data_frame: pd.DataFrame, required_columns: Sequence[str]) -> None:
    """Ensure *data_frame* contains all columns listed in *required_columns*.

    Args:
        data_frame: The DataFrame to validate.
        required_columns: Names that must be present.

    Raises:
        SystemExit: If any required column is missing.
    """
    missing_columns: List[str] = [col for col in required_columns if col not in data_frame.columns]
    if missing_columns:
        logging.error("Missing columns: %s", missing_columns)
        sys.exit(1)


def filter_data(
    data_frame: pd.DataFrame,
    routes: Sequence[str] | None = None,
    stop_ids: Sequence[int] | None = None,
    routes_exclude: Sequence[str] | None = None,
) -> pd.DataFrame:
    """Apply optional filters in a deterministic order.

    Filters are applied in-place (on a copy) with this precedence:

    1. Inclusive   ``routes``        → keep-only rows where ROUTE_NAME ∈ routes
    2. Exclusive   ``routes_exclude``→ drop-only  rows where ROUTE_NAME ∈ routes_exclude
    3. Inclusive   ``stop_ids``      → keep-only rows where STOP_ID   ∈ stop_ids

    Empty/None arguments are ignored.

    Args:
        data_frame: Original DataFrame.
        routes: Route names to *keep* (inclusive).
        stop_ids: Stop IDs to *keep* (inclusive).
        routes_exclude: Route names to *drop* (exclusive).

    Returns:
        A filtered DataFrame.
    """
    df: pd.DataFrame = data_frame.copy()

    if routes:
        df = df[df["ROUTE_NAME"].isin(routes)]

    if routes_exclude:
        df = df[~df["ROUTE_NAME"].isin(routes_exclude)]

    if stop_ids:
        df = df[df["STOP_ID"].isin(stop_ids)]

    return df


def write_to_excel(
    output_file: Path,
    filtered_data: pd.DataFrame,
    aggregated_peaks: Dict[str, pd.DataFrame],
    all_time_aggregated: pd.DataFrame,
    selection_summary: pd.DataFrame | None = None,
    clean_stops: pd.DataFrame | None = None,
) -> None:
    """Write the processed data sets to *output_file* in a sensible order.

    The workbook receives, in order:
        1. ``Summary``           – selection-vs-system totals (if provided)
        2. ``Stops Clean``       – per-stop summary with GTFS attributes
                                   (if provided)
        3. ``Original``          – raw (but optionally rounded) rows
        4. ``All Time Periods``  – aggregation across *all* rows
        5. One sheet per entry in ``aggregated_peaks``

    Args:
        output_file: Path to the workbook to create/overwrite.
        filtered_data: DataFrame for the "Original" sheet.
        aggregated_peaks: Mapping ``TIME_PERIOD → aggregated DataFrame``.
        all_time_aggregated: Aggregation across *all* rows.
        selection_summary: Optional summary DataFrame from
            :func:`build_selection_summary`.
        clean_stops: Optional clean per-stop sheet from
            :func:`build_clean_stops_sheet`.

    Raises:
        SystemExit: On any I/O error (permission, disk full, etc.).
    """
    try:
        with pd.ExcelWriter(output_file) as writer:
            if selection_summary is not None:
                selection_summary.to_excel(writer, sheet_name="Summary", index=False)
            if clean_stops is not None:
                clean_stops.to_excel(writer, sheet_name="Stops Clean", index=False)
            filtered_data.to_excel(writer, sheet_name="Original", index=False)
            all_time_aggregated.to_excel(writer, sheet_name="All Time Periods", index=False)
            for period, df_agg in aggregated_peaks.items():
                df_agg.to_excel(writer, sheet_name=period, index=False)

        adjust_excel_formatting(output_file)
        logging.info("The processed file has been saved as '%s'.", output_file)
    except (OSError, PermissionError) as exc:
        logging.error("Error writing the processed Excel file: %s", exc)
        sys.exit(1)


def adjust_excel_formatting(output_file: Path) -> None:
    """Bold the header row and auto-size columns for *output_file*.

    Args:
        output_file: Path to an existing XLSX workbook.

    Raises:
        SystemExit: If the workbook cannot be opened or saved.
    """
    try:
        workbook = load_workbook(output_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Bold the header row
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            # Auto-size column widths
            for col_idx, column_cells in enumerate(sheet.columns, start=1):
                max_length: int = 0
                col_letter: str = get_column_letter(col_idx)
                for cell in column_cells:
                    cell_val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_val))
                sheet.column_dimensions[col_letter].width = max_length + 2

        workbook.save(output_file)
    except (OSError, PermissionError) as exc:
        logging.error("Error adjusting Excel formatting: %s", exc)
        sys.exit(1)


def process_aggregations(
    filtered_data: pd.DataFrame,
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame], pd.DataFrame]:
    """Orchestrate rounding/bucketing and aggregation for output sheets.

    This function:
      1) Standardizes ``ROUTE_NAME`` as ``str``.
      2) Builds per-period slices if ``TIME_PERIODS`` is set.
      3) Aggregates for "All Time Periods" and each time-period slice.
         The cross-route behavior is controlled by
         ``AGGREGATE_ROUTES_TOGETHER``:
           * True  → one row per (STOP, STOP_ID) with ``ROUTES`` listing routes.
           * False → one row per (ROUTE_NAME, STOP, STOP_ID).
      4) Optionally rounds the "Original" sheet numeric columns.
      5) Applies binning or rounding to the aggregated totals.

    Args:
        filtered_data: The DataFrame *after* optional route/stop filtering.
            Must contain at least the columns listed in ``REQUIRED_COLUMNS``.

    Returns:
        A 3-tuple:
            * Final version of ``filtered_data`` (possibly rounded for "Original").
            * Mapping ``TIME_PERIOD → aggregated DataFrame`` (sheet-ready).
            * Aggregation across *all* rows and periods.
    """
    # ──────────────────────────────────────────────────────────────────
    # 0. Standardize ROUTE_NAME to string before any aggregation
    # ──────────────────────────────────────────────────────────────────
    filtered_data["ROUTE_NAME"] = filtered_data["ROUTE_NAME"].astype(str).str.strip()

    # ──────────────────────────────────────────────────────────────────
    # 1. Build per-period subsets (if any)
    # ──────────────────────────────────────────────────────────────────
    peak_data_dict: Dict[str, pd.DataFrame] = {}
    if TIME_PERIODS:
        for period in TIME_PERIODS:
            subset = filtered_data[filtered_data["TIME_PERIOD"] == period.upper()]
            # Keep only the columns needed for the "Original" per-period sheet seeds
            peak_data_dict[period] = subset[list(COLUMNS_TO_RETAIN)]

    # ──────────────────────────────────────────────────────────────────
    # 2. Aggregate (all-time + each slice); route roll-up is configurable
    # ──────────────────────────────────────────────────────────────────
    all_time_aggregated: pd.DataFrame = aggregate_by_stop(
        filtered_data, aggregate_routes_together=AGGREGATE_ROUTES_TOGETHER
    )

    aggregated_peaks: Dict[str, pd.DataFrame] = {}
    if TIME_PERIODS:
        for period, data_subset in peak_data_dict.items():
            aggregated_peaks[period] = aggregate_by_stop(
                data_subset, aggregate_routes_together=AGGREGATE_ROUTES_TOGETHER
            )

    # ──────────────────────────────────────────────────────────────────
    # 3. Optional rounding for "Original" sheet
    # ──────────────────────────────────────────────────────────────────
    if APPLY_ROUNDING:
        filtered_data[["BOARD_ALL", "ALIGHT_ALL"]] = filtered_data[
            ["BOARD_ALL", "ALIGHT_ALL"]
        ].round(1)

    # ──────────────────────────────────────────────────────────────────
    # 4. Format aggregated columns (bins OR decimal rounding)
    # ──────────────────────────────────────────────────────────────────
    all_aggregations: List[pd.DataFrame] = [all_time_aggregated] + list(aggregated_peaks.values())
    for df_agg in all_aggregations:
        if AGGREGATE_BIN_RANGES:
            for col in ("BOARD_ALL_TOTAL", "ALIGHT_ALL_TOTAL"):
                df_agg[col] = df_agg[col].apply(bin_ridership_value)
        elif APPLY_ROUNDING:
            df_agg[["BOARD_ALL_TOTAL", "ALIGHT_ALL_TOTAL"]] = df_agg[
                ["BOARD_ALL_TOTAL", "ALIGHT_ALL_TOTAL"]
            ].round(1)

    return filtered_data, aggregated_peaks, all_time_aggregated


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:  # noqa: D401 – imperative mood is OK for main entry point
    """Run the full read → filter → aggregate → (enrich) → write pipeline."""
    logging.basicConfig(
        level=LOG_LEVEL,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    _DEFAULT_INPUT = r"Path\To\Your\RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS).XLSX"
    _DEFAULT_OUTPUT_DIR = r"Path\To\Output\Folder"
    if str(INPUT_FILE_PATH) == _DEFAULT_INPUT or (
        OUTPUT_DIR is not None and str(OUTPUT_DIR) == _DEFAULT_OUTPUT_DIR
    ):
        logging.warning(
            "File paths are still set to their defaults. Update INPUT_FILE_PATH and "
            "OUTPUT_DIR in the CONFIGURATION section before running."
        )
        return

    if GTFS_JOIN_KEY not in {"stop_id", "stop_code"}:
        logging.error(
            "GTFS_JOIN_KEY must be 'stop_id' or 'stop_code'; got '%s'.", GTFS_JOIN_KEY,
        )
        sys.exit(1)

    input_file: Path = INPUT_FILE_PATH

    # Build output file path
    base_name: str = input_file.stem
    output_fname: str = f"{base_name}{OUTPUT_FILE_SUFFIX}{OUTPUT_FILE_EXTENSION}"
    if OUTPUT_DIR:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_file: Path = OUTPUT_DIR / output_fname
    else:
        output_file = input_file.parent / output_fname

    # Read and validate
    ridership_df: pd.DataFrame = read_excel_file(input_file)
    verify_required_columns(ridership_df, REQUIRED_COLUMNS)

    # Apply optional filters
    filtered_data: pd.DataFrame = filter_data(
        ridership_df,
        routes=ROUTES,
        stop_ids=STOP_IDS,
        routes_exclude=ROUTES_EXCLUDE,
    )

    # Build & log selection-vs-system summary (filtered rows vs. full input)
    selection_summary: pd.DataFrame = build_selection_summary(ridership_df, filtered_data)

    # Log missing optional STOP_IDS
    log_missing_stop_ids(STOP_IDS, filtered_data["STOP_ID"])

    # Standardise TIME_PERIOD values
    filtered_data["TIME_PERIOD"] = filtered_data["TIME_PERIOD"].astype(str).str.strip().str.upper()

    # Aggregate + format
    final_filtered, aggregated_peaks, all_time_aggregated = process_aggregations(filtered_data)

    # Optional GTFS enrichment
    clean_stops: pd.DataFrame | None = None
    if GTFS_PATH is not None:
        gtfs_stops: pd.DataFrame = load_gtfs_stops(GTFS_PATH)
        logging.info(
            "Loaded GTFS: %d stops; joining ridership STOP_ID against GTFS '%s'.",
            len(gtfs_stops), GTFS_JOIN_KEY,
        )
        all_time_aggregated = enrich_with_gtfs(all_time_aggregated, gtfs_stops, GTFS_JOIN_KEY)
        aggregated_peaks = {
            period: enrich_with_gtfs(df_agg, gtfs_stops, GTFS_JOIN_KEY)
            for period, df_agg in aggregated_peaks.items()
        }
        clean_stops = build_clean_stops_sheet(final_filtered, gtfs_stops, GTFS_JOIN_KEY)

    # Write to disk
    write_to_excel(
        output_file,
        final_filtered,
        aggregated_peaks,
        all_time_aggregated,
        selection_summary=selection_summary,
        clean_stops=clean_stops,
    )

    logging.info("Script completed successfully.")


if __name__ == "__main__":
    main()
