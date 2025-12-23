"""Flag bus trips that exceed route-specific load-factor thresholds.

This module reads raw trip-level ridership data, calculates load factors,
applies optional route filters, classifies trips by service period, and flags
load-factor violations.  It then exports:

- A combined CSV (machine-readable).
- A combined single-sheet Excel file (quick inspection).
- Per-route Excel workbooks (one sheet per direction).
- A plain-text violation log (optional).

Typical use cases
- Operational load monitoring.
- Compliance tracking against agency load-factor standards.
- Automated route-level reporting.
"""

from __future__ import annotations

import datetime as dt
import logging
import os
from typing import Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_FILE: Final[str] = r"\\File\Path\To\Your\STATISTICS_BY_ROUTE_AND_TRIP.XLSX"
OUTPUT_FILE: Final[str] = INPUT_FILE.replace(".XLSX", "_processed.xlsx")
BUS_CAPACITY: Final[int] = 39

# Routes that get the *higher* (1.25) load factor limit
HIGHER_LIMIT_ROUTES: Final[list[str]] = []

# Routes that get the *lower* (1.0) limit
LOWER_LIMIT_ROUTES: Final[list[str]] = []

LOWER_LOAD_FACTOR_LIMIT: Final[float] = 1.0
HIGHER_LOAD_FACTOR_LIMIT: Final[float] = 1.25

# Provide these as lists of route strings.
# Leave them empty if you do not want any filtering.
FILTER_IN_ROUTES: list[str] = []  # e.g. ["101", "202"]
FILTER_OUT_ROUTES: list[str] = []  # e.g. ["105", "106"]

# Specify how many decimals to round the LOAD_FACTOR to
DECIMAL_PLACES: Final[int] = 4

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

WRITE_VIOLATION_LOG: Final[bool] = True
VIOLATION_LOG_FILE: Final[str] = OUTPUT_FILE.replace(".xlsx", "_violations_log.txt")

# =============================================================================
# FUNCTIONS
# =============================================================================


def load_data(input_file: str) -> pd.DataFrame:
    """Load required columns from an Excel file.

    Args:
    ----------
    input_file :
        Absolute or relative path to ``STATISTICS_BY_ROUTE_AND_TRIP.XLSX`` or
        a similar ridership file.

    Returns:
    -------
    pandas.DataFrame
        A DataFrame containing *only* the columns needed for downstream
        processing.

    Raises:
    ------
    FileNotFoundError
        If *input_file* does not exist.
    ValueError
        If the expected columns are missing.
    """
    data_frame = pd.read_excel(input_file)
    selected_columns = [
        "SERIAL_NUMBER",
        "ROUTE_NAME",
        "DIRECTION_NAME",
        "TRIP_START_TIME",
        "BLOCK",
        "MAX_LOAD",
        "MAX_LOAD_P",
        "ALL_RECORDS_MAX_LOAD",
    ]
    return data_frame[selected_columns]


def assign_service_period(ts: pd.Timestamp | dt.time) -> str:
    """Map a trip’s start time to a service-period label.

    Args:
        ts: A pandas/NumPy time-like scalar (``pd.Timestamp``, ``datetime.time``,
            or ``None``/``NaT``).

    Returns:
        One of the seven period strings shown in the table below.
    """
    hour = ts.hour
    if 4 <= hour < 6:
        return "AM Early"
    elif 6 <= hour < 9:
        return "AM Peak"
    elif 9 <= hour < 15:
        return "Midday"
    elif 15 <= hour < 18:
        return "PM Peak"
    elif 18 <= hour < 21:
        return "PM Late"
    elif 21 <= hour < 24:
        return "PM Nite"
    else:
        return "Other"


def get_route_load_limit(route_name: str) -> float:
    """Return the applicable load-factor limit for *route_name*.

    Args:
    ----------
    route_name :
        The short route designator as it appears in the source file.

    Returns:
    -------
    float
        ``LOWER_LOAD_FACTOR_LIMIT`` if the route is in
        :data:`LOWER_LIMIT_ROUTES`, else ``HIGHER_LOAD_FACTOR_LIMIT``.
    """
    if route_name in LOWER_LIMIT_ROUTES:
        return LOWER_LOAD_FACTOR_LIMIT
    return HIGHER_LOAD_FACTOR_LIMIT


def check_load_factor_violation(row: pd.Series) -> str:
    """Flag a single row as a load-factor violation.

    Args:
    ----------
    row :
        A DataFrame row that already contains ``LOAD_FACTOR`` and
        ``ROUTE_NAME``.

    Returns:
    -------
    str
        ``"TRUE"`` if the row exceeds its route limit, otherwise ``"FALSE"``.
    """
    limit = get_route_load_limit(row["ROUTE_NAME"])
    return "TRUE" if row["LOAD_FACTOR"] > limit else "FALSE"


def determine_limit_type(route_name: str) -> str:
    """Label the limit type used by *route_name*.

    Args:
    ----------
    route_name :
        The short route designator.

    Returns:
    -------
    str
        ``"LOW"`` if the route uses the lower limit, else ``"HIGH"``.
    """
    if route_name in LOWER_LIMIT_ROUTES:
        return "LOW"
    return "HIGH"


def process_data(
    data_frame: pd.DataFrame,
    bus_capacity: int,
    filter_in_routes: list,
    filter_out_routes: list,
    decimals: int,
) -> pd.DataFrame:
    """Transform raw ridership data into an analysis-ready DataFrame.

    The transformation pipeline:

        1. Apply `filter_in_routes` and `filter_out_routes`.
        2. Add column ``SERVICE_PERIOD``.
        3. Compute and round ``LOAD_FACTOR``.
        4. Add ``LOAD_FACTOR_VIOLATION`` and ``ROUTE_LIMIT_TYPE``.
        5. Sort rows by descending ``LOAD_FACTOR``.

    Args:
        data_frame (pd.DataFrame):
            Raw trip-level ridership data.
        bus_capacity (int):
            Seated + crush load used as the divisor when calculating the
            load factor.
        filter_in_routes (list[str]):
            If truthy, **keep only** routes whose short name appears here.
        filter_out_routes (list[str]):
            If truthy, **drop** routes whose short name appears here.
        decimals (int):
            Number of decimal places to retain for ``LOAD_FACTOR``.

    Returns:
        pandas.DataFrame:
            A fully processed and neatly sorted DataFrame.
    """
    # 1) Apply filters
    if filter_in_routes:
        data_frame = data_frame[data_frame["ROUTE_NAME"].isin(filter_in_routes)]
    if filter_out_routes:
        data_frame = data_frame[~data_frame["ROUTE_NAME"].isin(filter_out_routes)]

    # 2) Assign service period and calculate load factor
    data_frame["SERVICE_PERIOD"] = data_frame["TRIP_START_TIME"].apply(assign_service_period)
    data_frame["LOAD_FACTOR"] = data_frame["MAX_LOAD"] / bus_capacity

    # 5) Round load factor to specified decimals
    data_frame["LOAD_FACTOR"] = data_frame["LOAD_FACTOR"].round(decimals)

    # 3) Mark whether load factor is violated
    data_frame["LOAD_FACTOR_VIOLATION"] = data_frame.apply(check_load_factor_violation, axis=1)

    # 4) Add column for route limit type
    data_frame["ROUTE_LIMIT_TYPE"] = data_frame["ROUTE_NAME"].apply(determine_limit_type)

    # Sort by 'LOAD_FACTOR' in descending order
    return data_frame.sort_values(by="LOAD_FACTOR", ascending=False)


def create_route_workbooks(data_frame: pd.DataFrame) -> None:
    """Generate one Excel workbook per route, with sheets per direction.

    Workbooks are written to the directory containing :pydata:`OUTPUT_FILE`.

    Parameters:
    ----------
    data_frame :
        The processed DataFrame returned by :func:`process_data`.
    """
    # Determine the directory in which to save per-route files
    output_dir = os.path.dirname(OUTPUT_FILE) or "."
    os.makedirs(output_dir, exist_ok=True)

    # Group by each ROUTE_NAME
    for route_name, route_df in data_frame.groupby("ROUTE_NAME", sort=False):
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Within each route, group by DIRECTION_NAME
        for direction_name, direction_df in route_df.groupby("DIRECTION_NAME", sort=False):
            # Sort trips by TRIP_START_TIME
            direction_df_sorted = direction_df.sort_values(
                by="TRIP_START_TIME", kind="mergesort"
            ).reset_index(drop=True)

            ws = wb.create_sheet(title=str(direction_name))

            # Write header row (bolded)
            headers = list(direction_df_sorted.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # Write each trip row
            for row_idx, (_, row) in enumerate(direction_df_sorted.iterrows(), start=2):
                for col_idx, header in enumerate(headers, start=1):
                    val = row[header]
                    if header == "TRIP_START_TIME":
                        # Preserve time formatting if possible
                        if hasattr(val, "strftime"):
                            cell_val = val
                        elif pd.isna(val):
                            cell_val = ""
                        else:
                            cell_val = val
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                        cell.number_format = "hh:mm"
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=val)

            # Adjust column widths based on content
            for idx, col in enumerate(headers, start=1):
                content_series = direction_df_sorted[col].astype(str)
                max_length = max(content_series.map(len).max(), len(str(col)))
                adjusted_width = max_length + 2
                column_letter = get_column_letter(idx)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook named after the route
        filename = f"{route_name}.xlsx"
        file_path = os.path.join(output_dir, filename)
        wb.save(file_path)
        logging.info("Saved workbook: %s", file_path)


def export_to_csv(data_frame: pd.DataFrame, csv_file_path: str) -> None:
    """Write *data_frame* to disk as a CSV.

    Args:
    ----------
    data_frame :
        The DataFrame to export.
    csv_file_path :
        Destination path for the CSV file.  Any existing file is overwritten.
    """
    data_frame.to_csv(csv_file_path, index=False)
    logging.info("Processed file saved to CSV: %s", csv_file_path)


def export_to_excel(data_frame: pd.DataFrame, output_file: str) -> None:
    """Export *data_frame* to a single-sheet Excel workbook.

    Column widths are auto-sized for readability.

    Args:
    ----------
    data_frame :
        The DataFrame to export.
    output_file :
        Destination ``.xlsx`` path.
    """
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        data_frame.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]

        # Adjust column widths based on the maximum length of the content in each column
        for idx, col in enumerate(data_frame.columns, 1):
            series = data_frame[col].astype(str)
            max_length = max(series.map(len).max(), len(str(col)))
            adjusted_width = max_length + 2  # Add extra space for clarity
            column_letter = get_column_letter(idx)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def print_high_load_trips(data_frame: pd.DataFrame) -> None:
    """Print trips whose ``MAX_LOAD`` exceeds an absolute threshold.

    Args:
    ----------
    data_frame :
        The processed ridership DataFrame.
    """
    high_load_trips = data_frame[data_frame["MAX_LOAD"] > 30]
    if not high_load_trips.empty:
        logging.info("Trips with MAX_LOAD over 30:\n%s", high_load_trips)


def write_violation_log(data_frame: pd.DataFrame, log_file_path: str) -> None:
    """Write a plain-text log of trips exceeding their load-factor limit.

    Args:
    ----------
    data_frame :
        The processed ridership DataFrame.
    log_file_path :
        Full path to the ``.txt`` file to create or overwrite.
    """
    # Filter rows where load‐factor is violated
    violations_df = data_frame[data_frame["LOAD_FACTOR_VIOLATION"] == "TRUE"]

    # Open (or create) the log file and overwrite any existing content
    with open(log_file_path, "w", encoding="utf-8") as log_file:
        if violations_df.empty:
            log_file.write(
                "No load‐factor violations found (all trips within permissible limits).\n"
            )
        else:
            # Write a header
            header = (
                "Trips with load‐factor violations (greater than route‐specific limit):\n\n"
                "ROUTE\tDIRECTION\tSTART_TIME\tMAX_LOAD\tLOAD_FACTOR\t"
                "SERVICE_PERIOD\tROUTE_LIMIT_TYPE\n"
            )
            log_file.write(header)

            # Write one line per violating trip
            for _, row in violations_df.iterrows():
                # Format TRIP_START_TIME (if it’s a time object)
                start_val = row.get("TRIP_START_TIME", None)
                if hasattr(start_val, "strftime"):  # datetime.time or pandas Timestamp
                    start_str = start_val.strftime("%H:%M")
                else:
                    start_str = "" if pd.isna(start_val) else str(start_val)

                line = (
                    f"{row.get('ROUTE_NAME', '')}\t"
                    f"{row.get('DIRECTION_NAME', '')}\t"
                    f"{start_str}\t"
                    f"{row.get('MAX_LOAD', '')}\t"
                    f"{row.get('LOAD_FACTOR', '')}\t"
                    f"{row.get('SERVICE_PERIOD', '')}\t"
                    f"{row.get('ROUTE_LIMIT_TYPE', '')}\n"
                )
                log_file.write(line)
    logging.info("Exported load‐factor violation log: %s", log_file_path)


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """Run the full ETL pipeline and create all exports."""
    # Load data
    data_frame = load_data(INPUT_FILE)

    # Process data with filtering and limit checks
    processed_data = process_data(
        data_frame,
        BUS_CAPACITY,
        FILTER_IN_ROUTES,
        FILTER_OUT_ROUTES,
        DECIMAL_PLACES,
    )

    # -------------------------------------------------------------------------
    # 1) EXPORT COMBINED CSV (good for programmatic consumption)
    # -------------------------------------------------------------------------
    combined_csv_path = INPUT_FILE.replace(".XLSX", "_processed.csv")
    export_to_csv(processed_data, combined_csv_path)

    # -------------------------------------------------------------------------
    # 2) EXPORT COMBINED EXCEL (good for a quick, single-sheet view)
    # -------------------------------------------------------------------------
    export_to_excel(processed_data, OUTPUT_FILE)
    logging.info("Processed file saved to Excel: %s", OUTPUT_FILE)

    # -------------------------------------------------------------------------
    # 3) EXPORT PER-ROUTE EXCEL WORKBOOKS (one .xlsx per route, sheets per direction)
    # -------------------------------------------------------------------------
    create_route_workbooks(processed_data)

    # -------------------------------------------------------------------------
    # PRINT HIGH-LOAD TRIPS TO CONSOLE
    # -------------------------------------------------------------------------
    print_high_load_trips(processed_data)

    # -------------------------------------------------------------------------
    # WRITE TEXT LOG OF VIOLATIONS (good for a human-readable, line-by-line summary)
    # -------------------------------------------------------------------------
    if WRITE_VIOLATION_LOG:
        write_violation_log(processed_data, VIOLATION_LOG_FILE)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
