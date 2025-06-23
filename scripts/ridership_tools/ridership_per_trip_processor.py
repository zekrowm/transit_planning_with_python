"""Ridership-by-trip/direction/route report generator.

This script converts a single Excel workbook containing trip-level ridership
statistics into *one* Excel workbook per route.  Each route workbook contains
a worksheet for every operating direction, with:

- Cleaned and renamed columns
- Per-trip ridership as a share of the route-total
- Optional bar-charts of ridership by trip-start time
- Optional highlighting of “ultra-low” trips
"""

from __future__ import annotations

import datetime
import os
from typing import Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_FILE: Final[str] = r"\\Path\To\Your\STATISTICS_BY_ROUTE_AND_TRIP.XLSX"
OUTPUT_FOLDER: Final[str] = r"\\Path\To\Your\Output\Folder"

DATE_TYPE: Final[str] = "Weekday"  # Label for the type of day

# COLUMN CONFIGURATION --------------------------------------------------------

# List of columns to retain; each entry may optionally include “: Custom Header”.
COLUMNS_CONFIG: Final[list[str]] = [
    "SERIAL_NUMBER: Trip ID",
    "TRIP_START_TIME: Start Time",
    "ROUTE_NAME: Route",
    "DIRECTION_NAME: Direction",
    "PASSENGERS_ON: Passengers",
]

# Process COLUMNS_CONFIG into two structures:
COLUMNS_TO_RETAIN: list[str] = []  # populated programmatically
COLUMN_RENAME_MAP: dict[str, str] = {}  # populated programmatically
for col_entry in COLUMNS_CONFIG:
    if ":" in col_entry:
        original, new_name = col_entry.split(":", 1)
        original = original.strip()
        new_name = new_name.strip()
        COLUMNS_TO_RETAIN.append(original)
        COLUMN_RENAME_MAP[original] = new_name
    else:
        col_name = col_entry.strip()
        COLUMNS_TO_RETAIN.append(col_name)

# OTHER OPTIONS ---------------------------------------------------------------
CREATE_CHARTS: Final[bool] = True
FLAG_ULTRA_LOW: Final[bool] = False
ULTRA_LOW_THRESHOLD: Final[float] = 1.0

# FILTERING OPTIONS -----------------------------------------------------------

FILTER_COLUMN_NAME: Final[str | None] = "ROUTE_NAME"
FILTER_IN_LIST: Final[list[str]] = []
FILTER_OUT_LIST: Final[list[str]] = []

# =============================================================================
# FUNCTIONS
# =============================================================================


def load_data(input_file: str, columns_to_retain: list[str]) -> pd.DataFrame:
    """Load and normalise the ridership workbook.

    The function reads the Excel workbook located at *input_file*, coerces
    ``TRIP_START_TIME`` to :class:`datetime.time`, and returns a DataFrame that
    contains only *columns_to_retain* (in the supplied order).

    Args:
        input_file: Fully-qualified path to the **source** Excel workbook.
        columns_to_retain: Ordered list of column names whose data must be kept.
            Any name absent from the workbook is silently ignored.

    Returns:
        A tidy :class:`pandas.DataFrame` with the requested columns.
        If ``TRIP_START_TIME`` is present, its dtype is guaranteed to be
        :class:`datetime.time` or *NaT*.

    Raises:
        FileNotFoundError: If *input_file* cannot be located.
        ValueError: If none of *columns_to_retain* exist in the workbook.

    Examples:
        >>> df = load_data("stats.xlsx", ["SERIAL_NUMBER", "TRIP_START_TIME"])
        >>> df.dtypes["TRIP_START_TIME"]
        dtype('O')  # actually python ``datetime.time`` objects
    """
    # 1. Read the Excel file
    df = pd.read_excel(input_file)

    # 2. Normalise 'TRIP_START_TIME' if it exists
    if "TRIP_START_TIME" in df.columns:
        series = df["TRIP_START_TIME"]

        if pd.api.types.is_datetime64_any_dtype(series):
            # pandas already gave us datetime64[ns]; convert safely for mypy
            df["TRIP_START_TIME"] = series.apply(lambda ts: ts.time() if not pd.isna(ts) else None)
        else:
            # Parse strings / other representations into datetime
            parsed = pd.to_datetime(
                series.astype(str).str.strip(),
                errors="coerce",  # invalid rows → NaT
            )
            # Extract the .time() without touching the .dt accessor (mypy-safe)
            df["TRIP_START_TIME"] = parsed.apply(lambda ts: ts.time() if not pd.isna(ts) else None)

    # 3. Retain only columns that actually exist
    existing_cols = [c for c in columns_to_retain if c in df.columns]
    if not existing_cols:
        raise ValueError("None of the requested columns exist in the provided workbook.")

    return df[existing_cols]


def create_output_folder(folder_path: str) -> None:
    """Create *folder_path* (and any parents) if it does not yet exist.

    Args:
        folder_path: Destination directory for the per-route workbooks.

    Returns:
        None

    Notes:
        This is essentially a thin wrapper around
        :pyfunc:`os.makedirs(..., exist_ok=True)` and thus never raises an
        exception if the directory already exists.
    """
    os.makedirs(folder_path, exist_ok=True)


def write_direction_sheet(
    wb: Workbook,
    direction_df: pd.DataFrame,
    direction_name: str,
    date_type: str,
    create_charts: bool,
    flag_ultra_low: bool,
    ultra_low_threshold: float,
) -> None:
    """Populate a single worksheet for *direction_name*.

    Steps performed:

    1. Sort *direction_df* chronologically by ``TRIP_START_TIME``.
    2. Compute each trip’s “Percent of Route Ridership”.
    3. Bold-face the trip with maximum passengers.
    4. Optionally flag “ultra-low” trips (≤ *ultra_low_threshold*) in **red**.
    5. Write data + headers to a new worksheet.
    6. Optionally attach an in-sheet bar chart.

    Args:
        wb: An *open* :class:`openpyxl.workbook.Workbook` to which the sheet
            will be added.
        direction_df: Sub-DataFrame containing only rows for the current
            route *and* direction.
        direction_name: Human-readable direction label (e.g. ``'Eastbound'``).
        date_type: Descriptor shown in the chart title (“Weekday”, “Saturday”,
            etc.).
        create_charts: If ``True``, embeds a ridership bar-chart.
        flag_ultra_low: If ``True``, highlights ultra-low trips.
        ultra_low_threshold: Numeric passenger cutoff that defines
            “ultra-low”.

    Returns:
        None.  *wb* is modified in-place.

    Raises:
        ValueError: If *direction_df* lacks required columns.

    Side Effects:
        Adds a new worksheet named *direction_name* to *wb*; writes cell
        formats, fonts, and (optionally) a chart.

    Example:
        >>> write_direction_sheet(wb, east_df, "East", "Weekday", True, False, 1.0)
    """
    # 1. Locally sort by TRIP_START_TIME (NaT last)
    direction_df = direction_df.sort_values("TRIP_START_TIME", na_position="last").reset_index(
        drop=True
    )

    # 2. Create a new worksheet named after this direction
    ws = wb.create_sheet(title=direction_name)

    # 3. Calculate total ridership and percentage share for each trip
    total_ridership = direction_df["PASSENGERS_ON"].sum()
    if total_ridership != 0:
        direction_df["Percent of Route Ridership"] = (
            direction_df["PASSENGERS_ON"] / total_ridership
        ).apply(lambda x: round(x * 100, 1) / 100)
    else:
        direction_df["Percent of Route Ridership"] = 0.0

    # 4. Round PASSENGERS_ON to 1 decimal
    direction_df["PASSENGERS_ON"] = direction_df["PASSENGERS_ON"].round(1)

    # 5. Determine the maximum ridership value (for bolding)
    max_ridership = direction_df["PASSENGERS_ON"].max()

    # 6. Write headers (using COLUMN_RENAME_MAP where provided)
    headers = list(direction_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        display = COLUMN_RENAME_MAP.get(header, header)
        ws.cell(row=1, column=col_idx, value=display).font = Font(bold=True)

    # 7. Write each data row
    for row_idx, (_, row_data) in enumerate(direction_df.iterrows(), start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if header == "TRIP_START_TIME":
                val = row_data[header]
                # If val is a Python time object
                if isinstance(val, type(datetime.datetime.now().time())):
                    cell.value = val
                    cell.number_format = "hh:mm"
                # If val is a pandas Timestamp (just in case)
                elif isinstance(val, pd.Timestamp):
                    cell.value = val.time()
                    cell.number_format = "hh:mm"
                else:
                    # Blank if unparseable / NaT
                    cell.value = None
            else:
                cell.value = row_data[header]

            # Format percentage column
            if header == "Percent of Route Ridership":
                cell.number_format = "0.0%"

        # 8. Bold the row if this trip has the maximum ridership
        if row_data["PASSENGERS_ON"] == max_ridership:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c_idx).font = Font(bold=True)

        # 9. Optionally flag ultra‐low trips (≤ ULTRA_LOW_THRESHOLD) in red
        if flag_ultra_low and row_data["PASSENGERS_ON"] <= ultra_low_threshold:
            for c_idx in range(1, len(headers) + 1):
                orig = ws.cell(row=row_idx, column=c_idx).font
                ws.cell(row=row_idx, column=c_idx).font = Font(
                    name=orig.name,
                    size=orig.size,
                    bold=orig.bold,
                    color="FF0000",  # red
                )

    # 10. Optionally create a bar chart of ‘Passengers’ vs. ‘Start Time’
    if create_charts:
        chart = BarChart()
        chart.title = f"Ridership by Time – {direction_name} ({date_type})"
        chart.x_axis.title = "Trip Start Time"
        chart.y_axis.title = "Ridership"

        pass_col_idx = headers.index("PASSENGERS_ON") + 1
        time_col_idx = headers.index("TRIP_START_TIME") + 1
        first_data_row = 2
        last_data_row = direction_df.shape[0] + 1

        data_series = Reference(
            ws,
            min_col=pass_col_idx,
            min_row=first_data_row,
            max_col=pass_col_idx,
            max_row=last_data_row,
        )
        chart.add_data(data_series, titles_from_data=False)

        categories = Reference(
            ws,
            min_col=time_col_idx,
            min_row=first_data_row,
            max_row=last_data_row,
        )
        chart.set_categories(categories)

        anchor_cell = f"{get_column_letter(len(headers) + 2)}2"
        ws.add_chart(chart, anchor_cell)


def create_route_workbook(
    route_name: str,
    route_df: pd.DataFrame,
    output_folder: str,
    date_type: str,
    create_charts: bool,
    flag_ultra_low: bool,
    ultra_low_threshold: float,
) -> None:
    """Create and save the per-route workbook.

    Args:
        route_name: Label used both for the Excel filename and for logging.
        route_df: DataFrame containing *all* trips for the route, regardless of
            direction.
        output_folder: Directory in which the workbook will be written.
        date_type: Forwarded to :func:`write_direction_sheet`.
        create_charts: Forwarded to :func:`write_direction_sheet`.
        flag_ultra_low: Forwarded to :func:`write_direction_sheet`.
        ultra_low_threshold: Forwarded to :func:`write_direction_sheet`.

    Returns:
        None

    Side Effects:
        * Writes ``{route_name}.xlsx`` to *output_folder*.
        * Prints a confirmation to ``stdout``.
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for direction_name, direction_df in route_df.groupby("DIRECTION_NAME", sort=False):
        write_direction_sheet(
            wb=wb,
            direction_df=direction_df,
            direction_name=str(direction_name),
            date_type=date_type,
            create_charts=create_charts,
            flag_ultra_low=flag_ultra_low,
            ultra_low_threshold=ultra_low_threshold,
        )

    output_path = os.path.join(output_folder, f"{route_name}.xlsx")
    wb.save(output_path)
    print(f"Saved workbook: {output_path}")


def main() -> None:
    """Entry point when the module is executed as a script.

    Workflow:

    1. Load and optionally filter the ridership data set.
    2. Sort rows to guarantee deterministic sheet order.
    3. Ensure the output directory exists.
    4. (Optional) Write a plain-text log of ultra-low trips.
    5. Generate one workbook per route.

    Returns:
        None

    Examples:
        >>> if __name__ == "__main__":
        ...     main()
    """
    # 1. Load
    df = load_data(INPUT_FILE, COLUMNS_TO_RETAIN)

    # 2. Optional filtering
    if FILTER_COLUMN_NAME:
        if FILTER_IN_LIST:
            df = df[df[FILTER_COLUMN_NAME].isin(FILTER_IN_LIST)]
        if FILTER_OUT_LIST:
            df = df[~df[FILTER_COLUMN_NAME].isin(FILTER_OUT_LIST)]

    # 3. Global sort – so that groupby("ROUTE_NAME") inherits chronological order
    df = df.sort_values(
        by=["ROUTE_NAME", "DIRECTION_NAME", "TRIP_START_TIME"],
        kind="mergesort",  # stable sort
    ).reset_index(drop=True)

    # 4. Ensure output folder exists
    create_output_folder(OUTPUT_FOLDER)

    # 5. If FLAG_ULTRA_LOW, write a log of all trips ≤ ULTRA_LOW_THRESHOLD
    if FLAG_ULTRA_LOW:
        low_df = df[df["PASSENGERS_ON"] <= ULTRA_LOW_THRESHOLD]
        log_path = os.path.join(OUTPUT_FOLDER, "ultra_low_ridership_log.txt")

        with open(log_path, "w", encoding="utf-8") as log_file:
            if low_df.empty:
                log_file.write(
                    f"No trips found with ultra-low ridership (≤ {ULTRA_LOW_THRESHOLD}).\n"
                )
            else:
                log_file.write(f"Trips with ultra-low ridership (≤ {ULTRA_LOW_THRESHOLD}):\n\n")
                for _, row in low_df.iterrows():
                    trip_id = row.get("SERIAL_NUMBER", "")
                    route = row.get("ROUTE_NAME", "")
                    direction = row.get("DIRECTION_NAME", "")
                    passengers = row.get("PASSENGERS_ON", "")
                    start_val = row.get("TRIP_START_TIME", None)
                    if isinstance(start_val, type(datetime.datetime.now().time())):
                        start_str = start_val.strftime("%H:%M")
                    else:
                        start_str = "" if pd.isna(start_val) else str(start_val)
                    log_file.write(
                        f"Route: {route}, Direction: {direction}, "
                        f"Trip ID: {trip_id}, Start Time: {start_str}, "
                        f"Passengers: {passengers}\n"
                    )
        print(f"Exported ultra-low ridership log: {log_path}")

    # 6. Create one workbook per route
    for route_name, route_df in df.groupby("ROUTE_NAME", sort=False):
        create_route_workbook(
            route_name=str(route_name),
            route_df=route_df,
            output_folder=OUTPUT_FOLDER,
            date_type=DATE_TYPE,
            create_charts=CREATE_CHARTS,
            flag_ultra_low=FLAG_ULTRA_LOW,
            ultra_low_threshold=ULTRA_LOW_THRESHOLD,
        )


if __name__ == "__main__":
    main()
