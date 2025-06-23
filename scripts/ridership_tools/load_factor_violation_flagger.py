"""Flag bus trips that exceed route-specific load-factor thresholds.

This module reads raw trip-level ridership data, calculates load factors,
applies optional route filters, classifies trips by service period, and flags
load-factor violations.  It then exports:

- A combined CSV (machine-readable).
- A combined single-sheet Excel file (quick inspection).
- Per-route Excel workbooks (one sheet per direction).
- A plain-text violation log (optional).

Typical use cases
- Operational load monitoring.
- Compliance tracking against agency load-factor standards.
- Automated route-level reporting.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Literal, Sequence, TypeAlias, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pandas import DataFrame, Series

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_FILE: Path = Path(r"\\File\Path\To\Your\STATISTICS_BY_ROUTE_AND_TRIP.XLSX")

OUTPUT_DIR: Path = Path(r"\\File\Path\To\Your")  # ← mirrors INPUT_FILE style
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Derived output filenames
OUTPUT_FILE: Path = OUTPUT_DIR / f"{INPUT_FILE.stem}_processed.xlsx"
VIOLATION_LOG_FILE: Path = OUTPUT_DIR / f"{INPUT_FILE.stem}_violations_log.txt"

# ---------------------------------------------------------------------------
# Other parameters
# ---------------------------------------------------------------------------
BUS_CAPACITY: int = 39

HIGHER_LIMIT_ROUTES: set[str] = {"101", "102", "103", "104"}
LOWER_LIMIT_ROUTES: set[str] = {"105", "106"}

LOWER_LOAD_FACTOR_LIMIT: float = 1.0
HIGHER_LOAD_FACTOR_LIMIT: float = 1.25

FILTER_IN_ROUTES: set[str] = set()  # e.g. {"101", "202"}
FILTER_OUT_ROUTES: set[str] = set()  # e.g. {"105", "106"}

DECIMAL_PLACES: int = 4
WRITE_VIOLATION_LOG: bool = True

# ----------------------------------------------------------------------------
# TYPE ALIASES
# ----------------------------------------------------------------------------

LoadFlag: TypeAlias = Literal["TRUE", "FALSE"]
LimitType: TypeAlias = Literal["LOW", "HIGH"]
ServicePeriod: TypeAlias = Literal[
    "AM Early",
    "AM Peak",
    "Midday",
    "PM Peak",
    "PM Late",
    "PM Nite",
    "Other",
]
TimeLike: TypeAlias = Union[pd.Timestamp, _dt.time, None]
Row: TypeAlias = Series

# ============================================================================
# FUNCTIONS
# ============================================================================


def load_data(input_file: Path) -> DataFrame:
    """Load and return only the columns required for processing."""
    required_cols: list[str] = [
        "SERIAL_NUMBER",
        "ROUTE_NAME",
        "DIRECTION_NAME",
        "TRIP_START_TIME",
        "BLOCK",
        "MAX_LOAD",
        "MAX_LOAD_P",
        "ALL_RECORDS_MAX_LOAD",
    ]
    df: DataFrame = pd.read_excel(input_file)
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Source file missing columns: {missing}")
    return df[required_cols]


def assign_service_period(ts: TimeLike) -> ServicePeriod:
    """Map a trip’s start time to a 7-bucket service-period label."""
    if ts is None or pd.isna(ts):
        return "Other"
    hour: int = ts.hour  # type: ignore[attr-defined]
    if 4 <= hour < 6:
        return "AM Early"
    if 6 <= hour < 9:
        return "AM Peak"
    if 9 <= hour < 15:
        return "Midday"
    if 15 <= hour < 18:
        return "PM Peak"
    if 18 <= hour < 21:
        return "PM Late"
    if 21 <= hour < 24:
        return "PM Nite"
    return "Other"


def get_route_load_limit(route_name: str) -> float:
    """Return the limit applicable to *route_name*."""
    return LOWER_LOAD_FACTOR_LIMIT if route_name in LOWER_LIMIT_ROUTES else HIGHER_LOAD_FACTOR_LIMIT


def check_load_factor_violation(row: Row) -> LoadFlag:
    """Flag a row as violating its route’s load-factor limit."""
    limit = get_route_load_limit(str(row["ROUTE_NAME"]))
    return "TRUE" if float(row["LOAD_FACTOR"]) > limit else "FALSE"


def determine_limit_type(route_name: str) -> LimitType:
    """Return 'LOW' or 'HIGH' limit type for *route_name*."""
    return "LOW" if route_name in LOWER_LIMIT_ROUTES else "HIGH"


def process_data(
    data_frame: DataFrame,
    bus_capacity: int,
    filter_in_routes: Sequence[str],
    filter_out_routes: Sequence[str],
    decimals: int,
) -> DataFrame:
    """Transform raw ridership data into an analysis-ready DataFrame."""
    df: DataFrame = data_frame.copy()

    if filter_in_routes:
        df = df[df["ROUTE_NAME"].isin(filter_in_routes)]
    if filter_out_routes:
        df = df[~df["ROUTE_NAME"].isin(filter_out_routes)]

    df["SERVICE_PERIOD"] = df["TRIP_START_TIME"].apply(assign_service_period)
    df["LOAD_FACTOR"] = (df["MAX_LOAD"] / bus_capacity).round(decimals)
    df["LOAD_FACTOR_VIOLATION"] = df.apply(check_load_factor_violation, axis=1)
    df["ROUTE_LIMIT_TYPE"] = df["ROUTE_NAME"].apply(determine_limit_type)

    return df.sort_values(by="LOAD_FACTOR", ascending=False).reset_index(drop=True)


def create_route_workbooks(data_frame: DataFrame) -> None:
    """Generate one Excel workbook per route, with sheets split by direction."""
    for route_name, route_df in data_frame.groupby("ROUTE_NAME", sort=False):
        wb = Workbook()
        wb.remove(wb.active)

        for direction_name, direction_df in route_df.groupby("DIRECTION_NAME", sort=False):
            sheet_df = direction_df.sort_values(by="TRIP_START_TIME", kind="mergesort").reset_index(
                drop=True
            )
            ws = wb.create_sheet(title=str(direction_name))

            headers = list(sheet_df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            for row_idx, (_, row) in enumerate(sheet_df.iterrows(), start=2):
                for col_idx, header in enumerate(headers, 1):
                    val = row[header]
                    if header == "TRIP_START_TIME" and hasattr(val, "strftime"):
                        cell_obj = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell_obj.number_format = "hh:mm"
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=val)

            for idx, col in enumerate(headers, 1):
                max_len = max(sheet_df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

        filepath: Path = OUTPUT_DIR / f"{route_name}.xlsx"
        wb.save(filepath)
        print(f"Saved workbook: {filepath}")


def export_to_csv(data_frame: DataFrame, csv_path: Path) -> None:
    """Export *data_frame* to a CSV file."""
    data_frame.to_csv(csv_path, index=False)
    print(f"Processed file saved to CSV: {csv_path}")


def export_to_excel(data_frame: DataFrame, output_file: Path) -> None:
    """Export *data_frame* to a single-sheet Excel workbook."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:  # type: ignore[attr-defined]  # noqa: E501
        data_frame.to_excel(writer, index=False, sheet_name="Sheet1")
        sheet = writer.sheets["Sheet1"]
        for idx, col in enumerate(data_frame.columns, 1):
            max_len = max(data_frame[col].astype(str).map(len).max(), len(col))
            sheet.column_dimensions[get_column_letter(idx)].width = max_len + 2


def print_high_load_trips(data_frame: DataFrame) -> None:
    """Emit trips whose MAX_LOAD exceeds an absolute threshold (30)."""
    high = data_frame[data_frame["MAX_LOAD"] > 30]
    if not high.empty:
        print("Trips with MAX_LOAD over 30:")
        print(high)


def write_violation_log(data_frame: DataFrame, log_path: Path) -> None:
    """Write a line-delimited log of load-factor violations."""
    violations = data_frame[data_frame["LOAD_FACTOR_VIOLATION"] == "TRUE"]

    with log_path.open("w", encoding="utf-8") as fh:
        if violations.empty:
            fh.write("No load-factor violations found (all trips within limits).\n")
            return

        fh.write(
            "Trips with load-factor violations (greater than route-specific limit):\n\n"
            "ROUTE\tDIRECTION\tSTART_TIME\tMAX_LOAD\tLOAD_FACTOR\tSERVICE_PERIOD\tROUTE_LIMIT_TYPE\n"  # noqa: E501
        )

        for _, row in violations.iterrows():
            ts = row["TRIP_START_TIME"]
            start = (
                ts.strftime("%H:%M")
                if hasattr(ts, "strftime")
                else ("" if pd.isna(ts) else str(ts))
            )
            fh.write(
                f"{row['ROUTE_NAME']}\t{row['DIRECTION_NAME']}\t{start}\t"
                f"{row['MAX_LOAD']}\t{row['LOAD_FACTOR']}\t"
                f"{row['SERVICE_PERIOD']}\t{row['ROUTE_LIMIT_TYPE']}\n"
            )
    print(f"Exported load-factor violation log: {log_path}")


# ============================================================================
# MAIN
# ============================================================================


def main() -> None:
    """Run the full ETL pipeline and create all exports."""
    df_raw = load_data(INPUT_FILE)

    df_proc = process_data(
        df_raw,
        BUS_CAPACITY,
        FILTER_IN_ROUTES,
        FILTER_OUT_ROUTES,
        DECIMAL_PLACES,
    )

    csv_path: Path = OUTPUT_DIR / f"{INPUT_FILE.stem}_processed.csv"
    export_to_csv(df_proc, csv_path)

    export_to_excel(df_proc, OUTPUT_FILE)
    print(f"Processed file saved to Excel: {OUTPUT_FILE}")

    create_route_workbooks(df_proc)

    print_high_load_trips(df_proc)

    if WRITE_VIOLATION_LOG:
        write_violation_log(df_proc, VIOLATION_LOG_FILE)


if __name__ == "__main__":
    main()
