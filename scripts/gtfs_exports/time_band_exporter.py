"""Export consolidated time-band tables from GTFS schedule data.

The script groups trips that share **both** an identical sequence of
time-points *and* identical segment run-times. For each unique
(time-point pattern + segment runtimes) combination it produces a single
row with:

- **FrTime / ToTime** – earliest and latest first-stop departures
- **Segment columns** – runtime (minutes) from the previous time-point
  (first cell is ``MISSING_TIME``)

Output structure
----------------
- One Excel workbook per ``(route_id, service_id)``
- One sheet per ``direction_id`` within that workbook
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

GTFS_FOLDER = Path(r"Path\To\Your\GTFS_Folder")
OUTPUT_FOLDER = Path(r"Path\To\Your\Output_Folder")

# Optional filters – leave empty to take everything
FILTER_IN_ROUTE_SHORT_NAMES: List[str] = []
FILTER_OUT_ROUTE_SHORT_NAMES: List[str] = []
FILTER_IN_SERVICE_IDS: List[str] = []
FILTER_OUT_SERVICE_IDS: List[str] = []

# --------------------------------------------------------------------------------------------------
# LOGGING
# --------------------------------------------------------------------------------------------------

LOG_LEVEL = logging.INFO
EXPORT_TIMEPOINTS_ONLY = True  # keep only stops where timepoint == 1
MISSING_TIME = "–"

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$")

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def hhmmss_to_min(t: Optional[str]) -> Optional[int]:
    """Convert *HH:MM* or *HH:MM:SS* to absolute minutes.

    A schedule can exceed 24 h (e.g. ``'27:35:00'``); those values are
    preserved as numbers ≥ 1440.

    Args:
        t: Time string in GTFS format or None.

    Returns:
        The number of minutes since 00:00 or None when t is missing or malformed.
    """
    if not isinstance(t, str):
        return None
    m = _TIME_RE.match(t.strip())
    if not m:
        return None
    h, mm, ss = m.groups()
    return int(h) * 60 + int(mm) + round(int(ss or 0) / 60)


def min_to_hhmm(mn: Optional[int]) -> str:
    """Convert minutes back to 'H:MM' (24 h-plus safe).

    Args:
        mn: Minutes since midnight or None.

    Returns:
        A string such as '5:07' or the sentinel MISSING_TIME when mn is None.
    """
    if mn is None:
        return MISSING_TIME
    h, m = divmod(mn, 60)
    return f"{h}:{m:02d}"


def safe_sheet(name: str) -> str:
    """Return an Excel-compliant sheet name.

    Excel forbids certain characters and limits names to 31
    characters. Invalid characters are replaced with “_”.

    Args:
        name: Desired sheet name.

    Returns:
        A sanitized sheet name that is never empty.
    """
    return re.sub(r"[:\\/*?\[\]]", "_", name)[:31] or "Sheet"


# ─────────────────────────  LOAD GTFS  ────────────────────────────── #
REQ = ["trips.txt", "stop_times.txt", "routes.txt", "stops.txt"]


def load_gtfs(folder: Path) -> Dict[str, pd.DataFrame]:
    """Load the four core GTFS text files into memory.

    Args:
        folder: Path that contains trips.txt, stop_times.txt,
            routes.txt and stops.txt.

    Returns:
        A mapping {'trips': df, 'stop_times': df, ...} with all
        columns typed as str except stop_sequence (int) and
        timepoint (float).

    Raises:
        FileNotFoundError: If any of the required files are missing.
        ValueError: If numeric coercion fails.
    """
    missing = [f for f in REQ if not (folder / f).exists()]
    if missing:
        raise FileNotFoundError("Missing GTFS file(s): " + ", ".join(missing))

    data = {f[:-4]: pd.read_csv(folder / f, dtype=str, low_memory=False) for f in REQ}

    st = data["stop_times"]
    st["stop_sequence"] = pd.to_numeric(st["stop_sequence"], errors="raise")
    if "timepoint" in st.columns:
        st["timepoint"] = pd.to_numeric(st["timepoint"], errors="coerce")
    return data


# ──────────────────  PER-TRIP PATTERN & RUNTIMES  ─────────────────── #
PatternTuple = Tuple[str, ...]
RuntimeSegTuple = Tuple[Union[int, str], ...]  # first element is "–"


def segment_runtimes(grp: pd.DataFrame) -> RuntimeSegTuple:
    """Calculate segment runtimes for a single trip.

    Args:
        grp: A stop_times subset for one trip, ordered by
            stop_sequence and containing arrival_time and/or departure_time.

    Returns:
        A tuple where element 0 is MISSING_TIME followed by the running
        time (minutes) between successive stops. An empty string ''
        marks segments where either time is missing.
    """
    times: List[Optional[int]] = []
    for _, row in grp.iterrows():
        dep = hhmmss_to_min(row.get("departure_time"))
        arr = hhmmss_to_min(row.get("arrival_time"))
        times.append(dep if dep is not None else arr)

    segs: List[Union[int, str]] = [MISSING_TIME]
    for a, b in zip(times, times[1:], strict=True):
        segs.append(b - a if a is not None and b is not None else "")
    return tuple(segs)
  

def build_index(
    gtfs: Dict[str, pd.DataFrame],
) -> Tuple[
    pd.DataFrame,
    Dict[int, PatternTuple],
    Dict[int, RuntimeSegTuple],
    Dict[int, List[str]],
]:
    """Build the master trip index and ancillary lookup tables.

    Filtering is applied according to the global FILTER_* lists.

    Args:
        gtfs: Output from load_gtfs().

    Returns:
        Tuple of:
            - trips_df: One row per trip with pattern/segment hashes.
            - stop_dict: Pattern hash → tuple of stop_id strings.
            - seg_dict: Segment hash → tuple of segment runtimes.
            - header_names: Pattern hash → list of 'Stop name (code)' strings.
    """
    trips = gtfs["trips"].copy()
    routes = gtfs["routes"][["route_id", "route_short_name"]]
    trips = trips.merge(routes, on="route_id", how="left")

    # Apply filters
    if FILTER_IN_ROUTE_SHORT_NAMES:
        trips = trips[trips["route_short_name"].isin(FILTER_IN_ROUTE_SHORT_NAMES)]
    if FILTER_OUT_ROUTE_SHORT_NAMES:
        trips = trips[~trips["route_short_name"].isin(FILTER_OUT_ROUTE_SHORT_NAMES)]
    if FILTER_IN_SERVICE_IDS:
        trips = trips[trips["service_id"].isin(FILTER_IN_SERVICE_IDS)]
    if FILTER_OUT_SERVICE_IDS:
        trips = trips[~trips["service_id"].isin(FILTER_OUT_SERVICE_IDS)]

    if trips.empty:
        return pd.DataFrame(), {}, {}, {}

    st = gtfs["stop_times"].merge(
        trips[["trip_id", "route_id", "service_id", "direction_id"]],
        on="trip_id",
        how="inner",
    )

    if EXPORT_TIMEPOINTS_ONLY and "timepoint" in st.columns:
        st = st[st["timepoint"] == 1]

    st = st.sort_values(["trip_id", "stop_sequence"])

    stop_dict: Dict[int, PatternTuple] = {}
    seg_dict: Dict[int, RuntimeSegTuple] = {}
    header_names: Dict[int, List[str]] = {}
    rows = []

    stops_lookup = (
        gtfs["stops"][["stop_id", "stop_name", "stop_code"]].set_index("stop_id").to_dict("index")
    )

    for _trip_id, grp in st.groupby("trip_id"):
        pattern: PatternTuple = tuple(grp["stop_id"])
        pat_hash = hash(pattern)
        stop_dict.setdefault(pat_hash, pattern)

        segs = segment_runtimes(grp)
        seg_hash = hash(segs)
        seg_dict.setdefault(seg_hash, segs)

        if pat_hash not in header_names:
            header_names[pat_hash] = [
                f"{rec['stop_name']} ({rec['stop_code']})"
                if (rec := stops_lookup.get(sid)) is not None
                else sid
                for sid in pattern
            ]

        start = hhmmss_to_min(grp.iloc[0].get("departure_time") or grp.iloc[0].get("arrival_time"))
        if start is None:
            continue

        rows.append(
            {
                "route_id": grp.iloc[0]["route_id"],
                "service_id": grp.iloc[0]["service_id"],
                "direction_id": grp.iloc[0]["direction_id"],
                "pattern_hash": pat_hash,
                "seg_hash": seg_hash,
                "start": start,
            }
        )

    return pd.DataFrame(rows), stop_dict, seg_dict, header_names


# ───────────────────  GROUP into TIME-BAND ROWS  ──────────────────── #
def make_bands(idx: pd.DataFrame) -> pd.DataFrame:
    """Aggregate trips into time-bands.

    Args:
        idx: DataFrame returned by build_index().

    Returns:
        A DataFrame where each row represents a unique
        (route_id, service_id, direction_id, pattern_hash, seg_hash)
        with FrTime, ToTime and Total (trip count).
    """
    gb = idx.groupby(["route_id", "service_id", "direction_id", "pattern_hash", "seg_hash"])
    out = (
        gb["start"]
        .agg(["min", "max", "count"])
        .rename(columns={"min": "FrTime", "max": "ToTime", "count": "Total"})
        .reset_index()
        .sort_values("FrTime")
    )
    return out


# ──────────────────────────  EXCEL EXPORT  ────────────────────────── #
def export_excel(
    bands: pd.DataFrame,
    stop_dict: Dict[int, PatternTuple],
    seg_dict: Dict[int, RuntimeSegTuple],
    header_names: Dict[int, List[str]],
    routes: pd.DataFrame,
) -> None:
    """Write one workbook per route/service with time-band tables.

    Args:
        bands: Output from make_bands().
        stop_dict: Pattern hash lookup from build_index().
        seg_dict: Segment hash lookup from build_index().
        header_names: Column headers keyed by pattern hash.
        routes: Original routes.txt table (for short names).

    Side Effects:
        Creates one .xlsx file per (route_id, service_id)
        in OUTPUT_FOLDER and logs each filename.
    """
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    route_short = routes.set_index("route_id")["route_short_name"].to_dict()

    for (rid, sid), grp_rs in bands.groupby(["route_id", "service_id"]):
        wb = Workbook()
        wb.remove(wb.active)

        for did, grp_dir in grp_rs.groupby("direction_id"):
            ws = wb.create_sheet(safe_sheet(f"Dir_{did or 'X'}"))

            # header uses first pattern in this direction
            first_pat = grp_dir.iloc[0]["pattern_hash"]
            hdr = ["Pattern", "FrTime", "ToTime", "Total", *header_names[first_pat]]
            ws.append(hdr)

            for _, r in grp_dir.iterrows():
                segs = seg_dict[r["seg_hash"]]
                ws.append(
                    [
                        r["pattern_hash"],
                        min_to_hhmm(int(r["FrTime"])),
                        min_to_hhmm(int(r["ToTime"])),
                        r["Total"],
                        *segs,
                    ]
                )

            for col in range(1, len(hdr) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

        fname = OUTPUT_FOLDER / (f"route_{route_short.get(rid, rid)}_cal{sid}_timeband_table.xlsx")
        wb.save(fname)
        logging.info("Wrote %s", fname)


# ==================================================================================================
# MAIN
# ==================================================================================================


def main() -> None:
    """Entry-point when the module is executed as a script."""
    logging.basicConfig(
        level=LOG_LEVEL,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.info("GTFS folder   : %s", GTFS_FOLDER)
    logging.info("Output folder : %s", OUTPUT_FOLDER)

    gtfs = load_gtfs(GTFS_FOLDER)
    idx, stop_dict, seg_dict, header_names = build_index(gtfs)

    if idx.empty:
        logging.warning("Nothing to export – check filters or timepoints.")
        return

    bands = make_bands(idx)
    export_excel(bands, stop_dict, seg_dict, header_names, gtfs["routes"])

    logging.info("Finished – %d band rows.", len(bands))


if __name__ == "__main__":
    main()
