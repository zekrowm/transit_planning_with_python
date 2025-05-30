"""
Script Name:
    bus_schedule_exporter.py

Purpose:
    Processes GTFS data to generate formatted Excel transit schedules per route,
    direction, and service ID. Imitates public transit schedules and is
    especially useful for processing for a large number of routes.

Inputs:
    1. GTFS text files (e.g., stops.txt, routes.txt, trips.txt, stop_times.txt,
       calendar.txt) located in GTFS_FOLDER_PATH.
    2. User-defined configurations:
       - GTFS_FOLDER_PATH: Path to the directory containing GTFS files.
       - BASE_OUTPUT_PATH: Path to the directory where output Excel files will be
         saved.
       - FILTER_SERVICE_IDS: List of service_ids to process (empty processes all).
       - FILTER_IN_ROUTES: List of route_short_names to include (empty includes all
         not filtered out).
       - FILTER_OUT_ROUTES: List of route_short_names to exclude.
       - TIME_FORMAT_OPTION: Time display format ('12' or '24' hour).

Outputs:
    1. Excel (.xlsx) files, each containing schedules for a specific route and
       service type (e.g., Weekday, Saturday), with separate sheets for each
       direction.
    2. Dynamically created output subfolders under BASE_OUTPUT_PATH, organized
       by service ID and its active days (e.g., "calendar_1_mon_tue_wed_thu_fri").

Dependencies:
    os, re, sys, logging, collections (defaultdict), pandas, openpyxl
"""

import logging
import os
import re
import sys
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_FOLDER_PATH = r"C:\Path\To\Your\GTFS_Folder"  # Contains GTFS .txt files

BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output_Folder"
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

FILTER_SERVICE_IDS = []  # e.g. ['1','2'] => only process these. Empty => process all
FILTER_IN_ROUTES = []  # If non-empty, only process these route short names
FILTER_OUT_ROUTES = []  # Exclude these route short names if non-empty

TIME_FORMAT_OPTION = "24"  # "12" or "24"
MISSING_TIME = "---"
MAX_COLUMN_WIDTH = 30

# =============================================================================
# FUNCTIONS
# =============================================================================

# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(gtfs_folder_path: str, files: list[str] = None, dtype=str):
    """
    Loads GTFS files into pandas DataFrames from the specified directory.
    This function uses the logging module for output.

    Parameters:
        gtfs_folder_path (str): Path to the directory containing GTFS files.
        files (list[str], optional): GTFS filenames to load. Default is all
            standard GTFS files:
            [
                "agency.txt",
                "stops.txt",
                "routes.txt",
                "trips.txt",
                "stop_times.txt",
                "calendar.txt",
                "calendar_dates.txt",
                "fare_attributes.txt",
                "fare_rules.txt",
                "feed_info.txt",
                "frequencies.txt",
                "shapes.txt",
                "transfers.txt"
            ]
        dtype (str or dict, optional): Pandas dtype to use. Default is str.

    Returns:
        dict[str, pd.DataFrame]: Dictionary keyed by file name without extension.

    Raises:
        OSError: If gtfs_folder_path doesn't exist or if any required file is missing.
        ValueError: If a file is empty or there's a parsing error.
        RuntimeError: For OS errors during file reading.
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = [
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        ]

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(
            f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}"
        )

    data = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info(f"Loaded {file_name} ({len(df)} records).")

        except pd.errors.EmptyDataError as exc:
            raise ValueError(
                f"File '{file_name}' in '{gtfs_folder_path}' is empty."
            ) from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc
    return data


# -----------------------------------------------------------------------------
# REGULAR FUNCTIONS
# -----------------------------------------------------------------------------


def time_to_minutes(time_str):
    """
    Converts a time string to total minutes since midnight.
    Supports 'HH:MM' and 'HH:MM AM/PM' formats, including hours >= 24
    in 24-hour mode. Returns None if invalid or if time_str == MISSING_TIME.
    """
    if not isinstance(time_str, str):
        return None
    if time_str.strip() == MISSING_TIME:
        return None

    result = None
    try:
        match = re.match(
            r"^(\d{1,2}):(\d{2})(?:\s*(AM|PM))?$", time_str.strip(), re.IGNORECASE
        )
        if match:
            hour_str, minute_str, period = match.groups()
            hour = int(hour_str)
            minute = int(minute_str)
            if period:
                period = period.upper()
                if period == "PM" and hour != 12:
                    hour += 12
                elif period == "AM" and hour == 12:
                    hour = 0
            result = hour * 60 + minute
    except (ValueError, TypeError, re.error):
        result = None

    return result


def adjust_time(time_str, time_format="24"):
    """
    Adjusts time strings to the desired format:
      - '12' => 12-hour with AM/PM
      - '24' => 24-hour

    Returns MISSING_TIME if input is MISSING_TIME, or None if invalid.
    """
    if not isinstance(time_str, str):
        return None
    if time_str.strip() == MISSING_TIME:
        return MISSING_TIME

    parts = time_str.strip().split(":")
    if len(parts) < 2:
        return None

    try:
        hours = int(parts[0])
        minutes = int(parts[1])
        if time_format == "12":
            # If hours >= 24, we can't fully convert properly, so keep as-is
            if hours >= 24:
                return time_str
            period = "AM" if hours < 12 else "PM"
            adjusted_hour = hours % 12
            if adjusted_hour == 0:
                adjusted_hour = 12
            return f"{adjusted_hour}:{minutes:02d} {period}"
        # 24-hour format:
        return f"{hours:02d}:{minutes:02d}"
    except ValueError:
        return None


def prepare_timepoints(stop_times):
    """
    Subset STOP_TIMES to timepoint=1 if available, else use all.
    Returns a DataFrame (TIMEPOINTS).
    """
    stop_times = stop_times.copy()
    stop_times["stop_sequence"] = pd.to_numeric(
        stop_times["stop_sequence"], errors="coerce"
    )
    if stop_times["stop_sequence"].isnull().any():
        # Using logging instead of print for consistency, though this could be print too.
        logging.warning(
            "Some 'stop_sequence' values could not be converted to numeric."
        )

    if "timepoint" in stop_times.columns:
        timepoints = stop_times[stop_times["timepoint"] == "1"].copy()
        logging.info("Filtered STOP_TIMES to rows with timepoint=1.")
    else:
        logging.warning(
            "Warning: 'timepoint' column not found. Using all stops as timepoints."
        )
        timepoints = stop_times

    return timepoints


def remove_empty_schedule_columns(input_df):
    """
    Drops any columns in input_df that are entirely '---' (MISSING_TIME).
    """
    schedule_cols = [col for col in input_df.columns if col.endswith("Schedule")]
    all_blank_cols = [
        col for col in schedule_cols if (input_df[col] == MISSING_TIME).all()
    ]
    input_df.drop(columns=all_blank_cols, inplace=True)
    return input_df


def check_schedule_order(
    input_df, ordered_stop_names, route_short_name, schedule_type, dir_id
):
    """
    Checks times in the DataFrame to ensure they increase across rows
    (within a trip) and down columns (across trips). Prints warnings
    if a violation is found.
    """
    # Row-wise check
    for _, row in input_df.iterrows():
        last_time = None
        for stop in ordered_stop_names:
            col_name = f"{stop} Schedule"
            if col_name not in row:
                continue
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                logging.warning(  # Changed print to logging.warning
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Trip '{row['Trip Headsign']}': '{stop}' time {time_str} "
                    "is earlier than the previous stop's time."
                )
                break
            last_time = current_time

    # Column-wise check
    for stop in ordered_stop_names:
        col_name = f"{stop} Schedule"
        if col_name not in input_df.columns:
            continue
        last_time = None
        for _, row in input_df.iterrows():
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                logging.warning(  # Changed print to logging.warning
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Stop '{stop}': time {time_str} is earlier than "
                    "the previous trip."
                )
                break
            last_time = current_time
    # Using logging.info for consistency
    logging.info(
        f"✅ Schedule order check passed for Route '{route_short_name}', Schedule '{schedule_type}', Direction '{dir_id}'."
    )


def safe_check_schedule_order(
    input_df, ordered_stop_names, route_short_name, schedule_type, dir_id
):
    """
    Wraps the check_schedule_order function in a try/except so that
    if there's a data error, we skip only that schedule order check.
    """
    try:
        check_schedule_order(
            input_df, ordered_stop_names, route_short_name, schedule_type, dir_id
        )
    except (ValueError, TypeError, KeyError) as error:
        logging.error(  # Changed print to logging.error
            f"❌ Skipping schedule order check for route '{route_short_name}', "
            f"schedule '{schedule_type}', direction '{dir_id}' due to error:\n  {error}"
        )


def map_service_id_to_schedule(service_row_local):
    """
    Maps a service_id row to a 'type' label based on the days it serves.
    """
    days = [
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
        "sunday",
    ]
    served_days = [day for day in days if service_row_local.get(day, "0") == "1"]

    weekday = {"monday", "tuesday", "wednesday", "thursday", "friday"}
    weekday_except_friday = {"monday", "tuesday", "wednesday", "thursday"}
    saturday = {"saturday"}
    sunday = {"sunday"}
    weekend = {"saturday", "sunday"}
    daily = set(days)

    schedule_label = "Holiday"
    if served_days:
        served_set = set(served_days)
        if served_set == weekday:
            schedule_label = "Weekday"
        elif served_set == weekday_except_friday:
            schedule_label = "Weekday_except_Friday"
        elif served_set == saturday:
            schedule_label = "Saturday"
        elif served_set == sunday:
            schedule_label = "Sunday"
        elif served_set == weekend:
            schedule_label = "Weekend"
        elif served_set == {"friday", "saturday"}:
            schedule_label = "Friday-Saturday"
        elif served_set == daily:
            schedule_label = "Daily"
        else:
            schedule_label = "Special"  # Could be a custom name based on days
    return schedule_label


def build_service_id_schedule_map(calendar_df):
    """
    Creates a dict: service_id -> schedule_type from CALENDAR.
    """
    service_id_schedule_map = {}
    for _, service_row_local in calendar_df.iterrows():
        sid_val = service_row_local["service_id"]
        stype_var = map_service_id_to_schedule(service_row_local)
        service_id_schedule_map[sid_val] = stype_var
    return service_id_schedule_map


def get_all_route_short_names(routes_df):
    """
    Returns a sorted list of all route_short_names found in ROUTES.
    """
    return sorted(routes_df["route_short_name"].dropna().unique().tolist())


def apply_in_out_filters(route_list):
    """
    Takes the list of all route short names.
    If FILTER_IN_ROUTES is non-empty, keep only those in that list.
    If FILTER_OUT_ROUTES is non-empty, remove those in that list.
    If both are empty, keep everything.
    """
    route_set = set(route_list)

    if FILTER_IN_ROUTES:
        route_set = route_set.intersection(set(FILTER_IN_ROUTES))

    if FILTER_OUT_ROUTES:
        route_set = route_set.difference(set(FILTER_OUT_ROUTES))

    return sorted(list(route_set))


def get_master_trip_stops(dir_id, relevant_trips_dir, timepoints, stops_df):
    """
    Among all trips in 'relevant_trips_dir' for direction=dir_id, pick the trip
    with the largest number of timepoints. Return a DataFrame of that trip's stops.
    """
    relevant_dir = relevant_trips_dir[relevant_trips_dir["direction_id"] == dir_id]
    if relevant_dir.empty:
        logging.warning(
            f"Warning: No trips found for direction_id '{dir_id}'."
        )  # Changed to logging
        return pd.DataFrame()

    dir_trip_ids = relevant_dir["trip_id"].unique()
    subset_tp = timepoints[timepoints["trip_id"].isin(dir_trip_ids)]
    if subset_tp.empty:
        logging.warning(
            f"Warning: No stop times found for direction_id '{dir_id}'."
        )  # Changed to logging
        return pd.DataFrame()

    # Find the "master" trip with the most timepoints
    trip_sizes = subset_tp.groupby("trip_id").size()
    if trip_sizes.empty:  # Added check for empty trip_sizes
        logging.warning(
            f"Warning: No trip sizes to determine master trip for direction_id '{dir_id}'."
        )
        return pd.DataFrame()
    master_trip_id = trip_sizes.idxmax()

    # Extract that trip’s stops in ascending sequence
    master_data = subset_tp[subset_tp["trip_id"] == master_trip_id].copy()
    master_data.sort_values("stop_sequence", inplace=True)

    # Merge to get stop_name
    master_data = master_data.merge(
        stops_df[["stop_id", "stop_name"]], how="left", on="stop_id"
    )

    # Count occurrence of repeated stops
    occurrence_counter = defaultdict(int)
    rows_data = []
    for _, row_2 in master_data.iterrows():
        sid = row_2["stop_id"]
        sseq = row_2["stop_sequence"]
        base_name = (
            row_2["stop_name"]
            if pd.notnull(row_2["stop_name"])
            else f"Unknown stop {sid}"
        )

        occurrence_counter[sid] += 1
        nth = occurrence_counter[sid]

        rows_data.append(
            {
                "stop_id": sid,
                "occurrence": nth,
                "stop_sequence": sseq,
                "base_stop_name": base_name,
            }
        )

    if not rows_data:  # Added check for empty rows_data
        logging.warning(
            f"Warning: No rows_data for master trip stops for direction_id '{dir_id}'."
        )
        return pd.DataFrame()

    out_df = pd.DataFrame(rows_data)
    # Build final_stop_name with repeated stops labeled
    name_occurrences = defaultdict(int)
    final_names = []
    for _, row_2 in out_df.iterrows():
        sid = row_2["stop_id"]
        name_occurrences[sid] += 1
        count_here = name_occurrences[sid]
        if count_here == 1:
            final_names.append(row_2["base_stop_name"])
        else:
            final_names.append(f"{row_2['base_stop_name']} ({count_here})")
    out_df["final_stop_name"] = final_names

    return out_df


def process_single_trip(trip_id, trip_stop_times, master_trip_stops, master_dict, ctx):
    """
    For each trip, produce a single schedule row. Uses a "forward-only" approach.
    """
    trips_df = ctx["trips"]
    routes_df = ctx["routes"]
    time_fmt = ctx["time_fmt"]

    trip_info_rows = trips_df[trips_df["trip_id"] == trip_id]
    if trip_info_rows.empty:
        logging.warning(f"Trip ID {trip_id} not found in trips_df. Skipping.")
        return [None] * (3 + len(master_trip_stops) + 1)  # Match expected row length
    trip_info = trip_info_rows.iloc[0]

    route_id = trip_info["route_id"]

    route_name_rows = routes_df[routes_df["route_id"] == route_id]["route_short_name"]
    if route_name_rows.empty:
        logging.warning(
            f"Route ID {route_id} for trip {trip_id} not found in routes_df. Using 'Unknown Route'."
        )
        route_name_val = "Unknown Route"
    else:
        route_name_val = route_name_rows.values[0]

    trip_headsign = trip_info.get("trip_headsign", "")
    direction_id = trip_info.get("direction_id", "")

    trip_stop_times = trip_stop_times.sort_values("stop_sequence")
    schedule_times = [MISSING_TIME] * len(master_trip_stops)
    valid_24h_times = []

    occurrence_ptr = defaultdict(int)
    for _, row_2 in trip_stop_times.iterrows():
        sid = row_2["stop_id"]
        if sid not in master_dict:
            continue

        arr_val = (row_2.get("arrival_time") or "").strip()
        dep_val = (row_2.get("departure_time") or "").strip()
        time_val = dep_val

        arr_m = time_to_minutes(arr_val)
        dep_m = time_to_minutes(dep_val)
        if arr_val and dep_val and arr_m is not None and dep_m is not None:
            # If arrival is earlier, use arrival as displayed time
            if arr_m < dep_m:
                time_val = arr_val
        elif arr_val and not dep_val:
            time_val = arr_val

        time_str_display = adjust_time(time_val, time_fmt)
        time_str_24 = adjust_time(time_val, "24")

        if not time_str_display or not time_str_24:
            continue

        oc_list = master_dict[sid]
        ptr = occurrence_ptr[sid]

        # Move through repeated stops (if any) in the master trip
        while ptr < len(oc_list) and oc_list[ptr][1] < row_2["stop_sequence"]:
            ptr += 1

        if ptr >= len(oc_list):
            continue

        (_, _, col_idx) = oc_list[ptr]
        schedule_times[col_idx] = time_str_display
        valid_24h_times.append(time_str_24)
        ptr += 1
        occurrence_ptr[sid] = ptr

    # Determine the final "sort_time" for sorting rows
    if valid_24h_times:
        try:
            # Ensure times are in HH:MM format before converting to timedelta
            formatted_24h_times = []
            for t in valid_24h_times:
                if re.match(r"^\d{1,2}:\d{2}$", t):  # Basic HH:MM check
                    # For times like "25:00", split and create timedelta
                    h, m = map(int, t.split(":"))
                    formatted_24h_times.append(pd.Timedelta(hours=h, minutes=m))
                elif re.match(r"^\d{1,2}:\d{2}:\d{2}$", t):  # HH:MM:SS check
                    formatted_24h_times.append(pd.to_timedelta(t))
                else:
                    logging.warning(
                        f"Invalid time format for sort_time: {t} in trip {trip_id}. Skipping this time."
                    )

            if formatted_24h_times:
                max_sort_time = max(formatted_24h_times)
            else:  # If all times were invalid
                max_sort_time = pd.Timedelta(days=999)  # Effectively last
        except (ValueError, TypeError) as e:
            logging.error(
                f"Error converting times for sorting in trip {trip_id}: {valid_24h_times}. Error: {e}"
            )
            max_sort_time = pd.Timedelta(days=999)  # Effectively last
    else:
        max_sort_time = pd.Timedelta(
            days=999
        )  # Effectively last, for trips with no valid times

    row_data = (
        [route_name_val, direction_id, trip_headsign] + schedule_times + [max_sort_time]
    )
    return row_data


def process_trips_for_direction(params):
    """
    Processes trips for a specific direction_id => returns a DataFrame for that direction.

    params expects a dictionary containing:
        - 'trips_dir'
        - 'master_trip_stops'
        - 'dir_id'
        - 'timepoints'
        - 'ctx' (the context dict with time_fmt, trips, routes)
        - 'route_short'
        - 'sched_type'
    """
    trips_dir = params["trips_dir"]
    master_trip_stops = params["master_trip_stops"]
    dir_id = params["dir_id"]
    timepoints = params["timepoints"]
    ctx = params["ctx"]
    route_short = params["route_short"]
    sched_type = params["sched_type"]

    if trips_dir.empty or master_trip_stops.empty:
        logging.info(
            f"No usable trips/stops for route '{route_short}', schedule '{sched_type}', direction '{dir_id}'. Skipping."
        )  # Changed to logging
        return pd.DataFrame()

    # Build a quick lookup so we know each stop_id's column index in final output
    master_dict = defaultdict(list)
    master_trip_stops = master_trip_stops.reset_index(drop=True)
    for i, row_2 in master_trip_stops.iterrows():
        sid = row_2["stop_id"]
        occ = row_2["occurrence"]
        mseq = row_2["stop_sequence"]
        master_dict[sid].append((occ, mseq, i))

    for sid in master_dict:
        # Sort by stop_sequence inside each list
        master_dict[sid].sort(key=lambda x: x[1])

    group_mask = timepoints["trip_id"].isin(trips_dir["trip_id"])
    timepoints_dir = timepoints[group_mask].copy()
    if timepoints_dir.empty:
        logging.info(
            f"No stop times for route '{route_short}', schedule '{sched_type}', direction '{dir_id}' in TIMEPOINTS. Skipping."
        )  # Changed to logging
        return pd.DataFrame()

    stop_names_ordered = master_trip_stops["final_stop_name"].tolist()
    col_names = ["Route Name", "Direction ID", "Trip Headsign"]
    col_names += [f"{sn} Schedule" for sn in stop_names_ordered]
    col_names.append("sort_time")

    rows = []
    for trip_id, grp in timepoints_dir.groupby("trip_id"):
        row_data = process_single_trip(
            trip_id=trip_id,
            trip_stop_times=grp,
            master_trip_stops=master_trip_stops,
            master_dict=master_dict,
            ctx=ctx,
        )
        if row_data[0] is not None:  # Check if trip processing was skipped
            rows.append(row_data)

    if not rows:
        return pd.DataFrame()

    out_df = pd.DataFrame(rows, columns=col_names)
    out_df.sort_values(by="sort_time", inplace=True)
    out_df.drop(columns=["sort_time"], inplace=True)

    safe_check_schedule_order(
        out_df, stop_names_ordered, route_short, sched_type, dir_id
    )
    remove_empty_schedule_columns(out_df)
    return out_df


def export_to_excel_multiple_sheets(df_dict, out_file):
    """
    Exports multiple DataFrames to an Excel file with each DataFrame on its own sheet
    using openpyxl.
    """
    if not df_dict:
        logging.info(f"No data to export to {out_file}.")  # Changed to logging
        return

    try:
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            for sheet_name, input_df in df_dict.items():
                if input_df.empty:
                    logging.info(
                        f"No data for sheet '{sheet_name}'. Skipping."
                    )  # Changed to logging
                    continue

                input_df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # Adjust column widths & alignment
                for col_num, _ in enumerate(input_df.columns, 1):
                    col_letter = get_column_letter(col_num)
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        cell.alignment = Alignment(horizontal="left")

                    # Calculate column width
                    column_cells = worksheet[col_letter]
                    try:
                        max_length = max(
                            len(str(cell.value)) for cell in column_cells if cell.value
                        )
                    except (
                        ValueError,
                        TypeError,
                    ):  # Handles case where column might be all None or empty
                        max_length = 10  # Default width
                    adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
        logging.info(f"Data exported to {out_file}")  # Changed to logging
    except Exception as e:
        logging.error(f"Failed to export to Excel file {out_file}. Error: {e}")


def format_service_id_folder_name(service_row):
    """
    Builds a subfolder name like "calendar_3_mon_tue_wed_thu_fri" based on:
    - service_id in the row
    - which days are marked '1'
    """
    service_id = service_row["service_id"]
    day_map = [
        ("monday", "mon"),
        ("tuesday", "tue"),
        ("wednesday", "wed"),
        ("thursday", "thu"),
        ("friday", "fri"),
        ("saturday", "sat"),
        ("sunday", "sun"),
    ]
    included_days = []
    for col, short_day in day_map:
        if service_row.get(col, "0") == "1":
            included_days.append(short_day)

    if included_days:
        day_str = "_".join(included_days)
    else:
        day_str = "none"  # or "holiday" if calendar_dates is used for this

    return f"calendar_{service_id}_{day_str}"


# -----------------------------------------------------------------------------
# SUB-STEPS
# -----------------------------------------------------------------------------


def filter_calendar_df(calendar_df):
    """
    Applies FILTER_SERVICE_IDS if any, returns filtered df or None if empty.
    """
    if FILTER_SERVICE_IDS:
        calendar_df = calendar_df[calendar_df["service_id"].isin(FILTER_SERVICE_IDS)]
    if calendar_df.empty:
        logging.warning(
            "No service_ids found after applying FILTER_SERVICE_IDS. Exiting."
        )  # Changed to logging
        return None
    return calendar_df


def process_route_service_combinations(ctx):
    """
    Loops over final routes + service rows and builds Excel exports.
    """
    routes_df = ctx["routes"]
    trips_df = ctx["trips"]
    calendar_df = ctx["calendar"]
    timepoints_df = ctx["timepoints"]
    service_id_schedule_map = ctx["service_id_schedule_map"]

    # 1) Determine final routes to process
    all_routes = get_all_route_short_names(routes_df)
    final_routes = apply_in_out_filters(all_routes)
    logging.info(
        f"Final route selection after filters: {final_routes}"
    )  # Changed to logging

    # 2) For each route, build schedules by direction & service_id
    for route_short_name in final_routes:
        logging.info(
            f"\nProcessing route '{route_short_name}'..."
        )  # Changed to logging
        route_ids = routes_df[routes_df["route_short_name"] == route_short_name][
            "route_id"
        ]
        if route_ids.empty:
            logging.error(
                f"Error: Route '{route_short_name}' not found in routes.txt."
            )  # Changed to logging
            continue

        for _, service_row in calendar_df.iterrows():
            service_id = service_row["service_id"]
            folder_name = format_service_id_folder_name(service_row)
            service_output_path = os.path.join(BASE_OUTPUT_PATH, folder_name)
            if not os.path.exists(service_output_path):
                os.makedirs(service_output_path)

            schedule_type = service_id_schedule_map.get(service_id, "Unknown")

            # Filter trips for this route + this service_id
            relevant_trips = trips_df[
                (trips_df["route_id"].isin(route_ids))
                & (trips_df["service_id"] == service_id)
            ]
            if relevant_trips.empty:
                logging.info(  # Changed to logging
                    f"  No trips for route='{route_short_name}' "
                    f"and service_id='{service_id}'."
                )
                continue

            direction_ids_local = (
                relevant_trips["direction_id"].dropna().unique()
            )  # Added dropna()
            df_sheets = {}
            for dir_id in direction_ids_local:
                logging.info(  # Changed to logging
                    f"    Building direction_id '{dir_id}' "
                    f"for service_id='{service_id}'..."
                )
                master_trip_stops = get_master_trip_stops(
                    dir_id, relevant_trips, timepoints_df, ctx["stops"]
                )
                if master_trip_stops.empty:
                    logging.info(
                        f"      No master trip stops for direction_id '{dir_id}'. Skipping this direction."
                    )
                    continue

                params_dict = {
                    "trips_dir": relevant_trips[
                        relevant_trips["direction_id"] == dir_id
                    ],
                    "master_trip_stops": master_trip_stops,
                    "dir_id": dir_id,
                    "timepoints": timepoints_df,
                    "ctx": ctx,
                    "route_short": route_short_name,
                    "sched_type": schedule_type,
                }
                output_df = process_trips_for_direction(params_dict)
                if not output_df.empty:
                    sheet_name = f"Direction_{dir_id}"
                    df_sheets[sheet_name] = output_df
                else:
                    logging.info(f"      No output data for direction_id '{dir_id}'.")

            if df_sheets:
                schedule_type_safe = (
                    schedule_type.replace(" ", "_").replace("-", "_").replace("/", "_")
                )
                out_file = os.path.join(
                    service_output_path,
                    f"route_{route_short_name}_schedule_{schedule_type_safe}.xlsx",
                )
                export_to_excel_multiple_sheets(df_sheets, out_file)
            else:
                logging.info(  # Changed to logging
                    f"  No data to export for service_id '{service_id}' "
                    f"on route '{route_short_name}'."
                )


# =============================================================================
# MAIN
# =============================================================================


def main():
    """
    Orchestrates the end-to-end GTFS schedule export process:
      1. Load GTFS input files directly via load_gtfs_data().
      2. Filter specific service IDs (calendar).
      3. Prepare timepoints.
      4. Build schedule map from calendar => service_id => schedule_type
      5. Process route-service-direction combos and export Excel.
    """

    # Configure basic logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # 1) Load GTFS data
    try:
        # Use the new load_gtfs_data, passing GTFS_FOLDER_PATH
        data = load_gtfs_data(GTFS_FOLDER_PATH, dtype=str)
        logging.info("Successfully loaded GTFS files overall.")  # Main success message
    except (
        OSError
    ) as error:  # Catches directory not found, missing files from new function
        logging.error(f"GTFS data loading error (OS): {error}")
        sys.exit(1)
    except ValueError as error:  # Catches empty file, parser errors from new function
        logging.error(f"GTFS data loading error (Value): {error}")
        sys.exit(1)
    except (
        RuntimeError
    ) as error:  # Catches OS errors during file reading from new function
        logging.error(f"GTFS data loading error (Runtime): {error}")
        sys.exit(1)
    except Exception as error:  # Catch any other unexpected error during loading
        logging.error(f"An unexpected error occurred while loading GTFS files: {error}")
        sys.exit(1)

    # 2) Filter service IDs if any
    if "calendar" not in data:
        logging.error("Error: 'calendar.txt' not found in loaded GTFS data. Exiting.")
        sys.exit(1)
    calendar_df_filtered = filter_calendar_df(data["calendar"])
    if calendar_df_filtered is None:
        # filter_calendar_df already logs a warning if it becomes empty.
        # We exit here because subsequent steps depend on it.
        logging.info("Exiting due to no service IDs to process.")
        return  # Exit main

    # 3) Prepare timepoints
    if "stop_times" not in data:
        logging.error("Error: 'stop_times.txt' not found in loaded GTFS data. Exiting.")
        sys.exit(1)
    timepoints_df = prepare_timepoints(data["stop_times"])

    # 4) Build service_id => schedule_type
    service_id_schedule_map = build_service_id_schedule_map(calendar_df_filtered)

    # 5) Build a context dict to reduce argument counts
    # Ensure all required data keys exist before adding to ctx
    required_data_keys = ["stops", "trips", "routes"]
    for key in required_data_keys:
        if key not in data:
            logging.error(f"Error: '{key}.txt' not found in loaded GTFS data. Exiting.")
            sys.exit(1)

    ctx = {
        "calendar": calendar_df_filtered,
        "timepoints": timepoints_df,
        "service_id_schedule_map": service_id_schedule_map,
        "time_fmt": TIME_FORMAT_OPTION,
        "stops": data["stops"],
        "trips": data["trips"],
        "routes": data["routes"],
    }

    # Final step: process route + service combos
    try:
        process_route_service_combinations(ctx)
        logging.info("Script finished successfully.")
    except Exception as e:
        logging.error(
            f"An error occurred during schedule processing: {e}", exc_info=True
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
