"""Generates Excel-formatted transit schedules from GTFS data.

This script processes GTFS files (e.g., trips.txt, stop_times.txt) to export
Excel schedule sheets by route, direction, and service type. Output mimics
print-style public schedules.

Typical usage is within ArcGIS Pro or a Jupyter notebook. Configuration is
manual: edit constants near the top before execution.

Outputs:
    - One Excel file per route × service type with schedule sheets by direction
"""

from __future__ import annotations

import logging
import os
import re
import sys
from collections import defaultdict
from typing import Any, Mapping, Optional, Sequence

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_FOLDER_PATH = r"C:\Path\To\Your\GTFS_Folder"  # Folder contains GTFS .txt files

BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output_Folder"
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

# Filter service id's can be found in calendar.txt file
FILTER_SERVICE_IDS: list[str] = ["1", "2", "3"]
FILTER_IN_ROUTES: list[str] = ["101", "202"]  # routes to keep
FILTER_OUT_ROUTES: list[str] = []  # routes to exclude

# Explicit schedule label overrides for the selected service_ids
# Assignments can be found in calendar.txt file
SERVICE_LABEL_OVERRIDES: dict[str, str] = {
    "1": "Weekday",
    "2": "Saturday",
    "3": "Sunday",
}

TIME_FORMAT_OPTION = "24"  # "12" or "24"
MISSING_TIME = "---"
MAX_COLUMN_WIDTH = 30

# Required and optional GTFS files
REQUIRED_GTFS_FILES: tuple[str, ...] = (
    "agency.txt",
    "stops.txt",
    "routes.txt",
    "trips.txt",
    "stop_times.txt",
    "calendar.txt",
)

OPTIONAL_GTFS_FILES: tuple[str, ...] = (
    "calendar_dates.txt",
    "feed_info.txt",
    "shapes.txt",
    "frequencies.txt",
    "transfers.txt",
    "fare_attributes.txt",
    "fare_rules.txt",
)

# =============================================================================
# FUNCTIONS
# =============================================================================


def format_output_folder_name(service_id: str, schedule_type: str) -> str:
    """Return an output folder name reflecting the schedule label and service_id.

    Example:
        Weekday (service_id=15) → 'weekday_sid_15'
    """
    return f"{_slugify(schedule_type)}_sid_{service_id}"


def _in_ipython() -> bool:
    """Return True if running under IPython/Jupyter."""
    return "IPYKERNEL" in os.environ or hasattr(sys, "ps1")


def _slugify(label: str) -> str:
    """Return a filesystem-friendly slug from a schedule label."""
    return str(label).strip().lower().replace(" ", "_").replace("-", "_").replace("/", "_")


def time_to_minutes(time_str: str) -> Optional[int]:
    """Convert a time string to minutes after midnight.

    Supports both *HH:MM* and *HH:MM AM/PM* tokens.  Times such as
    ``'25:15'`` (for past-midnight service) are accepted.  ``MISSING_TIME``
    or invalid inputs return ``None``.

    Args:
        time_str: Raw time string.

    Returns:
        Minutes after midnight, or ``None`` if the string is invalid.
    """
    if not isinstance(time_str, str):
        return None
    if time_str.strip() == MISSING_TIME:
        return None

    result = None
    try:
        match = re.match(r"^(\d{1,2}):(\d{2})(?:\s*(AM|PM))?$", time_str.strip(), re.IGNORECASE)
        if match:
            hour_str, minute_str, period = match.groups()
            hour = int(hour_str)
            minute = int(minute_str)
            if period:
                period = period.upper()
                if period == "PM" and hour != 12:
                    hour += 12
                elif period == "AM" and hour == 12:
                    hour = 0
            result = hour * 60 + minute
    except (ValueError, TypeError, re.error):
        result = None

    return result


def adjust_time(time_str: str, time_format: str = "24") -> Optional[str]:
    """Re-format a GTFS time string.

    Args:
        time_str: Raw *arrival_time* or *departure_time* field.
        time_format: Either ``"12"`` for 12-hour *h:MM AM/PM* output or
            ``"24"`` for zero-padded 24-hour output.

    Returns:
        The re-formatted time, ``MISSING_TIME`` if no time should be
        shown, or ``None`` for an unparsable input.
    """
    if not isinstance(time_str, str):
        return None
    if time_str.strip() == MISSING_TIME:
        return MISSING_TIME

    parts = time_str.strip().split(":")
    if len(parts) < 2:
        return None

    try:
        hours = int(parts[0])
        minutes = int(parts[1])
        if time_format == "12":
            # If hours >= 24, we can't fully convert properly, so keep as-is
            if hours >= 24:
                return time_str
            period = "AM" if hours < 12 else "PM"
            adjusted_hour = hours % 12
            if adjusted_hour == 0:
                adjusted_hour = 12
            return f"{adjusted_hour}:{minutes:02d} {period}"
        # 24-hour format:
        return f"{hours:02d}:{minutes:02d}"
    except ValueError:
        return None


def prepare_timepoints(stop_times: pd.DataFrame) -> pd.DataFrame:
    """Return only the rows in *stop_times* where ``timepoint == 1``.

    If the ``timepoint`` column is missing, the original DataFrame is
    returned unchanged.

    Args:
        stop_times: Raw *stop_times.txt* DataFrame.

    Returns:
        A DataFrame containing just the timepoint rows.
    """
    stop_times = stop_times.copy()
    stop_times["stop_sequence"] = pd.to_numeric(stop_times["stop_sequence"], errors="coerce")
    if stop_times["stop_sequence"].isnull().any():
        # Using logging instead of print for consistency, though this could be print too.
        logging.warning("Some 'stop_sequence' values could not be converted to numeric.")

    if "timepoint" in stop_times.columns:
        timepoints = stop_times[stop_times["timepoint"] == "1"].copy()
        logging.info("Filtered STOP_TIMES to rows with timepoint=1.")
    else:
        logging.warning("Warning: 'timepoint' column not found. Using all stops as timepoints.")
        timepoints = stop_times

    return timepoints


def remove_empty_schedule_columns(input_df: pd.DataFrame) -> pd.DataFrame:
    """Drop schedule columns whose entire contents equal ``MISSING_TIME``.

    Args:
        input_df: Schedule grid whose columns end with ``" Schedule"``.

    Returns:
        The same DataFrame, with all-blank schedule columns removed.
    """
    schedule_cols = [col for col in input_df.columns if col.endswith("Schedule")]
    all_blank_cols = [col for col in schedule_cols if (input_df[col] == MISSING_TIME).all()]
    input_df.drop(columns=all_blank_cols, inplace=True)
    return input_df


def check_schedule_order(
    input_df: pd.DataFrame,
    ordered_stop_names: list[str],
    route_short_name: str,
    schedule_type: str,
    dir_id: int,
) -> None:
    """Validate that times increase by stop (row) and by trip (column).

    Args:
        input_df (pandas.DataFrame): Schedule grid to inspect.
        ordered_stop_names (list[str]): Stops in their expected physical order.
        route_short_name (str): Human-readable route label.
        schedule_type (str): *Weekday*, *Saturday*, *Holiday*, etc.
        dir_id (int): ``direction_id`` value (0 or 1).
    """
    # Row-wise check
    for _, row in input_df.iterrows():
        last_time = None
        for stop in ordered_stop_names:
            col_name = f"{stop} Schedule"
            if col_name not in row:
                continue
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                logging.warning(  # Changed print to logging.warning
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Trip '{row['Trip Headsign']}': '{stop}' time {time_str} "
                    "is earlier than the previous stop's time."
                )
                break
            last_time = current_time

    # Column-wise check
    for stop in ordered_stop_names:
        col_name = f"{stop} Schedule"
        if col_name not in input_df.columns:
            continue
        last_time = None
        for _, row in input_df.iterrows():
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                logging.warning(  # Changed print to logging.warning
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Stop '{stop}': time {time_str} is earlier than "
                    "the previous trip."
                )
                break
            last_time = current_time
    # Using logging.info for consistency
    logging.info(
        f"✅ Schedule order check passed for Route '{route_short_name}', Schedule '{schedule_type}', Direction '{dir_id}'."
    )


def safe_check_schedule_order(
    input_df: pd.DataFrame,
    ordered_stop_names: list[str],
    route_short_name: str,
    schedule_type: str,
    dir_id: int,
) -> None:
    """Run :func:`check_schedule_order` but never propagate its errors."""
    try:
        check_schedule_order(input_df, ordered_stop_names, route_short_name, schedule_type, dir_id)
    except (ValueError, TypeError, KeyError) as error:
        logging.error(  # Changed print to logging.error
            f"❌ Skipping schedule order check for route '{route_short_name}', "
            f"schedule '{schedule_type}', direction '{dir_id}' due to error:\n  {error}"
        )


def map_service_id_to_schedule(service_row_local: pd.Series) -> str:
    """Translate one **calendar.txt** row into a schedule label.

    Args:
        service_row_local: A single record from *calendar.txt*.

    Returns:
        A label such as ``"Weekday"``, ``"Saturday"``, ``"Daily"``, etc.
    """
    days = [
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
        "sunday",
    ]
    served_days = [day for day in days if service_row_local.get(day, "0") == "1"]

    weekday = {"monday", "tuesday", "wednesday", "thursday", "friday"}
    weekday_except_friday = {"monday", "tuesday", "wednesday", "thursday"}
    saturday = {"saturday"}
    sunday = {"sunday"}
    weekend = {"saturday", "sunday"}
    daily = set(days)

    schedule_label = "Holiday"
    if served_days:
        served_set = set(served_days)
        if served_set == weekday:
            schedule_label = "Weekday"
        elif served_set == weekday_except_friday:
            schedule_label = "Weekday_except_Friday"
        elif served_set == saturday:
            schedule_label = "Saturday"
        elif served_set == sunday:
            schedule_label = "Sunday"
        elif served_set == weekend:
            schedule_label = "Weekend"
        elif served_set == {"friday", "saturday"}:
            schedule_label = "Friday-Saturday"
        elif served_set == daily:
            schedule_label = "Daily"
        else:
            schedule_label = "Special"  # Could be a custom name based on days
    return schedule_label


def build_service_id_schedule_map(calendar_df: pd.DataFrame) -> dict[str, str]:
    """Create service_id → schedule_type lookup with explicit overrides.

    Args:
        calendar_df: The parsed calendar.txt dataframe filtered to relevant services.

    Returns:
        Mapping from service_id to a human-friendly schedule label (e.g., "Weekday").
        Explicit overrides from SERVICE_LABEL_OVERRIDES take precedence; otherwise
        labels are inferred from the weekday/weekend bit flags.
    """
    service_id_schedule_map: dict[str, str] = {}
    for _, service_row_local in calendar_df.iterrows():
        sid_val = service_row_local["service_id"]
        stype_var = SERVICE_LABEL_OVERRIDES.get(
            sid_val, map_service_id_to_schedule(service_row_local)
        )
        service_id_schedule_map[sid_val] = stype_var
    return service_id_schedule_map


def get_all_route_short_names(routes_df: pd.DataFrame) -> list[str]:
    """Return every unique ``route_short_name`` in ascending order."""
    return sorted(routes_df["route_short_name"].dropna().unique().tolist())


def apply_in_out_filters(route_list: list[str]) -> list[str]:
    """Apply ``FILTER_IN_ROUTES`` and ``FILTER_OUT_ROUTES`` to *route_list*."""
    route_set = set(route_list)

    if FILTER_IN_ROUTES:
        route_set = route_set.intersection(set(FILTER_IN_ROUTES))

    if FILTER_OUT_ROUTES:
        route_set = route_set.difference(set(FILTER_OUT_ROUTES))

    return sorted(list(route_set))


def get_master_trip_stops(
    dir_id: int,
    relevant_trips_dir: pd.DataFrame,
    timepoints: pd.DataFrame,
    stops_df: pd.DataFrame,
) -> pd.DataFrame:
    """Select the trip with the most timepoints and return its ordered stops.

    Args:
        dir_id: Direction being processed (0 or 1).
        relevant_trips_dir: Trips for one route and one ``service_id``.
        timepoints: Pre-filtered timepoint rows.
        stops_df: Parsed *stops.txt* file.

    Returns:
        DataFrame describing the “master” trip’s stop sequence.
    """
    relevant_dir = relevant_trips_dir[relevant_trips_dir["direction_id"] == dir_id]
    if relevant_dir.empty:
        logging.warning(
            f"Warning: No trips found for direction_id '{dir_id}'."
        )  # Changed to logging
        return pd.DataFrame()

    dir_trip_ids = relevant_dir["trip_id"].unique()
    subset_tp = timepoints[timepoints["trip_id"].isin(dir_trip_ids)]
    if subset_tp.empty:
        logging.warning(
            f"Warning: No stop times found for direction_id '{dir_id}'."
        )  # Changed to logging
        return pd.DataFrame()

    # Find the "master" trip with the most timepoints
    trip_sizes = subset_tp.groupby("trip_id").size()
    if trip_sizes.empty:  # Added check for empty trip_sizes
        logging.warning(
            f"Warning: No trip sizes to determine master trip for direction_id '{dir_id}'."
        )
        return pd.DataFrame()
    master_trip_id = trip_sizes.idxmax()

    # Extract that trip’s stops in ascending sequence
    master_data = subset_tp[subset_tp["trip_id"] == master_trip_id].copy()
    master_data.sort_values("stop_sequence", inplace=True)

    # Merge to get stop_name
    master_data = master_data.merge(stops_df[["stop_id", "stop_name"]], how="left", on="stop_id")

    # Count occurrence of repeated stops
    occurrence_counter = defaultdict(int)
    rows_data = []
    for _, row_2 in master_data.iterrows():
        sid = row_2["stop_id"]
        sseq = row_2["stop_sequence"]
        base_name = row_2["stop_name"] if pd.notnull(row_2["stop_name"]) else f"Unknown stop {sid}"

        occurrence_counter[sid] += 1
        nth = occurrence_counter[sid]

        rows_data.append(
            {
                "stop_id": sid,
                "occurrence": nth,
                "stop_sequence": sseq,
                "base_stop_name": base_name,
            }
        )

    if not rows_data:  # Added check for empty rows_data
        logging.warning(f"Warning: No rows_data for master trip stops for direction_id '{dir_id}'.")
        return pd.DataFrame()

    out_df = pd.DataFrame(rows_data)
    # Build final_stop_name with repeated stops labeled
    name_occurrences = defaultdict(int)
    final_names = []
    for _, row_2 in out_df.iterrows():
        sid = row_2["stop_id"]
        name_occurrences[sid] += 1
        count_here = name_occurrences[sid]
        if count_here == 1:
            final_names.append(row_2["base_stop_name"])
        else:
            final_names.append(f"{row_2['base_stop_name']} ({count_here})")
    out_df["final_stop_name"] = final_names

    return out_df


def process_single_trip(
    trip_id: str,
    trip_stop_times: pd.DataFrame,
    master_trip_stops: pd.DataFrame,
    master_dict: dict[str, list[tuple[int, int, int]]],
    ctx: dict[str, Any],
) -> list[Any]:
    """Build one row of the schedule grid for *trip_id*."""
    trips_df = ctx["trips"]
    routes_df = ctx["routes"]
    time_fmt = ctx["time_fmt"]

    trip_info_rows = trips_df[trips_df["trip_id"] == trip_id]
    if trip_info_rows.empty:
        logging.warning(f"Trip ID {trip_id} not found in trips_df. Skipping.")
        return [None] * (3 + len(master_trip_stops) + 1)  # Match expected row length
    trip_info = trip_info_rows.iloc[0]

    route_id = trip_info["route_id"]

    route_name_rows = routes_df[routes_df["route_id"] == route_id]["route_short_name"]
    if route_name_rows.empty:
        logging.warning(
            f"Route ID {route_id} for trip {trip_id} not found in routes_df. Using 'Unknown Route'."
        )
        route_name_val = "Unknown Route"
    else:
        route_name_val = route_name_rows.values[0]

    trip_headsign = trip_info.get("trip_headsign", "")
    direction_id = trip_info.get("direction_id", "")

    trip_stop_times = trip_stop_times.sort_values("stop_sequence")
    schedule_times = [MISSING_TIME] * len(master_trip_stops)
    valid_24h_times = []

    occurrence_ptr = defaultdict(int)
    for _, row_2 in trip_stop_times.iterrows():
        sid = row_2["stop_id"]
        if sid not in master_dict:
            continue

        arr_val = (row_2.get("arrival_time") or "").strip()
        dep_val = (row_2.get("departure_time") or "").strip()
        time_val = dep_val

        arr_m = time_to_minutes(arr_val)
        dep_m = time_to_minutes(dep_val)
        if arr_val and dep_val and arr_m is not None and dep_m is not None:
            # If arrival is earlier, use arrival as displayed time
            if arr_m < dep_m:
                time_val = arr_val
        elif arr_val and not dep_val:
            time_val = arr_val

        time_str_display = adjust_time(time_val, time_fmt)
        time_str_24 = adjust_time(time_val, "24")

        if not time_str_display or not time_str_24:
            continue

        oc_list = master_dict[sid]
        ptr = occurrence_ptr[sid]

        # Move through repeated stops (if any) in the master trip
        while ptr < len(oc_list) and oc_list[ptr][1] < row_2["stop_sequence"]:
            ptr += 1

        if ptr >= len(oc_list):
            continue

        (_, _, col_idx) = oc_list[ptr]
        schedule_times[col_idx] = time_str_display
        valid_24h_times.append(time_str_24)
        ptr += 1
        occurrence_ptr[sid] = ptr

    # Determine the final "sort_time" for sorting rows
    if valid_24h_times:
        try:
            # Ensure times are in HH:MM format before converting to timedelta
            formatted_24h_times = []
            for t in valid_24h_times:
                if re.match(r"^\d{1,2}:\d{2}$", t):  # Basic HH:MM check
                    # For times like "25:00", split and create timedelta
                    h, m = map(int, t.split(":"))
                    formatted_24h_times.append(pd.Timedelta(hours=h, minutes=m))
                elif re.match(r"^\d{1,2}:\d{2}:\d{2}$", t):  # HH:MM:SS check
                    formatted_24h_times.append(pd.to_timedelta(t))
                else:
                    logging.warning(
                        f"Invalid time format for sort_time: {t} in trip {trip_id}. Skipping this time."
                    )

            if formatted_24h_times:
                max_sort_time = max(formatted_24h_times)
            else:  # If all times were invalid
                max_sort_time = pd.Timedelta(days=999)  # Effectively last
        except (ValueError, TypeError) as e:
            logging.error(
                f"Error converting times for sorting in trip {trip_id}: {valid_24h_times}. Error: {e}"
            )
            max_sort_time = pd.Timedelta(days=999)  # Effectively last
    else:
        max_sort_time = pd.Timedelta(days=999)  # Effectively last, for trips with no valid times

    row_data = [route_name_val, direction_id, trip_headsign] + schedule_times + [max_sort_time]
    return row_data


def process_trips_for_direction(params: dict[str, Any]) -> pd.DataFrame:
    """Generate the schedule grid for a single ``direction_id``."""
    trips_dir = params["trips_dir"]
    master_trip_stops = params["master_trip_stops"]
    dir_id = params["dir_id"]
    timepoints = params["timepoints"]
    ctx = params["ctx"]
    route_short = params["route_short"]
    sched_type = params["sched_type"]

    if trips_dir.empty or master_trip_stops.empty:
        logging.info(
            f"No usable trips/stops for route '{route_short}', schedule '{sched_type}', direction '{dir_id}'. Skipping."
        )  # Changed to logging
        return pd.DataFrame()

    # Build a quick lookup so we know each stop_id's column index in final output
    master_dict = defaultdict(list)
    master_trip_stops = master_trip_stops.reset_index(drop=True)
    for i, row_2 in master_trip_stops.iterrows():
        sid = row_2["stop_id"]
        occ = row_2["occurrence"]
        mseq = row_2["stop_sequence"]
        master_dict[sid].append((occ, mseq, i))

    for sid in master_dict:
        # Sort by stop_sequence inside each list
        master_dict[sid].sort(key=lambda x: x[1])

    group_mask = timepoints["trip_id"].isin(trips_dir["trip_id"])
    timepoints_dir = timepoints[group_mask].copy()
    if timepoints_dir.empty:
        logging.info(
            f"No stop times for route '{route_short}', schedule '{sched_type}', direction '{dir_id}' in TIMEPOINTS. Skipping."
        )  # Changed to logging
        return pd.DataFrame()

    stop_names_ordered = master_trip_stops["final_stop_name"].tolist()
    col_names = ["Route Name", "Direction ID", "Trip Headsign"]
    col_names += [f"{sn} Schedule" for sn in stop_names_ordered]
    col_names.append("sort_time")

    rows = []
    for trip_id, grp in timepoints_dir.groupby("trip_id"):
        row_data = process_single_trip(
            trip_id=trip_id,
            trip_stop_times=grp,
            master_trip_stops=master_trip_stops,
            master_dict=master_dict,
            ctx=ctx,
        )
        if row_data[0] is not None:  # Check if trip processing was skipped
            rows.append(row_data)

    if not rows:
        return pd.DataFrame()

    out_df = pd.DataFrame(rows, columns=col_names)
    out_df.sort_values(by="sort_time", inplace=True)
    out_df.drop(columns=["sort_time"], inplace=True)

    safe_check_schedule_order(out_df, stop_names_ordered, route_short, sched_type, dir_id)
    remove_empty_schedule_columns(out_df)
    return out_df


def export_to_excel_multiple_sheets(
    df_dict: dict[str, pd.DataFrame],
    out_file: str,
) -> None:
    """Write each DataFrame in *df_dict* to its own sheet in *out_file*."""
    if not df_dict:
        logging.info(f"No data to export to {out_file}.")  # Changed to logging
        return

    try:
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            for sheet_name, input_df in df_dict.items():
                if input_df.empty:
                    logging.info(
                        f"No data for sheet '{sheet_name}'. Skipping."
                    )  # Changed to logging
                    continue

                input_df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # Adjust column widths & alignment
                for col_num, _ in enumerate(input_df.columns, 1):
                    col_letter = get_column_letter(col_num)
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        cell.alignment = Alignment(horizontal="left")

                    # Calculate column width
                    column_cells = worksheet[col_letter]
                    try:
                        max_length = max(
                            len(str(cell.value)) for cell in column_cells if cell.value
                        )
                    except (
                        ValueError,
                        TypeError,
                    ):  # Handles case where column might be all None or empty
                        max_length = 10  # Default width
                    adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
        logging.info(f"Data exported to {out_file}")  # Changed to logging
    except Exception as e:
        logging.error(f"Failed to export to Excel file {out_file}. Error: {e}")


# -----------------------------------------------------------------------------
# SUB-STEPS
# -----------------------------------------------------------------------------


def filter_calendar_df(
    calendar_df: pd.DataFrame,
) -> Optional[pd.DataFrame]:
    """Filter *calendar_df* by ``FILTER_SERVICE_IDS``; return *None* if empty."""
    if FILTER_SERVICE_IDS:
        calendar_df = calendar_df[calendar_df["service_id"].isin(FILTER_SERVICE_IDS)]
    if calendar_df.empty:
        logging.warning(
            "No service_ids found after applying FILTER_SERVICE_IDS. Exiting."
        )  # Changed to logging
        return None
    return calendar_df


def process_route_service_combinations(ctx: dict[str, Any]) -> None:
    """Loop through every route × service_id and export schedule files.

    Args:
        ctx: Context with dataframes and configuration, including:
            - routes, trips, stops, calendar, timepoints
            - service_id_schedule_map
            - time_fmt
    """
    routes_df = ctx["routes"]
    trips_df = ctx["trips"]
    calendar_df = ctx["calendar"]
    timepoints_df = ctx["timepoints"]
    service_id_schedule_map = ctx["service_id_schedule_map"]

    # 1) Determine final routes to process
    all_routes = get_all_route_short_names(routes_df)
    final_routes = apply_in_out_filters(all_routes)
    logging.info("Final route selection after filters: %s", final_routes)

    # 2) For each route, build schedules by direction & service_id
    for route_short_name in final_routes:
        logging.info("Processing route '%s'...", route_short_name)
        route_ids = routes_df[routes_df["route_short_name"] == route_short_name]["route_id"]
        if route_ids.empty:
            logging.error("Error: Route '%s' not found in routes.txt.", route_short_name)
            continue

        for _, service_row in calendar_df.iterrows():
            service_id = service_row["service_id"]
            schedule_type = service_id_schedule_map.get(service_id, "Unknown")

            # Folder now reflects the override label (and the service_id)
            folder_name = format_output_folder_name(service_id, schedule_type)
            service_output_path = os.path.join(BASE_OUTPUT_PATH, folder_name)
            if not os.path.exists(service_output_path):
                os.makedirs(service_output_path)

            # Filter trips for this route + this service_id
            relevant_trips = trips_df[
                (trips_df["route_id"].isin(route_ids)) & (trips_df["service_id"] == service_id)
            ]
            if relevant_trips.empty:
                logging.info(
                    "  No trips for route='%s' and service_id='%s' (%s).",
                    route_short_name,
                    service_id,
                    schedule_type,
                )
                continue

            direction_ids_local = relevant_trips["direction_id"].dropna().unique()
            df_sheets: dict[str, pd.DataFrame] = {}
            for dir_id in direction_ids_local:
                logging.info(
                    "    Building direction_id '%s' for service_id='%s' (%s)...",
                    dir_id,
                    service_id,
                    schedule_type,
                )
                master_trip_stops = get_master_trip_stops(
                    dir_id, relevant_trips, timepoints_df, ctx["stops"]
                )
                if master_trip_stops.empty:
                    logging.info(
                        "      No master trip stops for direction_id '%s'. Skipping this direction.",
                        dir_id,
                    )
                    continue

                params_dict = {
                    "trips_dir": relevant_trips[relevant_trips["direction_id"] == dir_id],
                    "master_trip_stops": master_trip_stops,
                    "dir_id": dir_id,
                    "timepoints": timepoints_df,
                    "ctx": ctx,
                    "route_short": route_short_name,
                    "sched_type": schedule_type,
                }
                output_df = process_trips_for_direction(params_dict)
                if not output_df.empty:
                    sheet_name = f"Direction_{dir_id}"
                    df_sheets[sheet_name] = output_df
                else:
                    logging.info("      No output data for direction_id '%s'.", dir_id)

            if df_sheets:
                schedule_type_safe = _slugify(schedule_type)
                out_file = os.path.join(
                    service_output_path,
                    f"route_{route_short_name}_schedule_{schedule_type_safe}.xlsx",
                )
                export_to_excel_multiple_sheets(df_sheets, out_file)
            else:
                logging.info(
                    "  No data to export for service_id '%s' (%s) on route '%s'.",
                    service_id,
                    schedule_type,
                    route_short_name,
                )


# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Loads required GTFS files and any optional files that exist. Raises on missing
    required files but not on optional ones.

    Args:
        gtfs_folder_path: Path to folder containing the GTFS feed.
        files: Explicit sequence of file names to load. If None, defaults to
            REQUIRED_GTFS_FILES plus any OPTIONAL_GTFS_FILES found.
        dtype: Passed to pandas.read_csv(dtype=…) to control column dtypes.

    Returns:
        Mapping of file stem → dataframe, e.g., data["trips"].

    Raises:
        OSError: Folder missing or requested file not present.
        ValueError: Empty file or CSV parser failure.
        RuntimeError: OS error while reading a file.
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    # Default behavior: required + optional that actually exist
    if files is None:
        files = list(REQUIRED_GTFS_FILES) + [
            f for f in OPTIONAL_GTFS_FILES if os.path.exists(os.path.join(gtfs_folder_path, f))
        ]

    # Validate presence for the provided list
    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data: dict[str, pd.DataFrame] = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info("Loaded %s (%d records).", file_name, len(df))
        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc
        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc
        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

    return data


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """Coordinate the end-to-end GTFS → Excel workflow."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Build file list: required + optional that actually exist
    files_to_load = list(REQUIRED_GTFS_FILES)
    files_to_load += [
        f for f in OPTIONAL_GTFS_FILES if os.path.exists(os.path.join(GTFS_FOLDER_PATH, f))
    ]

    # Load GTFS data
    try:
        data = load_gtfs_data(GTFS_FOLDER_PATH, files=tuple(files_to_load), dtype=str)
        logging.info("Successfully loaded GTFS files overall.")
    except OSError as error:
        logging.error("GTFS data loading error (OS): %s", error)
        if _in_ipython():
            return
        raise
    except ValueError as error:
        logging.error("GTFS data loading error (Value): %s", error)
        if _in_ipython():
            return
        raise
    except RuntimeError as error:
        logging.error("GTFS data loading error (Runtime): %s", error)
        if _in_ipython():
            return
        raise
    except Exception as error:
        logging.error("An unexpected error occurred while loading GTFS files: %s", error)
        if _in_ipython():
            return
        raise

    # Ensure required tables are present
    if "calendar" not in data:
        logging.error("Error: 'calendar.txt' not found in loaded GTFS data. Exiting.")
        if _in_ipython():
            return
        raise RuntimeError("'calendar.txt' missing from loaded GTFS data.")

    calendar_df_filtered = filter_calendar_df(data["calendar"])
    if calendar_df_filtered is None:
        logging.info("Exiting due to no service IDs to process.")
        return

    if "stop_times" not in data:
        logging.error("Error: 'stop_times.txt' not found in loaded GTFS data. Exiting.")
        if _in_ipython():
            return
        raise RuntimeError("'stop_times.txt' missing from loaded GTFS data.")
    timepoints_df = prepare_timepoints(data["stop_times"])

    # Build service_id => schedule_type (with overrides)
    service_id_schedule_map = build_service_id_schedule_map(calendar_df_filtered)
    for sid in FILTER_SERVICE_IDS:
        logging.info("Service %s → %s", sid, service_id_schedule_map.get(sid, "<missing>"))

    # Build context
    required_data_keys = ["stops", "trips", "routes"]
    for key in required_data_keys:
        if key not in data:
            logging.error("Error: '%s.txt' not found in loaded GTFS data. Exiting.", key)
            if _in_ipython():
                return
            raise RuntimeError(f"'{key}.txt' missing from loaded GTFS data.")

    ctx = {
        "calendar": calendar_df_filtered,
        "timepoints": timepoints_df,
        "service_id_schedule_map": service_id_schedule_map,
        "time_fmt": TIME_FORMAT_OPTION,
        "stops": data["stops"],
        "trips": data["trips"],
        "routes": data["routes"],
    }

    # Produce schedules
    try:
        process_route_service_combinations(ctx)
        logging.info("Script finished successfully.")
    except Exception as e:
        logging.error("An error occurred during schedule processing: %s", e, exc_info=True)
        if _in_ipython():
            return
        raise


if __name__ == "__main__":
    main()
