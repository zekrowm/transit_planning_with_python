"""Exports GTFS timepoint-to-timepoint speeds into time-band tables for Excel.

Analyzes GTFS data to group trips by stop pattern and schedule. The script
generates an Excel workbook for each route and service ID, with sheets showing
segment speeds for different time bands and travel directions.

The script is designed for analysts and data scientists who need a tool for
batch processing GTFS data, and it is suitable for use in environments like
ArcGIS Pro or Jupyter Notebooks.

Configure the module-level constants in the CONFIGURATION section before running.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

GTFS_FOLDER = Path(
    r"G:\projects\dot\zkrohmal\_data\gtfs\connector_gtfs_2025_06_06"
)  # ←–– change me
OUTPUT_FOLDER = Path(
    r"\\S40SHAREPGC01\DOTWorking\zkrohm\data_requests\stop_pattern_exporter_test\output"
)  # ←–– change me

# Optional filters – leave empty to process everything
FILTER_IN_ROUTE_SHORT_NAMES: List[str] = ["101", "660"]
FILTER_OUT_ROUTE_SHORT_NAMES: List[str] = []
FILTER_IN_SERVICE_IDS: List[str] = ["3"]
FILTER_OUT_SERVICE_IDS: List[str] = []

EXPORT_TIMEPOINTS_ONLY = True
INPUT_DISTANCE_UNIT = "meters"  # "meters" or "feet"
MISSING_VAL = "–"

# --------------------------------------------------------------------------------------------------
# LOGGING
# --------------------------------------------------------------------------------------------------

LOG_LEVEL = logging.INFO

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$")

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def hhmmss_to_minutes(t: Optional[str]) -> Optional[int]:
    """Convert an ``HH:MM[:SS]`` string to minutes past midnight.

    The function is tolerant of either ``HH:MM`` or ``HH:MM:SS``; seconds are
    rounded to the nearest minute.  Invalid or *None* inputs propagate as
    *None*.

    Args:
        t: A time literal such as ``"7:05"``, ``"07:05:30"``, or
            ``"23:59:59"``.  Whitespace is stripped.

    Returns:
        The total minutes since 00 : 00 (local GTFS day), or ``None`` if
        parsing fails.
    """
    if not isinstance(t, str):
        return None
    m = _TIME_RE.match(t.strip())
    if not m:
        return None
    h, mm, ss = m.groups()
    return int(h) * 60 + int(mm) + round(int(ss or 0) / 60.0)


def minutes_to_hhmm(m: Optional[int]) -> str:
    """Convert minutes past midnight to a ``HH:MM`` display string.

    Args:
        m: Minutes since 00 : 00.  May be fractional—values are rounded
            to the nearest whole minute; *None* is allowed.

    Returns:
        A ``HH:MM`` string unless *m* is *None*, in which case the configured
        ``MISSING_VAL`` sentinel is returned.
    """
    if m is None:
        return MISSING_VAL
    h, mm = divmod(int(round(m)), 60)
    return f"{h}:{mm:02d}"


def convert_to_miles(dist: Union[str, float, int, None]) -> Optional[float]:
    """Convert a GTFS ``shape_dist_traveled`` value to miles.

    The input may be text, numeric, or missing.  The conversion factor is
    determined by the module-level constant ``INPUT_DISTANCE_UNIT``.

    Args:
        dist: Distance value in meters/feet (per *INPUT_DISTANCE_UNIT*) or a
            string representation thereof.

    Returns:
        The distance in statute miles, or ``None`` if the value is missing or
        cannot be coerced to ``float``.
    """
    if dist in ("", None) or pd.isna(dist):
        return None
    try:
        x = float(dist)
    except (TypeError, ValueError):
        return None
    if INPUT_DISTANCE_UNIT.lower() == "meters":
        return x / 1_609.344
    if INPUT_DISTANCE_UNIT.lower() == "feet":
        return x / 5_280.0
    return x


def mph(dist_miles: Optional[float], runtime_min: Optional[int]) -> Union[float, str]:
    """Compute miles-per-hour, guarding against divide-by-zero.

    Args:
        dist_miles: Segment length in miles, or *None*.
        runtime_min: Segment runtime in minutes; *None* or ``0`` yields the
            sentinel.

    Returns:
        The speed rounded to one decimal place, or ``MISSING_VAL`` if the
        inputs are not computable.
    """
    if dist_miles is None or runtime_min in (None, 0):
        return MISSING_VAL
    return round(dist_miles / (runtime_min / 60.0), 1)


def safe_sheet(name: str) -> str:
    """Sanitise an arbitrary label into a valid Excel worksheet name.

    Disallowed characters are replaced with underscores and the result is
    truncated to the Excel hard limit (31 chars).  An empty string maps to
    ``"Sheet"``.

    Args:
        name: Proposed worksheet name.

    Returns:
        A legal worksheet name suitable for ``openpyxl.Workbook.create_sheet``.
    """
    return re.sub(r"[:\\/*?\[\]]", "_", name)[:31] or "Sheet"


# ─────────────────────────  LOAD GTFS  ──────────────────────────── #
REQ_FILES = ["trips.txt", "stop_times.txt", "routes.txt", "stops.txt"]


def load_gtfs(folder: Path) -> Dict[str, pd.DataFrame]:
    """Read the core GTFS text files into memory.

    Required files are listed in ``REQ_FILES``.  Numeric columns are coerced
    to typed ``pandas`` dtypes where relevant.

    Args:
        folder: Directory containing a *complete* set of GTFS text files.

    Raises:
        FileNotFoundError: If any required file is missing.

    Returns:
        Mapping of file stem (e.g. ``"trips"``) to its loaded ``DataFrame``.
    """
    missing = [f for f in REQ_FILES if not (folder / f).exists()]
    if missing:
        raise FileNotFoundError(f"Missing GTFS file(s): {', '.join(missing)}")

    data = {
        f[:-4]: pd.read_csv(folder / f, dtype=str, low_memory=False) for f in REQ_FILES
    }

    st = data["stop_times"]
    st["stop_sequence"] = pd.to_numeric(st["stop_sequence"], errors="raise")
    if "timepoint" in st.columns:
        st["timepoint"] = pd.to_numeric(st["timepoint"], errors="coerce")
    if "shape_dist_traveled" in st.columns:
        st["shape_dist_traveled"] = pd.to_numeric(
            st["shape_dist_traveled"], errors="coerce"
        )

    return data


# ─────────────────  PATTERN- & SPEED-EXTRACTION  ────────────────── #
Pattern = Tuple[str, ...]
SegSpeeds = Tuple[Union[float, str], ...]


def _sum_numeric(seq):
    """Return the sum of numeric items, ignoring non-numeric elements.

    Args:
        seq: Iterable potentially containing numbers and non-numbers.

    Returns:
        Arithmetic sum of all ``int`` and ``float`` members.
    """
    return sum(x for x in seq if isinstance(x, (int, float)))


def segment_metrics(grp: pd.DataFrame):
    """Derive per-segment speeds plus trip totals for one trip instance.

    Args:
        grp: Consecutive rows from ``stop_times.txt`` for a single trip,
            sorted by ``stop_sequence``.

    Returns:
        Tuple ``(seg_speeds, dist_tot, time_tot)`` where

        * **seg_speeds** – tuple of segment-level mph values (string sentinel
          for any missing segment),
        * **dist_tot** – total distance in miles,
        * **time_tot** – total runtime in minutes.
    """
    times, dists = [], []
    for _, row in grp.iterrows():
        times.append(
            hhmmss_to_minutes(row.get("departure_time") or row.get("arrival_time"))
        )
        dists.append(convert_to_miles(row.get("shape_dist_traveled")))

    run_min, seg_mi, seg_mph = [MISSING_VAL], [MISSING_VAL], [MISSING_VAL]
    for (t0, t1), (d0, d1) in zip(zip(times, times[1:]), zip(dists, dists[1:])):
        run = None if None in (t0, t1) else t1 - t0
        dist = None if None in (d0, d1) else d1 - d0
        run_min.append(run if run is not None else MISSING_VAL)
        seg_mi.append(round(dist, 3) if dist is not None else MISSING_VAL)
        seg_mph.append(mph(dist, run))

    dist_tot = _sum_numeric(seg_mi[1:])
    time_tot = _sum_numeric(run_min[1:])
    return tuple(seg_mph), dist_tot, time_tot


def build_index(gtfs):
    """Create lookup tables linking trips → patterns → speed signatures.

    The heavy-lifting step that:

    1. Applies all user-defined filters,
    2. Identifies unique stop-patterns and speed arrays,
    3. Generates a long index of trips with start times for later banding.

    Args:
        gtfs: GTFS tables as returned by :func:`load_gtfs`.

    Returns:
        A four-tuple ``(index_df, pattern_lut, speed_lut, header_lut)``:

        * **index_df** – long table of trips (one row per trip),
        * **pattern_lut** – map ``pattern_hash → stop_id tuple``,
        * **speed_lut** – map ``speed_hash → {"seg_speeds", "mean_mph"}``,
        * **header_lut** – map ``pattern_hash → list[str]`` of human-readable
          column headers.
    """
    trips = gtfs["trips"].merge(
        gtfs["routes"][["route_id", "route_short_name"]], on="route_id", how="left"
    )

    if FILTER_IN_ROUTE_SHORT_NAMES:
        trips = trips[trips["route_short_name"].isin(FILTER_IN_ROUTE_SHORT_NAMES)]
    if FILTER_OUT_ROUTE_SHORT_NAMES:
        trips = trips[~trips["route_short_name"].isin(FILTER_OUT_ROUTE_SHORT_NAMES)]
    if FILTER_IN_SERVICE_IDS:
        trips = trips[trips["service_id"].isin(FILTER_IN_SERVICE_IDS)]
    if FILTER_OUT_SERVICE_IDS:
        trips = trips[~trips["service_id"].isin(FILTER_OUT_SERVICE_IDS)]

    if trips.empty:
        return pd.DataFrame(), {}, {}, {}

    st = (
        gtfs["stop_times"]
        .merge(
            trips[["trip_id", "route_id", "service_id", "direction_id"]], on="trip_id"
        )
        .sort_values(["trip_id", "stop_sequence"])
    )
    if EXPORT_TIMEPOINTS_ONLY and "timepoint" in st.columns:
        st = st[st["timepoint"] == 1]

    stop_meta = (
        gtfs["stops"][["stop_id", "stop_name", "stop_code"]]
        .set_index("stop_id")
        .to_dict("index")
    )

    pat_lut, speed_lut, header_lut, rows = {}, {}, {}, []

    for trip_id, grp in st.groupby("trip_id"):
        pattern = tuple(grp["stop_id"])
        pat_hash = hash(pattern)
        pat_lut.setdefault(pat_hash, pattern)

        seg_speeds, dist_tot, time_tot = segment_metrics(grp)
        spd_hash = hash(seg_speeds)
        if spd_hash not in speed_lut:
            speed_lut[spd_hash] = {
                "seg_speeds": seg_speeds,
                "mean_mph": mph(dist_tot, time_tot),
            }

        if pat_hash not in header_lut:
            header_lut[pat_hash] = [
                f"{m['stop_name']} ({m['stop_code']})"
                if (m := stop_meta.get(sid))
                else sid
                for sid in pattern
            ]

        first_seg = seg_speeds[1] if len(seg_speeds) > 1 else MISSING_VAL
        start_min = hhmmss_to_minutes(
            grp.iloc[0].get("departure_time") or grp.iloc[0].get("arrival_time")
        )

        rows.append(
            {
                "route_id": grp.iloc[0]["route_id"],
                "service_id": grp.iloc[0]["service_id"],
                "direction_id": grp.iloc[0]["direction_id"],
                "pattern_hash": pat_hash,
                "speed_hash": spd_hash,
                "first_seg_mph": None if first_seg == MISSING_VAL else first_seg,
                "start": start_min,
            }
        )

    return pd.DataFrame(rows), pat_lut, speed_lut, header_lut


# ───────────────────────  BAND-ROW COLLAPSE  ────────────────────── #
def band_rows(idx: pd.DataFrame) -> pd.DataFrame:
    """Collapse individual trips into time-of-day bands.

    Args:
        idx: Output from :func:`build_index`.

    Returns:
        A dataframe with one row per unique
        (route, service_id, direction, pattern_hash, speed_hash) *  time-band
        describing first/last departure and trip count.
    """
    gb = idx.groupby(
        ["route_id", "service_id", "direction_id", "pattern_hash", "speed_hash"]
    )

    out = (
        gb.agg(
            FrTime_min=("start", "min"),
            ToTime_min=("start", "max"),
            TripCount=("start", "count"),
        )
        .reset_index()
        .sort_values(["direction_id", "FrTime_min"])  # chronological
    )

    out["FrTime"] = out["FrTime_min"].apply(minutes_to_hhmm)
    out["ToTime"] = out["ToTime_min"].apply(minutes_to_hhmm)
    return out.drop(columns=["FrTime_min", "ToTime_min"])


# ────────────────────────  EXCEL EXPORT  ────────────────────────── #
def export_excel(
    bands: pd.DataFrame,
    pat_lut: Dict[int, Pattern],
    speed_lut: Dict[int, Dict[str, Union[SegSpeeds, float]]],
    header_lut: Dict[int, List[str]],
    routes: pd.DataFrame,
) -> None:
    """Write each (route × service_id) bundle to its own workbook.

    Each workbook contains one worksheet per direction.  Column widths are set
    to 14 characters for readability.

    Args:
        bands: Banded rows from :func:`band_rows`.
        pat_lut: Pattern lookup produced by :func:`build_index`.
        speed_lut: Speed lookup produced by :func:`build_index`.
        header_lut: Header lookup produced by :func:`build_index`.
        routes: ``routes.txt`` table; used to translate ``route_id`` to short
            names.

    Side Effects:
        • Creates *OUTPUT_FOLDER* if it does not exist.
        • Writes ``.xlsx`` files and logs progress via :pymod:`logging`.
    """
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    route_short = routes.set_index("route_id")["route_short_name"].to_dict()

    for (rid, sid), grp_rs in bands.groupby(["route_id", "service_id"]):
        wb = Workbook()
        wb.remove(wb.active)

        for did, grp_dir in grp_rs.groupby("direction_id"):
            ws = wb.create_sheet(safe_sheet(f"Dir_{did or 'X'}"))

            first_pat = grp_dir.iloc[0]["pattern_hash"]
            header = ["Pattern", "FrTime", "ToTime", "Mean_mph", *header_lut[first_pat]]
            ws.append(header)

            for _, row in grp_dir.iterrows():
                speed_rec = speed_lut[row["speed_hash"]]
                ws.append(
                    [
                        row["pattern_hash"],
                        row["FrTime"],
                        row["ToTime"],
                        speed_rec["mean_mph"],
                        *speed_rec["seg_speeds"],
                    ]
                )

            for col in range(1, len(header) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 14

        out = (
            OUTPUT_FOLDER
            / f"route_{route_short.get(rid, rid)}_cal{sid}_speed_table.xlsx"
        )
        wb.save(out)
        logging.info("Wrote %s", out)


# ==================================================================================================
# MAIN
# ==================================================================================================


def main() -> None:
    """CLI entry-point – orchestrates GTFS load, processing, and export."""
    logging.basicConfig(
        level=LOG_LEVEL,
        format="%(asctime)s  %(levelname)s  %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.info("GTFS folder:   %s", GTFS_FOLDER)
    logging.info("Output folder: %s", OUTPUT_FOLDER)

    gtfs = load_gtfs(GTFS_FOLDER)
    idx, pat_lut, speed_lut, header_lut = build_index(gtfs)

    if idx.empty:
        logging.warning("No trips after filters – nothing to export.")
        return

    bands = band_rows(idx)
    export_excel(bands, pat_lut, speed_lut, header_lut, gtfs["routes"])
    logging.info("Finished – %d band rows written.", len(bands))


if __name__ == "__main__":
    main()
