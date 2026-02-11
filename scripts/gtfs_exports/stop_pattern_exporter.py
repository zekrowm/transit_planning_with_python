"""Extract unique GTFS stop patterns and export them to Excel workbooks.

This script reads GTFS files and identifies unique stop sequences for each
(route_id, direction_id, service_id) combination. These "patterns" represent
distinct sequences of stops with distance calculations between them.

Each route/service pair is exported to a separate Excel file, with each
direction as a separate sheet. A “master” trip is chosen to define the
canonical stop order for each direction.

Typical usage:
    Adjust paths and options in the CONFIGURATION section, then run in
    ArcPro or standalone Python notebook.

Outputs:
    - Excel files written to OUTPUT_DIR with subfolders for each service_id.
"""

from __future__ import annotations

import logging
import os
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Literal, Mapping, Optional, Sequence, Tuple, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

# Folder containing the raw GTFS feed (must include at least stops.txt, trips.txt,
# stop_times.txt, and routes.txt). Use a raw string (r"") for Windows UNC paths.
INPUT_DIR: Path = Path(r"\\Path\\To\\Your\\GTFS_Folder")

# Folder where the output Excel files will be saved. Subfolders will be created
# for each service_id based on calendar.txt (if available).
OUTPUT_DIR: Path = Path(r"\\Path\\To\\Output_Folder")

# Include only these route_short_name values. Leave empty to include all routes.
FILTER_IN_ROUTE_SHORT_NAMES: List[str] = []

# Exclude these route_short_name values from processing. Leave empty to exclude none.
FILTER_OUT_ROUTE_SHORT_NAMES: List[str] = []

# Include only these service_id values (from trips.txt). Leave empty to include all.
FILTER_IN_CALENDAR_IDS: List[str] = []

# Used in the output Excel filename for labeling the signup period or data snapshot.
SIGNUP_NAME = "JAN_2025_Signup"

# Unit of shape_dist_traveled in GTFS feed. Choose either "meters" or "feet".
INPUT_DISTANCE_UNIT: Literal["meters", "feet"] = "meters"

# If True, convert all distance values to miles (from INPUT_DISTANCE_UNIT).
CONVERT_TO_MILES: bool = True

# If True, only include stops where timepoint == 1 in stop_times.txt.
# Useful for focusing on scheduled timing stops rather than all intermediate stops.
EXPORT_TIMEPOINTS_ONLY: bool = True

# If True and EXPORT_TIMEPOINTS_ONLY is enabled, validate that the sum of segment
# distances roughly matches the trip's full shape_dist_traveled (within a small margin).
VALIDATE_TIMEPOINT_DISTANCE: bool = True

# -----------------------------------------------------------------------------
# LOGGING
# -----------------------------------------------------------------------------

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# =============================================================================
# FUNCTIONS
# =============================================================================


def is_number(value: Any) -> bool:
    """Check if the input value can be converted to a float.

    Args:
        value: Any Python object.

    Returns:
        True if value can be cast to float, else False.
    """
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def convert_dist_to_miles(distance: Any, input_unit: str) -> Any:
    """Convert a distance value to miles.

    Args:
        distance: Raw distance value as string or number.
        input_unit: The unit of input ('feet' or 'meters').

    Returns:
        Distance in miles, or original value if conversion is not possible.

    Warns:
        Logs a warning if the unit is invalid or the value can't be converted.
    """
    if not CONVERT_TO_MILES or pd.isna(distance) or distance == "":
        return distance
    try:
        num_distance = float(distance)
    except ValueError:
        logging.warning(f"Could not convert distance '{distance}' to float. Returning as-is.")
        return distance

    if input_unit.lower() == "feet":
        conv = 5280.0
    elif input_unit.lower() == "meters":
        conv = 1609.34
    else:
        logging.warning("Unknown distance unit '%s'. No conversion done.", input_unit)
        conv = 1.0
    return num_distance / conv


def parse_time_to_minutes(time_str: str) -> Optional[float]:
    """Convert HH:MM:SS to minutes past midnight.

    Args:
        time_str: A time string in HH:MM:SS format. Hours can exceed 24.

    Returns:
        Float minutes since midnight, or None if parsing fails.
    """
    if not isinstance(time_str, str):
        return None
    parts = time_str.strip().split(":")
    if len(parts) != 3:
        return None
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2])
        return hours * 60 + minutes + seconds / 60.0
    except (TypeError, ValueError):
        return None


def minutes_to_hhmm(minutes_val: Optional[float]) -> str:
    """Convert minutes past midnight to HH:MM 24-hour format.

    Args:
        minutes_val: Minutes since midnight.

    Returns:
        Time string in HH:MM format, or empty string if invalid.
    """
    if minutes_val is None or pd.isna(minutes_val):
        return ""
    total_minutes = int(round(minutes_val))  # Round to nearest minute
    hours = total_minutes // 60
    mins = total_minutes % 60
    return f"{hours:02d}:{mins:02d}"


def format_service_id_folder_name(service_id: str, calendar_df: Optional[pd.DataFrame]) -> str:
    """Generate a subfolder name from service_id and weekday info.

    Args:
        service_id: The GTFS service_id.
        calendar_df: DataFrame from calendar.txt.

    Returns:
        Folder name like 'calendar_10_mon_tue' or 'calendar_5_none'.
    """
    if calendar_df is None or calendar_df.empty:
        return f"calendar_{service_id}"

    # Ensure service_id in calendar_df['service_id'] is compared as string if it's not already
    calendar_df["service_id"] = calendar_df["service_id"].astype(str)
    row = calendar_df[calendar_df["service_id"] == str(service_id)]
    if row.empty:
        return f"calendar_{service_id}"

    row = row.iloc[0]  # Take the first matching row
    day_map = [
        ("monday", "mon"),
        ("tuesday", "tue"),
        ("wednesday", "wed"),
        ("thursday", "thu"),
        ("friday", "fri"),
        ("saturday", "sat"),
        ("sunday", "sun"),
    ]
    served_days = []
    for col, short_name in day_map:
        # calendar_df is loaded with dtype=str, so values are '0' or '1'
        if row.get(col, "0") == "1":
            served_days.append(short_name)

    if served_days:
        day_str = "_".join(served_days)
    else:
        day_str = "none"

    return f"calendar_{service_id}_{day_str}"


def filter_trips(
    trips_df: pd.DataFrame, routes_df: pd.DataFrame, cal_ids: List[str]
) -> pd.DataFrame:
    """Filter trips by route_short_name and service_id.

    Args:
        trips_df: GTFS trips table.
        routes_df: GTFS routes table.
        cal_ids: List of service_ids to include.

    Returns:
        Filtered DataFrame of trips.
    """
    merged = pd.merge(
        trips_df, routes_df[["route_id", "route_short_name"]], on="route_id", how="left"
    )
    # Filter in
    if FILTER_IN_ROUTE_SHORT_NAMES:
        merged = merged[merged["route_short_name"].isin(FILTER_IN_ROUTE_SHORT_NAMES)]
    # Filter out
    if FILTER_OUT_ROUTE_SHORT_NAMES:
        merged = merged[~merged["route_short_name"].isin(FILTER_OUT_ROUTE_SHORT_NAMES)]
    # Filter calendar
    if cal_ids and "service_id" in merged.columns:
        # Ensure cal_ids are strings for comparison if service_id is string
        cal_ids_str = [str(c) for c in cal_ids]
        merged = merged[merged["service_id"].astype(str).isin(cal_ids_str)]
    elif cal_ids:
        logging.warning("No service_id in data for filtering by calendar_id.")
    return merged


# -----------------------------------------------------------------------------
# BUILD PATTERNS
# -----------------------------------------------------------------------------


def generate_unique_patterns(
    trips_df: pd.DataFrame, stop_times_df: pd.DataFrame, stops_df: pd.DataFrame
) -> Dict[Tuple, Dict[str, Any]]:
    """Identify unique stop patterns grouped by route, direction, and service.

    Args:
        trips_df: GTFS trips DataFrame.
        stop_times_df: GTFS stop_times DataFrame.
        stops_df: GTFS stops DataFrame.

    Returns:
        Dictionary mapping unique (route, direction, service, stops) to metadata.
    """
    tmp = pd.merge(
        stop_times_df,
        trips_df[["trip_id", "route_id", "direction_id", "service_id"]],
        on="trip_id",
        how="inner",
    )
    if "shape_dist_traveled" not in tmp.columns:
        tmp["shape_dist_traveled"] = np.nan  # Will be float due to np.nan

    tmp = pd.merge(tmp, stops_df[["stop_id", "stop_name"]], on="stop_id", how="left")

    # Ensure stop_sequence is numeric for correct sorting
    if "stop_sequence" in tmp.columns:
        # This conversion should ideally happen once after loading stop_times_df
        # but ensuring it here defensively if df is passed around.
        if not pd.api.types.is_numeric_dtype(tmp["stop_sequence"]):
            tmp["stop_sequence"] = pd.to_numeric(tmp["stop_sequence"], errors="raise")
    else:
        logging.error(
            "generate_unique_patterns: 'stop_sequence' column missing in merged stop_times. "
            "Critical for sorting."
        )
        return {}

    tmp.sort_values(["trip_id", "stop_sequence"], inplace=True)

    # If we only want timepoints
    # timepoint column was converted to numeric earlier if it exists
    if EXPORT_TIMEPOINTS_ONLY and "timepoint" in tmp.columns:
        # Ensure 'timepoint' is numeric for comparison if not already
        if not pd.api.types.is_numeric_dtype(tmp["timepoint"]):
            tmp["timepoint"] = pd.to_numeric(tmp["timepoint"], errors="coerce")
        tmp = tmp[tmp["timepoint"] == 1]  # Comparison with numeric 1

    # For distance validation
    trip_distances: dict[str, Any] = {}
    if VALIDATE_TIMEPOINT_DISTANCE and EXPORT_TIMEPOINTS_ONLY:
        for trip_id_val, group_sub in tmp.groupby("trip_id"):
            # Ensure 'shape_dist_traveled' is numeric for comparison if not already
            if "shape_dist_traveled" in group_sub.columns and not pd.api.types.is_numeric_dtype(
                group_sub["shape_dist_traveled"]
            ):
                group_sub["shape_dist_traveled"] = pd.to_numeric(
                    group_sub["shape_dist_traveled"], errors="coerce"
                )

            group_sub_dist_valid = group_sub.dropna(subset=["shape_dist_traveled"])
            if group_sub_dist_valid.empty:
                trip_distances[str(trip_id_val)] = None
            else:
                dist_val = (
                    group_sub_dist_valid.iloc[-1]["shape_dist_traveled"]
                    - group_sub_dist_valid.iloc[0]["shape_dist_traveled"]
                )
                trip_distances[str(trip_id_val)] = convert_dist_to_miles(
                    dist_val, INPUT_DISTANCE_UNIT
                )

    # Build patterns
    patterns_list: list[dict[str, Any]] = []
    for trip_id_val, group_sub in tmp.groupby("trip_id"):
        group_sub = group_sub.sort_values("stop_sequence")
        stops_for_trip: list[tuple[str, str]] = []
        prev_dist_val = None  # numeric after first valid shape_dist_traveled

        for _, row in group_sub.iterrows():
            stop_id_val = row["stop_id"]
            shape_val = row["shape_dist_traveled"]  # numeric or NaN

            if prev_dist_val is None:
                dist_str = "-"
            else:
                if pd.notnull(shape_val) and pd.notnull(prev_dist_val):
                    diff_numeric = shape_val - prev_dist_val
                    diff_miles = convert_dist_to_miles(diff_numeric, INPUT_DISTANCE_UNIT)
                    if pd.notnull(diff_miles) and isinstance(diff_miles, (int, float)):
                        dist_str = f"{diff_miles:.2f}" if diff_miles != 0 else "0.00"
                    else:
                        dist_str = ""
                else:
                    dist_str = ""

            stops_for_trip.append((stop_id_val, dist_str))

            if pd.notnull(shape_val):
                prev_dist_val = shape_val

        # Validate timepoint distance
        if VALIDATE_TIMEPOINT_DISTANCE and EXPORT_TIMEPOINTS_ONLY:
            sum_seg = 0.0
            for _, dist_segment_str in stops_for_trip:
                if is_number(dist_segment_str):
                    sum_seg += float(dist_segment_str)

            full_trip_dist_miles = trip_distances.get(trip_id_val, None)
            if (
                full_trip_dist_miles is not None
                and isinstance(full_trip_dist_miles, (int, float))
                and abs(sum_seg - full_trip_dist_miles) > 0.02
            ):
                logging.warning(
                    "Trip %s sum of segments=%.2f vs. full=%.2f mismatch >0.02",
                    trip_id_val,
                    sum_seg,
                    full_trip_dist_miles,
                )

        if group_sub.empty:
            logging.warning("Empty group for trip_id %s in pattern generation.", trip_id_val)
            continue

        first_row = group_sub.iloc[0]
        route_id_val = first_row["route_id"]
        direction_id_val = first_row["direction_id"]
        service_id_val = first_row["service_id"]

        patterns_list.append(
            {
                "trip_id": trip_id_val,
                "route_id": route_id_val,
                "direction_id": direction_id_val,
                "service_id": service_id_val,
                "pattern_stops": tuple(stops_for_trip),
            }
        )

    # Accumulate unique patterns
    patterns_dict: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in patterns_list:
        key = (
            record["route_id"],
            record["direction_id"],
            record["service_id"],
            record["pattern_stops"],
        )
        if key not in patterns_dict:
            patterns_dict[key] = {
                "route_id": record["route_id"],
                "direction_id": record["direction_id"],
                "service_id": record["service_id"],
                "pattern_stops": record["pattern_stops"],
                "trip_count": 0,
                "trip_ids": [],
            }
        patterns_dict[key]["trip_count"] += 1
        patterns_dict[key]["trip_ids"].append(record["trip_id"])

    logging.info("Found %d unique patterns.", len(patterns_dict))
    return patterns_dict


def assign_pattern_ids(patterns_dict: Dict[Tuple, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Assign numeric pattern IDs to each unique stop sequence.

    Args:
        patterns_dict: Dictionary of unique patterns.

    Returns:
        List of enriched pattern records with assigned pattern_id.
    """
    group_map = defaultdict(list)
    for pattern_val in patterns_dict.values():
        route_id_val = pattern_val["route_id"]
        dir_id_val = pattern_val["direction_id"]  # string "0", "1"
        srv_id_val = pattern_val["service_id"]  # string
        group_map[(route_id_val, srv_id_val, dir_id_val)].append(pattern_val)

    out: list[dict[str, Any]] = []
    for (_route_id_val, _service_id_val, _direction_id_val), recs in group_map.items():
        recs = sorted(recs, key=lambda x: x["pattern_stops"])
        for idx, pattern_rec in enumerate(recs, 1):
            pattern_rec["pattern_id"] = idx
            out.append(
                {
                    "route_id": pattern_rec["route_id"],
                    "direction_id": pattern_rec["direction_id"],
                    "service_id": pattern_rec["service_id"],
                    "pattern_stops": pattern_rec["pattern_stops"],
                    "trip_count": pattern_rec["trip_count"],
                    "trip_ids": pattern_rec["trip_ids"],
                    "pattern_id": idx,
                }
            )

    logging.info("Assigned pattern IDs to pattern records.")
    return out


# -----------------------------------------------------------------------------
# EARLIEST START TIME
# -----------------------------------------------------------------------------


def compute_earliest_start_times(
    pattern_records: List[Dict[str, Any]], stop_times_df: pd.DataFrame
) -> None:
    """Add earliest start time (in minutes and HH:MM) to pattern records.

    Args:
        pattern_records: List of pattern records with trip_ids.
        stop_times_df: GTFS stop_times DataFrame.

    Side Effects:
        Updates pattern_records in-place with new time fields.
    """
    if (
        "arrival_time" not in stop_times_df.columns
        and "departure_time" not in stop_times_df.columns
    ):
        logging.warning(
            "Cannot compute earliest start times: arrival_time and departure_time columns missing."
        )
        for rec in pattern_records:
            rec["earliest_time_minutes"] = None
            rec["earliest_time_str"] = ""
        return

    # Ensure stop_sequence is numeric for sorting (should have been done post-load)
    if "stop_sequence" in stop_times_df.columns and not pd.api.types.is_numeric_dtype(
        stop_times_df["stop_sequence"]
    ):
        logging.warning(
            "compute_earliest_start_times: stop_sequence in stop_times_df is not numeric. "
            "Attempting conversion."
        )
        stop_times_df["stop_sequence"] = pd.to_numeric(
            stop_times_df["stop_sequence"], errors="raise"
        )

    stop_times_by_trip = stop_times_df.groupby("trip_id")
    for rec in pattern_records:
        trip_ids = rec["trip_ids"]
        earliest_val = None  # in minutes

        for t_id in trip_ids:
            if t_id not in stop_times_by_trip.groups:
                continue

            group_2 = stop_times_by_trip.get_group(t_id).sort_values("stop_sequence")
            if group_2.empty:
                continue

            first_stop_event = group_2.iloc[0]
            arr_str = first_stop_event.get("arrival_time", None)
            dep_str = first_stop_event.get("departure_time", None)

            arr_minutes = parse_time_to_minutes(arr_str) if arr_str else None
            dep_minutes = parse_time_to_minutes(dep_str) if dep_str else None

            candidates: list[float] = []
            if arr_minutes is not None:
                candidates.append(arr_minutes)
            if dep_minutes is not None:
                candidates.append(dep_minutes)

            if not candidates:
                continue

            this_min = min(candidates)
            if earliest_val is None or this_min < earliest_val:
                earliest_val = this_min

        rec["earliest_time_minutes"] = earliest_val
        rec["earliest_time_str"] = minutes_to_hhmm(earliest_val)


# -----------------------------------------------------------------------------
# MASTER TRIP
# -----------------------------------------------------------------------------


def find_master_trip_stops(
    route_id_val: str,
    direction_id_val: str,
    relevant_trips: pd.DataFrame,
    stop_times_df: pd.DataFrame,
    stops_df: pd.DataFrame,
) -> List[Tuple[str, str]]:
    """Identify the trip with the most timepoints as the master for a direction.

    Args:
        route_id_val: Route ID of interest.
        direction_id_val: Direction (0 or 1).
        relevant_trips: Subset of trips for the route and direction.
        stop_times_df: GTFS stop_times DataFrame.
        stops_df: GTFS stops DataFrame.

    Returns:
        List of (stop_id, stop_name) tuples for the master trip.
    """
    if relevant_trips.empty:
        return []
    st_sub = stop_times_df[stop_times_df["trip_id"].isin(relevant_trips["trip_id"])]

    # timepoint column is numeric if it exists (after initial conversion)
    if "timepoint" in st_sub.columns and EXPORT_TIMEPOINTS_ONLY:
        # Ensure numeric for comparison
        if not pd.api.types.is_numeric_dtype(st_sub["timepoint"]):
            st_sub["timepoint"] = pd.to_numeric(st_sub["timepoint"], errors="coerce")
        st_sub = st_sub[st_sub["timepoint"] == 1]

    sizes = st_sub.groupby("trip_id").size()
    if sizes.empty:
        return []
    best_trip_id = sizes.idxmax()

    # stop_sequence is numeric
    best_group = st_sub[st_sub["trip_id"] == best_trip_id].sort_values("stop_sequence")
    best_group = pd.merge(best_group, stops_df[["stop_id", "stop_name"]], on="stop_id", how="left")

    out_list = []
    for _, row in best_group.iterrows():
        stop_id_val = row["stop_id"]
        stop_name_val = row.get("stop_name", "Unknown")
        out_list.append((stop_id_val, stop_name_val))
    return out_list


def forward_match_pattern_to_master(
    pattern_stops: List[Tuple[str, str]], master_stops: List[Tuple[str, str]]
) -> List[str]:
    """Align pattern segment distances to the master stop order.

    Args:
        pattern_stops: List of (stop_id, dist_str) tuples from a pattern.
        master_stops: List of (stop_id, stop_name) from master trip.

    Returns:
        List of distances aligned to master_stops.
    """
    result = [""] * len(master_stops)
    i = 0  # master_stops index
    j = 0  # pattern_stops index

    # The first stop of a pattern conceptually has no preceding segment *within that pattern*.
    # Its distance value in pattern_stops (e.g., "-") should be placed.
    # If pattern_stops[0] is ("stop_A", "-"), and master_stops[0] is ("stop_A", "Stop A Name"),
    # then result[0] should be "-".

    while i < len(master_stops) and j < len(pattern_stops):
        master_sid = master_stops[i][0]
        pat_sid, dist_str = pattern_stops[j]

        if master_sid == pat_sid:
            result[i] = dist_str  # Place the distance string from the pattern
            i += 1
            j += 1
        else:
            i += 1
    return result


# -----------------------------------------------------------------------------
# EXCEL EXPORT
# -----------------------------------------------------------------------------


def create_workbook() -> Workbook:
    """Create a new openpyxl Workbook with the default sheet removed.

    Returns:
        An openpyxl Workbook instance.
    """
    workbook = Workbook()
    if workbook.active:
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    return workbook


def fill_worksheet_for_direction(
    workbook: Workbook,
    sheet_title: str,
    route_short_name: str,
    direction_id_val: str,
    service_id_val: str,
    pattern_records_dir: List[Dict[str, Any]],
    master_stops: List[Tuple[str, str]],
) -> None:
    """Add a sheet to the workbook with patterns for one route-direction.

    Args:
        workbook: The Excel workbook object.
        sheet_title: Name for the worksheet.
        route_short_name: Short name of the route.
        direction_id_val: Direction ID (0 or 1).
        service_id_val: GTFS service_id.
        pattern_records_dir: List of patterns for this direction.
        master_stops: List of (stop_id, stop_name) tuples for the master trip.

    Side Effects:
        Adds a worksheet to the workbook.
    """
    try:
        worksheet = workbook.create_sheet(title=sheet_title)
    except ValueError:
        worksheet = workbook.create_sheet(title=f"{sheet_title[:28]}_X")

    if not master_stops:
        worksheet.append(
            [f"No master stops found for Route {route_short_name} Dir {direction_id_val}."]
        )
        return

    header = [
        "Route",
        "Direction",
        "Calendar (service_id)",
        "Pattern ID",
        "Trip Count",
        "Earliest Start Time",
    ]
    for _, stop_name_val in master_stops:
        header.append(stop_name_val)
    worksheet.append(header)

    pattern_records_dir = sorted(
        pattern_records_dir,
        key=lambda rec: (
            rec.get("earliest_time_minutes") is None,
            rec.get("earliest_time_minutes", float("inf")),
            rec.get("pattern_id", 0),
        ),
    )

    for rec in pattern_records_dir:
        pat_id = rec["pattern_id"]
        trip_count = rec["trip_count"]
        earliest_str = rec.get("earliest_time_str", "")
        pattern_stops_tuple = rec["pattern_stops"]
        row_distances = forward_match_pattern_to_master(pattern_stops_tuple, master_stops)

        row_data = [
            route_short_name,
            str(direction_id_val),
            str(service_id_val),
            pat_id,
            trip_count,
            earliest_str,
        ]
        row_data.extend(row_distances)
        worksheet.append(row_data)

    for col_index, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_index)
        max_len = len(str(header[col_index - 1]))
        # Iterate through data for this column to find max length for slightly better auto-fit
        for row_idx in range(2, worksheet.max_row + 1):  # Start from data row 2
            cell_value = worksheet.cell(row=row_idx, column=col_index).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        worksheet.column_dimensions[col_letter].width = max(12, min(max_len + 2, 50))


def export_patterns_to_excel(
    pattern_records: List[Dict[str, Any]],
    routes_df: pd.DataFrame,
    trips_df: pd.DataFrame,
    stop_times_df: pd.DataFrame,
    stops_df: pd.DataFrame,
    calendar_df: Optional[pd.DataFrame] = None,
) -> None:
    """Export pattern data to Excel workbooks by route and service_id.

    Args:
        pattern_records: List of enriched pattern records.
        routes_df: GTFS routes DataFrame.
        trips_df: GTFS trips DataFrame.
        stop_times_df: GTFS stop_times DataFrame.
        stops_df: GTFS stops DataFrame.
        calendar_df: Optional calendar DataFrame.

    Side Effects:
        Writes Excel files to disk.
    """
    group_map = defaultdict(list)
    for pat_rec in pattern_records:
        rid_val = str(pat_rec["route_id"])
        sid_val = str(pat_rec["service_id"])
        group_map[(rid_val, sid_val)].append(pat_rec)

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logging.info(f"Created output directory: {OUTPUT_DIR}")

    for (rid_val, sid_val), group_list_for_route_service in group_map.items():
        route_info = routes_df[routes_df["route_id"] == rid_val]
        if not route_info.empty:
            short_name = route_info.iloc[0].get("route_short_name", f"Route_{rid_val}")
        else:
            short_name = f"Route_{rid_val}"
            logging.warning(f"No route_short_name found for route_id {rid_val}.")

        sane_short_name = (
            "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in str(short_name))
            .strip()
            .replace(" ", "_")
        )

        folder_name = format_service_id_folder_name(sid_val, calendar_df)
        folder_path = os.path.join(OUTPUT_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        workbook = create_workbook()
        dir_map = defaultdict(list)
        for rec_dir in group_list_for_route_service:
            d_id_val = str(rec_dir["direction_id"])
            dir_map[d_id_val].append(rec_dir)

        if not dir_map:
            logging.warning(
                f"No patterns to export for Route {sane_short_name}, Service ID {sid_val}."
            )
            continue

        for direction_val_str, recs_for_direction in dir_map.items():
            all_trip_ids_for_master = set()
            for pattern_rec in (
                recs_for_direction
            ):  # recs_for_direction contains patterns for specific route,service,direction
                all_trip_ids_for_master.update(pattern_rec["trip_ids"])

            # Filter trips_df for current route, service, direction to find master stops
            # Ensure IDs are strings for comparison, as they are loaded as such.
            relevant_trips_for_master = trips_df[
                (trips_df["route_id"].astype(str) == rid_val)
                & (trips_df["service_id"].astype(str) == sid_val)
                & (trips_df["direction_id"].astype(str) == direction_val_str)
                & (
                    trips_df["trip_id"].isin(all_trip_ids_for_master)
                )  # trip_id is already in the correct set
            ]

            current_master_stops = find_master_trip_stops(
                rid_val,
                direction_val_str,
                relevant_trips_for_master,
                stop_times_df,
                stops_df,
            )

            sheet_title = f"Dir{direction_val_str}"
            fill_worksheet_for_direction(
                workbook,
                sheet_title,
                sane_short_name,
                direction_val_str,
                sid_val,
                recs_for_direction,
                current_master_stops,
            )

        if not workbook.sheetnames:
            logging.warning(
                f"Workbook for Route {sane_short_name}, Service {sid_val} is empty. Not saving."
            )
            continue

        filename = f"{sane_short_name}_{sid_val}_{SIGNUP_NAME}.xlsx"
        filename = "".join(c if c.isalnum() or c in (".", "_", "-") else "_" for c in filename)

        full_filepath = os.path.join(folder_path, filename)
        try:
            workbook.save(full_filepath)
            logging.info("Saved workbook: %s", full_filepath)
        except Exception as exc:
            logging.error("Could not save workbook '%s': %s", filename, exc)


# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_folder_path: Absolute or relative path to the folder
            containing the GTFS feed.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Folder missing or one of *files* not present.
        ValueError: Empty file or CSV parser failure.
        RuntimeError: Generic OS error while reading a file.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data: dict[str, pd.DataFrame] = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=cast("Any", dtype), low_memory=False)
            data[key] = df
            logging.info("Loaded %s (%d records).", file_name, len(df))

        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

    return data


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """Main script function for generating GTFS pattern exports."""
    if not os.path.exists(OUTPUT_DIR):
        try:
            os.makedirs(OUTPUT_DIR)
            logging.info("Created output directory: %s", OUTPUT_DIR)
        except OSError as exc:
            logging.error("Could not create output directory %s: %s", OUTPUT_DIR, exc)
            return

    # Load calendar.txt (optional)
    calendar_df = None
    calendar_file_path = str(INPUT_DIR / "calendar.txt")
    if os.path.exists(calendar_file_path):
        try:
            calendar_df = pd.read_csv(calendar_file_path, dtype=str)

            if calendar_df.empty and os.path.getsize(calendar_file_path) > 0:
                logging.warning(
                    "Loaded calendar.txt but it appears to have headers only or is malformed "
                    "(%d records).",
                    len(calendar_df),
                )
            elif not calendar_df.empty:
                logging.info("Loaded calendar.txt successfully (%d records).", len(calendar_df))
            else:
                logging.info("calendar.txt (%s) exists but is empty.", calendar_file_path)
                calendar_df = None

        except pd.errors.EmptyDataError:
            logging.warning(
                "calendar.txt (%s) is empty. Proceeding without calendar day names.",
                calendar_file_path,
            )
            calendar_df = None

        except Exception as exc:
            logging.warning("Could not load calendar.txt from %s: %s", calendar_file_path, exc)
            calendar_df = None

    else:
        logging.info(
            "No calendar.txt found at %s; subfolders will be 'calendar_<service_id>' only.",
            calendar_file_path,
        )

    # Load essential GTFS files
    essential_files = [
        "stops.txt",
        "trips.txt",
        "stop_times.txt",
        "routes.txt",
    ]
    try:
        gtfs_data = load_gtfs_data(str(INPUT_DIR), files=essential_files, dtype=str)
        logging.info("Loaded essential GTFS files successfully using load_gtfs_data().")

    except (OSError, ValueError, RuntimeError) as exc:
        logging.error("Failed to load essential GTFS files: %s", exc)
        return

    stops_df = gtfs_data.get("stops")
    trips_df = gtfs_data.get("trips")
    stop_times_df = gtfs_data.get("stop_times")
    routes_df = gtfs_data.get("routes")

    if stops_df is None or trips_df is None or stop_times_df is None or routes_df is None:
        logging.error("One or more essential GTFS DataFrames could not be loaded. Exiting.")
        return

    if any(df.empty for df in [stops_df, trips_df, stop_times_df, routes_df]):
        empty_files = [
            name
            for name, df in zip(
                ["stops", "trips", "stop_times", "routes"],
                [stops_df, trips_df, stop_times_df, routes_df],
                strict=True,
            )
            if df.empty
        ]
        logging.error(
            "Essential GTFS DataFrame(s) are empty: %s. This may indicate issues with the input "
            "files. Exiting.",
            ", ".join(empty_files),
        )
        return

    # --- Type Conversions due to dtype=str from load_gtfs_data ---
    if "stop_sequence" in stop_times_df.columns:
        stop_times_df["stop_sequence"] = pd.to_numeric(
            stop_times_df["stop_sequence"], errors="raise"
        )
    else:
        logging.error(
            "'stop_sequence' column is missing from stop_times.txt. This is critical. Exiting."
        )
        return

    if "timepoint" in stop_times_df.columns:
        stop_times_df["timepoint"] = pd.to_numeric(stop_times_df["timepoint"], errors="coerce")

    if "shape_dist_traveled" in stop_times_df.columns:
        stop_times_df["shape_dist_traveled"] = pd.to_numeric(
            stop_times_df["shape_dist_traveled"], errors="coerce"
        )
    # --- End Type Conversions ---

    filtered_trips = filter_trips(trips_df, routes_df, cal_ids=FILTER_IN_CALENDAR_IDS)
    if filtered_trips.empty:
        logging.info("No trips after filtering. This may be expected based on filters. Exiting.")
        return

    patterns_dict = generate_unique_patterns(filtered_trips, stop_times_df, stops_df)
    if not patterns_dict:
        logging.warning("No patterns found. Exiting.")
        return

    pattern_records = assign_pattern_ids(patterns_dict)
    if not pattern_records:
        logging.warning("No pattern records after assigning IDs. Exiting.")
        return

    compute_earliest_start_times(pattern_records, stop_times_df)
    export_patterns_to_excel(
        pattern_records, routes_df, trips_df, stop_times_df, stops_df, calendar_df
    )

    logging.info("Processing complete.")


if __name__ == "__main__":
    main()
