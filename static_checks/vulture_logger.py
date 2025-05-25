"""
Script Name:
    vulture_logger.py

Purpose:
    Scans Python files using Vulture to detect dead (unused) code.
    The script gathers .py files from specified target locations and
    executes Vulture on each file. Finally, it logs the results and
    generates an Excel summary report of Vulture's findings with
    confidence levels.

Inputs:
    1. Target Python files/folders (specified by `FILES_OR_FOLDERS` constant):
       A list of paths to Python files or directories to be scanned.
    2. Skip paths (specified by `SKIP_PATHS` constant):
       A list of file or directory paths to be excluded from the scan.
    3. Configuration constants defined within the script:
       - `OUTPUT_FOLDER`: Directory path to save the detailed log and Excel summary.
       - `LOG_LEVEL`: Logging level for console output (e.g., `logging.INFO`).
       - `DETAILED_LOG_FILENAME_PREFIX`: Prefix for the detailed log file name.
       - `VULTURE_MIN_CONFIDENCE`: The minimum confidence level (0-100) for Vulture
         to report an item as dead code.

Outputs:
    1. Detailed Log File: A timestamped .log file (e.g.,
       `vulture_detailed_log_YYYYMMDD_HHMMSS.log`) in `OUTPUT_FOLDER`,
       containing the complete stdout and stderr from Vulture for each
       processed file.
    2. Excel Summary File: A timestamped .xlsx file (e.g.,
       `vulture_results_YYYYMMDD_HHMMSS.xlsx`) in `OUTPUT_FOLDER`,
       with a sheet "Vulture Dead Code Summary" detailing for each file:
       Script Name, Immediate Folder, Dead Code Items Count, Dead Code Details,
       Vulture Stderr, and Full Path.
    3. Console output: Status messages, progress updates, and a summary of
       the dead code analysis.

Dependencies:
    - logging, os, re, subprocess, sys, datetime, pathlib, typing, openpyxl
    - vulture (external command-line tool, must be installed and accessible)
"""

from __future__ import annotations

import logging
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ==================================================================================================
# CONFIGURATION
# ==================================================================================================

FILES_OR_FOLDERS: List[str] = [
    # Example: mix .py files and/or folders
    # r"C:\Path\to\your\python_project_folder",
    # r"C:\Path\to\another_script.py",
    r"C:\Temp\TestScripts",  # Replace with your actual paths
]

SKIP_PATHS: List[str] = [
    # r"C:\Path\to\folder_to_skip",
    # r"C:\Path\to\file_to_skip.py",
    r"C:\Path\to\your\python_project_folder\venv",
    r"C:\Temp\TestScripts\skip_this_script.py",  # Replace with your actual paths
]

OUTPUT_FOLDER: str = r"C:\Temp\VultureReports"  # Replace with your desired output path

LOG_LEVEL: int = logging.INFO  # set to logging.DEBUG for more detail
DETAILED_LOG_FILENAME_PREFIX: str = "vulture_detailed_log"
VULTURE_MIN_CONFIDENCE: int = 60  # Vulture's default is 60

# ==================================================================================================
# LOGGING SETUP
# ==================================================================================================
logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
console_logger = logging.getLogger(__name__)

# ==================================================================================================
# FUNCTIONS
# ==================================================================================================


def setup_detailed_logger(
    output_folder: str, filename_prefix: str, log_level: int
) -> Tuple[logging.Logger, str]:
    """Create a dedicated file logger and return (logger, log_filepath)."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"{filename_prefix}_{timestamp}.log"
    Path(output_folder).mkdir(parents=True, exist_ok=True)
    log_filepath = os.path.join(output_folder, log_filename)

    detail_logger = logging.getLogger("VultureDetailLogger")
    detail_logger.setLevel(log_level)
    detail_logger.propagate = False
    detail_logger.handlers.clear()

    file_handler = logging.FileHandler(log_filepath, mode="w", encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(message)s"))  # Log raw tool output
    detail_logger.addHandler(file_handler)

    return detail_logger, log_filepath


def gather_python_files_in_folder(folder_path: str) -> List[str]:
    """Recursively list *.py inside folder_path."""
    return [
        os.path.join(root, f)
        for root, _, files in os.walk(folder_path)
        for f in files
        if f.endswith(".py")
    ]


def is_skipped(path_str: str, skip_list: List[str]) -> bool:
    """Return True if path_str matches or is inside any path in skip_list."""
    norm_path = Path(path_str).resolve().as_posix().lower()
    for s in skip_list:
        norm_skip = Path(s).resolve().as_posix().lower()
        if norm_path == norm_skip or norm_path.startswith(norm_skip + "/"):
            return True
    return False


# --------------------------- Vulture helpers ------------------------------------------------------


def run_vulture_on_file(
    py_file: str, min_confidence: int
) -> Tuple[str | None, str | None]:
    """Run Vulture and return (stdout, stderr)."""
    try:
        command = [
            sys.executable,
            "-m",
            "vulture",
            py_file,
            "--min-confidence",
            str(min_confidence),
        ]
        console_logger.debug("Running command: %s", " ".join(command))
        proc = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,  # Vulture exits with non-zero if dead code is found
        )
        return proc.stdout, proc.stderr
    except FileNotFoundError:
        console_logger.error(
            "python -m vulture not found. Is Vulture installed and in your PATH?"
        )
        return None, "Execution Error: Vulture not found."
    except Exception as exc:
        console_logger.error("Unexpected error running Vulture on %s: %s", py_file, exc)
        return None, f"Execution Error: {exc}"


def parse_vulture_output(
    vulture_stdout: str | None, file_being_scanned: str
) -> List[Dict[str, str]]:
    """
    Parse Vulture's stdout to extract dead code items.
    Example line: C:\Path\file.py:10: unused variable 'x' (60% confidence)
    """
    if not vulture_stdout:
        return []

    found_items: List[Dict[str, str]] = []
    # Regex to capture: path (optional), line, item description, confidence
    # Vulture output might sometimes include the full path again, or just line:item (confidence)
    # Making the path part flexible.
    item_re = re.compile(r"^(?:.+?:)?(\d+):\s*(.+?)\s*\((\d+%)\s*confidence\)$")

    for line in vulture_stdout.strip().splitlines():
        line = line.strip()
        match = item_re.match(line)
        if match:
            line_num, item_desc, confidence_str = match.groups()
            found_items.append(
                {
                    "file_path": file_being_scanned,  # Use the path of the file Vulture was run on
                    "line": line_num,
                    "item": item_desc.strip(),
                    "confidence": confidence_str,
                }
            )
        elif (
            line
        ):  # Non-empty line that didn't match, could be an error or other message from Vulture
            console_logger.debug("Unparsed Vulture output line: %s", line)
            # You could decide to include these in stderr or a separate notes field
    return found_items


# ==================================================================================================
# MAIN VULTURE WORKFLOW
# ==================================================================================================


def scan_for_dead_code_and_create_outputs(
    files_or_folders: List[str],
    skip_list: List[str],
    output_folder: str,
    min_confidence: int,
) -> None:
    """Run Vulture, log details, build Excel summary."""
    detail_logger, log_filepath = setup_detailed_logger(
        output_folder,
        DETAILED_LOG_FILENAME_PREFIX,
        logging.DEBUG,  # Log all Vulture output
    )
    console_logger.info("Detailed Vulture log: %s", log_filepath)

    # -------------------------------------------------------------------------
    # Collect .py files
    # -------------------------------------------------------------------------
    collected: List[str] = []
    for entry in files_or_folders:
        abs_entry = Path(entry).resolve()
        if not abs_entry.exists():
            console_logger.warning("Path not found, skipping: %s", abs_entry)
            continue
        if abs_entry.is_dir():
            console_logger.info("Scanning folder: %s", abs_entry)
            collected.extend(gather_python_files_in_folder(abs_entry.as_posix()))
        elif abs_entry.as_posix().lower().endswith(".py"):
            collected.append(abs_entry.as_posix())
        else:
            console_logger.debug("Non-Python path skipped: %s", abs_entry)

    # Remove duplicates (preserve order) and apply skip list
    unique_files = list(dict.fromkeys(collected))
    final_files = [f for f in unique_files if not is_skipped(f, skip_list)]

    if not final_files:
        console_logger.info("No Python files to scan after applying skips.")
        detail_logger.info("No Python files to scan after applying skips.")
        return

    console_logger.info("Scanning %d Python file(s) with Vulture...", len(final_files))
    detail_logger.info(
        "Scanning %d Python file(s) with Vulture (min_confidence=%d%%)...",
        len(final_files),
        min_confidence,
    )

    # -------------------------------------------------------------------------
    # Scan each file with Vulture
    # -------------------------------------------------------------------------
    results_summary: List[Dict[str, Any]] = []
    for idx, py_file_path_str in enumerate(final_files, start=1):
        py_file = Path(py_file_path_str)
        console_logger.info("[%d/%d] Scanning: %s", idx, len(final_files), py_file)
        detail_logger.info("\n" + "=" * 80)
        detail_logger.info("FILE: %s", py_file)
        detail_logger.info("Time: %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        detail_logger.info("=" * 80 + "\n")

        stdout_vulture, stderr_vulture = run_vulture_on_file(
            py_file.as_posix(), min_confidence
        )

        detail_logger.info("--- Vulture STDOUT ---\n%s", stdout_vulture or "<empty>")
        if (
            stderr_vulture
        ):  # Vulture often uses stderr for info messages even on success
            detail_logger.info("\n--- Vulture STDERR ---\n%s", stderr_vulture)

        dead_code_items = parse_vulture_output(stdout_vulture, py_file.as_posix())
        formatted_dead_code = "; ".join(
            [
                f"{item['item']} (line {item['line']}, {item['confidence']})"
                for item in dead_code_items
            ]
        )

        results_summary.append(
            {
                "script_name": py_file.name,
                "immediate_folder": py_file.parent.name,
                "full_path": py_file.as_posix(),
                "dead_code_items_count": len(dead_code_items),
                "dead_code_details": (
                    formatted_dead_code if dead_code_items else "None found"
                ),
                "vulture_stderr": stderr_vulture.strip() if stderr_vulture else "",
            }
        )

    # -------------------------------------------------------------------------
    # Excel summary
    # -------------------------------------------------------------------------
    if not results_summary:
        console_logger.info("No results to write to Excel.")
        detail_logger.info("No results to write to Excel.")
        return

    console_logger.info("Building Excel summary for Vulture findings...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulture Dead Code Summary"

    headers = [
        "Script Name",
        "Immediate Folder",
        "Dead Code Items Count",
        "Dead Code Details",
        "Vulture Stderr",
        "Full Path",
    ]
    ws.append(headers)

    for row_data in results_summary:
        ws.append(
            [
                row_data["script_name"],
                row_data["immediate_folder"],
                row_data["dead_code_items_count"],
                row_data["dead_code_details"],
                row_data["vulture_stderr"],
                row_data["full_path"],
            ]
        )

    # Autofit columns (simple approach)
    for i, column_cells in enumerate(ws.columns):
        max_length = 0
        column = get_column_letter(i + 1)
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:  # pylint: disable=bare-except
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = min(adjusted_width, 70)  # Cap width

    excel_name = f"vulture_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_path = Path(output_folder) / excel_name
    try:
        wb.save(excel_path.as_posix())
        console_logger.info("Excel summary written: %s", excel_path)
        detail_logger.info("\nExcel summary written: %s", excel_path)
    except Exception as e:
        console_logger.error("Failed to save Excel file: %s", e)
        detail_logger.error("Failed to save Excel file: %s", e)

    # Close file handlers for the detail_logger
    for handler in detail_logger.handlers:
        handler.close()
        detail_logger.removeHandler(handler)

    console_logger.info("Vulture scan and reporting complete.")


# ==================================================================================================
# MAIN
# ==================================================================================================


def main_cli() -> None:
    """Main command-line interface function."""
    # Ensure output folder exists
    Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)

    scan_for_dead_code_and_create_outputs(
        FILES_OR_FOLDERS, SKIP_PATHS, OUTPUT_FOLDER, VULTURE_MIN_CONFIDENCE
    )


if __name__ == "__main__":
    # For testing: Create dummy files and folders if they don't exist
    # You would remove this in your actual usage
    if FILES_OR_FOLDERS == [
        r"C:\Temp\TestScripts"
    ]:  # Only if using default example path
        Path(r"C:\Temp\TestScripts").mkdir(parents=True, exist_ok=True)
        Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)
        with open(
            r"C:\Temp\TestScripts\sample_script_1.py", "w", encoding="utf-8"
        ) as f:
            f.write(
                "import os\n\ndef unused_function():\n    pass\n\nMY_UNUSED_VAR = 123\n\nprint('Hello')\n"
            )
        with open(
            r"C:\Temp\TestScripts\sample_script_2.py", "w", encoding="utf-8"
        ) as f:
            f.write(
                "def used_function():\n    print('This is used')\n\nused_function()\n"
            )
        with open(
            r"C:\Temp\TestScripts\skip_this_script.py", "w", encoding="utf-8"
        ) as f:
            f.write("SHOULD_BE_SKIPPED = True\n")

    main_cli()
