"""
Script Name:
    static_code_auditor.py

Purpose:
    Performs a comprehensive static analysis sweep across specified Python files and folders.
    This includes type checking (mypy), dead code detection (vulture), style and quality enforcement
    (pylint + isort), and docstring structure validation. Each tool's output is logged in detail
    and summarized in a separate Excel file.

Inputs:
    1. FILES_OR_FOLDERS (List[str]):
       A list of file and folder paths to scan. Directories are searched recursively.
    2. SKIP_PATHS (List[str]):
       Paths to exclude from scanning. Matching is based on full path containment.
    3. Configuration flags within the script:
       - ENABLE_MYPY (bool):        Whether to run mypy.
       - ENABLE_VULTURE (bool):     Whether to run vulture.
       - ENABLE_LINT (bool):        Whether to run pylint, isort, and docstring checks.
       - OUTPUT_FOLDER (str):       Directory to write logs and reports.
       - LOG_LEVEL (int):           Logging verbosity (e.g. logging.INFO).
       - MYPY_ADDITIONAL_ARGS (List[str]): Extra flags to pass to mypy.
       - VULTURE_MIN_CONFIDENCE (int): Minimum confidence threshold for vulture.
       - REQUIRED_DOC_HEADERS (Tuple[str, ...]): Headers expected in each module-level docstring.

Outputs:
    1. Detailed Log Files:
       One .log file per enabled tool (e.g., mypy_detailed_log_YYYYMMDD_HHMMSS.log),
       containing full stdout and stderr output from the tool.
    2. Excel Summary Files:
       One .xlsx file per tool with a timestamped name and structured summary sheet:
           - mypy_results_YYYYMMDD_HHMMSS.xlsx
           - vulture_results_YYYYMMDD_HHMMSS.xlsx
           - lint_results_YYYYMMDD_HHMMSS.xlsx
    3. Console Output:
       Progress messages and final summary.

Dependencies:
        - os, sys, re, subprocess, logging, pathlib, datetime, typing, ast
        - openpyxl, mypy, vulture, pylint, isort
"""

from __future__ import annotations

import ast
import logging
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION – toggle these for notebook usage
# ============================================================================

FILES_OR_FOLDERS: List[str] = [
    r"C:\Path\to\Main\Input\Folder", # Replace with your folder or file path
]
SKIP_PATHS: List[str] = [
    # e.g. r"C:\Path\to\project\venv",
]

OUTPUT_FOLDER: str = r"C:\Path\to\Output\Folder" # Replace with your folder path
LOG_LEVEL: int = logging.INFO

ENABLE_MYPY:    bool = True
ENABLE_VULTURE: bool = True
ENABLE_LINT:    bool = True  # pylint + isort + docstring

# mypy
MYPY_ADDITIONAL_ARGS: List[str] = []  # e.g. ["--ignore-missing-imports"]

# vulture
VULTURE_MIN_CONFIDENCE: int = 60      # default is 60

# lint (pylint + isort + docstring)
REQUIRED_DOC_HEADERS: Tuple[str, ...] = (
    "Script Name:",
    "Purpose:",
    "Inputs:",
    "Outputs:",
    "Dependencies:",
)

# ============================================================================
# SHARED HELPERS
# ============================================================================

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
CONSOLE = logging.getLogger(__name__)


def setup_detailed_logger(
    out_folder: str, prefix: str, level: int = logging.DEBUG
) -> Tuple[logging.Logger, str]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = Path(out_folder).expanduser().resolve()
    folder.mkdir(parents=True, exist_ok=True)
    logfile = str(folder / f"{prefix}_{ts}.log")

    lg = logging.getLogger(prefix)
    lg.setLevel(level)
    lg.propagate = False
    lg.handlers.clear()

    fh = logging.FileHandler(logfile, mode="w", encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(message)s"))
    lg.addHandler(fh)
    return lg, logfile


def gather_python_files(
    paths: List[str], skip_list: List[str]
) -> List[str]:
    """Collect .py files, respecting the skip list."""
    collected: List[str] = []

    def is_skipped(p: Path) -> bool:
        norm = p.resolve().as_posix().lower()
        for s in skip_list:
            ns = Path(s).resolve().as_posix().lower()
            if norm == ns or norm.startswith(ns + "/"):
                return True
        return False

    for entry in paths:
        p = Path(entry).expanduser()
        if not p.exists():
            CONSOLE.warning("Path not found – skipping: %s", p)
            continue
        if p.is_dir():
            for root, _, files in os.walk(p):
                for f in files:
                    if f.endswith(".py"):
                        fpath = Path(root, f)
                        if not is_skipped(fpath):
                            collected.append(fpath.as_posix())
        elif p.suffix.lower() == ".py" and not is_skipped(p):
            collected.append(p.as_posix())

    # dedupe while preserving order
    return list(dict.fromkeys(collected))


# ============================================================================
# 1) MYPY CHECK
# ============================================================================

_MYPY_RE = re.compile(r"Found\s+(\d+)\s+error")


def run_mypy(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "mypy_detailed_log")
    CONSOLE.info("mypy log → %s", log_fp)

    results = []
    for idx, py in enumerate(files, 1):
        logger.info("\n%s\nFILE %d/%d → %s\n%s",
                    "="*80, idx, len(files), py, "="*80)
        try:
            proc = subprocess.run(
                [sys.executable, "-m", "mypy",
                 "--show-error-codes", "--no-color-output",
                 *MYPY_ADDITIONAL_ARGS, py],
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            out, err = proc.stdout, proc.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m mypy not found – aborting mypy pass.")
            return len(files)
        logger.info(out or "")
        if err:
            logger.info("\n--- stderr ---\n%s", err)

        if "Success: no issues found" in (out or ""):
            errors = 0
        elif m := _MYPY_RE.search(out or ""):
            errors = int(m.group(1))
        else:
            errors = len([l for l in (out or "").splitlines() if ": error:" in l])

        results.append({
            "script": Path(py).name,
            "folder": Path(py).parent.name,
            "path": py,
            "errors": errors,
            "success": "Yes" if errors == 0 else "No",
            "stderr": err or "",
        })

    # Excel summary
    wb = Workbook()
    ws = wb.active
    ws.title = "mypy Summary"
    ws.append(["Script", "Folder", "# Errors", "Success", "Path", "Stderr"])
    for r in results:
        ws.append([r["script"], r["folder"], r["errors"],
                   r["success"], r["path"], r["stderr"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = \
            min(max(len(str(c.value or "")) for c in col) + 2, 80)

    xlsx = Path(OUTPUT_FOLDER) / f"mypy_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("mypy summary → %s", xlsx)

    return sum(1 for r in results if r["errors"] > 0)


# ============================================================================
# 2) VULTURE CHECK
# ============================================================================

_VULTURE_LINE = re.compile(r"^(?:.+?:)?(\d+):\s*(.+?)\s*\((\d+%) confidence\)$")


def run_vulture(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "vulture_detailed_log")
    CONSOLE.info("vulture log → %s", log_fp)

    summary = []
    for idx, py in enumerate(files, 1):
        logger.info("\n%s\nFILE %d/%d → %s\n%s",
                    "="*80, idx, len(files), py, "="*80)
        try:
            proc = subprocess.run(
                [sys.executable, "-m", "vulture", py,
                 "--min-confidence", str(VULTURE_MIN_CONFIDENCE)],
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            out, err = proc.stdout, proc.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m vulture not found – aborting vulture pass.")
            return len(files)
        logger.info(out or "")
        if err:
            logger.info("\n--- stderr ---\n%s", err)

        items = []
        for line in (out or "").splitlines():
            m = _VULTURE_LINE.match(line.strip())
            if m:
                items.append(f"{m.group(2)} (line {m.group(1)}, {m.group(3)})")

        summary.append({
            "script": Path(py).name,
            "folder": Path(py).parent.name,
            "path": py,
            "count": len(items),
            "details": "; ".join(items) if items else "None",
            "stderr": (err or "").strip(),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Vulture Dead Code"
    ws.append(["Script", "Folder", "Items", "Details", "Stderr", "Path"])
    for r in summary:
        ws.append([r["script"], r["folder"], r["count"],
                   r["details"], r["stderr"], r["path"]])
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            min(max(len(str(c.value or "")) for c in col) + 2, 70)

    xlsx = Path(OUTPUT_FOLDER) / f"vulture_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("vulture summary → %s", xlsx)

    return sum(1 for r in summary if r["count"] > 0)


# ============================================================================
# 3) PYLINT + ISORT + DOCSTRING CHECK
# ============================================================================

_PYLINT_ISSUE = re.compile(r":\s*[CREFWE]\d{4}:")


def _docstring_ok(py_file: str) -> bool:
    try:
        with open(py_file, "r", encoding="utf-8") as fh:
            tree = ast.parse(fh.read())
    except (SyntaxError, OSError):
        return False
    doc = ast.get_docstring(tree)
    if not doc:
        return False
    want = [h.lower() for h in REQUIRED_DOC_HEADERS]
    idx = 0
    for ln in (l.strip().lower() for l in doc.splitlines() if l.strip()):
        if ln.startswith(want[idx]):
            idx += 1
            if idx == len(want):
                return True
    return False


def run_lint(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "lint_detailed_log")
    CONSOLE.info("lint log → %s", log_fp)

    summary = []
    for idx, py in enumerate(files, 1):
        logger.info("\n%s\nFILE %d/%d → %s\n%s",
                    "="*80, idx, len(files), py, "="*80)

        # pylint
        try:
            pr = subprocess.run(
                [sys.executable, "-m", "pylint", py],
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            pout, perr = pr.stdout, pr.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m pylint not found – aborting lint pass.")
            return len(files)
        logger.info("--- pylint stdout ---\n%s", pout or "")
        if perr:
            logger.info("\n--- pylint stderr ---\n%s", perr)

        score = 0.0
        issues = 0
        for line in (pout or "").splitlines():
            if _PYLINT_ISSUE.search(line):
                issues += 1
            if "Your code has been rated at" in line:
                try:
                    score = float(line.split("rated at")[1].split("/")[0])
                except Exception:
                    pass

        # isort
        try:
            ir = subprocess.run(
                [sys.executable, "-m", "isort", "--check-only", "--diff", py],
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            iout, ierr = ir.stdout, ir.stderr
            needs_isort = ir.returncode != 0
        except FileNotFoundError:
            CONSOLE.error("python -m isort not found – aborting lint pass.")
            return len(files)
        logger.info("\n--- isort diff ---\n%s", iout or "")
        if ierr:
            logger.info("\n--- isort stderr ---\n%s", ierr)

        # docstring
        doc_ok = _docstring_ok(py)

        summary.append({
            "script": Path(py).name,
            "folder": Path(py).parent.name,
            "score": score,
            "issues": issues,
            "isort": "Yes" if needs_isort else "No",
            "doc":   "Yes" if doc_ok else "No",
            "stderr": (perr or "") + (ierr or ""),
            "path": py,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Lint Summary"
    ws.append([
        "Script", "Folder", "Pylint Score", "Pylint Issues",
        "Needs isort", "Docstring OK", "Stderr", "Path"
    ])
    for r in summary:
        ws.append([
            r["script"], r["folder"], r["score"], r["issues"],
            r["isort"], r["doc"], r["stderr"], r["path"]
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = \
            min(max(len(str(c.value or "")) for c in col) + 2, 80)

    xlsx = Path(OUTPUT_FOLDER) / f"lint_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("lint summary → %s", xlsx)

    def _has_issues(r: dict) -> bool:
        return (r["issues"] > 0) or (r["isort"] == "Yes") or (r["doc"] == "No")

    return sum(1 for r in summary if _has_issues(r))


# ============================================================================
# MAIN
# ============================================================================

def main() -> None:
    files = gather_python_files(FILES_OR_FOLDERS, SKIP_PATHS)
    if not files:
        CONSOLE.warning("No Python files found – nothing to do.")
        return

    Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)
    overall_failures = 0

    if ENABLE_MYPY:
        CONSOLE.info("\n=== Running mypy pass ===")
        overall_failures += run_mypy(files)

    if ENABLE_VULTURE:
        CONSOLE.info("\n=== Running vulture pass ===")
        overall_failures += run_vulture(files)

    if ENABLE_LINT:
        CONSOLE.info("\n=== Running pylint/isort pass ===")
        overall_failures += run_lint(files)

    if overall_failures == 0:
        CONSOLE.info("🎉 All enabled checks passed with flying colours.")
    else:
        CONSOLE.warning(
            "⚠️ Static analysis detected issues in %d file(s).", overall_failures
        )


if __name__ == "__main__":
    main()
