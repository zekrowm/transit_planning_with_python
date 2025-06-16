"""Runs static analysis tools on Python files and exports detailed logs and summaries.

Executes multiple code quality tools—`mypy`, `vulture`, `pylint`, and `pydocstyle`
—on specified Python files and directories. Generates per-tool log files
and Excel summaries, intended for systematic QA in large or shared codebases.

Outputs:
    - Log files: One per tool with stdout/stderr from each scan
    - Excel summaries: One per tool with issue counts and details
    - Console output: Progress updates and summary of results
"""

from __future__ import annotations

import logging
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple  # Added Dict and Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

FILES_OR_FOLDERS: List[str] = [
    r"C:\Path\to\Main\Input\Folder",  # Replace with your folder or file path
]
SKIP_PATHS: List[str] = [
    # e.g. r"C:\Path\to\project\venv",
]
OUTPUT_FOLDER: str = r"C:\Path\to\Output\Folder"  # Replace with your folder path
LOG_LEVEL: int = logging.INFO

ENABLE_MYPY: bool = True
ENABLE_VULTURE: bool = True
ENABLE_PYLINT: bool = True
ENABLE_PYDOCSTYLE: bool = True

# mypy
MYPY_ADDITIONAL_ARGS: List[str] = []  # e.g. ["--ignore-missing-imports"]
# vulture
VULTURE_MIN_CONFIDENCE: int = 60  # default is 60
# pydocstyle
PYDOCSTYLE_ADDITIONAL_ARGS: List[str] = ["--convention=google"]
# pylint
PYLINT_ADDITIONAL_ARGS: list[str] = ["--errors-only"]  # == "-E"

# ----------------------------------------------------------------------------
# LOGGING
# ----------------------------------------------------------------------------

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
CONSOLE = logging.getLogger(__name__)

# ============================================================================
# FUNCTIONS
# ============================================================================


def setup_detailed_logger(
    out_folder: str, prefix: str, level: int = logging.DEBUG
) -> Tuple[logging.Logger, str]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = Path(out_folder).expanduser().resolve()
    folder.mkdir(parents=True, exist_ok=True)
    logfile = str(folder / f"{prefix}_{ts}.log")
    lg = logging.getLogger(prefix)
    lg.setLevel(level)
    lg.propagate = False
    lg.handlers.clear()
    fh = logging.FileHandler(logfile, mode="w", encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(message)s"))
    lg.addHandler(fh)
    return lg, logfile


def gather_python_files(paths: List[str], skip_list: List[str]) -> List[str]:
    """Collect .py files, respecting the skip list."""
    collected: List[str] = []

    def is_skipped(p: Path) -> bool:
        norm = p.resolve().as_posix().lower()
        for s in skip_list:
            ns = Path(s).resolve().as_posix().lower()
            if norm == ns or norm.startswith(ns + "/"):
                return True
        return False

    for entry in paths:
        p = Path(entry).expanduser()
        if not p.exists():
            CONSOLE.warning("Path not found – skipping: %s", p)
            continue
        if p.is_dir():
            for root, _, files in os.walk(p):
                for f in files:
                    if f.endswith(".py"):
                        fpath = Path(root, f)
                        if not is_skipped(fpath):
                            collected.append(fpath.as_posix())
        elif p.suffix.lower() == ".py" and not is_skipped(p):
            collected.append(p.as_posix())
    return list(dict.fromkeys(collected))


# ----------------------------------------------------------------------------
# 1) MYPY CHECK
# ----------------------------------------------------------------------------
_MYPY_RE = re.compile(r"Found\s+(\d+)\s+error")


def run_mypy(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "mypy_detailed_log")
    CONSOLE.info("mypy log → %s", log_fp)
    results: List[Dict[str, Any]] = []  # Explicitly typed
    for idx, py in enumerate(files, 1):
        logger.info(
            "\n%s\nFILE %d/%d → %s\n%s", "=" * 80, idx, len(files), py, "=" * 80
        )
        try:
            proc = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "mypy",
                    "--show-error-codes",
                    "--no-color-output",
                    *MYPY_ADDITIONAL_ARGS,
                    py,
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            out, err = proc.stdout, proc.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m mypy not found – aborting mypy pass.")
            return len(files)  # Treat all files as failures for this pass
        logger.info(out or "")
        if err:
            logger.info("\n--- stderr ---\n%s", err)
        if "Success: no issues found" in (out or ""):
            errors = 0
        elif m := _MYPY_RE.search(out or ""):
            errors = int(m.group(1))
        else:
            errors = len([l for l in (out or "").splitlines() if ": error:" in l])
        results.append(
            {
                "script": Path(py).name,
                "folder": Path(py).parent.name,
                "path": py,
                "errors": errors,
                "success": "Yes" if errors == 0 else "No",
                "stderr": err or "",
            }
        )
    # Excel summary
    wb = Workbook()
    ws = wb.active
    if ws is None:  # Should not happen with wb.active
        CONSOLE.error("Failed to create active worksheet for mypy.")
        return len(results)  # Or some other error indication
    ws.title = "mypy Summary"
    ws.append(["Script", "Folder", "# Errors", "Success", "Path", "Stderr"])
    for r in results:
        ws.append(
            [
                r["script"],
                r["folder"],
                r["errors"],
                r["success"],
                r["path"],
                r["stderr"],
            ]
        )
    for col_cells in ws.columns:  # Changed variable name for clarity
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)

    xlsx = Path(OUTPUT_FOLDER) / f"mypy_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("mypy summary → %s", xlsx)
    return sum(1 for r in results if r["errors"] > 0)


# ----------------------------------------------------------------------------
# 2) VULTURE CHECK
# ----------------------------------------------------------------------------
_VULTURE_LINE = re.compile(r"^(?:.+?:)?(\d+):\s*(.+?)\s*\((\d+%) confidence\)$")


def run_vulture(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "vulture_detailed_log")
    CONSOLE.info("vulture log → %s", log_fp)
    summary: List[Dict[str, Any]] = []  # Explicitly typed
    for idx, py in enumerate(files, 1):
        logger.info(
            "\n%s\nFILE %d/%d → %s\n%s", "=" * 80, idx, len(files), py, "=" * 80
        )
        try:
            proc = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "vulture",
                    py,
                    "--min-confidence",
                    str(VULTURE_MIN_CONFIDENCE),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            out, err = proc.stdout, proc.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m vulture not found – aborting vulture pass.")
            return len(files)
        logger.info(out or "")
        if err:
            logger.info("\n--- stderr ---\n%s", err)
        items = []
        for line in (out or "").splitlines():
            m = _VULTURE_LINE.match(line.strip())
            if m:
                items.append(f"{m.group(2)} (line {m.group(1)}, {m.group(3)})")
        summary.append(
            {
                "script": Path(py).name,
                "folder": Path(py).parent.name,
                "path": py,
                "count": len(items),
                "details": "; ".join(items) if items else "None",
                "stderr": (err or "").strip(),
            }
        )
    wb = Workbook()
    ws = wb.active
    if ws is None:
        CONSOLE.error("Failed to create active worksheet for vulture.")
        return len(summary)
    ws.title = "Vulture Dead Code"
    ws.append(["Script", "Folder", "Items", "Details", "Stderr", "Path"])
    for r in summary:
        ws.append(
            [r["script"], r["folder"], r["count"], r["details"], r["stderr"], r["path"]]
        )
    for i, col_cells in enumerate(ws.columns, 1):  # Changed variable name
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 70)

    xlsx = Path(OUTPUT_FOLDER) / f"vulture_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("vulture summary → %s", xlsx)
    return sum(1 for r in summary if r["count"] > 0)


# ----------------------------------------------------------------------------
# 3) PYLINT
# ----------------------------------------------------------------------------

_PYLINT_ISSUE = re.compile(r":\s*[EF]\d{4}:")  # only E(rror)/F(atal) codes

def run_pylint(files: list[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "pylint_detailed_log")
    CONSOLE.info("pylint (errors-only) log → %s", log_fp)

    summary: list[dict[str, Any]] = []
    for idx, py in enumerate(files, 1):
        logger.info(
            "\n%s\nFILE %d/%d → %s\n%s",
            "=" * 80, idx, len(files), py, "=" * 80,
        )
        try:
            proc = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "pylint",
                    *PYLINT_ADDITIONAL_ARGS,  # <-- -E
                    py,
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
        except FileNotFoundError:
            CONSOLE.error("python -m pylint not found – aborting pylint pass.")
            return len(files)

        out, err = proc.stdout, proc.stderr
        logger.info(out or "")
        if err:
            logger.info("\n--- pylint stderr ---\n%s", err)

        issues = sum(1 for l in (out or "").splitlines() if _PYLINT_ISSUE.search(l))
        score = (
            float(out.split("rated at")[1].split("/")[0])
            if "rated at" in out
            else 0.0
        )

        summary.append(
            {
                "script": Path(py).name,
                "folder": Path(py).parent.name,
                "score": score,
                "issues": issues,
                "stderr": err or "",
                "path": py,
            }
        )

    # --- Excel summary (unchanged structure but no isort column) -------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Pylint (errors only)"
    ws.append(["Script", "Folder", "Score", "# Errors", "Stderr", "Path"])
    for r in summary:
        ws.append(
            [
                r["script"],
                r["folder"],
                r["score"],
                r["issues"],
                r["stderr"],
                r["path"],
            ]
        )
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    xlsx = Path(OUTPUT_FOLDER) / f"pylint_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(xlsx)
    CONSOLE.info("pylint summary → %s", xlsx)

    return sum(1 for r in summary if r["issues"] > 0)


# ----------------------------------------------------------------------------
# 4) PYDOCSTYLE CHECK
# ----------------------------------------------------------------------------
def run_pydocstyle(files: List[str]) -> int:
    logger, log_fp = setup_detailed_logger(OUTPUT_FOLDER, "pydocstyle_detailed_log")
    CONSOLE.info("pydocstyle log → %s", log_fp)
    summary: List[Dict[str, Any]] = []
    files_with_issues = 0

    for idx, py_file_path in enumerate(files, 1):
        logger.info(
            "\n%s\nFILE %d/%d → %s\n%s",
            "=" * 80,
            idx,
            len(files),
            py_file_path,
            "=" * 80,
        )
        try:
            proc = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "pydocstyle",
                    *PYDOCSTYLE_ADDITIONAL_ARGS,
                    py_file_path,
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            # pydocstyle prints errors to stdout and exits with 0 if files are processed,
            # non-zero for operational errors (e.g., file not found, bad option).
            # Actual docstring violations don't cause a non-zero exit code by default.
            stdout, stderr = proc.stdout, proc.stderr
        except FileNotFoundError:
            CONSOLE.error("python -m pydocstyle not found – aborting pydocstyle pass.")
            return len(files)  # Treat all files as failures for this pass

        logger.info(stdout or "")
        if stderr:
            logger.info("\n--- pydocstyle stderr ---\n%s", stderr)

        issue_lines = [line for line in (stdout or "").splitlines() if line.strip()]
        num_issues = len(issue_lines)
        issue_details = "\n".join(issue_lines)  # Store all issues

        if num_issues > 0:
            files_with_issues += 1

        summary.append(
            {
                "script": Path(py_file_path).name,
                "folder": Path(py_file_path).parent.name,
                "path": py_file_path,
                "count": num_issues,
                "details": issue_details if num_issues > 0 else "None",
                "stderr": (stderr or "").strip(),
            }
        )

    # Excel summary
    wb = Workbook()
    ws = wb.active
    if ws is None:
        CONSOLE.error("Failed to create active worksheet for pydocstyle.")
        return files_with_issues  # Or len(summary)
    ws.title = "Pydocstyle Summary"
    ws.append(["Script", "Folder", "# Issues", "Issue Details", "Stderr", "Path"])
    for r in summary:
        ws.append(
            [r["script"], r["folder"], r["count"], r["details"], r["stderr"], r["path"]]
        )

    # Adjust column widths
    ws.column_dimensions[get_column_letter(1)].width = 30  # Script
    ws.column_dimensions[get_column_letter(2)].width = 30  # Folder
    ws.column_dimensions[get_column_letter(3)].width = 10  # # Issues
    ws.column_dimensions[
        get_column_letter(4)
    ].width = 70  # Issue Details (allow more space)
    ws.column_dimensions[get_column_letter(5)].width = 40  # Stderr
    ws.column_dimensions[get_column_letter(6)].width = 60  # Path

    # Auto-wrap text in "Issue Details" column for better readability
    for row in ws.iter_rows(min_row=2, max_col=4, min_col=4):  # Start from second row
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)

    xlsx_path = (
        Path(OUTPUT_FOLDER) / f"pydocstyle_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
    wb.save(xlsx_path)
    CONSOLE.info("pydocstyle summary → %s", xlsx_path)

    return files_with_issues


# ============================================================================
# MAIN
# ============================================================================
def main() -> None:
    files = gather_python_files(FILES_OR_FOLDERS, SKIP_PATHS)
    if not files:
        CONSOLE.warning("No Python files found – nothing to do.")
        return

    Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)
    overall_tools_failed_files = 0  # Renamed for clarity

    if ENABLE_MYPY:
        CONSOLE.info("\n=== Running mypy pass ===")
        failures = run_mypy(files)
        if failures > 0:
            CONSOLE.warning("Mypy found issues in %d file(s).", failures)
            overall_tools_failed_files += (
                failures  # Or simply increment a "failed tools" counter
            )
        else:
            CONSOLE.info("Mypy pass: No issues found.")

    if ENABLE_VULTURE:
        CONSOLE.info("\n=== Running vulture pass ===")
        failures = run_vulture(files)
        if failures > 0:
            CONSOLE.warning(
                "Vulture found potential dead code in %d file(s).", failures
            )
            overall_tools_failed_files += failures
        else:
            CONSOLE.info("Vulture pass: No issues found.")

    if ENABLE_PYINT:
        CONSOLE.info("\n=== Running pylint pass ===")
        failures = run_pylint(files)
        if failures > 0:
            CONSOLE.warning("Pylint/isort found issues in %d file(s).", failures)
            overall_tools_failed_files += failures
        else:
            CONSOLE.info("Pylint/isort pass: No issues found.")

    if ENABLE_PYDOCSTYLE:
        CONSOLE.info("\n=== Running pydocstyle pass ===")
        failures = run_pydocstyle(files)
        if failures > 0:
            CONSOLE.warning("Pydocstyle found issues in %d file(s).", failures)
            overall_tools_failed_files += failures
        else:
            CONSOLE.info("Pydocstyle pass: No issues found.")

    CONSOLE.info("\n" + "=" * 80)
    if (
        overall_tools_failed_files == 0
    ):  # Check if any tool reported any file with issues
        CONSOLE.info(
            "🎉 All enabled static analysis checks passed successfully for all files."
        )
    else:
        # This count might be misleading if a single file fails multiple checks.
        # The current `overall_tools_failed_files` sums up counts of files failing EACH tool.
        # A more accurate summary might be the number of unique files that had *any* issue.
        # For now, keeping it as is, representing the sum of "files with issues" from each tool.
        CONSOLE.warning(
            "⚠️ Static analysis detected issues. Total count of 'files with issues' across all tools: %d. "
            "Review individual tool logs and summaries for details.",
            overall_tools_failed_files,
        )
    CONSOLE.info(
        "Detailed logs and Excel summaries are in: %s", Path(OUTPUT_FOLDER).resolve()
    )


if __name__ == "__main__":
    main()
