"""
A script to process ridership data, with optional filtering based on a column,
split into route/direction sheets, highlight and calculate key stats,
and optionally produce bar charts.

Additional features:
- Computes percent_of_route_ridership (lowercase) as the percentage of route ridership,
  rounding to one decimal on the percentage scale.
- Allows a combined configuration for columns: if a colon is used,
  the left-hand side is the column to retain and the right-hand side is
  the custom header to display.
"""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------
#                         CONFIGURATION SECTION
# ---------------------------------------------------------------------
INPUT_FILE = r'\\Path\To\Your\STATISTICS_BY_ROUTE_AND_TRIP.XLSX'
OUTPUT_FOLDER = r'\\Path\To\Your\Output\Folder'

# User can indicate something like 'Weekday', 'Saturday', 'Sunday', etc.
DATE_TYPE = 'Weekday'

# ------------------ COLUMN CONFIGURATION ---------------------------
# Define a list of columns to retain.
# Each entry can be either:
#   "COLUMN_NAME"  (no renaming) or
#   "COLUMN_NAME: Custom Header"
# The script will split these into the columns to retain (in order)
# and a mapping of column names to display names.
COLUMNS_CONFIG = [
    "SERIAL_NUMBER: Trip ID",
    "TRIP_START_TIME: Start Time",
    "ROUTE_NAME: Route",
    "DIRECTION_NAME: Direction",
    "PASSENGERS_ON: Passengers"
]

# Process COLUMNS_CONFIG to produce COLUMNS_TO_RETAIN and COLUMN_RENAME_MAP.
COLUMNS_TO_RETAIN = []
COLUMN_RENAME_MAP = {}
for col_entry in COLUMNS_CONFIG:
    if ":" in col_entry:
        original, new_name = col_entry.split(":", 1)
        original = original.strip()
        new_name = new_name.strip()
        COLUMNS_TO_RETAIN.append(original)
        COLUMN_RENAME_MAP[original] = new_name
    else:
        col_entry = col_entry.strip()
        COLUMNS_TO_RETAIN.append(col_entry)
# ---------------------------------------------------------------------
#                         OTHER OPTIONS
# ---------------------------------------------------------------------

# Whether to create bar charts (True/False)
CREATE_CHARTS = True

# Whether to flag ultra-low trips (True/False)
FLAG_ULTRA_LOW = False
ULTRA_LOW_THRESHOLD = 1  # e.g., 1 average rider

# ---------------------- FILTERING OPTIONS ----------------------------
# Column name to filter on (e.g. 'ROUTE_NAME' or 'DIRECTION_NAME').
# If you don't want to filter, leave this as None or an empty string.
FILTER_COLUMN_NAME = 'ROUTE_NAME'

# If not empty, only rows whose FILTER_COLUMN_NAME is in this list will remain
FILTER_IN_LIST = ["101", "202"
]

# If not empty, rows whose FILTER_COLUMN_NAME is in this list will be excluded
FILTER_OUT_LIST = [
]

# ---------------------------------------------------------------------
#                          HELPER FUNCTIONS
# ---------------------------------------------------------------------

def load_data(input_file: str, columns_to_retain: list[str]) -> pd.DataFrame:
    """
    Load the dataset using pandas and keep only the requested columns.

    Parameters
    ----------
    input_file : str
        Path to the input Excel file.
    columns_to_retain : list of str
        Columns to keep in the final DataFrame, in the desired order.

    Returns
    -------
    pd.DataFrame
        The DataFrame with only the requested columns.
    """
    df = pd.read_excel(input_file)
    # Ensure we only keep columns that exist in the dataset
    existing_cols = [col for col in columns_to_retain if col in df.columns]
    df = df[existing_cols]
    return df


def create_output_folder(folder_path: str) -> None:
    """
    Create the output folder if it doesn't already exist.

    Parameters
    ----------
    folder_path : str
        Path to the output folder.
    """
    os.makedirs(folder_path, exist_ok=True)


def write_direction_sheet(
    wb: Workbook,
    direction_df: pd.DataFrame,
    direction_name: str,
    date_type: str,
    create_charts: bool,
    flag_ultra_low: bool,
    ultra_low_threshold: float
) -> None:
    """
    Create a sheet in the given workbook for the specified direction,
    bold the row with the highest ridership, calculate ridership share,
    optionally create a bar chart, and optionally flag ultra-low ridership.

    Parameters
    ----------
    wb : openpyxl Workbook
        Workbook object to which the new sheet will be added.
    direction_df : pd.DataFrame
        Subset of data for this direction.
    direction_name : str
        Name of the direction (will be used as the sheet name).
    date_type : str
        Used in chart titles (e.g., 'Weekday', 'Saturday').
    create_charts : bool
        Whether or not to create a bar chart for this sheet.
    flag_ultra_low : bool
        Whether or not to highlight ultra-low ridership trips.
    ultra_low_threshold : float
        Threshold for ultra-low ridership (inclusive).
    """
    # Create a new worksheet
    ws = wb.create_sheet(title=direction_name)

    # Compute total ridership and the percentage share for each trip.
    total_ridership = direction_df['PASSENGERS_ON'].sum()
    # Compute as a fraction; then round the percentage to one decimal on the percent scale.
    direction_df['Percent of Route Ridership'] = (
        direction_df['PASSENGERS_ON'] / total_ridership if total_ridership != 0 else 0
    ).apply(lambda x: round(100 * x, 1) / 100)

    # Round PASSENGERS_ON to 1 decimal.
    direction_df['PASSENGERS_ON'] = direction_df['PASSENGERS_ON'].round(1)

    # Find the maximum ridership for bolding.
    max_ridership = direction_df['PASSENGERS_ON'].max()

    # Write headers with optional renaming.
    headers = list(direction_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        display_header = COLUMN_RENAME_MAP.get(header, header)
        ws.cell(row=1, column=col_idx, value=display_header).font = Font(bold=True)

    # Write data rows.
    for row_idx, (_, row_data) in enumerate(direction_df.iterrows(), start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Special formatting for TRIP_START_TIME column.
            if header == "TRIP_START_TIME":
                try:
                    if hasattr(row_data[header], "strftime"):
                        cell.value = row_data[header]
                    else:
                        cell.value = pd.to_datetime(row_data[header])
                    cell.number_format = "hh:mm"
                except Exception:
                    cell.value = row_data[header]
            else:
                cell.value = row_data[header]
            # Format percent_of_route_ridership as percentage with 1 decimal.
            if header == "Percent of Route Ridership":
                cell.number_format = "0.0%"
        # Bold the row with the max ridership.
        if row_data['PASSENGERS_ON'] == max_ridership:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
        # Optionally flag ultra-low ridership trips in red.
        if flag_ultra_low and row_data['PASSENGERS_ON'] <= ultra_low_threshold:
            for col_idx in range(1, len(headers) + 1):
                original_font = ws.cell(row=row_idx, column=col_idx).font
                ws.cell(row=row_idx, column=col_idx).font = Font(
                    name=original_font.name,
                    size=original_font.size,
                    bold=original_font.bold,
                    color="FF0000"  # red
                )

    # Optionally create a bar chart of ridership by trip time.
    if create_charts:
        chart = BarChart()
        chart.title = f"Ridership by Time - {direction_name} ({date_type})"
        chart.x_axis.title = "Trip Start Time"
        chart.y_axis.title = "Ridership"

        # Locate columns for PASSENGERS_ON and TRIP_START_TIME.
        pass_on_col = headers.index('PASSENGERS_ON') + 1
        time_col = headers.index('TRIP_START_TIME') + 1
        min_row = 2
        max_row = direction_df.shape[0] + 1

        # Add data series to the chart.
        values = Reference(ws, min_col=pass_on_col, min_row=min_row,
                           max_col=pass_on_col, max_row=max_row)
        chart.add_data(values, titles_from_data=False)

        # Set categories based on TRIP_START_TIME.
        categories = Reference(ws, min_col=time_col, min_row=min_row,
                               max_row=max_row)
        chart.set_categories(categories)

        # Position the chart (e.g., two columns to the right of the data).
        chart_anchor = f"{get_column_letter(len(headers) + 2)}2"
        ws.add_chart(chart, chart_anchor)


def create_route_workbook(
    route_name: str,
    route_df: pd.DataFrame,
    output_folder: str,
    date_type: str,
    create_charts: bool,
    flag_ultra_low: bool,
    ultra_low_threshold: float
) -> None:
    """
    Create a workbook for a given route, with sheets for each direction.

    Parameters
    ----------
    route_name : str
        The route name (will be used in the output filename).
    route_df : pd.DataFrame
        Subset of data for this route (all directions).
    output_folder : str
        The folder where the Excel file should be saved.
    date_type : str
        Used in chart titles (e.g., 'Weekday', 'Saturday').
    create_charts : bool
        Whether or not to create bar charts in each sheet.
    flag_ultra_low : bool
        Whether or not to highlight ultra-low ridership trips.
    ultra_low_threshold : float
        Threshold for ultra-low ridership (inclusive).
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Group by direction and create a sheet for each.
    for direction_name, direction_df in route_df.groupby('DIRECTION_NAME'):
        write_direction_sheet(
            wb=wb,
            direction_df=direction_df,
            direction_name=str(direction_name),
            date_type=date_type,
            create_charts=create_charts,
            flag_ultra_low=flag_ultra_low,
            ultra_low_threshold=ultra_low_threshold
        )

    output_path = os.path.join(output_folder, f"{route_name}.xlsx")
    wb.save(output_path)
    print(f"Saved workbook: {output_path}")


def main() -> None:
    """
    Main execution function.
    1. Loads data from the configured INPUT_FILE.
    2. Optionally filters data based on FILTER_COLUMN_NAME, FILTER_IN_LIST, FILTER_OUT_LIST.
    3. Creates/ensures OUTPUT_FOLDER exists.
    4. Groups data by ROUTE_NAME and creates a workbook for each group.
    """
    # 1. Load data
    df = load_data(INPUT_FILE, COLUMNS_TO_RETAIN)

    # 2. Optional filtering
    if FILTER_COLUMN_NAME:
        if FILTER_IN_LIST:
            df = df[df[FILTER_COLUMN_NAME].isin(FILTER_IN_LIST)]
        if FILTER_OUT_LIST:
            df = df[~df[FILTER_COLUMN_NAME].isin(FILTER_OUT_LIST)]

    # 3. Create output folder if needed
    create_output_folder(OUTPUT_FOLDER)

    # 4. For each route, generate a workbook
    for route_name, route_df in df.groupby('ROUTE_NAME'):
        create_route_workbook(
            route_name=str(route_name),
            route_df=route_df,
            output_folder=OUTPUT_FOLDER,
            date_type=DATE_TYPE,
            create_charts=CREATE_CHARTS,
            flag_ultra_low=FLAG_ULTRA_LOW,
            ultra_low_threshold=ULTRA_LOW_THRESHOLD
        )


if __name__ == "__main__":
    main()
