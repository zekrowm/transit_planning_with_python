"""
Ridership by Route and Stop Processor

This script processes ridership data from an input Excel file by filtering specific routes and stop IDs,
aggregating the data for defined time periods, and exporting the results to a new Excel file with multiple
formatted sheets. It supports two main "formatting" controls:

1. APPLY_ROUNDING = True/False
    - Whether we round BOARD_ALL and ALIGHT_ALL to 1 decimal place in the "Original" tab
    - If AGGREGATE_BIN_RANGES is False, also round the aggregated totals to 1 decimal place
2. AGGREGATE_BIN_RANGES = True/False
    - Whether we convert the aggregated totals (BOARD_ALL_TOTAL, ALIGHT_ALL_TOTAL)
      into text ranges ("0-4.9", "5-24.9", "25 or more") in each aggregated sheet
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# ==========================
# Configuration Section
# ==========================

INPUT_FILE_PATH = r'\\Path\To\Your\RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS).XLSX'
OUTPUT_FILE_SUFFIX = '_processed'
OUTPUT_FILE_EXTENSION = '.xlsx'

# ROUTES, STOP_IDS, and TIME_PERIODS can be left empty
# If empty, the script will skip filtering or time-period breakdown for these lists.
ROUTES = []  # e.g. [] means skip route filter
STOP_IDS = [
    1067, 1069, 2122, 2174, 3208, 3220
]  # e.g. [] means skip stop filter
TIME_PERIODS = ['AM EARLY', 'AM PEAK', 'MIDDAY', 'PM PEAK', 'PM LATE', 'PM NITE']  # e.g. [] means skip time-period breakdown

# If True, ridership columns in the "Original" data are rounded to 1 decimal place.
# Also, if AGGREGATE_BIN_RANGES = False, aggregated totals get rounded instead of binned.
APPLY_ROUNDING = True

# If True, aggregated totals (BOARD_ALL_TOTAL, ALIGHT_ALL_TOTAL) are converted to
# "0-4.9", "5-24.9", or "25 or more". If False, they remain numeric and
# get rounded to 1 decimal place only if APPLY_ROUNDING = True.
AGGREGATE_BIN_RANGES = True

REQUIRED_COLUMNS = ['TIME_PERIOD', 'ROUTE_NAME', 'STOP', 'STOP_ID', 'BOARD_ALL', 'ALIGHT_ALL']
COLUMNS_TO_RETAIN = ['ROUTE_NAME', 'STOP', 'STOP_ID', 'BOARD_ALL', 'ALIGHT_ALL']

# ==========================
# End of Configuration
# ==========================


def bin_ridership_value(value):
    """
    Categorize a ridership value into a range (e.g., "0-4.9", "5-24.9", "25 or more").
    """
    if value < 5:
        return "0-4.9"
    elif value < 25:
        return "5-24.9"
    else:
        return "25 or more"


def aggregate_by_stop(data_subset):
    """
    Summarize ridership by stop, totaling boardings and alightings, and listing unique routes.
    """
    aggregated = data_subset.groupby(['STOP', 'STOP_ID'], as_index=False).agg({
        'BOARD_ALL': 'sum',
        'ALIGHT_ALL': 'sum',
        'ROUTE_NAME': lambda x: ', '.join(sorted(x.unique()))
    })
    aggregated.rename(columns={
        'BOARD_ALL': 'BOARD_ALL_TOTAL',
        'ALIGHT_ALL': 'ALIGHT_ALL_TOTAL',
        'ROUTE_NAME': 'ROUTES'
    }, inplace=True)
    return aggregated


def read_excel_file(input_file):
    """
    Load an Excel file into a DataFrame. Exits if the file is missing or unreadable.
    """
    try:
        return pd.read_excel(input_file)
    except FileNotFoundError:
        print(f"Error: The file '{input_file}' does not exist.")
        sys.exit(1)
    except pd.errors.ExcelFileError as error:
        print(f"Error reading the Excel file: {error}")
        sys.exit(1)


def verify_required_columns(data_frame, required_columns):
    """
    Check if all required columns exist in the DataFrame. Exits if any are missing.
    """
    missing_columns = [col for col in required_columns if col not in data_frame.columns]
    if missing_columns:
        print(f"Error: Missing columns: {missing_columns}")
        sys.exit(1)


def filter_data(data_frame, routes, stop_ids):
    """
    Filter the data by route names and stop IDs. Returns the filtered DataFrame.
    """
    filtered_df = data_frame.copy()
    if routes:
        filtered_df = filtered_df[filtered_df['ROUTE_NAME'].isin(routes)]
    if stop_ids:
        filtered_df = filtered_df[filtered_df['STOP_ID'].isin(stop_ids)]
    return filtered_df


def write_to_excel(output_file, filtered_data, aggregated_peaks, all_time_aggregated):
    """
    Save processed ridership data to an Excel file with multiple sheets.
    """
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            filtered_data.to_excel(writer, sheet_name='Original', index=False)

            # Write each time period's aggregated data
            for period, df_agg in aggregated_peaks.items():
                df_agg.to_excel(writer, sheet_name=period, index=False)

            # Always write the all-time aggregated data
            all_time_aggregated.to_excel(writer, sheet_name='All Time Periods', index=False)

            writer.save()

        adjust_excel_formatting(output_file)
        print(f"Success: The processed file has been saved as '{output_file}'.")
    except Exception as error:
        print(f"Error writing the processed Excel file: {error}")
        sys.exit(1)


def adjust_excel_formatting(output_file):
    """
    Format an Excel file by bolding headers and adjusting column widths.
    """
    try:
        workbook = load_workbook(output_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Bold the header row
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            # Auto-adjust column widths
            for column_cells in sheet.columns:
                max_length = 0
                col_letter = column_cells[0].column_letter
                for cell in column_cells:
                    cell_val = str(cell.value) if cell.value is not None else ''
                    max_length = max(max_length, len(cell_val))
                sheet.column_dimensions[col_letter].width = max_length + 2
        workbook.save(output_file)
    except Exception as error:
        print(f"Error adjusting Excel formatting: {error}")
        sys.exit(1)


def main():
    """
    Process ridership data: read, filter, aggregate, apply formatting, and save to Excel.
    """
    input_file = INPUT_FILE_PATH
    base, ext = os.path.splitext(input_file)
    ext = ext.lower()
    if ext != OUTPUT_FILE_EXTENSION:
        print(f"Warning: The input file has extension '{ext}'. Using '{OUTPUT_FILE_EXTENSION}' for output.")
        ext = OUTPUT_FILE_EXTENSION
    output_file = f"{base}{OUTPUT_FILE_SUFFIX}{ext}"

    # Read and verify the Excel data
    ridership_df = read_excel_file(input_file)
    verify_required_columns(ridership_df, REQUIRED_COLUMNS)

    # Apply optional filters
    filtered_data = filter_data(ridership_df, ROUTES, STOP_IDS)

    # Standardize 'TIME_PERIOD' values
    filtered_data['TIME_PERIOD'] = filtered_data['TIME_PERIOD'].astype(str).str.strip().str.upper()

    # Only build time-period sheets if TIME_PERIODS is non-empty
    peak_data_dict = {}
    if TIME_PERIODS:
        for period in TIME_PERIODS:
            period_upper = period.upper()
            subset = filtered_data[filtered_data['TIME_PERIOD'] == period_upper]
            peak_data_dict[period] = subset[COLUMNS_TO_RETAIN]
    else:
        peak_data_dict = {}

    # Aggregate data for all rows (regardless of time period)
    all_time_aggregated = aggregate_by_stop(filtered_data)

    # Aggregate data for each time period (if any)
    aggregated_peaks = {}
    if TIME_PERIODS:
        for period, data_subset in peak_data_dict.items():
            aggregated_peaks[period] = aggregate_by_stop(data_subset)

    # 1) Round the original ridership columns if requested
    if APPLY_ROUNDING:
        for col in ['BOARD_ALL', 'ALIGHT_ALL']:
            if col in filtered_data.columns:
                filtered_data[col] = filtered_data[col].round(1)

    # 2) Format aggregated columns
    for df_agg in [all_time_aggregated] + list(aggregated_peaks.values()):

        # If we want to bin the aggregated columns into categories:
        if AGGREGATE_BIN_RANGES:
            # Convert numeric aggregated totals into bins
            for col in ['BOARD_ALL_TOTAL', 'ALIGHT_ALL_TOTAL']:
                if col in df_agg.columns:
                    df_agg[col] = df_agg[col].apply(bin_ridership_value)

        else:
            # If not binning, but rounding is desired, do decimal rounding
            if APPLY_ROUNDING:
                for col in ['BOARD_ALL_TOTAL', 'ALIGHT_ALL_TOTAL']:
                    if col in df_agg.columns:
                        df_agg[col] = df_agg[col].round(1)
            else:
                # If neither binning nor rounding, do nothing to aggregated columns
                pass

    # Write the data to Excel
    write_to_excel(output_file, filtered_data, aggregated_peaks, all_time_aggregated)


if __name__ == "__main__":
    main()
