"""
Module for generating GTFS bus arrivals checklists in printable Excel format.
"""

import os

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Output directory
BASE_OUTPUT_PATH = r'\\your_file_path\here\\'

# Define columns to read as strings
DTYPE_DICT = {
    'stop_id': str,
    'trip_id': str,
    'route_id': str,
    'service_id': str,
    # Add other ID fields as needed
}

# Input file paths to load GTFS files with specified dtypes
BASE_INPUT_PATH = r'\\your_file_path\here\\'

# List of required GTFS files
GTFS_FILES = ['trips.txt', 'stop_times.txt', 'routes.txt', 'stops.txt', 'calendar.txt']

# Check for existence of input directory
if not os.path.exists(BASE_INPUT_PATH):
    raise FileNotFoundError(f"The input directory {BASE_INPUT_PATH} does not exist.")

# Load GTFS files with specified dtypes
for file_name in GTFS_FILES:
    file_path = os.path.join(BASE_INPUT_PATH, file_name)
    if not os.path.exists(file_path):
        raise FileNotFoundError(
            f"The required GTFS file {file_name} does not exist in {BASE_INPUT_PATH}."
        )

trips = pd.read_csv(os.path.join(BASE_INPUT_PATH, 'trips.txt'), dtype=DTYPE_DICT)
stop_times = pd.read_csv(os.path.join(BASE_INPUT_PATH, 'stop_times.txt'), dtype=DTYPE_DICT)
routes = pd.read_csv(os.path.join(BASE_INPUT_PATH, 'routes.txt'), dtype=DTYPE_DICT)
stops = pd.read_csv(os.path.join(BASE_INPUT_PATH, 'stops.txt'), dtype=DTYPE_DICT)
calendar = pd.read_csv(os.path.join(BASE_INPUT_PATH, 'calendar.txt'), dtype=DTYPE_DICT)

# Define clusters with stop IDs (e.g., bus centers with multiple nearby stops)
# Format: {'Cluster Name': ['stop_id1', 'stop_id2', ...]}
CLUSTERS = {
    'Your Cluster 1': ['1', '2', '3'],   # Replace with your cluster name and stop IDs
    'Your Cluster 2': ['4', '5', '6'],
    'Your Cluster 3': ['7', '8', '9', '10'],
    # Add more clusters as needed
}

# Define schedule types and corresponding days in the calendar
# Format: {'Schedule Type': ['day1', 'day2', ...]}
SCHEDULE_TYPES = {
    'Weekday': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday': ['sunday'],
    # 'Friday': ['friday'],  # Uncomment if you have a unique Friday schedule
}

# Time windows for filtering trips
# Format: {'Schedule Type': {'Time Window Name': ('Start Time', 'End Time')}}
# Times should be in 'HH:MM' 24-hour format
TIME_WINDOWS = {
    'Weekday': {
        'morning': ('06:00', '09:59'),
        'afternoon': ('14:00', '17:59'),
        # 'evening': ('18:00', '21:59'),  # Add as needed
    },
    'Saturday': {
        'midday': ('10:00', '13:59'),
        # Add more time windows for Saturday if needed
    },
    # 'Sunday': {  # Uncomment and customize for Sunday if needed
    #     'morning': ('08:00', '11:59'),
    #     'afternoon': ('12:00', '15:59'),
    # },
}

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

# Create the output directory if it doesn't exist
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

def fix_time_format(time_str):
    """
    Convert the given time to HH:MM format, ignoring seconds if present.

    Parameters:
        time_str (str): Time string in 'HH:MM:SS' or 'HH:MM' format.

    Returns:
        str: Time string in 'HH:MM' format.
    """
    parts = time_str.split(":")
    hours = int(parts[0])
    minutes = int(parts[1])
    if hours >= 24:
        hours -= 24
    return f"{hours:02}:{minutes:02}"

# Ensure 'stop_id' is string in stops DataFrame
stops['stop_id'] = stops['stop_id'].astype(str)

# Process each schedule type
for schedule_name, days in SCHEDULE_TYPES.items():
    print(f"Processing schedule: {schedule_name}")
    # Filter services for the current schedule type
    service_mask = calendar[days].astype(bool).all(axis=1)
    relevant_service_ids = calendar.loc[service_mask, 'service_id']

    # Filter trips to include only those that match the relevant service IDs
    trips_filtered = trips[trips['service_id'].isin(relevant_service_ids)]

    if trips_filtered.empty:
        print(f"No trips found for {schedule_name} schedule. Skipping.")
        continue

    # Merge trips with stop_times and routes to include route_short_name and block_id
    merged_data = pd.merge(stop_times, trips_filtered, on='trip_id')
    merged_data = pd.merge(merged_data, routes[['route_id', 'route_short_name']], on='route_id')

    # Ensure 'stop_id' is string in merged_data
    merged_data['stop_id'] = merged_data['stop_id'].astype(str)

    # Create a new column sequence_long
    merged_data['sequence_long'] = 'middle'

    # Assign "start" to sequence_long for rows with stop_sequence 1
    merged_data.loc[merged_data['stop_sequence'] == 1, 'sequence_long'] = 'start'

    # Get the highest stop_sequence number value for each trip
    max_sequence = merged_data.groupby('trip_id')['stop_sequence'].transform('max')

    # Assign "last" to sequence_long for rows with the highest stop_sequence number
    merged_data.loc[merged_data['stop_sequence'] == max_sequence, 'sequence_long'] = 'last'

    # Process each cluster
    for cluster_name, cluster_stop_ids in CLUSTERS.items():
        print(f"Processing cluster: {cluster_name} for {schedule_name} schedule")
        # Ensure cluster_stop_ids are strings
        cluster_stop_ids = [str(sid) for sid in cluster_stop_ids]

        # Filter merged_data by stop_id for the current cluster
        cluster_data = merged_data[merged_data['stop_id'].isin(cluster_stop_ids)]

        if cluster_data.empty:
            print(f"No data found for {cluster_name} on {schedule_name} schedule. Skipping.")
            continue

        # Apply the function to the time columns
        cluster_data['arrival_time'] = cluster_data['arrival_time'].apply(fix_time_format)
        cluster_data['departure_time'] = cluster_data['departure_time'].apply(fix_time_format)

        # Ensure times are strings in HH:MM
        cluster_data['arrival_time'] = cluster_data['arrival_time'].astype(str)
        cluster_data['departure_time'] = cluster_data['departure_time'].astype(str)

        # Sort by arrival_time using a temporary datetime conversion
        cluster_data['arrival_sort'] = pd.to_datetime(cluster_data['arrival_time'], format='%H:%M')
        cluster_data = cluster_data.sort_values(by='arrival_sort').drop(columns='arrival_sort')

        # Add 'act_arrival' and 'act_departure' columns with placeholders
        cluster_data.insert(
            cluster_data.columns.get_loc('arrival_time') + 1,
            'act_arrival',
            '________'
        )
        cluster_data.insert(
            cluster_data.columns.get_loc('departure_time') + 1,
            'act_departure',
            '________'
        )
        cluster_data.insert(
            cluster_data.columns.get_loc('block_id') + 1,
            'act_block',
            '________'
        )

        # Modify 'act_arrival' where 'sequence_long' is 'start'
        cluster_data.loc[cluster_data['sequence_long'] == 'start', 'act_arrival'] = '__XXXX__'

        # Modify 'act_departure' where 'sequence_long' is 'last'
        cluster_data.loc[cluster_data['sequence_long'] == 'last', 'act_departure'] = '__XXXX__'

        # Add 'bus_number' column with underscores
        cluster_data['bus_number'] = '________'

        # Add 'comments' column with underscores
        cluster_data['comments'] = '________________'

        # Add 'stop_name' column next to 'stop_id'
        cluster_data = pd.merge(
            cluster_data,
            stops[['stop_id', 'stop_name']],
            on='stop_id',
            how='left'
        )

        # Move specified columns to desired positions
        first_columns = [
            'route_short_name', 'trip_headsign', 'stop_sequence', 'sequence_long',
            'stop_id', 'stop_name', 'arrival_time', 'act_arrival',
            'departure_time', 'act_departure', 'block_id', 'act_block', 'bus_number', 'comments'
        ]
        other_columns = [col for col in cluster_data.columns if col not in first_columns]
        cluster_data = cluster_data[first_columns + other_columns]

        # Drop unnecessary columns
        cluster_data = cluster_data.drop(
            columns=[
                'shape_dist_traveled', 'shape_id', 'route_id', 'service_id',
                'trip_id', 'timepoint', 'direction_id', 'stop_headsign', 'pickup_type',
                'drop_off_type', 'wheelchair_accessible', 'bikes_allowed', 'trip_short_name'
            ],
            errors='ignore'
        )

        # Define the output file name for all trips
        output_file_name = f'{cluster_name}_{schedule_name}_data.xlsx'
        output_file = os.path.join(BASE_OUTPUT_PATH, output_file_name)

        # Export all cluster data to Excel with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:  # pylint: disable=abstract-class-instantiated
            cluster_data.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Align all headers to the left
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='left')

            # Adjust the width of all columns
            for idx, col in enumerate(cluster_data.columns, 1):  # 1-based indexing for Excel columns
                column_letter = get_column_letter(idx)
                max_length = max(
                    cluster_data[col].astype(str).map(len).max(),  # Maximum length of column entries
                    len(str(col))  # Length of the column header
                ) + 2  # Adding extra space for better readability
                worksheet.column_dimensions[column_letter].width = max_length

        print(f"Processed and exported data for {cluster_name} on {schedule_name} schedule.")

        # Now, check if there are time windows for this schedule
        if schedule_name in TIME_WINDOWS:
            for time_window_name, time_range in TIME_WINDOWS[schedule_name].items():
                start_time_str, end_time_str = time_range

                # Parse the start and end times in HH:MM format
                start_dt = pd.to_datetime(start_time_str, format='%H:%M').time()
                end_dt = pd.to_datetime(end_time_str, format='%H:%M').time()

                # Convert arrival_time strings (HH:MM) to datetime.time for filtering
                arrival_times = pd.to_datetime(cluster_data['arrival_time'], format='%H:%M').dt.time
                filtered_data = cluster_data[
                    (arrival_times >= start_dt) & (arrival_times <= end_dt)
                ]

                if filtered_data.empty:
                    print(
                        f"No data found for {cluster_name} on {schedule_name} schedule in "
                        f"{time_window_name} time window. Skipping."
                    )
                    continue

                # Define the output file name for the time window
                output_file_name = f'{cluster_name}_{schedule_name}_{time_window_name}_data.xlsx'
                output_file = os.path.join(BASE_OUTPUT_PATH, output_file_name)

                # Export filtered data to Excel with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:  # pylint: disable=abstract-class-instantiated
                    filtered_data.to_excel(writer, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Align all headers to the left
                    for cell in worksheet[1]:
                        cell.alignment = Alignment(horizontal='left')

                    # Adjust the width of all columns
                    for idx, col in enumerate(filtered_data.columns, 1):  # 1-based indexing for Excel columns
                        column_letter = get_column_letter(idx)
                        max_length = max(
                            filtered_data[col].astype(str).map(len).max(),  # Maximum length of column entries
                            len(str(col))  # Length of the column header
                        ) + 2  # Adding extra space for better readability
                        worksheet.column_dimensions[column_letter].width = max_length

                print(
                    f"Processed and exported data for {cluster_name} on {schedule_name} schedule "
                    f"in {time_window_name} time window."
                )

print("All clusters and schedules have been processed and exported.")
