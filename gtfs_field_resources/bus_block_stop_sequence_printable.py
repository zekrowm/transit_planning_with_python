"""
GTFS Block Schedule Printable

This script processes GTFS data to create printable tables for each block.
Each table includes all trips within a block, even if they belong to different routes.
This allows field checkers to see the complete schedule and behavior of a bus block.

Each table includes:
- Block ID
- Route
- Direction
- Trip ID
- Trip Start Time
- Stop Sequence
- Timepoint
- Stop ID
- Stop Name
- Scheduled Time
- Actual Time (blank)
- Boardings (blank)
- Alightings (blank)
- Comments (blank)

The blank columns are placeholders for manual data entry during ride checks.
"""

import os
import math

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ================================
# CONFIGURATION
# ================================
BASE_INPUT_PATH = r"C:\Path\To\Your\Input\Folder"    # <<< EDIT HERE
BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output\Folder"  # <<< EDIT HERE

STOP_TIMES_FILE = "stop_times.txt"
TRIPS_FILE = "trips.txt"
CALENDAR_FILE = "calendar.txt"
STOPS_FILE = "stops.txt"
ROUTES_FILE = "routes.txt"

# If you only want certain service IDs or route short names, specify them here:
FILTER_SERVICE_IDS = []          # e.g. ['1', '2'] or [] to include all
FILTER_ROUTE_SHORT_NAMES = []    # e.g. ['101', '202'] or [] to include all

# Placeholder values for printing:
MISSING_TIME = "________"
MISSING_VALUE = "_____"

# Maximum column width for neat Excel formatting:
MAX_COLUMN_WIDTH = 35

# ================================
# HELPER FUNCTIONS
# ================================

def time_to_seconds(time_str):
    """
    Converts a 'HH:MM:SS' or 'HH:MM' string into total seconds.
    Handles hours >= 24 by rolling over (e.g., 25:10:00 -> 1:10:00).
    """
    if pd.isnull(time_str):
        return math.nan

    parts = time_str.strip().split(':')
    if len(parts) < 2:
        return math.nan

    try:
        hours = int(parts[0]) % 24  # Roll over hours >= 24
        minutes = int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
    except ValueError:
        return math.nan

    return hours * 3600 + minutes * 60 + seconds

def format_hhmm(total_seconds):
    """
    Given a time in total seconds, returns a 'HH:MM' string (24-hour).
    If invalid, returns an empty string.
    """
    if pd.isnull(total_seconds) or total_seconds < 0:
        return ""
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"

def export_to_excel(data_frame, output_file):
    """
    Exports a DataFrame to an Excel file and applies basic formatting:
    - Left alignment
    - Sets column widths
    - Text wrapping for headers
    """
    if data_frame.empty:
        print(f"No data to export to {output_file}")
        return

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Pylint warns about abstract-class-instantiated for ExcelWriter.
    # Typically, pd.ExcelWriter with 'openpyxl' is fine for most uses.
    # If you prefer, you can suppress the warning or switch to a specific writer.
    # pylint: disable=abstract-class-instantiated
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        data_frame.to_excel(writer, index=False, sheet_name='Schedule')
        worksheet = writer.sheets['Schedule']

        # Adjust columns
        for col_i, col_name in enumerate(data_frame.columns, 1):
            col_letter = get_column_letter(col_i)

            # Header alignment and text wrap
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.alignment = Alignment(horizontal='left', wrap_text=True)

            # Data alignment
            for row_i in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_i}"]
                cell.alignment = Alignment(horizontal='left')

            # Set column width based on max content length, capped at MAX_COLUMN_WIDTH
            max_len = max(len(str(col_name)), 10)  # Minimum width
            for row_i in range(2, worksheet.max_row + 1):
                val = worksheet[f"{col_letter}{row_i}"].value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            worksheet.column_dimensions[col_letter].width = min(max_len + 2, MAX_COLUMN_WIDTH)

    print(f"Exported: {output_file}")


def load_data():
    """
    Loads GTFS data from CSV files. Returns multiple DataFrames.
    Raises FileNotFoundError or other exceptions if something is wrong.
    """
    trips_path = os.path.join(BASE_INPUT_PATH, TRIPS_FILE)
    stop_times_path = os.path.join(BASE_INPUT_PATH, STOP_TIMES_FILE)
    stops_path = os.path.join(BASE_INPUT_PATH, STOPS_FILE)
    routes_path = os.path.join(BASE_INPUT_PATH, ROUTES_FILE)

    # Removed calendar_df since it was unused (fixes the unused-variable warning).
    trips_df = pd.read_csv(trips_path, dtype=str)
    stop_times_df = pd.read_csv(stop_times_path, dtype=str)
    stops_df = pd.read_csv(stops_path, dtype=str)
    routes_df = pd.read_csv(routes_path, dtype=str)

    return trips_df, stop_times_df, stops_df, routes_df


def filter_data(trips_df, stop_times_df, routes_df):
    """
    Merges and filters trips/routes, then filters stop_times accordingly.
    Returns updated trips_df, stop_times_df.
    """

    # Merge route_short_name into trips
    routes_subset = routes_df[['route_id', 'route_short_name']]
    trips_df = trips_df.merge(routes_subset, on='route_id', how='left')

    # Apply Route Filtering
    if FILTER_ROUTE_SHORT_NAMES:
        blocks_for_selected_routes = (
            trips_df[trips_df['route_short_name'].isin(FILTER_ROUTE_SHORT_NAMES)]
            ['block_id'].dropna().unique()
        )
        if len(blocks_for_selected_routes) == 0:
            print("No blocks found with the specified route short names.")
            return pd.DataFrame(), pd.DataFrame()

        trips_df = trips_df[trips_df['block_id'].isin(blocks_for_selected_routes)]

    # Apply Service ID Filtering
    if FILTER_SERVICE_IDS:
        trips_df = trips_df[trips_df['service_id'].isin(FILTER_SERVICE_IDS)]

    # Filter stop_times to only include relevant trips
    stop_times_df = stop_times_df[
        stop_times_df['trip_id'].isin(trips_df['trip_id'])
    ]

    return trips_df, stop_times_df


def prepare_stop_times(trips_df, stop_times_df, stops_df):
    """
    Adds and formats columns in stop_times_df with scheduling and stop info.
    Returns the updated stop_times_df.
    """
    # If 'timepoint' does not exist, create a new column with 0.
    if 'timepoint' not in stop_times_df.columns:
        stop_times_df['timepoint'] = 0
    else:
        # Convert to numeric, fill NaN with 0
        stop_times_df['timepoint'] = (
            pd.to_numeric(stop_times_df['timepoint'], errors='coerce')
            .fillna(0).astype(int)
        )

    # Merge essential trip columns into stop_times
    needed_trip_cols = ['trip_id', 'block_id', 'route_short_name', 'direction_id']
    stop_times_df = stop_times_df.merge(
        trips_df[needed_trip_cols],
        on='trip_id',
        how='left'
    )

    # Convert arrival/departure times to seconds and format
    stop_times_df['arrival_seconds'] = stop_times_df['arrival_time'].apply(time_to_seconds)
    stop_times_df['departure_seconds'] = stop_times_df['departure_time'].apply(time_to_seconds)
    stop_times_df['scheduled_time_hhmm'] = stop_times_df['departure_seconds'].apply(format_hhmm)

    # Merge in stop names
    stop_name_map = stops_df.set_index('stop_id')['stop_name'].to_dict()
    stop_times_df['stop_name'] = stop_times_df['stop_id'].map(stop_name_map).fillna("Unknown Stop")

    # Sort by block, trip, and stop_sequence
    stop_times_df = stop_times_df.dropna(subset=['block_id'])
    stop_times_df['stop_sequence'] = pd.to_numeric(stop_times_df['stop_sequence'], errors='coerce')
    stop_times_df = stop_times_df.dropna(subset=['stop_sequence'])
    stop_times_df.sort_values(['block_id', 'trip_id', 'stop_sequence'], inplace=True)

    return stop_times_df


def export_blocks(stop_times_df):
    """
    Groups rows by block_id and exports each block to a separate Excel file.
    """
    all_blocks = stop_times_df['block_id'].unique()
    print(f"Found {len(all_blocks)} blocks to export.\n")

    for block_id in all_blocks:
        block_subset = stop_times_df[stop_times_df['block_id'] == block_id].copy()
        if block_subset.empty:
            continue

        # For each trip_id within this block, find earliest departure
        first_departures = (
            block_subset.groupby('trip_id')['departure_seconds']
            .min()
            .reset_index(name='trip_start_seconds')
        )
        first_departures['trip_start_hhmm'] = first_departures['trip_start_seconds'].apply(format_hhmm)
        block_subset = block_subset.merge(first_departures, on='trip_id', how='left')

        block_subset['Trip Start Time'] = block_subset['trip_start_hhmm']

        # Select and rename columns for clarity
        out_cols = [
            'block_id',
            'route_short_name',
            'direction_id',
            'trip_id',
            'Trip Start Time',
            'stop_sequence',
            'timepoint',
            'stop_id',
            'stop_name',
            'scheduled_time_hhmm',
        ]
        final_df = block_subset[out_cols].copy()
        final_df.rename(columns={
            'block_id': 'Block ID',
            'route_short_name': 'Route',
            'direction_id': 'Direction',
            'trip_id': 'Trip ID',
            'stop_sequence': 'Stop Sequence',
            'timepoint': 'Timepoint',
            'stop_id': 'Stop ID',
            'stop_name': 'Stop Name',
            'scheduled_time_hhmm': 'Scheduled Time',
        }, inplace=True)

        # Insert placeholders
        final_df['Actual Time'] = MISSING_TIME
        final_df['Boardings'] = MISSING_VALUE
        final_df['Alightings'] = MISSING_VALUE
        final_df['Comments'] = MISSING_VALUE

        # Reorder columns to place 'Timepoint' after 'Stop Sequence'
        final_df = final_df[[
            'Block ID',
            'Route',
            'Direction',
            'Trip ID',
            'Trip Start Time',
            'Stop Sequence',
            'Timepoint',
            'Stop ID',
            'Stop Name',
            'Scheduled Time',
            'Actual Time',
            'Boardings',
            'Alightings',
            'Comments'
        ]]

        final_df.sort_values(
            by=['Trip Start Time', 'Trip ID', 'Stop Sequence'],
            inplace=True
        )

        filename = f"block_{block_id}_schedule_printable.xlsx"
        output_path = os.path.join(BASE_OUTPUT_PATH, filename)
        export_to_excel(final_df, output_path)


def main():
    """
    Main function to load, filter, process, and export printable bus block schedules.
    """
    try:
        trips_df, stop_times_df, stops_df, routes_df = load_data()
    except FileNotFoundError as file_not_found_err:
        print(f"File not found: {file_not_found_err}")
        return
    except Exception as err:  # Consider catching more specific exceptions if possible
        print(f"Error reading GTFS files: {err}")
        return

    trips_df, stop_times_df = filter_data(trips_df, stop_times_df, routes_df)
    if trips_df.empty or stop_times_df.empty:
        return

    stop_times_df = prepare_stop_times(trips_df, stop_times_df, stops_df)
    export_blocks(stop_times_df)

    print("\nAll blocks have been processed and exported.")


if __name__ == "__main__":
    # Disabling these if you prefer not to refactor into smaller functions:
    # pylint: disable=too-many-locals, too-many-statements
    main()
