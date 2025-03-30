"""
GTFS Schedule Exporter Module.

Processes GTFS (General Transit Feed Specification) files to generate
clearly formatted Excel schedules per route and direction, grouped
by service ID. Allows filtering of routes and service IDs through
configurable lists, improving flexibility for targeted analysis.

Key Features:
- Filters routes using `FILTER_IN_ROUTES` and `FILTER_OUT_ROUTES`.
- Filters services using `FILTER_SERVICE_IDS`. When non-empty, only these
  service IDs are processed.
- Dynamically generates descriptive output folders based on `service_id`
  and active days of week (e.g., "calendar_2_sat").
- Supports configurable time formatting (`12-hour` or `24-hour`) for
  schedule readability.
- Provides data validation checks on schedule times to detect ordering issues.
- Handles GTFS files robustly with error checking and clear messaging.

Configuration adjustments are made within the CONFIGURATION SECTION at the top
of the module.
"""
import os
import re
import sys
from collections import defaultdict

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION
# ==============================

BASE_INPUT_PATH = r"C:\Path\To\Your\System\GTFS_Data"
BASE_OUTPUT_PATH = r"C:\Path\To\Your\Output_Folder"
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

FILTER_SERVICE_IDS = []  # e.g. ['1','2'] => only process these. Empty => process all
FILTER_IN_ROUTES = []    # If non-empty, only process these route short names
FILTER_OUT_ROUTES = []   # Exclude these route short names if non-empty

TIME_FORMAT_OPTION = '24'  # '12' or '24'
MISSING_TIME = "---"
MAX_COLUMN_WIDTH = 30

trips_file = os.path.join(BASE_INPUT_PATH, "trips.txt")
stop_times_file = os.path.join(BASE_INPUT_PATH, "stop_times.txt")
routes_file = os.path.join(BASE_INPUT_PATH, "routes.txt")
stops_file = os.path.join(BASE_INPUT_PATH, "stops.txt")
calendar_file = os.path.join(BASE_INPUT_PATH, "calendar.txt")

# Globals recognized by linters
TRIPS = None
STOP_TIMES = None
ROUTES = None
STOPS = None
CALENDAR = None
TIMEPOINTS = None


# ==============================
# UTILITY FUNCTIONS
# ==============================

def time_to_minutes(time_str):
    """
    Converts a time string to total minutes since midnight.
    Supports 'HH:MM' and 'HH:MM AM/PM' formats, including hours >= 24
    in 24-hour mode. Returns None if the format is invalid or time_str
    == MISSING_TIME.
    """
    if not isinstance(time_str, str):
        return None
    if time_str.strip() == MISSING_TIME:
        return None

    result = None
    try:
        match = re.match(
            r'^(\d{1,2}):(\d{2})(?:\s*(AM|PM))?$',
            time_str.strip(),
            re.IGNORECASE
        )
        if match:
            hour_str, minute_str, period = match.groups()
            hour = int(hour_str)
            minute = int(minute_str)
            if period:
                period = period.upper()
                if period == 'PM' and hour != 12:
                    hour += 12
                elif period == 'AM' and hour == 12:
                    hour = 0
            result = hour * 60 + minute
    except (ValueError, TypeError, re.error):
        result = None

    return result


def adjust_time(time_str, time_format='24'):
    """
    Adjusts time strings to the desired format:
      - '12' => 12-hour with AM/PM
      - '24' => 24-hour

    Returns MISSING_TIME if input is MISSING_TIME, or None if invalid.
    """
    if not isinstance(time_str, str):
        return None

    if time_str.strip() == MISSING_TIME:
        return MISSING_TIME

    result = None
    parts = time_str.strip().split(":")
    if len(parts) >= 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            if time_format == '12':
                # If hours >= 24, we can't fully convert, so keep as-is
                if hours >= 24:
                    result = time_str
                else:
                    period = 'AM' if hours < 12 else 'PM'
                    adjusted_hour = hours % 12
                    if adjusted_hour == 0:
                        adjusted_hour = 12
                    result = f"{adjusted_hour}:{minutes:02d} {period}"
            else:
                # 24-hour format
                result = f"{hours:02d}:{minutes:02d}"
        except ValueError:
            result = None
    return result


def load_gtfs_files():
    """
    Loads all GTFS data into global dataframes.
    """
    global TRIPS, STOP_TIMES, ROUTES, STOPS, CALENDAR
    try:
        TRIPS = pd.read_csv(trips_file, dtype=str)
        STOP_TIMES = pd.read_csv(stop_times_file, dtype=str)
        ROUTES = pd.read_csv(routes_file, dtype=str)
        STOPS = pd.read_csv(stops_file, dtype=str)
        CALENDAR = pd.read_csv(calendar_file, dtype=str)
        print("Successfully loaded all GTFS files.")
    except FileNotFoundError as error:
        print(f"Error: {error}")
        sys.exit(1)
    except (pd.errors.EmptyDataError, pd.errors.ParserError) as error:
        print(f"Error reading GTFS files: {error}")
        sys.exit(1)
    except Exception as error:
        # Here we'd ideally catch more specific errors if known, rather than broad Exception
        print(f"An unexpected error occurred while reading GTFS files: {error}")
        sys.exit(1)


def prepare_timepoints():
    """
    Subset STOP_TIMES to timepoint=1 if available, else use all.
    Attempt to limit to major stops only, unless 'timepoint' is missing.
    """
    STOP_TIMES['stop_sequence'] = pd.to_numeric(
        STOP_TIMES['stop_sequence'],
        errors='coerce'
    )
    if STOP_TIMES['stop_sequence'].isnull().any():
        print("Warning: Some 'stop_sequence' values could not be converted to numeric.")

    if 'timepoint' in STOP_TIMES.columns:
        global TIMEPOINTS
        TIMEPOINTS = STOP_TIMES[STOP_TIMES['timepoint'] == '1'].copy()
        print("Filtered STOP_TIMES to rows with timepoint=1.")
    else:
        print("Warning: 'timepoint' column not found. Using all stops as timepoints.")
        TIMEPOINTS = STOP_TIMES.copy()


def remove_empty_schedule_columns(input_df):
    """
    Drops any columns in input_df that are entirely '---' (MISSING_TIME).
    """
    schedule_cols = [col for col in input_df.columns if col.endswith("Schedule")]
    all_blank_cols = [
        col for col in schedule_cols
        if (input_df[col] == MISSING_TIME).all()
    ]
    input_df.drop(columns=all_blank_cols, inplace=True)
    return input_df


def check_schedule_order(input_df, ordered_stop_names,
                         route_short_name, schedule_type, dir_id):
    """
    Checks times in the DataFrame to ensure they increase across rows (within a trip)
    and down columns (across trips). Prints warnings if a violation is found.
    """
    # Row-wise check
    for _, row in input_df.iterrows():
        last_time = None
        for stop in ordered_stop_names:
            col_name = f"{stop} Schedule"
            if col_name not in row:
                continue
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                print(
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Trip '{row['Trip Headsign']}': '{stop}' time {time_str} "
                    "is earlier than the previous stop's time."
                )
                break
            last_time = current_time

    # Column-wise check
    for stop in ordered_stop_names:
        col_name = f"{stop} Schedule"
        if col_name not in input_df.columns:
            continue
        last_time = None
        for _, row in input_df.iterrows():
            time_str = row[col_name]
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                print(
                    f"⚠️ Time order violation in Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{dir_id}', "
                    f"Stop '{stop}': time {time_str} is earlier than "
                    "the previous trip."
                )
                break
            last_time = current_time

    print("✅ Schedule order check passed.")


def safe_check_schedule_order(input_df, ordered_stop_names,
                              route_short_name, schedule_type, dir_id):
    """
    Wraps the check_schedule_order function in a try/except so that
    if there's a data error, we skip only that schedule order check.
    """
    try:
        check_schedule_order(
            input_df,
            ordered_stop_names,
            route_short_name,
            schedule_type,
            dir_id
        )
    except Exception as error:
        # In real usage, consider more specific exceptions
        print(
            f"❌ Skipping schedule order check for route '{route_short_name}', "
            f"schedule '{schedule_type}', direction '{dir_id}' due to error:\n   {error}"
        )


def map_service_id_to_schedule(service_row_local):
    """
    Maps a service_id row to a 'type' label based on the days it serves.
    """
    days = [
        'monday', 'tuesday', 'wednesday',
        'thursday', 'friday', 'saturday', 'sunday'
    ]
    served_days = [day for day in days if service_row_local.get(day, '0') == '1']

    weekday = {'monday', 'tuesday', 'wednesday', 'thursday', 'friday'}
    weekday_except_friday = {'monday', 'tuesday', 'wednesday', 'thursday'}
    saturday = {'saturday'}
    sunday = {'sunday'}
    weekend = {'saturday', 'sunday'}
    daily = set(days)

    schedule_label = 'Holiday'
    if served_days:
        served_set = set(served_days)
        if served_set == weekday:
            schedule_label = 'Weekday'
        elif served_set == weekday_except_friday:
            schedule_label = 'Weekday_except_Friday'
        elif served_set == saturday:
            schedule_label = 'Saturday'
        elif served_set == sunday:
            schedule_label = 'Sunday'
        elif served_set == weekend:
            schedule_label = 'Weekend'
        elif served_set == {'friday', 'saturday'}:
            schedule_label = 'Friday-Saturday'
        elif served_set == daily:
            schedule_label = 'Daily'
        else:
            schedule_label = 'Special'

    return schedule_label


def build_service_id_schedule_map():
    """
    Creates a dict: service_id -> schedule_type from CALENDAR.
    """
    service_id_schedule_map = {}
    for _, service_row_local in CALENDAR.iterrows():
        sid_val = service_row_local['service_id']
        stype_var = map_service_id_to_schedule(service_row_local)
        service_id_schedule_map[sid_val] = stype_var

    return service_id_schedule_map


def get_all_route_short_names():
    """
    Returns a sorted list of all route_short_names found in ROUTES.
    """
    return sorted(ROUTES['route_short_name'].dropna().unique().tolist())


def apply_in_out_filters(route_list):
    """
    Takes the list of all route short names.
    If FILTER_IN_ROUTES is non-empty, keep only those in that list.
    If FILTER_OUT_ROUTES is non-empty, remove those in that list.
    If both are empty, we keep everything.
    """
    route_set = set(route_list)

    if FILTER_IN_ROUTES:
        route_set = route_set.intersection(set(FILTER_IN_ROUTES))

    if FILTER_OUT_ROUTES:
        route_set = route_set.difference(set(FILTER_OUT_ROUTES))

    return sorted(route_set)


def get_master_trip_stops(dir_id, relevant_trips_dir):
    """
    Among all trips in 'relevant_trips_dir' for direction=dir_id, pick the trip with
    the largest number of timepoints. Return a DataFrame of that trip's stops.
    """
    relevant_dir = relevant_trips_dir[relevant_trips_dir['direction_id'] == dir_id]
    if relevant_dir.empty:
        print(f"Warning: No trips found for direction_id '{dir_id}'.")
        return pd.DataFrame()

    dir_trip_ids = relevant_dir['trip_id'].unique()
    subset_tp = TIMEPOINTS[TIMEPOINTS['trip_id'].isin(dir_trip_ids)]
    if subset_tp.empty:
        print(f"Warning: No stop times found for direction_id '{dir_id}'.")
        return pd.DataFrame()

    # Find the "master" trip with the most timepoints
    trip_sizes = subset_tp.groupby('trip_id').size()
    master_trip_id = trip_sizes.idxmax()

    # Extract that trip’s stops in ascending sequence
    master_data = subset_tp[subset_tp['trip_id'] == master_trip_id].copy()
    master_data.sort_values('stop_sequence', inplace=True)

    # Merge to get stop_name
    master_data = master_data.merge(
        STOPS[['stop_id', 'stop_name']],
        how='left',
        on='stop_id'
    )

    # Count occurrence of repeated stops
    occurrence_counter = defaultdict(int)
    rows = []
    for _, row_2 in master_data.iterrows():
        sid = row_2['stop_id']
        sseq = row_2['stop_sequence']
        base_name = (
            row_2['stop_name']
            if pd.notnull(row_2['stop_name'])
            else f"Unknown stop {sid}"
        )

        occurrence_counter[sid] += 1
        nth = occurrence_counter[sid]

        rows.append({
            'stop_id': sid,
            'occurrence': nth,
            'stop_sequence': sseq,
            'base_stop_name': base_name
        })

    out_df = pd.DataFrame(rows)
    # Build final_stop_name with repeated stops labeled
    name_occurrences = defaultdict(int)
    final_names = []
    for _, row_2 in out_df.iterrows():
        sid = row_2['stop_id']
        name_occurrences[sid] += 1
        count_here = name_occurrences[sid]
        if count_here == 1:
            final_names.append(row_2['base_stop_name'])
        else:
            final_names.append(f"{row_2['base_stop_name']} ({count_here})")
    out_df['final_stop_name'] = final_names

    return out_df


def process_single_trip(trip_id, trip_stop_times, master_trip_stops,
                        master_dict, time_fmt):
    """
    For each trip, produce a single schedule row. Uses a "forward-only" approach.
    """
    trip_info = TRIPS[TRIPS['trip_id'] == trip_id].iloc[0]
    route_id = trip_info['route_id']
    route_name_val = ROUTES[ROUTES['route_id'] == route_id]['route_short_name'].values[0]
    trip_headsign = trip_info.get('trip_headsign', '')
    direction_id = trip_info.get('direction_id', '')

    trip_stop_times = trip_stop_times.sort_values('stop_sequence')
    schedule_times = [MISSING_TIME] * len(master_trip_stops)
    valid_24h_times = []

    occurrence_ptr = defaultdict(int)
    for _, row_2 in trip_stop_times.iterrows():
        sid = row_2['stop_id']
        real_seq = row_2['stop_sequence']
        if sid not in master_dict:
            continue

        arr_val = (row_2.get('arrival_time') or "").strip()
        dep_val = (row_2.get('departure_time') or "").strip()
        time_val = dep_val

        arr_m = time_to_minutes(arr_val)
        dep_m = time_to_minutes(dep_val)
        if arr_val and dep_val and arr_m is not None and dep_m is not None:
            if arr_m < dep_m:
                time_val = arr_val
        elif arr_val and not dep_val:
            time_val = arr_val

        time_str_display = adjust_time(time_val, time_fmt)
        time_str_24 = adjust_time(time_val, '24')

        if not time_str_display or not time_str_24:
            continue

        oc_list = master_dict[sid]
        ptr = occurrence_ptr[sid]

        # Advance ptr while master_seq < real_seq
        while ptr < len(oc_list) and oc_list[ptr][1] < real_seq:
            ptr += 1

        if ptr >= len(oc_list):
            continue

        (_, _, col_idx) = oc_list[ptr]
        schedule_times[col_idx] = time_str_display
        valid_24h_times.append(time_str_24)
        ptr += 1
        occurrence_ptr[sid] = ptr

    if valid_24h_times:
        try:
            timedeltas = [pd.to_timedelta(f"{t}:00") for t in valid_24h_times]
            max_sort_time = max(timedeltas)
        except Exception:
            max_sort_time = pd.to_timedelta('00:00')
    else:
        max_sort_time = pd.to_timedelta('9999:00:00')

    row_data = [
        route_name_val,
        direction_id,
        trip_headsign
    ] + schedule_times + [max_sort_time]
    return row_data


def process_trips_for_direction(params):
    """
    Processes trips for a specific direction_id => returns a DataFrame for that direction.
    """
    trips_dir = params["trips_dir"]
    master_trip_stops = params["master_trip_stops"]
    time_fmt = params["time_fmt"]
    route_short = params["route_short"]
    sched_type = params["sched_type"]
    dir_id = params["dir_id"]

    if trips_dir.empty or master_trip_stops.empty:
        print(f"No usable trips/stops for direction {dir_id}. Skipping.")
        return pd.DataFrame()

    master_dict = defaultdict(list)
    master_trip_stops = master_trip_stops.reset_index(drop=True)
    for i, row_2 in master_trip_stops.iterrows():
        sid = row_2['stop_id']
        occ = row_2['occurrence']
        mseq = row_2['stop_sequence']
        master_dict[sid].append((occ, mseq, i))

    for sid in master_dict:
        master_dict[sid].sort(key=lambda x: x[1])  # sort by stop_sequence

    group_mask = TIMEPOINTS['trip_id'].isin(trips_dir['trip_id'])
    timepoints_dir = TIMEPOINTS[group_mask].copy()
    if timepoints_dir.empty:
        print(f"No stop times for direction {dir_id} in TIMEPOINTS. Skipping.")
        return pd.DataFrame()

    stop_names_ordered = master_trip_stops['final_stop_name'].tolist()
    col_names = ["Route Name", "Direction ID", "Trip Headsign"]
    col_names += [f"{sn} Schedule" for sn in stop_names_ordered]
    col_names.append("sort_time")

    rows = []
    for trip_id, grp in timepoints_dir.groupby('trip_id'):
        row_data = process_single_trip(
            trip_id=trip_id,
            trip_stop_times=grp,
            master_trip_stops=master_trip_stops,
            master_dict=master_dict,
            time_fmt=time_fmt
        )
        rows.append(row_data)

    if not rows:
        return pd.DataFrame()

    out_df = pd.DataFrame(rows, columns=col_names)
    out_df.sort_values(by='sort_time', inplace=True)
    out_df.drop(columns=['sort_time'], inplace=True)

    safe_check_schedule_order(
        out_df,
        stop_names_ordered,
        route_short,
        sched_type,
        dir_id
    )
    remove_empty_schedule_columns(out_df)

    return out_df


def export_to_excel_multiple_sheets(df_dict, out_file):
    """
    Exports multiple DataFrames to an Excel file with each DataFrame on its own sheet
    using openpyxl.
    """
    if not df_dict:
        print(f"No data to export to {out_file}.")
        return

    # Despite occasionally triggering E0110 in Pylint, openpyxl is used here.
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        for sheet_name, input_df in df_dict.items():
            if input_df.empty:
                print(f"No data for sheet '{sheet_name}'. Skipping.")
                continue

            input_df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            # Adjust column widths & alignment
            for col_num, _ in enumerate(input_df.columns, 1):
                col_letter = get_column_letter(col_num)
                header_cell = worksheet[f'{col_letter}1']
                header_cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row_num}']
                    cell.alignment = Alignment(horizontal='left')

                # Calculate column width
                column_cells = worksheet[col_letter]
                try:
                    max_length = max(
                        len(str(cell.value))
                        for cell in column_cells
                        if cell.value
                    )
                except (ValueError, TypeError):
                    max_length = 10
                adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
                worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Data exported to {out_file}")


def format_service_id_folder_name(service_row):
    """
    Builds a subfolder name like "calendar_3_mon_tue_wed_thu_fri" based on:
    - service_id in the row
    - which days are marked '1'
    """
    service_id = service_row['service_id']
    day_map = [
        ('monday', 'mon'),
        ('tuesday', 'tue'),
        ('wednesday', 'wed'),
        ('thursday', 'thu'),
        ('friday', 'fri'),
        ('saturday', 'sat'),
        ('sunday', 'sun'),
    ]
    included_days = []
    for col, short_day in day_map:
        if service_row.get(col, '0') == '1':
            included_days.append(short_day)

    if included_days:
        day_str = "_".join(included_days)
    else:
        day_str = "none"  # or "holiday"

    return f"calendar_{service_id}_{day_str}"


def main():
    """
    Orchestrates the end-to-end GTFS schedule export process:
      1. Load GTFS input files.
      2. Filter specific routes and service IDs, if configured.
      3. Identify key stops (timepoints) for constructing schedules.
      4. Generate schedules by direction for each route-service combination.
      5. Validate times for ordering, remove columns with all missing times.
      6. Export to Excel, one workbook per service ID/route combination.
    """
    load_gtfs_files()
    prepare_timepoints()
    service_id_schedule_map = build_service_id_schedule_map()

    global CALENDAR
    if FILTER_SERVICE_IDS:
        CALENDAR = CALENDAR[CALENDAR['service_id'].isin(FILTER_SERVICE_IDS)]
    if CALENDAR.empty:
        print("No service_ids found after applying FILTER_SERVICE_IDS. Exiting.")
        return

    all_routes = get_all_route_short_names()
    final_routes = apply_in_out_filters(all_routes)
    print(f"Final route selection after filters: {final_routes}")

    for route_short_name in final_routes:
        print(f"\nProcessing route '{route_short_name}'...")
        route_ids = ROUTES[ROUTES['route_short_name'] == route_short_name]['route_id']
        if route_ids.empty:
            print(f"Error: Route '{route_short_name}' not found in routes.txt.")
            continue

        for _, service_row in CALENDAR.iterrows():
            service_id = service_row['service_id']
            folder_name = format_service_id_folder_name(service_row)
            service_output_path = os.path.join(BASE_OUTPUT_PATH, folder_name)
            if not os.path.exists(service_output_path):
                os.makedirs(service_output_path)

            schedule_type = service_id_schedule_map.get(service_id, "Unknown")

            # Filter trips for this route + this service_id
            relevant_trips = TRIPS[
                (TRIPS['route_id'].isin(route_ids)) &
                (TRIPS['service_id'] == service_id)
            ]
            if relevant_trips.empty:
                print(
                    f"  No trips for route='{route_short_name}' "
                    f"and service_id='{service_id}'."
                )
                continue

            direction_ids_local = relevant_trips['direction_id'].unique()
            df_sheets = {}
            for dir_id in direction_ids_local:
                print(
                    f"    Building direction_id '{dir_id}' "
                    f"for service_id='{service_id}'..."
                )
                master_trip_stops = get_master_trip_stops(dir_id, relevant_trips)
                if master_trip_stops.empty:
                    continue

                params_dict = {
                    "trips_dir": relevant_trips[relevant_trips['direction_id'] == dir_id],
                    "master_trip_stops": master_trip_stops,
                    "time_fmt": TIME_FORMAT_OPTION,
                    "route_short": route_short_name,
                    "sched_type": schedule_type,
                    "dir_id": dir_id
                }
                output_df = process_trips_for_direction(params_dict)
                if not output_df.empty:
                    sheet_name = f"Direction_{dir_id}"
                    df_sheets[sheet_name] = output_df

            if df_sheets:
                schedule_type_safe = (
                    schedule_type.replace(' ', '_')
                                 .replace('-', '_')
                                 .replace('/', '_')
                )
                out_file = os.path.join(
                    service_output_path,
                    f"route_{route_short_name}_schedule_{schedule_type_safe}.xlsx"
                )
                export_to_excel_multiple_sheets(df_sheets, out_file)
            else:
                print(
                    f"  No data to export for service_id '{service_id}' "
                    f"on route '{route_short_name}'."
                )


if __name__ == "__main__":
    main()
