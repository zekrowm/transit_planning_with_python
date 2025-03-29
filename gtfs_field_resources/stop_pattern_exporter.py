"""
This module extracts unique stop patterns from GTFS data and exports them to Excel workbooks, organized by route, direction, and service ID.
It optionally loads calendar data to create structured subfolders based on days of service.

Key features:
- Route-based, direction-based, and service-based (calendar) filtering.
- Distance unit conversions (meters or feet to miles).
- Export of timepoint-only stop patterns with optional validation of distances.
- Computation and inclusion of earliest departure times for each pattern.
- Excel outputs organized into subfolders by service_id, labeled clearly with service days (e.g., calendar_123_mon_tue).
- Clear master-trip structure for visualizing stop patterns across different trips.
"""
import logging
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# ===============================
# CONFIGURATION
# ===============================
INPUT_DIR = r'\\Path\\To\\Your\\GTFS'
OUTPUT_DIR = r'\\Path\\To\\Output'
TRIPS_FILE = 'trips.txt'
STOP_TIMES_FILE = 'stop_times.txt'
STOPS_FILE = 'stops.txt'
ROUTES_FILE = 'routes.txt'
CALENDAR_FILE = 'calendar.txt'  # to be used for subfolder naming if available

FILTER_IN_ROUTE_SHORT_NAMES = []
FILTER_OUT_ROUTE_SHORT_NAMES = []
FILTER_IN_CALENDAR_IDS = []

SIGNUP_NAME = "January2025Signup"

INPUT_DISTANCE_UNIT = "meters"
CONVERT_TO_MILES = True
EXPORT_TIMEPOINTS_ONLY = True
VALIDATE_TIMEPOINT_DISTANCE = True

# ------------------------------------------------------------
# Helper Functions
# ------------------------------------------------------------
def is_number(x):
    """
    Check if a value can be converted to a float.
    """
    try:
        float(x)
        return True
    except:
        return False

def convert_dist_to_miles(distance, input_unit):
    """
    Convert 'distance' to miles if configured. Otherwise return as is.
    """
    if not CONVERT_TO_MILES or pd.isna(distance):
        return distance
    if input_unit.lower() == "feet":
        conv = 5280.0
    elif input_unit.lower() == "meters":
        conv = 1609.34
    else:
        logging.warning("Unknown distance unit '%s'. No conversion done.", input_unit)
        conv = 1.0
    return distance / conv

def parse_time_to_minutes(timestr):
    """
    Parse HH:MM:SS (GTFS style) to float minutes. Return None if invalid.
    Allows hours >= 24 for service passing midnight.
    """
    if not isinstance(timestr, str):
        return None
    parts = timestr.strip().split(':')
    if len(parts) != 3:
        return None
    try:
        hh = int(parts[0])
        mm = int(parts[1])
        ss = int(parts[2])
        return hh*60 + mm + ss/60.0
    except:
        return None

def minutes_to_hhmm(m):
    """Convert float minutes to 'HH:MM' (24-hour)."""
    if m is None or pd.isna(m):
        return ""
    total_m = int(m)
    hh = total_m // 60
    mm = total_m % 60
    return f"{hh:02d}:{mm:02d}"

def format_service_id_folder_name(service_id, calendar_df):
    """
    Build a subfolder name like 'calendar_3_mon_tue_fri' or 'calendar_10_none' if no days are served.
    If calendar_df is None or doesn't contain the service_id, fallback to 'calendar_<service_id>'.
    """
    if calendar_df is None or calendar_df.empty:
        return f"calendar_{service_id}"

    row = calendar_df[calendar_df['service_id'] == str(service_id)]
    if row.empty:
        return f"calendar_{service_id}"

    row = row.iloc[0]  # Take the first matching row
    day_map = [
        ('monday', 'mon'),
        ('tuesday', 'tue'),
        ('wednesday', 'wed'),
        ('thursday', 'thu'),
        ('friday', 'fri'),
        ('saturday', 'sat'),
        ('sunday', 'sun'),
    ]
    served_days = []
    for col, short_name in day_map:
        if str(row.get(col, '0')) == '1':
            served_days.append(short_name)

    if served_days:
        day_str = "_".join(served_days)
    else:
        day_str = "none"

    return f"calendar_{service_id}_{day_str}"

# ------------------------------------------------------------
# Loading GTFS
# ------------------------------------------------------------
def load_gtfs_files(input_dir):
    if not os.path.exists(input_dir):
        raise FileNotFoundError(f"Input dir does not exist: {input_dir}")

    stops = pd.read_csv(os.path.join(input_dir, STOPS_FILE))
    trips = pd.read_csv(os.path.join(input_dir, TRIPS_FILE))
    stop_times = pd.read_csv(os.path.join(input_dir, STOP_TIMES_FILE))
    routes = pd.read_csv(os.path.join(input_dir, ROUTES_FILE))

    logging.info("Loaded GTFS files successfully.")
    return {
        'stops': stops,
        'trips': trips,
        'stop_times': stop_times,
        'routes': routes
    }

def filter_trips(trips_df, routes_df, cal_ids):
    """
    Filter trips by desired route short names and calendar service IDs.

    The filter includes trips whose route_short_name is in FILTER_IN_ROUTE_SHORT_NAMES,
    excludes those in FILTER_OUT_ROUTE_SHORT_NAMES, and further filters by any
    specified service IDs.
    """
    # Merge to get route_short_name
    merged = pd.merge(
        trips_df,
        routes_df[['route_id','route_short_name']],
        on='route_id', how='left'
    )
    # Filter in
    if FILTER_IN_ROUTE_SHORT_NAMES:
        merged = merged[merged['route_short_name'].isin(FILTER_IN_ROUTE_SHORT_NAMES)]
    # Filter out
    if FILTER_OUT_ROUTE_SHORT_NAMES:
        merged = merged[~merged['route_short_name'].isin(FILTER_OUT_ROUTE_SHORT_NAMES)]
    # Filter calendar
    if cal_ids and 'service_id' in merged.columns:
        merged = merged[merged['service_id'].isin(cal_ids)]
    elif cal_ids:
        logging.warning("No service_id in data for filtering.")
    return merged

# ------------------------------------------------------------
# Build Patterns
# ------------------------------------------------------------
def generate_unique_patterns(trips_df, stop_times_df, stops_df):
    """
    Creates unique patterns keyed by (route_id, direction_id, service_id, patternOfStops).
    Each pattern is a tuple of stops: [ (stop_id, distanceFromPrevious), ...]
    """
    # Merge to get route/direction
    tmp = pd.merge(
        stop_times_df,
        trips_df[['trip_id','route_id','direction_id','service_id']],
        on='trip_id', how='inner'
    )
    if 'shape_dist_traveled' not in tmp.columns:
        tmp['shape_dist_traveled'] = np.nan

    # Merge stop names
    tmp = pd.merge(tmp, stops_df[['stop_id','stop_name']], on='stop_id', how='left')
    tmp.sort_values(['trip_id','stop_sequence'], inplace=True)

    # If we want timepoint-only
    if EXPORT_TIMEPOINTS_ONLY and 'timepoint' in tmp.columns:
        tmp = tmp[tmp['timepoint'] == 1]

    # For distance validation
    trip_dist = {}
    if VALIDATE_TIMEPOINT_DISTANCE and EXPORT_TIMEPOINTS_ONLY:
        # compute original full-trip dist for each trip
        for tid, grp in tmp.groupby('trip_id'):
            grp = grp.dropna(subset=['shape_dist_traveled'])
            if grp.empty:
                trip_dist[tid] = None
            else:
                dist_val = grp.iloc[-1]['shape_dist_traveled'] - grp.iloc[0]['shape_dist_traveled']
                trip_dist[tid] = convert_dist_to_miles(dist_val, INPUT_DISTANCE_UNIT)

    # Build patterns
    patterns_list = []
    for tid, grp in tmp.groupby('trip_id'):
        grp = grp.sort_values('stop_sequence')
        stops_for_this_trip = []
        prevDistVal = None
        for _, row in grp.iterrows():
            sid = row['stop_id']
            shapeVal = row['shape_dist_traveled']
            if prevDistVal is None:
                dStr = "-"
            else:
                if pd.notnull(shapeVal) and pd.notnull(prevDistVal):
                    diff = convert_dist_to_miles(shapeVal - prevDistVal, INPUT_DISTANCE_UNIT)
                    dStr = f"{diff:.2f}" if diff else ""
                else:
                    dStr = ""
            stops_for_this_trip.append( (sid, dStr) )
            prevDistVal = shapeVal if pd.notnull(shapeVal) else None

        # Validate if desired
        if VALIDATE_TIMEPOINT_DISTANCE and EXPORT_TIMEPOINTS_ONLY:
            sumSeg = 0.0
            for (_, ds) in stops_for_this_trip:
                if ds not in ("-", "", None) and is_number(ds):
                    sumSeg += float(ds)
            fullTrip = trip_dist.get(tid, None)
            if fullTrip is not None and abs(sumSeg - fullTrip) > 0.02:
                logging.warning(
                    "Trip %s sum of segments=%.2f vs. full=%.2f mismatch >0.02",
                    tid, sumSeg, fullTrip
                )

        # Store record
        firstRow = grp.iloc[0]
        rid = firstRow['route_id']
        did = firstRow['direction_id']
        sid_ = firstRow['service_id']
        patterns_list.append({
            'trip_id': tid,
            'route_id': rid,
            'direction_id': did,
            'service_id': sid_,
            'pattern_stops': tuple(stops_for_this_trip)
        })

    # Accumulate unique patterns
    patterns_dict = {}
    for rec in patterns_list:
        key = (rec['route_id'], rec['direction_id'], rec['service_id'], rec['pattern_stops'])
        if key not in patterns_dict:
            patterns_dict[key] = {
                'route_id': rec['route_id'],
                'direction_id': rec['direction_id'],
                'service_id': rec['service_id'],
                'pattern_stops': rec['pattern_stops'],
                'trip_count': 0,
                'trip_ids': []
            }
        patterns_dict[key]['trip_count'] += 1
        patterns_dict[key]['trip_ids'].append(rec['trip_id'])

    logging.info("Found %d unique patterns.", len(patterns_dict))
    return patterns_dict

def assign_pattern_ids(patterns_dict):
    """
    For each route/service/direction, assign pattern_id in ascending order.
    """
    group_map = defaultdict(list)
    for v in patterns_dict.values():
        route_id = v['route_id']
        dir_id = v['direction_id']
        srv_id = v['service_id']
        group_map[(route_id, srv_id, dir_id)].append(v)

    out = []
    for (rid, sid, did), recs in group_map.items():
        # sort stable by pattern_stops
        recs = sorted(recs, key=lambda x: x['pattern_stops'])
        for i, pattern_rec in enumerate(recs, 1):
            pattern_rec['pattern_id'] = i
            out.append({
                'route_id': pattern_rec['route_id'],
                'direction_id': pattern_rec['direction_id'],
                'service_id': pattern_rec['service_id'],
                'pattern_stops': pattern_rec['pattern_stops'],
                'trip_count': pattern_rec['trip_count'],
                'trip_ids': pattern_rec['trip_ids'],
                'pattern_id': i
            })
    logging.info("Assigned pattern IDs to pattern records.")
    return out

# ------------------------------------------------------------
# EARLIEST START TIME
# ------------------------------------------------------------
def compute_earliest_start_times(pattern_records, stop_times_df):
    """
    For each pattern, find the earliest arrival/departure time for the *first* stop
    of each trip in that pattern. Then store the min as earliest_time_minutes & earliest_time_str.
    """
    if 'arrival_time' not in stop_times_df.columns:
        for r in pattern_records:
            r['earliest_time_minutes'] = None
            r['earliest_time_str'] = ""
        return

    st_by_trip = stop_times_df.groupby('trip_id')
    for rec in pattern_records:
        trip_ids = rec['trip_ids']
        earliest_val = None
        for tid in trip_ids:
            if tid not in st_by_trip.groups:
                continue
            g2 = st_by_trip.get_group(tid).sort_values('stop_sequence')
            if g2.empty:
                continue
            arr = g2.iloc[0].get('arrival_time','')
            dep = g2.iloc[0].get('departure_time','')
            arrm = parse_time_to_minutes(str(arr)) if arr else None
            depm = parse_time_to_minutes(str(dep)) if dep else None
            candidates = []
            if arrm is not None:
                candidates.append(arrm)
            if depm is not None:
                candidates.append(depm)
            if not candidates:
                continue
            thisMin = min(candidates)
            if earliest_val is None or thisMin < earliest_val:
                earliest_val = thisMin
        rec['earliest_time_minutes'] = earliest_val
        rec['earliest_time_str'] = minutes_to_hhmm(earliest_val) if earliest_val else ""

# ------------------------------------------------------------
# MASTER TRIP
# ------------------------------------------------------------
def find_master_trip_stops(route_id, direction_id, relevant_trips, stop_times_df, stops_df):
    """
    Among 'relevant_trips' for route+dir, find the trip with the largest # of stops/timepoints.
    Return [ (stop_id, stop_name), ... ]
    """
    if relevant_trips.empty:
        return []
    st_sub = stop_times_df[stop_times_df['trip_id'].isin(relevant_trips['trip_id'])]
    if 'timepoint' in st_sub.columns and EXPORT_TIMEPOINTS_ONLY:
        st_sub = st_sub[st_sub['timepoint'] == 1]

    sizes = st_sub.groupby('trip_id').size()
    if sizes.empty:
        return []
    best_tid = sizes.idxmax()
    best_grp = st_sub[st_sub['trip_id'] == best_tid].sort_values('stop_sequence')
    best_grp = pd.merge(best_grp, stops_df[['stop_id','stop_name']], on='stop_id', how='left')

    out_list = []
    for _, row in best_grp.iterrows():
        sid = row['stop_id']
        sname = row.get('stop_name','Unknown')
        out_list.append((sid, sname))
    return out_list

def forward_match_pattern_to_master(pattern_stops, master_stops):
    """
    Forward-only match: place pattern distances in the matching columns (by stop_id).
    The first matched pattern stop => '-'.
    """
    result = [""] * len(master_stops)
    i = 0
    j = 0
    while i < len(master_stops) and j < len(pattern_stops):
        master_sid = master_stops[i][0]
        pat_sid = pattern_stops[j][0]
        dist_str = pattern_stops[j][1]
        if master_sid == pat_sid:
            result[i] = dist_str
            i += 1
            j += 1
        else:
            i += 1
    return result

# ------------------------------------------------------------
# EXCEL EXPORT
# ------------------------------------------------------------
def create_workbook():
    """
    Create and return a new openpyxl Workbook instance with the default sheet removed.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    return wb

def fill_worksheet_for_direction(
    wb, sheet_title, route_short_name, direction_id, service_id,
    pattern_records_dir, master_stops
):
    """
    Create one Excel sheet for route/direction. Each pattern is one row, columns
    are "Route, Direction, Calendar (service_id), Pattern ID, Trip Count, Earliest Start Time"
    plus one column per master_stop.
    """
    try:
        ws = wb.create_sheet(title=sheet_title)
    except ValueError:
        # If there's a naming conflict or length issue, slightly rename
        ws = wb.create_sheet(title=f"{sheet_title[:25]}_X")

    if not master_stops:
        ws.append(["No master stops found for direction."])
        return

    # Build header
    header = [
        "Route",
        "Direction",
        "Calendar (service_id)",
        "Pattern ID",
        "Trip Count",
        "Earliest Start Time"
    ]
    for (_, sname) in master_stops:
        header.append(sname)
    ws.append(header)

    # sort pattern records by earliest_time_minutes
    pattern_records_dir = sorted(
        pattern_records_dir,
        key=lambda r: (r.get('earliest_time_minutes') is None, r.get('earliest_time_minutes', 9999999))
    )

    for rec in pattern_records_dir:
        pat_id = rec['pattern_id']
        tc = rec['trip_count']
        e_str = rec.get('earliest_time_str', "")
        pattern_stops = rec['pattern_stops']
        row_distances = forward_match_pattern_to_master(pattern_stops, master_stops)

        row = [
            route_short_name,
            direction_id,
            service_id,
            pat_id,
            tc,
            e_str
        ]
        row.extend(row_distances)
        ws.append(row)

    # Column widths
    for col_index, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30

def export_patterns_to_excel(
    pattern_records,
    routes_df,
    stop_times_df,
    stops_df,
    calendar_df=None
):
    """
    For each (route_id, service_id), group patterns by direction, create a subfolder
    named for that service_id's days (if calendar.txt loaded), and save an Excel workbook
    with one sheet per direction.
    """
    group_map = defaultdict(list)
    for pr in pattern_records:
        rid = pr['route_id']
        sid = pr['service_id']
        group_map[(rid, sid)].append(pr)

    for (rid, sid), group_list in group_map.items():
        # find route_short_name
        route_info = routes_df[routes_df['route_id'] == rid]
        if not route_info.empty:
            short_name = route_info.iloc[0].get('route_short_name', f"Route_{rid}")
        else:
            short_name = f"Route_{rid}"

        # Create subfolder for this service_id, e.g. "calendar_10_mon_tue"
        folder_name = format_service_id_folder_name(sid, calendar_df)
        folder_path = os.path.join(OUTPUT_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Create a workbook
        wb = create_workbook()

        # group by direction
        dir_map = defaultdict(list)
        for r2 in group_list:
            dir_id = r2['direction_id']
            dir_map[dir_id].append(r2)

        for d_id, recs_dir in dir_map.items():
            # gather all trip_ids
            all_trip_ids = set()
            for r3 in recs_dir:
                all_trip_ids.update(r3['trip_ids'])

            # subset stop_times to those trip_ids
            st_sub = stop_times_df[stop_times_df['trip_id'].isin(all_trip_ids)]
            if 'timepoint' in st_sub.columns and EXPORT_TIMEPOINTS_ONLY:
                st_sub = st_sub[st_sub['timepoint'] == 1]

            # pick trip with max # stops => master_stops
            sizes = st_sub.groupby('trip_id').size()
            if sizes.empty:
                master_stops = []
            else:
                best_tid = sizes.idxmax()
                mg = st_sub[st_sub['trip_id'] == best_tid].sort_values('stop_sequence')
                mg = pd.merge(mg, stops_df[['stop_id','stop_name']], on='stop_id', how='left')
                master_stops = []
                for _, rowz in mg.iterrows():
                    master_stops.append( (rowz['stop_id'], rowz.get('stop_name','Unknown')) )

            # fill sheet
            sheet_title = f"Dir{d_id}"
            fill_worksheet_for_direction(
                wb, sheet_title, short_name, d_id, sid, recs_dir, master_stops
            )

        # Save to the subfolder
        fn = f"{short_name}_{sid}_{SIGNUP_NAME}.xlsx"
        full_fp = os.path.join(folder_path, fn)
        try:
            wb.save(full_fp)
            logging.info("Saved workbook: %s", full_fp)
        except Exception as e:
            logging.error("Could not save workbook '%s': %s", fn, e)

# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
def main():
    """
    Main entry point for generating and exporting unique GTFS stop patterns.

    Steps:
    1) Create the output directory if it doesn’t exist.
    2) Optionally load calendar.txt for subfolder naming by service day.
    3) Load the core GTFS tables (stops, trips, stop_times, routes).
    4) Filter trips if necessary based on route short names and calendar IDs.
    5) Build unique stop patterns for each trip group.
    6) Compute earliest start times for each pattern.
    7) Export the patterns to Excel, creating subfolders grouped by service_id.
    """
    # 1) Ensure output dir
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    # 2) Attempt to load calendar.txt for day-of-week subfolder naming
    calendar_df = None
    cal_path = os.path.join(INPUT_DIR, CALENDAR_FILE)
    if os.path.exists(cal_path):
        try:
            calendar_df = pd.read_csv(cal_path, dtype=str)
            logging.info("Loaded calendar.txt successfully.")
        except Exception as e:
            logging.warning("Could not load calendar.txt: %s", e)
            calendar_df = None
    else:
        logging.info("No calendar.txt found; subfolders will be 'calendar_<service_id>' only.")

    # 3) Load GTFS files
    try:
        gtfs_data = load_gtfs_files(INPUT_DIR)
    except Exception as e:
        logging.error("Failed to load GTFS: %s", e)
        return

    stops_df = gtfs_data['stops']
    trips_df = gtfs_data['trips']
    stop_times_df = gtfs_data['stop_times']
    routes_df = gtfs_data['routes']

    # 4) Filter trips if desired
    filtered_trips = filter_trips(
        trips_df,
        routes_df,
        cal_ids=FILTER_IN_CALENDAR_IDS
    )
    if filtered_trips.empty:
        logging.error("No trips after filtering. Exiting.")
        return

    # 5) Generate unique patterns
    patterns_dict = generate_unique_patterns(filtered_trips, stop_times_df, stops_df)
    if not patterns_dict:
        logging.warning("No patterns found. Exiting.")
        return

    pattern_records = assign_pattern_ids(patterns_dict)
    if not pattern_records:
        logging.warning("No pattern records. Exiting.")
        return

    # 6) Earliest Start Times
    compute_earliest_start_times(pattern_records, stop_times_df)

    # 7) Export patterns to Excel (with subfolders for each service_id)
    export_patterns_to_excel(
        pattern_records,
        routes_df,
        stop_times_df,
        stops_df,
        calendar_df=calendar_df
    )

if __name__ == "__main__":
    main()
