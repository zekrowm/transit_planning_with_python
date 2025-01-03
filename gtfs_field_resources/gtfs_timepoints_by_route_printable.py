"""
gtfs_timepoints_by_route_printable.py

This module processes GTFS (General Transit Feed Specification) data to generate
printable timepoints by route. It reads GTFS files, filters trips based on
schedule types and time windows, and exports the data to formatted Excel files.
These exports can be printed and used for ride checks.
"""

import os
import sys
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Input file paths for GTFS files (add your own paths here)
BASE_INPUT_PATH = r'\\your_file_path\here\\'
TRIPS_FILE = os.path.join(BASE_INPUT_PATH, "trips.txt")
STOP_TIMES_FILE = os.path.join(BASE_INPUT_PATH, "stop_times.txt")
ROUTES_FILE = os.path.join(BASE_INPUT_PATH, "routes.txt")
STOPS_FILE = os.path.join(BASE_INPUT_PATH, "stops.txt")
CALENDAR_FILE = os.path.join(BASE_INPUT_PATH, "calendar.txt")

# Output directory (add your own path here)
BASE_OUTPUT_PATH = r'\\your_file_path\here\\'
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

# List of route short names to process
ROUTE_SHORT_NAMES = ['101', '102']  # Modify as needed; check your 'routes.txt' GTFS file if unsure

# Define schedule types and corresponding days in the calendar
# Format: {'Schedule Type': ['day1', 'day2', ...]}
SCHEDULE_TYPES = {
    'Weekday': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday': ['sunday'],
    # 'Friday': ['friday'],  # Uncomment if you have a unique Friday schedule
}

# Time windows for filtering trips
# Format: {'Schedule Type': {'Time Window Name': ('Start Time', 'End Time')}}
# Times should be in 'HH:MM:SS' 24-hour format
TIME_WINDOWS = {
    'Weekday': {
        'full_day': ('00:00', '23:59'),
        'morning': ('06:00', '09:59'),
        'afternoon': ('14:00', '17:59'),
        # 'evening': ('18:00', '21:59'),  # Add as needed
    },
    'Saturday': {
        'midday': ('10:00', '13:59'),
        # Add more time windows for Saturday if needed
    },
    # 'Sunday': {  # Uncomment and customize for Sunday if needed
    #     'morning': ('08:00', '11:59'),
    #     'afternoon': ('12:00', '15:59'),
    # },
}

# Placeholder values
MISSING_TIME = "________"
COMMENTS_PLACEHOLDER = "__________________"

# Maximum column width for Excel output (used to wrap long headers)
MAX_COLUMN_WIDTH = 30  # Adjust as needed

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

# Load GTFS files with basic error handling
try:
    trips = pd.read_csv(TRIPS_FILE)
    stop_times = pd.read_csv(STOP_TIMES_FILE)
    routes = pd.read_csv(ROUTES_FILE)
    stops = pd.read_csv(STOPS_FILE)
    calendar = pd.read_csv(CALENDAR_FILE)
except FileNotFoundError as e:
    print(f"Error: {e}")
    print("Please check your input file paths in the configuration section.")
    sys.exit(1)
except Exception as e:
    print(f"An unexpected error occurred while reading GTFS files: {e}")
    sys.exit(1)

# Check for 'timepoint' column and filter timepoints
if 'timepoint' in stop_times.columns:
    timepoints = stop_times[stop_times['timepoint'] == 1]
else:
    print("Warning: 'timepoint' column not found. Using all stops as timepoints.")
    timepoints = stop_times.copy()


def adjust_time(time_str):
    """
    Adjusts the time string to ensure hours are within 0-23.

    Parameters:
        time_str (str): Time string in 'HH:MM:SS' format.

    Returns:
        str or None: Adjusted time string in 'HH:MM' format or None if invalid.
    """
    parts = time_str.strip().split(":")
    if len(parts) >= 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            if hours >= 24:
                hours -= 24
            return f"{hours:02}:{minutes:02}"
        except ValueError:
            return None
    else:
        return None


def get_ordered_stops(direction_id, relevant_trips):
    """
    Retrieves ordered stop names and IDs for a given direction and set of trips.

    Parameters:
        direction_id (int): The direction ID to filter trips.
        relevant_trips (DataFrame): DataFrame containing relevant trips.

    Returns:
        tuple: A tuple containing a list of ordered stop names and a list of ordered stop IDs.
    """
    relevant_trips_direction = relevant_trips[relevant_trips['direction_id'] == direction_id]

    if relevant_trips_direction.empty:
        print(f"No trips found for direction_id '{direction_id}'.")
        return [], []

    # Get all stops for this direction in order
    all_stops = timepoints[timepoints['trip_id'].isin(relevant_trips_direction['trip_id'])]
    all_stops = all_stops.sort_values(['trip_id', 'stop_sequence'])

    if all_stops.empty:
        print(f"No stop times found for direction_id '{direction_id}'.")
        return [], []

    # Get unique stops in order
    unique_stops = all_stops.drop_duplicates('stop_id')[['stop_id', 'stop_sequence']]
    unique_stops = unique_stops.sort_values('stop_sequence')

    # Get stop names
    stop_names = stops[stops['stop_id'].isin(unique_stops['stop_id'])]
    stop_names = stop_names.set_index('stop_id')['stop_name']

    # Create ordered list of stop names
    ordered_stop_names = [
        stop_names.get(stop_id, f"Unknown Stop ID {stop_id}") for stop_id in unique_stops['stop_id']
    ]

    return ordered_stop_names, unique_stops['stop_id'].tolist()


def process_direction(relevant_trips_direction, direction_id, ordered_stop_names, ordered_stop_ids):
    """
    Processes trips for a specific direction and compiles schedule data.

    Parameters:
        relevant_trips_direction (DataFrame): Trips relevant to the current direction and time window.
        direction_id (int): The direction ID being processed.
        ordered_stop_names (list): Ordered list of stop names.
        ordered_stop_ids (list): Ordered list of stop IDs.

    Returns:
        DataFrame: Compiled schedule data for export.
    """
    if relevant_trips_direction.empty:
        print(f"Skipping direction_id '{direction_id}' due to lack of data.")
        return pd.DataFrame()

    output_data = []

    for trip_id, group in timepoints[
        timepoints['trip_id'].isin(relevant_trips_direction['trip_id'])
    ].groupby('trip_id'):
        trip_info = relevant_trips_direction[
            relevant_trips_direction['trip_id'] == trip_id
        ].iloc[0]
        route_name = routes[routes['route_id'] == trip_info['route_id']]['route_short_name'].values[0]
        trip_headsign = trip_info.get('trip_headsign', '')

        row = [MISSING_TIME, trip_id, route_name, direction_id, trip_headsign]

        # Create a dictionary to store schedule times for each stop
        schedule_times = {stop_id: MISSING_TIME for stop_id in ordered_stop_ids}

        for _, stop in group.iterrows():
            time_str = stop['departure_time'].strip()
            # Fix time format if hours >= 24
            time_str = adjust_time(time_str)
            if time_str is None:
                print(
                    f"Invalid time format '{stop['departure_time']}' in trip_id '{trip_id}' at stop_id '{stop['stop_id']}'"
                )
                continue
            schedule_times[stop['stop_id']] = time_str

        # Add schedule times and blank actual times for each stop
        for stop_id in ordered_stop_ids:
            row.append(schedule_times[stop_id])
            row.append(MISSING_TIME)

        row.append(COMMENTS_PLACEHOLDER)
        output_data.append(row)

    # Create DataFrame
    columns = ['Date', 'Trip ID', 'Route Name', 'Direction ID', 'Trip Headsign']
    for stop_name in ordered_stop_names:
        columns.extend([f'{stop_name} Schedule', f'{stop_name} Actual'])
    columns.append('Comments')

    df = pd.DataFrame(output_data, columns=columns)

    # Use pd.to_timedelta to sort without affecting the original time columns
    schedule_columns = [col for col in df.columns if col.endswith('Schedule')]
    if schedule_columns:
        first_schedule_col = schedule_columns[0]
        # Convert to timedelta for sorting
        df['sort_time'] = pd.to_timedelta(df[first_schedule_col] + ':00', errors='coerce')
        # Sort by the first schedule column
        df = df.sort_values(by='sort_time')
        # Drop the temporary sort_time column
        df = df.drop(columns=['sort_time'])
    else:
        print(f"No schedule columns found for direction_id '{direction_id}'.")

    return df


def export_to_excel(df, output_file):
    """
    Exports the DataFrame to an Excel file with formatting.

    Parameters:
        df (DataFrame): The DataFrame to export.
        output_file (str): The path to the output Excel file.
    """
    if df.empty:
        print(f"No data to export to {output_file}.")
        return

    # Create a Pandas Excel writer using openpyxl as the engine.
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        worksheet = writer.sheets['Sheet1']

        # Apply alignment to headers and adjust column widths
        for col_num, col_name in enumerate(df.columns, 1):  # 1-based indexing for openpyxl
            # Get the column letter
            col_letter = get_column_letter(col_num)

            # Set alignment to left and enable text wrapping for the header
            header_cell = worksheet[f'{col_letter}1']
            header_cell.alignment = Alignment(horizontal='left', wrap_text=True)

            # Set alignment to left for all data cells in the column
            for row_num in range(2, worksheet.max_row + 1):  # Start from row 2 to skip header
                cell = worksheet[f'{col_letter}{row_num}']
                cell.alignment = Alignment(horizontal='left')

            # Calculate the maximum width for this column
            column_cells = worksheet[col_letter]
            max_length = max(
                len(str(cell.value)) for cell in column_cells if cell.value is not None
            )

            # Adjust the column width, limiting it to the maximum column width
            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Data exported to {output_file}")


# Process each schedule type
for schedule_type, days in SCHEDULE_TYPES.items():
    print(f"\nProcessing schedule type '{schedule_type}'...")

    if not all(day in calendar.columns for day in days):
        print(
            f"Error: One or more day columns not found in calendar.txt for schedule type '{schedule_type}'."
        )
        continue

    # Filter for services that are available on all specified days
    relevant_service_ids = calendar[calendar[days].all(axis=1)]['service_id']

    if relevant_service_ids.empty:
        print(f"No services found for schedule type '{schedule_type}'.")
        continue

    # Process each route in the list
    for route_short_name in ROUTE_SHORT_NAMES:
        print(f"Processing route {route_short_name}...")

        # Get route_ids for the current route_short_name
        route_ids = routes[routes['route_short_name'] == route_short_name]['route_id']
        if route_ids.empty:
            print(f"Error: Route '{route_short_name}' not found in routes.txt.")
            continue  # Skip to next route

        relevant_trips = trips[
            (trips['route_id'].isin(route_ids)) &
            (trips['service_id'].isin(relevant_service_ids))
        ]

        if relevant_trips.empty:
            print(
                f"No trips found for route '{route_short_name}' with the specified service IDs for schedule type '{schedule_type}'."
            )
            continue  # Skip to next route

        # Process each direction_id present in the data
        direction_ids = relevant_trips['direction_id'].unique()
        for direction_id in direction_ids:
            print(f"Processing direction_id {direction_id}...")

            # Get ordered stops
            ordered_stop_names, ordered_stop_ids = get_ordered_stops(
                direction_id, relevant_trips
            )

            # Process each time window for the current schedule type
            for time_window_name, (start_time_str, end_time_str) in TIME_WINDOWS.get(
                schedule_type, {}
            ).items():
                print(f"Processing time window '{time_window_name}'...")

                # Filter trips that depart within the specified time window
                # Merge relevant_trips with stop_times to get the first departure time
                trips_with_times = relevant_trips.merge(
                    stop_times[['trip_id', 'departure_time', 'stop_sequence']],
                    on='trip_id'
                )

                # Adjust departure times
                trips_with_times['adjusted_departure_time'] = trips_with_times[
                    'departure_time'
                ].apply(adjust_time)
                trips_with_times['departure_timedelta'] = pd.to_timedelta(
                    trips_with_times['adjusted_departure_time'] + ':00', errors='coerce'
                )

                # Get the first stop of each trip (minimum stop_sequence)
                first_stops = trips_with_times.groupby('trip_id')['stop_sequence'].min().reset_index()
                first_departures = trips_with_times.merge(
                    first_stops, on=['trip_id', 'stop_sequence']
                )

                # Convert start_time and end_time to timedelta
                start_time = pd.to_timedelta(start_time_str + ':00')
                end_time = pd.to_timedelta(end_time_str + ':00')

                # Filter trips within the time window and correct direction_id
                trips_in_time_window = first_departures[
                    (first_departures['departure_timedelta'] >= start_time) &
                    (first_departures['departure_timedelta'] <= end_time) &
                    (first_departures['direction_id'] == direction_id)
                ]

                trips_in_time_window_ids = trips_in_time_window['trip_id'].unique()

                if len(trips_in_time_window_ids) == 0:
                    print(
                        f"No trips found in time window '{time_window_name}' for direction_id '{direction_id}'."
                    )
                    continue

                # Filter relevant_trips_direction to these trips
                relevant_trips_direction = relevant_trips[
                    relevant_trips['trip_id'].isin(trips_in_time_window_ids)
                ]

                # Process direction
                df = process_direction(
                    relevant_trips_direction,
                    direction_id,
                    ordered_stop_names,
                    ordered_stop_ids
                )

                if df.empty:
                    print(
                        f"No data to export for route {route_short_name}, schedule '{schedule_type}', "
                        f"time window '{time_window_name}', direction_id '{direction_id}'."
                    )
                    continue

                # Export DataFrame to Excel with formatting using openpyxl
                output_file = os.path.join(
                    BASE_OUTPUT_PATH,
                    f"route_{route_short_name}_{schedule_type}_{time_window_name}_direction_{direction_id}.xlsx"
                )

                # Export to Excel with formatting
                export_to_excel(df, output_file)
