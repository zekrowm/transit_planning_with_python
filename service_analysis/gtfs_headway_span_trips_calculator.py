"""
Route Vital Signs Calculator

This script processes General Transit Feed Specification (GTFS) data to generate
key performance indicators (KPIs) and essential service-level metrics for transit routes.
It computes core statistics such as service span, headways, average run times, route distances,
and operating speeds. Additionally, the script identifies interlined routes and validates
trip data consistency across defined time-blocks.

Core Functionalities:
    - Filters out non-service ("fake") routes based on configurable criteria.
    - Calculates essential metrics including:
        - Service span (first and last trip times)
        - Time-block based headways and trip counts
        - Average and median route distances
        - Average run times and operating speeds
    - Supports flexible units (meters, feet, or miles) for distance calculations.
    - Identifies interlined routes to provide reference information without impacting route metrics.
    - Includes built-in validation checks to ensure data consistency and accuracy.

Configuration:
    - Input and output directories
    - GTFS files to be loaded
    - Definitions for customizable time-blocks
    - List of route identifiers to exclude from analysis
    - Unit specification for distance measurement

Usage:
    Place required GTFS files in the specified input directory, configure script parameters
    as needed, and run the script. Outputs will be exported to Excel, formatted for clarity
    and ease of route comparison.

This tool is intended to simplify and standardize route performance assessment, making
comparative analysis straightforward for planners, analysts, and transit agencies.
"""

import os
from datetime import timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_INPUT_PATH = r'\\your_folder_path\here\\'
OUTPUT_PATH = r'\\your_folder_path\here\\'
gtfs_files = [
    'routes.txt',
    'trips.txt',
    'stop_times.txt',
    'calendar.txt',
    'calendar_dates.txt'
]

OUTPUT_EXCEL = "route_schedule_headway_with_modes.xlsx"

time_blocks_config = {
    'AM Early': ('04:00', '05:59'),
    'AM Peak':  ('06:00', '08:59'),
    'Midday':   ('09:00', '14:59'),
    'PM Peak':  ('15:00', '17:59'),
    'PM Late':  ('18:00', '20:59'),
    'PM Nite':  ('21:00', '23:59'),
    'Other':    ('24:00', '27:59'),
}

schedule_types = {
    'Weekday':  ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday':   ['sunday'],
}

FAKE_ROUTES = ["9999A","9999B","9999C"]

# Set this to 'meters', 'feet', or 'miles'
DISTANCE_UNIT = 'meters'

# -----------------------------------------------------------------------------
# FUNCTIONS
# -----------------------------------------------------------------------------

def get_distance_conversion_factor(unit: str) -> float:
    """
    Return the factor to multiply shape_dist_traveled by
    so that we end up with miles.
    """
    unit_lower = unit.lower()
    if unit_lower == 'meters':
        return 1 / 1609.344
    elif unit_lower == 'feet':
        return 1 / 5280.0
    elif unit_lower == 'miles':
        return 1.0
    else:
        raise ValueError(f"Unknown distance unit={unit}. Must be 'meters','feet','miles'.")

def check_input_files(base_path, files):
    """Verify that the GTFS input directory and all needed files exist."""
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"Input directory {base_path} does not exist.")
    for f in files:
        full = os.path.join(base_path, f)
        if not os.path.exists(full):
            raise FileNotFoundError(f"Missing GTFS file: {f} in {base_path}")


def load_gtfs_data(base_path, files):
    """Load each GTFS file as a pandas DataFrame and return a dict of them."""
    data = {}
    for f in files:
        path = os.path.join(base_path, f)
        name = f.replace('.txt','')
        df = pd.read_csv(path)
        data[name] = df
        print(f"Loaded {f} with {len(df)} rows.")
    return data


def parse_time_blocks(tb_config):
    """Convert 'HH:MM' strings to timedelta-based intervals."""
    from datetime import timedelta
    res = {}
    for block_name, (start_str, end_str) in tb_config.items():
        h1, m1 = map(int, start_str.split(':'))
        h2, m2 = map(int, end_str.split(':'))
        res[block_name] = (
            timedelta(hours=h1, minutes=m1),
            timedelta(hours=h2, minutes=m2)
        )
    return res


def assign_time_block(td, blocks_dict):
    """Given a single timedelta (td), find which time block it falls into."""
    for block_name, (start, end) in blocks_dict.items():
        if start <= td <= end:
            return block_name
    return 'other'


def format_timedelta(td):
    """Format a timedelta as 'HH:MM' or None if invalid."""
    if pd.isna(td):
        return None
    secs = int(td.total_seconds())
    hh = secs // 3600
    mm = (secs % 3600)//60
    return f"{hh:02}:{mm:02}"


def calculate_run_time_and_distance(stop_times_sub):
    """
    Summarize one trip's run time (in minutes) and distance (in miles).
    """
    arrival_max = stop_times_sub['arrival_time'].max()
    depart_min = stop_times_sub['departure_time'].min()
    if pd.notnull(arrival_max) and pd.notnull(depart_min):
        run_td = arrival_max - depart_min
        run_min = run_td.total_seconds() / 60.0
    else:
        run_min = None

    if 'shape_dist_traveled' in stop_times_sub.columns:
        dist = stop_times_sub['shape_dist_traveled'].max()
    else:
        dist = None

    return pd.Series({'trip_run_minutes': run_min, 'trip_distance_miles': dist})


def calculate_headways_for_block(departure_series):
    """Compute the modal headway (in min) for consecutive departures."""
    d_sorted = departure_series.sort_values()
    diffs = d_sorted.diff().dropna().apply(lambda x: x.total_seconds() / 60.0)
    if diffs.empty:
        return None
    return diffs.mode()[0]


def save_to_excel(df, out_dir, out_filename, sheet_name="Route_Schedule_Headway"):
    """Save the final DataFrame to an Excel file, auto-fit columns, center text."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header
    ws.append(df.columns.tolist())

    # Write rows
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    # Auto-fit column widths, center alignment
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max(len(str(x.value)) for x in col if x.value is not None)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 3
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

    os.makedirs(out_dir, exist_ok=True)
    full_out = os.path.join(out_dir, out_filename)
    wb.save(full_out)
    print(f"Saved {full_out}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    try:
        # 1) Determine distance conversion factor from user config
        dist_factor = get_distance_conversion_factor(DISTANCE_UNIT)

        print("Checking files...")
        check_input_files(GTFS_INPUT_PATH, gtfs_files)
        print("All input files found.\n")

        print("Loading GTFS data...")
        data = load_gtfs_data(GTFS_INPUT_PATH, gtfs_files)
        print("Loaded all GTFS.\n")

        # Extract DataFrames
        routes_df = data['routes']
        trips_df = data['trips']
        stop_times_df = data['stop_times']
        calendar_df = data['calendar']
        # calendar_dates_df = data['calendar_dates']  # If needed.

        # 2) Filter out "fake" routes
        routes_df = routes_df[~routes_df['route_short_name'].isin(FAKE_ROUTES)]
        print(f"Filtered out fake routes: {FAKE_ROUTES}")
        print(f"Remaining routes: {len(routes_df)}\n")

        # 3) Convert times to timedelta
        stop_times_df['arrival_time'] = pd.to_timedelta(
            stop_times_df['arrival_time'], errors='coerce'
        )
        stop_times_df['departure_time'] = pd.to_timedelta(
            stop_times_df['departure_time'], errors='coerce'
        )

        # 4) Convert shape_dist_traveled to miles, if present
        if 'shape_dist_traveled' in stop_times_df.columns:
            stop_times_df['shape_dist_traveled'] *= dist_factor
            print(f"Converted shape_dist_traveled to miles (assuming {DISTANCE_UNIT}).")

        # 5) Build block->routes and route->interlines map
        trips_w_routename = trips_df.merge(
            routes_df[['route_id','route_short_name']],
            on='route_id',
            how='left'
        )
        block_to_routes = (
            trips_w_routename.groupby('block_id')['route_short_name']
            .apply(lambda s: set(s.dropna()))
            .to_dict()
        )
        route_to_interlines = {}
        for blk, rtset in block_to_routes.items():
            for rt in rtset:
                route_to_interlines.setdefault(rt, set()).update(rtset - {rt})

        # 6) Compute route-level distance/speed metrics first
        trip_metrics = (
            stop_times_df.groupby('trip_id', as_index=False)
            .apply(calculate_run_time_and_distance)
        )
        # Attach route_id, direction_id
        trip_metrics = trip_metrics.merge(
            trips_df[['trip_id','route_id','direction_id']],
            on='trip_id',
            how='left'
        )

        agg_dict = {
            'trip_run_minutes': 'mean',
            'trip_distance_miles': ['mean','median']
        }

        route_dir_metrics = (
            trip_metrics
            .groupby(['route_id','direction_id'], as_index=False)
            .agg(agg_dict)
        )
        # Flatten columns
        route_dir_metrics.columns = [
            'route_id','direction_id',
            'avg_run_minutes','avg_distance_miles','median_distance_miles'
        ]
        # Speed
        route_dir_metrics['avg_speed_mph'] = (
            route_dir_metrics['avg_distance_miles']
            / (route_dir_metrics['avg_run_minutes']/60.0)
        )

        # Merge route_short_name, route_long_name
        route_dir_metrics = route_dir_metrics.merge(
            routes_df[['route_id','route_short_name','route_long_name']],
            on='route_id',
            how='left'
        )

        # 7) Parse time blocks
        time_blocks = parse_time_blocks(time_blocks_config)

        # 8) For each schedule type
        for sched_type, days in schedule_types.items():
            print(f"Processing: {sched_type} => {days}")

            # (a) Filter valid service_ids
            mask = pd.Series([True]*len(calendar_df))
            for d in days:
                mask &= (calendar_df[d] == 1)
            valid_sids = calendar_df[mask]['service_id'].unique()

            # (b) Filter trips to those service_ids + route_ids
            valid_routeids = routes_df['route_id'].unique()
            trips_filt = trips_df[
                trips_df['service_id'].isin(valid_sids)
                & trips_df['route_id'].isin(valid_routeids)
            ]
            if trips_filt.empty:
                print(f"No trips for {sched_type}\n")
                continue

            # (c) Get first-stop departure_time for each trip
            st_filt = stop_times_df[
                stop_times_df['trip_id'].isin(trips_filt['trip_id'].unique())
            ]
            first_stops = st_filt[st_filt['stop_sequence'] == 1].copy()
            first_stops.dropna(subset=['departure_time'], inplace=True)

            # Attach route info
            trip_info = trips_filt[['trip_id','route_id','direction_id']].merge(
                routes_df[['route_id','route_short_name','route_long_name']],
                on='route_id', how='left'
            )
            first_stops = first_stops.merge(
                trip_info,
                on='trip_id',
                how='left'
            )

            if first_stops.empty:
                print(f"No first-stop data for {sched_type}\n")
                continue

            # (d) Assign time blocks
            def block_assigner(td):
                for bn, (start, end) in time_blocks.items():
                    if start <= td <= end:
                        return bn
                return 'other'

            first_stops['time_block'] = first_stops['departure_time'].apply(block_assigner)

            # (e) HEADWAYS: group by (route_id, direction_id, time_block)
            headway_df = (
                first_stops.groupby(['route_id','direction_id','time_block'])['departure_time']
                .apply(calculate_headways_for_block)
                .reset_index(name='headway_minutes')
            )
            headway_pivot = headway_df.pivot(
                index=['route_id','direction_id'],
                columns='time_block',
                values='headway_minutes'
            ).reset_index()
            # rename pivot columns -> e.g. "AM Early" => "am_early_headway"
            rename_hw = {
                tb: tb.lower().replace(' ','_') + '_headway'
                for tb in time_blocks_config.keys()
            }
            headway_pivot.rename(columns=rename_hw, inplace=True)

            # (f) Binned trip counts with pivot_table(..., fill_value=0)
            tb_counts = (
                first_stops
                .groupby(['route_id','direction_id','time_block'])['trip_id']
                .nunique()
                .reset_index(name='trip_count')
            )
            tbc_pivot = tb_counts.pivot_table(
                index=['route_id','direction_id'],
                columns='time_block',
                values='trip_count',
                fill_value=0
            ).reset_index()

            rename_counts = {
                tb: tb.lower().replace(' ','_') + '_trips'
                for tb in time_blocks_config.keys()
            }
            tbc_pivot.rename(columns=rename_counts, inplace=True)

            # (g) Compute first/last trip times
            def route_trip_times(grp):
                min_t = grp['departure_time'].min()
                max_t = grp['departure_time'].max()
                return pd.Series({
                    'first_trip_time': format_timedelta(min_t),
                    'last_trip_time':  format_timedelta(max_t),
                    'total_trips':     grp['trip_id'].nunique()
                })
            rtimes = (
                first_stops.groupby(['route_id','direction_id'], as_index=False)
                .apply(route_trip_times)
            )

            # (h) Merge these route-level lumps
            final_data = rtimes.merge(
                headway_pivot, on=['route_id','direction_id'], how='left'
            )
            final_data = final_data.merge(
                tbc_pivot, on=['route_id','direction_id'], how='left'
            )

            # attach route_short_name & route_long_name
            route_lookup = routes_df[['route_id','route_short_name','route_long_name']].drop_duplicates()
            final_data = final_data.merge(
                route_lookup,
                on='route_id', how='left'
            )

            # attach computed run time/distance/speed
            final_data = final_data.merge(
                route_dir_metrics[[
                    'route_id','direction_id','avg_run_minutes',
                    'avg_distance_miles','median_distance_miles','avg_speed_mph'
                ]],
                on=['route_id','direction_id'],
                how='left'
            )

            # (i) Build "interlined_routes" column from route_to_interlines
            def get_interlines(rt_short):
                s = route_to_interlines.get(rt_short, set())
                return ", ".join(sorted(s))
            final_data['interlined_routes'] = final_data['route_short_name'].apply(get_interlines)

            # (j) Service span
            def compute_service_span(row):
                ft = row['first_trip_time']
                lt = row['last_trip_time']
                if pd.notnull(ft) and pd.notnull(lt):
                    td_start = pd.to_timedelta(ft + ":00")
                    td_end   = pd.to_timedelta(lt + ":00")
                    return format_timedelta(td_end - td_start)
                return None
            final_data['service_span'] = final_data.apply(compute_service_span, axis=1)

            # Convert avg_run_minutes -> HH:MM
            def minutes_to_hhmm(x):
                if pd.isna(x):
                    return None
                td = timedelta(minutes=x)
                return format_timedelta(td)
            final_data['avg_run_time'] = final_data['avg_run_minutes'].apply(minutes_to_hhmm)

            # (k) Reorder columns
            desired_cols = [
                'route_short_name','route_long_name','direction_id','interlined_routes',
                'service_span','first_trip_time','last_trip_time','total_trips',
                'am_early_trips','am_peak_trips','midday_trips','pm_peak_trips','pm_late_trips',
                'pm_nite_trips','other_trips',
                'am_early_headway','am_peak_headway','midday_headway','pm_peak_headway',
                'pm_late_headway','pm_nite_headway','other_headway',
                'avg_run_time','avg_distance_miles','median_distance_miles','avg_speed_mph'
            ]
            existing = [c for c in desired_cols if c in final_data.columns]
            extra = [c for c in final_data.columns if c not in existing]
            final_data = final_data[existing + extra].copy()

            # (l) Round distance/speed columns
            for c in ['avg_distance_miles','median_distance_miles','avg_speed_mph']:
                if c in final_data.columns:
                    final_data[c] = final_data[c].round(1)

            # (m) Validation check: Compare total_trips to sum of the 7 time-block columns
            time_block_cols = [
                'am_early_trips','am_peak_trips','midday_trips','pm_peak_trips',
                'pm_late_trips','pm_nite_trips','other_trips'
            ]
            for idx, row in final_data.iterrows():
                # Sum only the columns that exist and are not null
                binned_sum = 0
                for col in time_block_cols:
                    if col in final_data.columns and pd.notnull(row[col]):
                        binned_sum += row[col]

                if pd.notnull(row['total_trips']) and binned_sum != row['total_trips']:
                    print(
                        f"WARNING: For {sched_type} schedule, route "
                        f"{row['route_short_name']} dir {row['direction_id']}, "
                        f"binned sum={binned_sum} but total_trips={row['total_trips']}."
                    )

            # Drop helper columns if present
            for col in ['route_id','avg_run_minutes','other_x','other_y']:
                if col in final_data.columns:
                    final_data.drop(columns=col, inplace=True)

            # (n) Save to Excel
            out_file = f"{sched_type}_{OUTPUT_EXCEL}"
            save_to_excel(final_data, OUTPUT_PATH, out_file)

        print("All schedule types processed successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
