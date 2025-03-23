"""
gtfs_trips_hourly_reporter.py

This module processes General Transit Feed Specification (GTFS) data to generate
reports of trips for selected routes based on a specified time interval and exports
the results to Excel workbooks.
"""

from datetime import time
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Input directory containing GTFS files
BASE_INPUT_PATH = r'\\your_file_path\here\\'

# Output directory for the Excel file
BASE_OUTPUT_PATH = r'\\your_file_path\here\\'

# GTFS files to load
gtfs_files = [
    'trips.txt',
    'stop_times.txt',
    'routes.txt',
    'stops.txt',
    'calendar.txt'
]

# Routes and directions to process
route_directions = [
    # Replace with your route name(s) and desired direction
    # Can be filtered to direction 0, 1, or None for no filter
    {'route_short_name': '310', 'direction_id': 0},   # Process only direction 0 for route 310
    {'route_short_name': '101', 'direction_id': None} # Process all directions for route 101
]

# **New Configuration: Time Interval in Minutes**
TIME_INTERVAL_MINUTES = 60  # Users can change this to 30, 15, etc.

# **New Configuration: Calendar Filter Days**
# Specify the days to filter services. For example, to include only weekdays:
CALENDAR_FILTER_DAYS = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
# To process all services separately, set it to an empty list:
# CALENDAR_FILTER_DAYS = []

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

def check_input_files(base_path, files):
    """Ensure all required GTFS files exist in the input directory."""
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"The input directory {base_path} does not exist.")
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        if not os.path.exists(file_path):
            raise FileNotFoundError(
                f"The required GTFS file {file_name} does not exist in {base_path}."
            )

def load_gtfs_data(base_path, files):
    """Load GTFS files into Pandas DataFrames."""
    data = {}
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        data_name = file_name.replace('.txt', '')
        data[data_name] = pd.read_csv(file_path)
    return data

def fix_time_format(time_str):
    """
    Fix time formats by:
    - Adding leading zeros to single-digit hours
    - Converting hours greater than 23 by subtracting 24
    """
    if pd.isna(time_str):
        return time_str
    parts = time_str.split(":")

    # Add leading zero if the hour is a single digit
    if len(parts[0]) == 1:
        parts[0] = '0' + parts[0]

    # Correct times where the hour exceeds 23 (indicating next day service)
    if int(parts[0]) >= 24:
        parts[0] = str(int(parts[0]) - 24).zfill(2)

    return ":".join(parts)

def get_time_bin(t, interval):
    """
    Assigns a time object to a specific time bin based on the interval.

    Args:
        t (datetime.time): The time to bin.
        interval (int): The interval in minutes.

    Returns:
        str: A string representing the time bin (e.g., "08:00-08:29").
    """
    total_minutes = t.hour * 60 + t.minute
    bin_start = (total_minutes // interval) * interval
    bin_end = bin_start + interval - 1
    if bin_end >= 1440:
        bin_end -= 1440  # Wrap around if necessary
    start_hour, start_min = divmod(bin_start, 60)
    end_hour, end_min = divmod(bin_end, 60)
    return f"{str(start_hour).zfill(2)}:{str(start_min).zfill(2)}-{str(end_hour).zfill(2)}:{str(end_min).zfill(2)}"

def process_and_export(data, route_dirs, output_path, interval_minutes, calendar_filter_days):
    """
    Process the GTFS data and export trips per specified interval to
    individual Excel files for each route.
    """
    trips = data['trips']
    stop_times = data['stop_times']
    routes = data['routes']
    calendar = data['calendar']

    # Apply Calendar Filter
    if calendar_filter_days:
        # Create a boolean mask for the specified days
        day_columns = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        mask = True
        for day in day_columns:
            if day in calendar_filter_days:
                mask &= calendar[day] == 1
            else:
                mask &= calendar[day] == 0
        relevant_service_ids = calendar[mask]['service_id']
        print("Relevant service IDs based on calendar filter:")
        print(relevant_service_ids)

        # Filter trips to include only those with relevant service IDs
        trips_filtered = trips[trips['service_id'].isin(relevant_service_ids)]
        print("Filtered trips based on calendar:")
        print(trips_filtered.head())
    else:
        # If no filter is applied, prepare to process all services separately
        trips_filtered = trips.copy()
        print("No calendar filter applied. Processing all services separately.")

    # Merge stop_times with trips and routes
    merged_data = pd.merge(stop_times, trips_filtered, on='trip_id')
    merged_data = pd.merge(
        merged_data,
        routes[['route_id', 'route_short_name']],
        on='route_id'
    )

    # Apply time format correction
    merged_data['arrival_time'] = merged_data['arrival_time'].apply(fix_time_format)
    merged_data['departure_time'] = merged_data['departure_time'].apply(fix_time_format)

    # Convert to datetime.time objects
    merged_data['arrival_time'] = pd.to_datetime(
        merged_data['arrival_time'].str.strip(), format='%H:%M:%S', errors='coerce'
    ).dt.time
    merged_data['departure_time'] = pd.to_datetime(
        merged_data['departure_time'].str.strip(), format='%H:%M:%S', errors='coerce'
    ).dt.time

    print("Merged data:")
    print(merged_data.head())

    # Generate all possible time bins
    time_bins = []
    for hour in range(24):
        for minute in range(0, 60, interval_minutes):
            bin_start = hour * 60 + minute
            bin_end = bin_start + interval_minutes - 1
            if bin_end >= 1440:
                bin_end -= 1440  # Wrap around if necessary
            start_hour, start_min = divmod(bin_start, 60)
            end_hour, end_min = divmod(bin_end, 60)
            bin_label = f"{str(start_hour).zfill(2)}:{str(start_min).zfill(2)}-{str(end_hour).zfill(2)}:{str(end_min).zfill(2)}"
            time_bins.append(bin_label)

    print("Generated time bins:")
    print(time_bins)

    # Process each route and direction
    for rd in route_dirs:
        route_short = rd['route_short_name']
        direction_id = rd['direction_id']

        if direction_id is not None:
            filtered = merged_data[
                (merged_data['route_short_name'] == route_short) &
                (merged_data['direction_id'] == direction_id)
            ]
        else:
            filtered = merged_data[merged_data['route_short_name'] == route_short]

        print(f"Filtered data for route {route_short} direction {direction_id}:")
        print(filtered.head())

        # Further filter to starting stops (stop_sequence == 1)
        start_times = filtered[filtered['stop_sequence'] == 1]
        print(f"Start times for route {route_short} direction {direction_id}:")
        print(start_times.head())

        # Drop rows with invalid departure_time
        start_times = start_times.dropna(subset=['departure_time'])

        if calendar_filter_days:
            # Assign each departure_time to a time bin
            start_times['time_bin'] = start_times['departure_time'].apply(
                lambda t: get_time_bin(t, interval_minutes)
            )

            print(f"Time bins for route {route_short} direction {direction_id}:")
            print(start_times[['departure_time', 'time_bin']].head())

            # Count trips per time bin
            trips_per_bin = start_times.groupby('time_bin').size().reset_index(
                name='trip_count'
            )

            # Ensure all time bins are included
            trips_per_bin = pd.DataFrame({'time_bin': time_bins}).merge(
                trips_per_bin, on='time_bin', how='left'
            ).fillna(0)
            trips_per_bin['trip_count'] = trips_per_bin['trip_count'].astype(int)

            print(f"Trips per time bin for route {route_short} direction {direction_id}:")
            print(trips_per_bin)
        else:
            # Process each service_id separately
            service_ids = filtered['service_id'].unique()
            for service_id in service_ids:
                service_filtered = filtered[filtered['service_id'] == service_id]
                service_start_times = service_filtered[service_filtered['stop_sequence'] == 1].dropna(subset=['departure_time'])

                # Assign each departure_time to a time bin
                service_start_times['time_bin'] = service_start_times['departure_time'].apply(
                    lambda t: get_time_bin(t, interval_minutes)
                )

                print(f"Time bins for service {service_id}, route {route_short}, direction {direction_id}:")
                print(service_start_times[['departure_time', 'time_bin']].head())

                # Count trips per time bin
                trips_per_bin = service_start_times.groupby('time_bin').size().reset_index(
                    name='trip_count'
                )

                # Ensure all time bins are included
                trips_per_bin = pd.DataFrame({'time_bin': time_bins}).merge(
                    trips_per_bin, on='time_bin', how='left'
                ).fillna(0)
                trips_per_bin['trip_count'] = trips_per_bin['trip_count'].astype(int)

                print(f"Trips per time bin for service {service_id}, route {route_short}, direction {direction_id}:")
                print(trips_per_bin)

                # Create a new Excel workbook for the current service, route, and direction
                wb = Workbook()
                ws = wb.active
                ws.title = f"Service_{service_id}_Route_{route_short}_Dir_{direction_id}" if direction_id is not None else f"Service_{service_id}_Route_{route_short}_All_Directions"

                # Write headers
                ws.append(trips_per_bin.columns.tolist())

                # Write data rows
                for row in trips_per_bin.itertuples(index=False, name=None):
                    ws.append(row)

                # Adjust column widths and alignments
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in col) + 2
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max_length
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center')

                # Ensure output directory exists
                os.makedirs(output_path, exist_ok=True)

                # Define the output filename
                file_name = (
                    f"Trips_Per_{interval_minutes}Min_Service_{service_id}_Route_{route_short}_Dir_{direction_id}.xlsx"
                    if direction_id is not None
                    else f"Trips_Per_{interval_minutes}Min_Service_{service_id}_Route_{route_short}_All_Directions.xlsx"
                )
                output_file = os.path.join(output_path, file_name)

                # Save the workbook
                wb.save(output_file)
                print(f"Trips per {interval_minutes} minutes for {file_name} successfully exported!")

            # Continue to next route/direction
            continue

        # Create a new Excel workbook for the current route and direction
        wb = Workbook()
        ws = wb.active
        ws.title = (
            f"Route_{route_short}_Dir_{direction_id}"
            if direction_id is not None
            else f"Route_{route_short}_All_Directions"
        )

        # Write headers
        ws.append(trips_per_bin.columns.tolist())

        # Write data rows
        for row in trips_per_bin.itertuples(index=False, name=None):
            ws.append(row)

        # Adjust column widths and alignments
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length
            for cell in col:
                cell.alignment = Alignment(horizontal='center')

        # Ensure output directory exists
        os.makedirs(output_path, exist_ok=True)

        # Define the output filename
        file_name = (
            f"Trips_Per_{interval_minutes}Min_Route_{route_short}_Dir_{direction_id}.xlsx"
            if direction_id is not None
            else f"Trips_Per_{interval_minutes}Min_Route_{route_short}_All_Directions.xlsx"
        )
        output_file = os.path.join(output_path, file_name)

        # Save the workbook
        wb.save(output_file)
        print(f"Trips per {interval_minutes} minutes for {file_name} successfully exported!")

def main():
    """Main function to execute the script."""
    try:
        # Validate TIME_INTERVAL_MINUTES
        if not isinstance(TIME_INTERVAL_MINUTES, int) or TIME_INTERVAL_MINUTES <= 0:
            raise ValueError("TIME_INTERVAL_MINUTES must be a positive integer.")
        if 1440 % TIME_INTERVAL_MINUTES != 0:
            print(f"Warning: {TIME_INTERVAL_MINUTES} does not evenly divide into 1440 minutes. Some time bins may overlap or miss.")

        # Check if all input files exist
        check_input_files(BASE_INPUT_PATH, gtfs_files)

        # Load GTFS data
        data = load_gtfs_data(BASE_INPUT_PATH, gtfs_files)

        # Process data and export to Excel
        process_and_export(
            data,
            route_directions,
            BASE_OUTPUT_PATH,
            TIME_INTERVAL_MINUTES,
            CALENDAR_FILTER_DAYS
        )

    except FileNotFoundError as fnf_error:
        print(f"File not found error: {fnf_error}")
    except pd.errors.ParserError as parse_error:
        print(f"Parsing error while reading GTFS files: {parse_error}")
    except PermissionError as perm_error:
        print(f"Permission error: {perm_error}")
    except ValueError as ve:
        print(f"Value error: {ve}")

if __name__ == "__main__":
    main()
